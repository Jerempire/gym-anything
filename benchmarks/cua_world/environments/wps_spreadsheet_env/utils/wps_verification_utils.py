#!/usr/bin/env python3
"""Verification utilities for WPS Spreadsheet tasks."""

import os
import sys
import json
import logging
import tempfile
import shutil
from pathlib import Path

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def copy_and_parse_spreadsheet(container_path, copy_from_env, file_format='xlsx'):
    """
    Copy a spreadsheet file from the container and parse it.

    Args:
        container_path: Path to the file in the container (e.g., /home/ga/Documents/data.xlsx)
        copy_from_env: Function to copy files from container to host
        file_format: Format of the file ('xlsx', 'xls', 'csv', 'ods')

    Returns:
        tuple: (success, workbook/data, error_message, temp_dir)
    """
    temp_dir = None

    try:
        # Create temp directory for file extraction
        temp_dir = tempfile.mkdtemp(prefix='wps_verify_')

        # Copy the file from container
        temp_file = os.path.join(temp_dir, f'input.{file_format}')

        try:
            copy_from_env(container_path, temp_file)
        except Exception as e:
            # Try alternative extensions
            for alt_ext in ['xlsx', 'xls', 'csv', 'ods', 'et']:
                if alt_ext == file_format:
                    continue
                try:
                    alt_path = container_path.rsplit('.', 1)[0] + '.' + alt_ext
                    copy_from_env(alt_path, temp_file)
                    file_format = alt_ext
                    logger.info(f"Successfully copied with alternative extension: {alt_ext}")
                    break
                except:
                    continue
            else:
                return False, None, f"Failed to copy file: {e}", temp_dir

        # Parse based on file format
        if file_format == 'xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(temp_file, data_only=False)
                return True, wb, None, temp_dir
            except Exception as e:
                return False, None, f"Failed to parse XLSX: {e}", temp_dir

        elif file_format == 'xls':
            try:
                import xlrd
                # For .xls files, we need to read them differently
                # Return the path and let caller handle it
                return True, temp_file, None, temp_dir
            except Exception as e:
                return False, None, f"Failed to parse XLS: {e}", temp_dir

        elif file_format == 'csv':
            try:
                import pandas as pd
                df = pd.read_csv(temp_file)
                return True, df, None, temp_dir
            except Exception as e:
                return False, None, f"Failed to parse CSV: {e}", temp_dir

        elif file_format == 'ods':
            try:
                import pandas as pd
                df = pd.read_excel(temp_file, engine='odf')
                return True, df, None, temp_dir
            except Exception as e:
                return False, None, f"Failed to parse ODS: {e}", temp_dir

        else:
            return False, None, f"Unsupported format: {file_format}", temp_dir

    except Exception as e:
        logger.error(f"Error in copy_and_parse_spreadsheet: {e}", exc_info=True)
        return False, None, str(e), temp_dir


def get_spreadsheet_text(wb):
    """
    Extract all text from a spreadsheet workbook.

    Args:
        wb: openpyxl Workbook object

    Returns:
        str: All text from the spreadsheet
    """
    text_parts = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")

        for row in sheet.iter_rows():
            row_text = []
            for cell in row:
                if cell.value is not None:
                    row_text.append(str(cell.value))
            if row_text:
                text_parts.append(' | '.join(row_text))

    return '\n'.join(text_parts)


def get_cell_value(wb, sheet_name, row, col):
    """
    Get value from a specific cell.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)

    Returns:
        Cell value or None
    """
    try:
        sheet = wb[sheet_name]
        cell = sheet.cell(row=row, column=col)
        return cell.value
    except:
        return None


def get_cell_formula(wb, sheet_name, row, col):
    """
    Get formula from a specific cell.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)

    Returns:
        Formula string or None
    """
    try:
        sheet = wb[sheet_name]
        cell = sheet.cell(row=row, column=col)
        return cell.value if cell.data_type == 'f' else None
    except:
        return None


def check_cell_formatting(wb, sheet_name, row, col, **format_attrs):
    """
    Check formatting attributes of a cell.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        **format_attrs: Attributes to check (bold, italic, font_size, etc.)

    Returns:
        bool: True if all specified attributes match
    """
    try:
        sheet = wb[sheet_name]
        cell = sheet.cell(row=row, column=col)

        font = cell.font

        for attr, expected in format_attrs.items():
            if attr == 'bold':
                if font.bold != expected:
                    return False
            elif attr == 'italic':
                if font.italic != expected:
                    return False
            elif attr == 'font_size':
                if font.size != expected:
                    return False
            elif attr == 'font_color':
                if font.color and font.color.rgb:
                    # Compare RGB values
                    actual = str(font.color.rgb)
                    if not actual.endswith(expected):
                        return False
                else:
                    return False

        return True

    except Exception as e:
        logger.error(f"Error checking cell formatting: {e}")
        return False


def check_column_width(sheet, col_letter, min_width=None, max_width=None):
    """
    Check column width.

    Args:
        sheet: openpyxl Worksheet object
        col_letter: Column letter (e.g., 'A', 'B')
        min_width: Minimum expected width
        max_width: Maximum expected width

    Returns:
        bool: True if width is within range
    """
    try:
        from openpyxl.utils import column_index_from_string

        col_idx = column_index_from_string(col_letter)
        width = sheet.column_dimensions[col_letter].width

        if min_width and width < min_width:
            return False
        if max_width and width > max_width:
            return False

        return True

    except:
        return False


def check_row_height(sheet, row_num, min_height=None, max_height=None):
    """
    Check row height.

    Args:
        sheet: openpyxl Worksheet object
        row_num: Row number
        min_height: Minimum expected height
        max_height: Maximum expected height

    Returns:
        bool: True if height is within range
    """
    try:
        height = sheet.row_dimensions[row_num].height

        if min_height and height < min_height:
            return False
        if max_height and height > max_height:
            return False

        return True

    except:
        return False


def check_cell_background(wb, sheet_name, row, col, expected_color=None):
    """
    Check cell background color.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        expected_color: Expected color RGB (e.g., 'FFFF00' for yellow)

    Returns:
        bool: True if color matches
    """
    try:
        sheet = wb[sheet_name]
        cell = sheet.cell(row=row, column=col)

        if cell.fill and cell.fill.fgColor:
            actual = str(cell.fill.fgColor.rgb)
            if expected_color:
                return actual.endswith(expected_color)
            else:
                # Just check if there's a fill color
                return actual != '00000000' and actual != 'FFFFFFFF'
        return False

    except:
        return False


def check_conditional_formatting(sheet, range_start, range_end, rule_type=None):
    """
    Check if conditional formatting exists in a range.

    Args:
        sheet: openpyxl Worksheet object
        range_start: Starting cell (e.g., 'A1')
        range_end: Ending cell (e.g., 'D10')
        rule_type: Type of rule to check (e.g., 'cellIs', 'expression')

    Returns:
        bool: True if conditional formatting exists
    """
    try:
        if not hasattr(sheet, 'conditional_formatting'):
            return False

        cf = sheet.conditional_formatting

        # Check if there's any conditional formatting in the range
        for range_string, rules in cf._cf_rules.items():
            if range_string and range_start in range_string:
                if rule_type:
                    # Check for specific rule type
                    for rule in rules:
                        if hasattr(rule, 'type') and rule.type == rule_type:
                            return True
                    return False
                return True

        return False

    except:
        return False


def get_chart_data(wb, sheet_name, chart_title=None):
    """
    Extract data from charts in a sheet.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        chart_title: Optional title to filter by

    Returns:
        list: List of chart information dictionaries
    """
    charts_info = []

    try:
        sheet = wb[sheet_name]

        if hasattr(sheet, '_charts'):
            for chart in sheet._charts:
                info = {
                    'title': chart.title if hasattr(chart, 'title') else None,
                    'type': type(chart).__name__,
                    'width': chart.width if hasattr(chart, 'width') else None,
                    'height': chart.height if hasattr(chart, 'height') else None,
                }

                if chart_title is None or (info['title'] and chart_title in info['title']):
                    charts_info.append(info)

        return charts_info

    except Exception as e:
        logger.error(f"Error extracting chart data: {e}")
        return []


def check_pivot_table(wb, sheet_name, expected_fields=None):
    """
    Check if pivot table exists and has expected fields.

    Args:
        wb: openpyxl Workbook object
        sheet_name: Name of the sheet
        expected_fields: List of expected field names

    Returns:
        dict: Information about pivot table
    """
    # Note: openpyxl has limited pivot table support
    # This is a basic check
    result = {
        'exists': False,
        'sheet_name': sheet_name,
    }

    try:
        # Check if sheet name suggests pivot table
        if 'pivot' in sheet_name.lower():
            result['exists'] = True
            result['note'] = 'Sheet name suggests pivot table (openpyxl limited support)'

        return result

    except:
        return result


def check_data_validation(sheet, cell_address, validation_type=None):
    """
    Check if data validation exists for a cell.

    Args:
        sheet: openpyxl Worksheet object
        cell_address: Cell address (e.g., 'A1')
        validation_type: Type of validation ('whole', 'decimal', 'list', 'date', 'time', 'textLength', 'custom')

    Returns:
        bool: True if validation exists
    """
    try:
        if not hasattr(sheet, 'data_validations'):
            return False

        dv = sheet.data_validations

        for validation in dv.dataValidation:
            if cell_address in validation.sqref:
                if validation_type is None:
                    return True
                if validation.type == validation_type:
                    return True

        return False

    except:
        return False


def cleanup_verification_temp(temp_dir):
    """
    Clean up temporary directory used for verification.

    Args:
        temp_dir: Path to temporary directory
    """
    if temp_dir and os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
        except Exception as e:
            logger.warning(f"Failed to cleanup temp dir {temp_dir}: {e}")


def vlm_verify_screenshot(env_info, traj, prompt_template):
    """
    Query the VLM with the final screenshot and a structured prompt.
    Returns parsed JSON from VLM response, or None if VLM is unavailable.

    Args:
        env_info: Environment info dict (contains query_vlm, get_final_screenshot).
        traj: Trajectory dict.
        prompt_template: Structured prompt asking VLM to evaluate the screenshot.

    Returns:
        Parsed dict from VLM response, or None.
    """
    query_vlm = env_info.get('query_vlm')
    get_final_screenshot = env_info.get('get_final_screenshot')

    if not query_vlm or not get_final_screenshot:
        logger.warning("VLM verification unavailable: missing query_vlm or get_final_screenshot")
        return None

    final_frame = get_final_screenshot(traj)
    if not final_frame:
        logger.warning("VLM verification unavailable: no final frame")
        return None

    try:
        result = query_vlm(prompt=prompt_template, image=final_frame)
        if result and result.get("success"):
            return result.get("parsed", {})
        logger.warning(f"VLM query failed: {result.get('error', 'unknown') if result else 'no result'}")
    except Exception as e:
        logger.warning(f"VLM query exception: {e}")


def check_sort_range(sheet, expected_start_row=None, expected_columns=None):
    """
    Check if data in a range appears to be sorted.

    Args:
        sheet: openpyxl Worksheet object
        expected_start_row: Expected row where data starts
        expected_columns: List of column letters that should be sorted

    Returns:
        dict: Sort status information
    """
    result = {
        'is_sorted': False,
        'sorted_columns': [],
        'unsorted_columns': [],
    }

    try:
        if expected_columns is None:
            # Try to detect columns with data
            expected_columns = []

            # Get first row to detect columns
            first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            if first_row:
                for i, val in enumerate(first_row[0]):
                    if val is not None:
                        expected_columns.append(chr(65 + i))  # A, B, C, ...

        for col_letter in expected_columns:
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter)

            # Get values in this column
            values = []
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                if row[0] is not None:
                    values.append(row[0])

            if len(values) > 1:
                # Check if sorted (ascending)
                is_sorted = all(values[i] <= values[i+1] for i in range(len(values)-1))

                if is_sorted:
                    result['sorted_columns'].append(col_letter)
                else:
                    result['unsorted_columns'].append(col_letter)

        result['is_sorted'] = len(result['unsorted_columns']) == 0

        return result

    except Exception as e:
        logger.error(f"Error checking sort: {e}")
        return result


def check_filter_exists(sheet):
    """
    Check if autoFilter is applied to a sheet.

    Args:
        sheet: openpyxl Worksheet object

    Returns:
        bool: True if filter exists
    """
    try:
        return sheet.auto_filter is not None and sheet.auto_filter.ref is not None
    except:
        return False


def get_filter_criteria(sheet):
    """
    Get filter criteria applied to a sheet.

    Args:
        sheet: openpyxl Worksheet object

    Returns:
        dict: Filter information
    """
    result = {
        'has_filter': False,
        'filter_range': None,
        'filtered_columns': [],
    }

    try:
        if sheet.auto_filter and sheet.auto_filter.ref:
            result['has_filter'] = True
            result['filter_range'] = sheet.auto_filter.ref

            # Check for column filter details
            if hasattr(sheet.auto_filter, 'filterColumn'):
                for fc in sheet.auto_filter.filterColumn:
                    result['filtered_columns'].append(fc.colId)

        return result

    except Exception as e:
        logger.error(f"Error getting filter criteria: {e}")
        return result

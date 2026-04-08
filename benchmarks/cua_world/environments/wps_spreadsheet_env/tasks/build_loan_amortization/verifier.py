#!/usr/bin/env python3
"""Verifier for build_loan_amortization task."""

import os
import sys
import json
import logging
import tempfile
import shutil

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_loan_amortization(traj, env_info, task_info):
    """
    Verify the loan amortization task using a hybrid approach.
    Uses openpyxl for structural/formula checking and VLM for chart & visual summary verification.
    """
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_pmt = metadata.get('expected_pmt', 9662.65)
    expected_month1_int = metadata.get('expected_month1_interest', 2930.21)

    score = 0
    feedback_parts = []
    temp_dir = tempfile.mkdtemp(prefix='loan_verify_')
    
    try:
        # 1. Retrieve JSON metadata
        result_json_path = os.path.join(temp_dir, 'task_result.json')
        copy_from_env('/tmp/task_result.json', result_json_path)
        with open(result_json_path, 'r') as f:
            result = json.load(f)

        if not result.get('file_exists'):
            return {"passed": False, "score": 0, "feedback": "Target spreadsheet file missing."}

        file_modified = (result.get('current_hash') != result.get('initial_hash')) and \
                        (result.get('file_mtime', 0) >= result.get('task_start_time', 0))
        
        if file_modified:
            score += 10
            feedback_parts.append("✅ File modified successfully")
        else:
            feedback_parts.append("❌ File was NOT modified (hash matches initial state)")
            return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

        # 2. Extract and parse the spreadsheet
        xlsx_path = os.path.join(temp_dir, 'loan.xlsx')
        copy_from_env('/home/ga/Documents/loan_amortization.xlsx', xlsx_path)
        
        from openpyxl import load_workbook
        # Load twice: one for formulas (data_only=False), one for values (data_only=True)
        wb_formulas = load_workbook(xlsx_path, data_only=False)
        wb_values = load_workbook(xlsx_path, data_only=True)
        ws_f = wb_formulas.active
        ws_v = wb_values.active

        # Check Table Structure and Formulas
        data_rows = 0
        formula_count = 0
        balances = []
        pmt_values = []
        int_values = []
        
        # Analyze rows 20 to 80 (60 months)
        for r in range(20, 81):
            if ws_f.cell(row=r, column=1).value is not None:  # Month number
                data_rows += 1
            
            # Count formula usage across calculation columns
            for c in range(3, 9):
                cell_f = ws_f.cell(row=r, column=c)
                if isinstance(cell_f.value, str) and str(cell_f.value).startswith('='):
                    formula_count += 1

            # Extract computed cached values for accuracy checking
            val_pmt = ws_v.cell(row=r, column=4).value
            val_int = ws_v.cell(row=r, column=5).value
            val_bal = ws_v.cell(row=r, column=7).value
            
            if isinstance(val_pmt, (int, float)): pmt_values.append(val_pmt)
            if isinstance(val_int, (int, float)): int_values.append(val_int)
            if isinstance(val_bal, (int, float)): balances.append(val_bal)

        # Evaluate Data Population (15 pts)
        if data_rows >= 58:
            score += 15
            feedback_parts.append(f"✅ {data_rows} monthly rows populated")
        elif data_rows > 10:
            score += 5
            feedback_parts.append(f"⚠️ Partially populated ({data_rows} rows)")
        else:
            feedback_parts.append(f"❌ Table mostly empty ({data_rows} rows)")

        # Evaluate Formula Usage (20 pts)
        if formula_count >= 150:
            score += 20
            feedback_parts.append("✅ Extensive use of financial formulas")
        elif formula_count >= 50:
            score += 10
            feedback_parts.append("⚠️ Partial formula usage (hardcoded detected?)")
        else:
            feedback_parts.append("❌ Very few/no formulas used (data hardcoded)")

        # Evaluate Financial Accuracy (15 pts)
        if pmt_values and int_values:
            pmt_match = any(abs(abs(p) - expected_pmt) < 5.0 for p in pmt_values[:5])
            int_match = any(abs(abs(i) - expected_month1_int) < 5.0 for i in int_values[:5])
            
            if pmt_match and int_match:
                score += 15
                feedback_parts.append("✅ Accurate PMT and Interest calculations")
            elif pmt_match or int_match:
                score += 7
                feedback_parts.append("⚠️ Computations partially accurate")
            else:
                feedback_parts.append("❌ Computations do not match expected SBA metrics")
        else:
            feedback_parts.append("⚠️ Could not extract cached values for computation check")

        # Evaluate Balance Monotonicity (15 pts)
        if len(balances) >= 55:
            monotonic = all(balances[i] > balances[i+1] for i in range(len(balances)-1))
            hits_zero = abs(balances[-1]) < 5.0
            
            if monotonic and hits_zero:
                score += 15
                feedback_parts.append("✅ Balance decreases correctly to 0")
            elif monotonic:
                score += 10
                feedback_parts.append("⚠️ Balance decreases but does not hit 0 properly")
            else:
                feedback_parts.append("❌ Balance does not decrease monotonically")

        # 3. VLM Verification for Visual Chart and Summary Area (25 pts)
        if query_vlm:
            from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
            
            frames = sample_trajectory_frames(traj, n=3)
            final_img = get_final_screenshot(traj)
            images = frames + [final_img] if final_img else frames

            prompt = """
            Analyze these screenshots of an agent using WPS Spreadsheet to build an amortization schedule.
            Respond in JSON format:
            {
                "has_summary_totals": true/false,
                "has_chart_visible": true/false
            }
            1. Are there summary totals for the loan (Total Interest, Total Principal, Total Payments) visible at the bottom of the data?
            2. Is there a chart (line chart, bar chart, area chart) visible that visualizes the loan balance over time?
            """
            
            vlm_res = query_vlm(images=images, prompt=prompt)
            if vlm_res and vlm_res.get("success"):
                parsed = vlm_res.get("parsed", {})
                
                if parsed.get("has_summary_totals"):
                    score += 10
                    feedback_parts.append("✅ Summary totals visually confirmed")
                else:
                    feedback_parts.append("❌ Summary totals missing in screenshots")
                
                if parsed.get("has_chart_visible"):
                    score += 15
                    feedback_parts.append("✅ Balance chart visually confirmed")
                else:
                    feedback_parts.append("❌ Chart missing in screenshots")
            else:
                feedback_parts.append("⚠️ VLM analysis failed, skipping visual checks")
        else:
            feedback_parts.append("⚠️ query_vlm unavailable")

        # Final Score check
        passed = score >= 60 and formula_count > 30

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification script exception: {e}"}
        
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
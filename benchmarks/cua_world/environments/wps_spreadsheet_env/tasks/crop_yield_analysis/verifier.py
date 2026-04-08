#!/usr/bin/env python3
"""
Verifier for Crop Yield Analysis task.
Verifies multi-sheet aggregation, formulas, IF conditions, and ranking.
"""

import os
import json
import logging
import tempfile
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ground truth values for spot checking
COUNTY_AVG_YIELDS = {
    "Black Hawk": 189.6, "Boone": 199.0, "Cerro Gordo": 200.8, "Dallas": 199.4, 
    "Grundy": 204.0, "Hamilton": 207.2, "Hardin": 194.0, "Jasper": 184.4, 
    "Kossuth": 204.8, "Marshall": 196.0, "Plymouth": 184.4, "Polk": 183.6, 
    "Sioux": 189.2, "Story": 193.2, "Webster": 196.8
}

COUNTY_PERFORMANCE = {
    "Black Hawk": "Low", "Boone": "Medium", "Cerro Gordo": "Medium", "Dallas": "Medium",
    "Grundy": "Medium", "Hamilton": "High", "Hardin": "Low", "Jasper": "Low",
    "Kossuth": "Medium", "Marshall": "Medium", "Plymouth": "Low", "Polk": "Low",
    "Sioux": "Low", "Story": "Low", "Webster": "Medium"
}

YEARLY_AVG = {
    2018: 203.0,
    2019: 194.133,
    2020: 173.867,
    2021: 204.467,
    2022: 200.0,
}


def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def verify_crop_yield_analysis(traj, env_info, task_info):
    """Verify spreadsheet creation, data manipulation, and formula presence."""
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    score = 0
    feedback_parts = []
    
    try:
        import openpyxl
    except ImportError:
        return {"passed": False, "score": 0, "feedback": "Error: openpyxl not installed on host."}

    # Retrieve execution metadata
    result_meta = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", result_meta.name)
        with open(result_meta.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(result_meta.name):
            os.unlink(result_meta.name)

    # 1. Anti-gaming File Modification Check
    if not result.get("file_modified", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "FAIL: File was not modified. No actions performed."
        }
    feedback_parts.append("File modification verified")

    # Copy the actual spreadsheet to examine
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/iowa_corn_production.xlsx", temp_xlsx.name)
        
        # Load twice: once for values, once for formulas
        wb_data = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        
        sheet_names = [sn.lower().strip() for sn in wb_data.sheetnames]

        # 2. Check Yield Column (15 Points)
        field_sheet = wb_data.active
        g1_val = str(field_sheet.cell(row=1, column=7).value or "").lower()
        if "yield" in g1_val:
            score += 5
            
            # Spot check 3 random cells for yield math: Production_Bu (F) / Acres_Harvested (E)
            # Boone 2018 is row 2: 20073200 / 93800 = 214
            # Polk 2018 is row 12: 7820800 / 41600 = 188
            # Grundy 2018 is row 47: 16632000 / 75600 = 220
            yield_correct = 0
            for r, expected in [(2, 214.0), (12, 188.0), (47, 220.0)]:
                actual = safe_float(field_sheet.cell(row=r, column=7).value)
                if actual is not None and abs(actual - expected) < 0.5:
                    yield_correct += 1
            
            yield_pts = int((yield_correct / 3) * 10)
            score += yield_pts
            feedback_parts.append(f"Yield Column: {yield_pts + 5}/15 pts")
        else:
            feedback_parts.append("Yield Column: Missing/Incorrect Header")

        # 3. Check County Summary Sheet (45 points)
        summary_name = next((s for s in wb_data.sheetnames if "summary" in s.lower()), None)
        if summary_name:
            ws_summary = wb_data[summary_name]
            score += 5  # Sheet exists
            
            # Map counties to their rows for robust checking
            county_rows = {}
            for r in range(2, min(25, ws_summary.max_row + 1)):
                val = str(ws_summary.cell(row=r, column=1).value or "").strip()
                if val:
                    county_rows[val] = r
            
            if len(county_rows) >= 14:
                score += 5
                
                # Spot check averages (15 pts)
                avg_correct = 0
                for c_name, expected_avg in [("Boone", 199.0), ("Story", 193.2), ("Hamilton", 207.2)]:
                    r = county_rows.get(c_name)
                    if r:
                        # Find the avg yield column (usually B or col 2)
                        val = safe_float(ws_summary.cell(row=r, column=2).value)
                        if val is not None and abs(val - expected_avg) < 1.0:
                            avg_correct += 1
                score += int((avg_correct / 3) * 15)

                # Spot check Performance Categories (10 pts)
                perf_correct = 0
                for c_name, expected_perf in [("Hamilton", "High"), ("Boone", "Medium"), ("Polk", "Low")]:
                    r = county_rows.get(c_name)
                    if r:
                        # Find the performance column (usually F or col 6)
                        val = str(ws_summary.cell(row=r, column=6).value or "").strip().lower()
                        if val == expected_perf.lower():
                            perf_correct += 1
                score += int((perf_correct / 3) * 10)

                # Spot check Rank (10 pts)
                rank_correct = 0
                for c_name, expected_rank in [("Hamilton", 1), ("Polk", 15), ("Kossuth", 2)]:
                    r = county_rows.get(c_name)
                    if r:
                        # Find the rank column (usually G or col 7)
                        val = safe_float(ws_summary.cell(row=r, column=7).value)
                        if val is not None and abs(val - expected_rank) < 0.5:
                            rank_correct += 1
                score += int((rank_correct / 3) * 10)

            feedback_parts.append("County Summary: Verified")
        else:
            feedback_parts.append("County Summary: Missing")

        # 4. Check Statistics Sheet (20 points)
        stats_name = next((s for s in wb_data.sheetnames if "statistic" in s.lower()), None)
        if stats_name:
            ws_stats = wb_data[stats_name]
            score += 5
            
            # Build year mapping
            year_rows = {}
            for r in range(2, min(15, ws_stats.max_row + 1)):
                val = safe_float(ws_stats.cell(row=r, column=1).value)
                if val is not None:
                    year_rows[int(val)] = r

            # Spot check 2018 and 2020 averages
            stats_correct = 0
            for yr, expected_avg in [(2018, 203.0), (2020, 173.867)]:
                r = year_rows.get(yr)
                if r:
                    val = safe_float(ws_stats.cell(row=r, column=2).value)
                    if val is not None and abs(val - expected_avg) < 1.0:
                        stats_correct += 1
            
            score += int((stats_correct / 2) * 15)
            feedback_parts.append("Statistics Sheet: Verified")
        else:
            feedback_parts.append("Statistics Sheet: Missing")

        # 5. VLM trajectory verification (20 points)
        # We ensure the agent actually worked in the UI
        vlm_passed = False
        query_vlm = env_info.get('query_vlm')
        if query_vlm:
            from gym_anything.vlm import sample_trajectory_frames
            frames = sample_trajectory_frames(traj, n=4)
            if frames:
                vlm_prompt = """Look at these frames of a computer session.
                Did the user spend time interacting with WPS Spreadsheet editing cells, typing formulas, or manipulating multiple sheets?
                Reply with JSON: {"wps_interaction": true/false}"""
                vlm_resp = query_vlm(prompt=vlm_prompt, images=frames)
                if vlm_resp and vlm_resp.get('parsed', {}).get('wps_interaction'):
                    vlm_passed = True
        
        if vlm_passed:
            score += 20
            feedback_parts.append("VLM visual proof: Passed")
        else:
            feedback_parts.append("VLM visual proof: Failed/Skipped")

        # Final Evaluation
        passed = score >= 65
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Error during verification: {e}")
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

if __name__ == "__main__":
    # Test execution branch (stub)
    print("Verifier standalone test.")
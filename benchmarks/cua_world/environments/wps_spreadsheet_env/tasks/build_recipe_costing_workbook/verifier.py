#!/usr/bin/env python3
"""
Verifier for build_recipe_costing_workbook task.

VERIFICATION STRATEGY:
Checks for the presence and validity of the formulas required for the cost buildup.
Uses partial string matching on formulas to ensure users actually applied VLOOKUP/SUM
instead of typing out numbers directly. Verifies the new Menu_Pricing sheet creation
along with its formulas and visualization.
"""

import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def vlm_check_chart(env_info, traj):
    """Fallback check for the chart via VLM on trajectory."""
    try:
        from gym_anything.vlm import get_final_screenshot, sample_trajectory_frames
        query_vlm = env_info.get('query_vlm')
        if not query_vlm:
            return False
            
        frames = sample_trajectory_frames(traj, n=3)
        final = get_final_screenshot(traj)
        images = frames + ([final] if final else [])
        if not images:
            return False
            
        res = query_vlm(
            images=images,
            prompt="""Analyze these screenshots of a spreadsheet. 
Respond in JSON format: {"has_bar_chart": true/false}
Does any screenshot show a bar chart (likely displaying Gross Profit or Menu Pricing)?"""
        )
        if res and res.get("parsed"):
            return res["parsed"].get("has_bar_chart", False)
    except Exception as e:
        logger.warning(f"VLM check failed: {e}")
    return False

def verify_recipe_costing(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Ensure openpyxl is available
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
        import openpyxl

    # Read exported metadata
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "Workbook not found. Agent did not save the file."}

    # Copy workbook from environment
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/restaurant_recipe_costing.xlsx", temp_wb.name)
        wb_formulas = openpyxl.load_workbook(temp_wb.name, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse workbook: {e}"}
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)

    score = 0
    feedback_parts = []
    
    # Anti-gaming file check
    if result.get("file_created_during_task", False):
        score += 5
        feedback_parts.append("File saved (+5)")
    else:
        feedback_parts.append("File not modified during task")

    sheets = wb_formulas.sheetnames
    if "Recipes" not in sheets:
        return {"passed": False, "score": score, "feedback": "Recipes sheet missing"}

    ws_r = wb_formulas["Recipes"]
    
    lookup_count = 0
    extended_count = 0
    sum_count = 0
    cps_count = 0

    # Analyze Recipes sheet formulas
    for row in range(1, ws_r.max_row + 1):
        cell_a = ws_r.cell(row=row, column=1).value
        cell_c = ws_r.cell(row=row, column=3).value
        
        # Determine if it's an ingredient row
        if cell_a and isinstance(cell_a, str) and cell_c and isinstance(cell_c, str):
            exclude_list = ["Ingredient", "Classic Beef Burger", "Caesar Salad", "Spaghetti Bolognese", 
                            "Grilled Chicken Sandwich", "New England Clam Chowder", "Garden Vegetable Stir-Fry", 
                            "Total Cost:", "Cost Per Serving:"]
            if cell_a not in exclude_list:
                u_cost = ws_r.cell(row=row, column=4).value
                e_cost = ws_r.cell(row=row, column=5).value
                
                if u_cost and isinstance(u_cost, str) and u_cost.startswith("="):
                    up = u_cost.upper()
                    if "VLOOKUP" in up or "INDEX" in up or "MATCH" in up or "INGREDIENT" in up:
                        lookup_count += 1
                        
                if e_cost and isinstance(e_cost, str) and e_cost.startswith("="):
                    if "*" in e_cost or "PRODUCT" in e_cost.upper():
                        extended_count += 1
                        
        # Check Total Cost row
        if cell_a == "Total Cost:" or ws_r.cell(row=row, column=4).value == "Total Cost:":
            tot_cell = ws_r.cell(row=row, column=5).value
            if tot_cell and isinstance(tot_cell, str) and tot_cell.startswith("="):
                if "SUM" in tot_cell.upper() or "+" in tot_cell:
                    sum_count += 1
                    
        # Check Cost Per Serving row
        if cell_a == "Cost Per Serving:" or ws_r.cell(row=row, column=4).value == "Cost Per Serving:":
            cps_cell = ws_r.cell(row=row, column=5).value
            if cps_cell and isinstance(cps_cell, str) and cps_cell.startswith("="):
                if "/" in cps_cell:
                    cps_count += 1

    # Score Recipes Sheet
    if lookup_count >= 35:
        score += 15
        feedback_parts.append("Lookups valid (+15)")
    else:
        feedback_parts.append(f"Lookups incomplete ({lookup_count}/44)")
        
    if extended_count >= 35:
        score += 15
        feedback_parts.append("Extended costs valid (+15)")
    else:
        feedback_parts.append(f"Extended costs incomplete ({extended_count}/44)")
        
    if sum_count >= 5:
        score += 15
        feedback_parts.append("Totals valid (+15)")
    else:
        feedback_parts.append(f"Totals incomplete ({sum_count}/6)")
        
    if cps_count >= 5:
        score += 10
        feedback_parts.append("Cost per serving valid (+10)")
    else:
        feedback_parts.append(f"CPS incomplete ({cps_count}/6)")

    # Score Menu_Pricing Sheet
    menu_pricing_exists = False
    ws_m = None
    for sheet in sheets:
        if "menu" in sheet.lower() and "pricing" in sheet.lower():
            menu_pricing_exists = True
            ws_m = wb_formulas[sheet]
            break
            
    if menu_pricing_exists:
        score += 10
        feedback_parts.append("Menu_Pricing sheet found (+10)")
        
        sugg_count = 0
        gp_count = 0
        
        for row in range(2, min(15, ws_m.max_row + 1)):
            sugg_cell = ws_m.cell(row=row, column=4).value
            gp_cell = ws_m.cell(row=row, column=5).value
            
            if sugg_cell is not None:
                if isinstance(sugg_cell, str) and sugg_cell.startswith("="):
                    if "/" in sugg_cell or "*" in sugg_cell:
                        sugg_count += 1
                elif isinstance(sugg_cell, (int, float)):
                    sugg_count += 1
                    
            if gp_cell is not None:
                if isinstance(gp_cell, str) and gp_cell.startswith("="):
                    if "-" in gp_cell:
                        gp_count += 1
                elif isinstance(gp_cell, (int, float)):
                    gp_count += 1
                    
        if sugg_count >= 5:
            score += 10
            feedback_parts.append("Suggested price calculated (+10)")
        else:
            feedback_parts.append(f"Suggested price incomplete ({sugg_count}/6)")
            
        if gp_count >= 5:
            score += 10
            feedback_parts.append("Gross profit calculated (+10)")
        else:
            feedback_parts.append(f"Gross profit incomplete ({gp_count}/6)")
            
        # Chart Verification
        has_chart = False
        if hasattr(ws_m, '_charts') and len(ws_m._charts) > 0:
            has_chart = True
        elif len(wb_formulas.chartsheets) > 0:
            has_chart = True
            
        if has_chart:
            score += 10
            feedback_parts.append("Chart detected via API (+10)")
        else:
            if vlm_check_chart(env_info, traj):
                score += 10
                feedback_parts.append("Chart detected via VLM (+10)")
            else:
                feedback_parts.append("Chart not found")
    else:
        feedback_parts.append("Menu_Pricing sheet missing")

    key_criteria_met = lookup_count >= 20 and menu_pricing_exists
    passed = score >= 60 and key_criteria_met
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
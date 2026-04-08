#!/usr/bin/env python3
"""
Verifier for build_supplier_scorecard task.
"""

import os
import json
import logging
import tempfile

# Framework imports
from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

# wps_verification_utils is expected to be available in the environment's python path via sys.path injection
# but we will implement safe robust fallback parsing just in case.
try:
    from utils.wps_verification_utils import copy_and_parse_spreadsheet
    HAS_WPS_UTILS = True
except ImportError:
    HAS_WPS_UTILS = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

VERIFICATION_PROMPT = """You are verifying if a user successfully built a Supplier Performance Scorecard in WPS Spreadsheet.
Look at these screenshots from their session.
1. Is there a spreadsheet sheet visible that looks like a "Scorecard"?
2. Does it have columns for Vendor, Shipments, On Time Rate, Price, Total Value, Score, and Rank?
3. Are there formulas being edited or visible in the formula bar (e.g., COUNTIF, SUMIF, RANK)?
4. Does the table contain calculated values (not just empty cells or raw data)?

Answer in JSON format:
{
    "has_scorecard_sheet": true/false,
    "has_correct_columns": true/false,
    "shows_formula_usage": true/false,
    "has_calculated_values": true/false,
    "reasoning": "brief explanation"
}
"""

def get_wps_file(copy_from_env, path, extract_format='xlsx'):
    """Helper to get file and load with openpyxl."""
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, f'doc.{extract_format}')
    try:
        copy_from_env(path, temp_file)
        import openpyxl
        wb = openpyxl.load_workbook(temp_file, data_only=False)
        return True, wb, temp_dir
    except Exception as e:
        logger.error(f"Failed to copy/parse {path}: {e}")
        return False, None, temp_dir

def verify_scorecard(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "System error: copy_from_env missing"}

    score = 0
    feedback_parts = []
    
    # Check exported stats
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            export_result = json.load(f)
    except:
        export_result = {}
    finally:
        if os.path.exists(temp_res.name): os.unlink(temp_res.name)

    # 1. Anti-gaming: File must be modified
    if export_result.get("file_modified_during_task"):
        feedback_parts.append("File saved successfully")
        score += 5
    else:
        feedback_parts.append("File was NOT saved/modified during task")
        return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

    # 2. Extract Spreadsheet & Ground Truth
    success, wb, tdir = get_wps_file(copy_from_env, "/home/ga/Documents/supplier_shipments.xlsx")
    
    temp_truth = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    truth_data = {}
    try:
        copy_from_env("/var/lib/task_ground_truth/scorecard_truth.json", temp_truth.name)
        with open(temp_truth.name, 'r') as f:
            truth_data = json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load ground truth: {e}")
    finally:
        if os.path.exists(temp_truth.name): os.unlink(temp_truth.name)

    if not success or not wb:
        feedback_parts.append("Failed to read modified spreadsheet")
        return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}

    try:
        # Check Sheet Existence
        sheets = wb.sheetnames
        if "Scorecard" in sheets:
            score += 10
            feedback_parts.append("Scorecard sheet found")
            ws = wb["Scorecard"]
        else:
            feedback_parts.append(f"Scorecard sheet missing. Found: {sheets}")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}

        # Check Headers
        expected_headers = ["Vendor", "Total_Shipments", "On_Time_Rate", "Avg_Unit_Price", "Total_Value", "Weighted_Score", "Rank"]
        actual_headers = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, 8)]
        
        headers_match = sum(1 for e, a in zip(expected_headers, actual_headers) if e.lower() == a.lower() or e.replace('_', ' ').lower() == a.replace('_', ' ').lower())
        if headers_match >= 6:
            score += 10
            feedback_parts.append("Headers match")
        else:
            feedback_parts.append(f"Headers issue: expected {expected_headers}, got {actual_headers[:7]}")

        # Check Vendors
        expected_vendors = set(truth_data.get("vendors", []))
        actual_vendors = set()
        for r in range(2, 20):
            val = ws.cell(row=r, column=1).value
            if val and isinstance(val, str):
                actual_vendors.add(val.strip())
        
        vendor_overlap = expected_vendors.intersection(actual_vendors)
        if len(vendor_overlap) >= 12:
            score += 15
            feedback_parts.append(f"Vendors listed ({len(vendor_overlap)}/15)")
        else:
            feedback_parts.append(f"Missing vendors: only {len(vendor_overlap)} found")

        # Check Formulas Usage (Columns B through G)
        formula_scores = {
            "COUNTIF": False,
            "SUMIF": False,
            "AVERAGEIF": False,
            "MAX": False,
            "RANK": False
        }
        
        for r in range(2, 17):
            for c in range(2, 8):
                cell_val = ws.cell(row=r, column=c).value
                if isinstance(cell_val, str) and cell_val.startswith("="):
                    upper_val = cell_val.upper()
                    if "COUNTIF" in upper_val: formula_scores["COUNTIF"] = True
                    if "SUMIF" in upper_val: formula_scores["SUMIF"] = True
                    if "AVERAGEIF" in upper_val: formula_scores["AVERAGEIF"] = True
                    if "MAX" in upper_val: formula_scores["MAX"] = True
                    if "RANK" in upper_val: formula_scores["RANK"] = True

        formulas_found = sum(1 for v in formula_scores.values() if v)
        score += (formulas_found * 8) # Up to 40 points
        feedback_parts.append(f"Formulas used: {formulas_found}/5")
        
        for k, v in formula_scores.items():
            if not v: feedback_parts.append(f"Missing {k} formula")

        # VLM Trajectory Verification
        if query_vlm and traj:
            frames = sample_trajectory_frames(traj, n=4)
            final_frame = get_final_screenshot(traj)
            if final_frame:
                frames.append(final_frame)
                
            if frames:
                vlm_res = query_vlm(prompt=VERIFICATION_PROMPT, images=frames)
                if vlm_res.get("success"):
                    parsed = vlm_res.get("parsed", {})
                    if parsed.get("has_scorecard_sheet") and parsed.get("has_calculated_values"):
                        score += 20
                        feedback_parts.append("VLM: Workflow confirmed visually")
                    else:
                        feedback_parts.append("VLM: Visual confirmation failed")
                else:
                    feedback_parts.append("VLM Error")

    except Exception as e:
        logger.error(f"Error inspecting workbook: {e}", exc_info=True)
        feedback_parts.append(f"Verification error: {e}")
    finally:
        import shutil
        shutil.rmtree(tdir, ignore_errors=True)

    passed = score >= 60
    return {
        "passed": passed,
        "score": min(score, 100),
        "feedback": " | ".join(feedback_parts)
    }
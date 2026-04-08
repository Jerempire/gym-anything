#!/usr/bin/env python3
"""
Verifier for build_compensation_analysis task.
Checks formulas, calculations, formatting, and spreadsheet layout.
Uses robust verification with partial scoring.
"""

import json
import os
import tempfile
import logging
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_compensation_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    score = 0
    feedback_parts = []
    
    # 1. Read export result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not export_result.get("output_exists") or not export_result.get("file_modified_during_task"):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Failed: Expected output file was not created or modified during the task."
        }
        
    score += 5
    feedback_parts.append("File exists and modified")

    # 2. Get Ground Truth
    temp_gt = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/var/lib/task_ground_truth/expected_values.json", temp_gt.name)
        with open(temp_gt.name, 'r') as f:
            ground_truth = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read ground truth: {e}"}
    finally:
        if os.path.exists(temp_gt.name):
            os.unlink(temp_gt.name)

    target_file = export_result.get("target_file", "/home/ga/Documents/hr_analysis_complete.xlsx")

    # 3. Pull Spreadsheet to examine locally
    # We copy it twice to use openpyxl with and without data_only
    temp_xlsx_formulas = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_xlsx_values = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(target_file, temp_xlsx_formulas.name)
        copy_from_env(target_file, temp_xlsx_values.name)
        
        # Verify openpyxl is available
        try:
            import openpyxl
        except ImportError:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            import openpyxl

        # Load workbooks
        wb_f = openpyxl.load_workbook(temp_xlsx_formulas.name, data_only=False)
        wb_v = openpyxl.load_workbook(temp_xlsx_values.name, data_only=True)
        
        # ==========================================
        # VERIFY EMPLOYEES SHEET (31 points)
        # ==========================================
        if "Employees" in wb_f.sheetnames:
            ws_f = wb_f["Employees"]
            ws_v = wb_v["Employees"]
            
            # Check headers
            header_O = str(ws_v.cell(row=1, column=15).value).strip()
            header_P = str(ws_v.cell(row=1, column=16).value).strip()
            
            if header_O.lower() == "annualincome":
                score += 3
                feedback_parts.append("Col O header correct")
            if header_P.lower() == "experiencecategory":
                score += 3
                feedback_parts.append("Col P header correct")
                
            # Check formulas presence in sample rows
            has_annual_formulas = all(ws_f.cell(row=r, column=15).data_type == 'f' for r in [2, 10, 100])
            has_exp_formulas = all(ws_f.cell(row=r, column=16).data_type == 'f' for r in [2, 10, 100])
            
            if has_annual_formulas:
                score += 8
                feedback_parts.append("AnnualIncome formulas present")
            if has_exp_formulas:
                score += 8
                feedback_parts.append("ExperienceCategory formulas present")
                
            # Check calculated values against ground truth (spot checks)
            annual_correct = 0
            exp_correct = 0
            spot_checks_a = ground_truth["spot_checks_annual"]
            spot_checks_c = ground_truth["spot_checks_category"]
            
            for row_str, exp_ann in spot_checks_a.items():
                r = int(row_str)
                # Allow minor rounding differences if they used division/multiplication weirdly
                val = ws_v.cell(row=r, column=15).value
                if val is not None and type(val) in [int, float] and abs(val - exp_ann) < 5:
                    annual_correct += 1
            
            for row_str, exp_cat in spot_checks_c.items():
                r = int(row_str)
                val = ws_v.cell(row=r, column=16).value
                if val is not None and str(val).strip().lower() == exp_cat.lower():
                    exp_correct += 1
                    
            if annual_correct == len(spot_checks_a):
                score += 9
                feedback_parts.append("AnnualIncome values correct")
            elif annual_correct > 0:
                score += 4
                feedback_parts.append("AnnualIncome values partially correct")
                
            if exp_correct == len(spot_checks_c):
                score += 9
                feedback_parts.append("ExperienceCategory values correct")
            elif exp_correct > 0:
                score += 4
                feedback_parts.append("ExperienceCategory values partially correct")
        else:
            feedback_parts.append("Employees sheet missing or renamed")
            
        # ==========================================
        # VERIFY SUMMARY SHEET (54 points)
        # ==========================================
        if "Summary" in wb_f.sheetnames:
            ws_f_sum = wb_f["Summary"]
            ws_v_sum = wb_v["Summary"]
            score += 5
            
            # Check Headers
            req_headers = ["Department", "HeadCount", "AvgMonthlyIncome", "TotalMonthlyPayroll", "AttritionCount", "AttritionRate"]
            headers_found = [str(ws_v_sum.cell(row=1, column=c).value).strip().lower() for c in range(1, 7)]
            headers_match = sum(1 for h in req_headers if any(h.lower() in found for found in headers_found))
            if headers_match >= 5:
                score += 5
                feedback_parts.append("Summary headers correct")
                
            # Map departments to rows (in case they sorted them differently)
            dept_rows = {}
            for r in range(2, 6):
                dept_val = str(ws_v_sum.cell(row=r, column=1).value).strip()
                if "Human" in dept_val:
                    dept_rows["Human Resources"] = r
                elif "Research" in dept_val or "R&D" in dept_val:
                    dept_rows["Research & Development"] = r
                elif "Sales" in dept_val:
                    dept_rows["Sales"] = r
                    
            if len(dept_rows) == 3:
                score += 5
                feedback_parts.append("Departments listed correctly")
                
                # Check metrics & formulas
                formulas_used = False
                hc_correct = 0
                avg_correct = 0
                tot_correct = 0
                att_c_correct = 0
                att_r_correct = 0
                
                for dept, r in dept_rows.items():
                    gt = ground_truth["departments"][dept]
                    
                    # Check if they actually used formulas for the aggregations
                    if ws_f_sum.cell(row=r, column=2).data_type == 'f' or ws_f_sum.cell(row=r, column=3).data_type == 'f':
                        formulas_used = True
                    
                    # 1. HeadCount
                    hc = ws_v_sum.cell(row=r, column=2).value
                    if hc is not None and type(hc) in [int, float] and abs(hc - gt["HeadCount"]) < 1:
                        hc_correct += 1
                        
                    # 2. AvgMonthlyIncome
                    avg = ws_v_sum.cell(row=r, column=3).value
                    if avg is not None and type(avg) in [int, float]:
                        if gt["AvgMonthlyIncome"] > 0 and abs((avg - gt["AvgMonthlyIncome"]) / gt["AvgMonthlyIncome"]) < 0.05:
                            avg_correct += 1
                            
                    # 3. TotalMonthlyPayroll
                    tot = ws_v_sum.cell(row=r, column=4).value
                    if tot is not None and type(tot) in [int, float]:
                        if gt["TotalMonthlyPayroll"] > 0 and abs((tot - gt["TotalMonthlyPayroll"]) / gt["TotalMonthlyPayroll"]) < 0.05:
                            tot_correct += 1
                            
                    # 4. AttritionCount
                    att_c = ws_v_sum.cell(row=r, column=5).value
                    if att_c is not None and type(att_c) in [int, float] and abs(att_c - gt["AttritionCount"]) < 1:
                        att_c_correct += 1
                        
                    # 5. AttritionRate
                    att_r = ws_v_sum.cell(row=r, column=6).value
                    if att_r is not None and type(att_r) in [int, float]:
                        # Handle percentages (0.10 vs 10%)
                        if abs(att_r - gt["AttritionRate"]) < 0.02 or abs(att_r - (gt["AttritionRate"]*100)) < 2:
                            att_r_correct += 1

                if formulas_used:
                    score += 5
                    feedback_parts.append("Formulas used in Summary")
                    
                if hc_correct == 3: score += 8
                elif hc_correct > 0: score += 4
                
                if avg_correct == 3: score += 8
                elif avg_correct > 0: score += 4
                
                if tot_correct == 3: score += 7
                elif tot_correct > 0: score += 3
                
                if att_c_correct == 3: score += 6
                elif att_c_correct > 0: score += 3
                
                if att_r_correct == 3: score += 5
                elif att_r_correct > 0: score += 2
                
                if sum([hc_correct, avg_correct, tot_correct, att_c_correct, att_r_correct]) == 15:
                    feedback_parts.append("All summary metrics perfectly calculated")
                else:
                    feedback_parts.append("Some summary metrics incorrect or missing")
            else:
                feedback_parts.append("Departments not listed correctly in Summary sheet")
        else:
            feedback_parts.append("Summary sheet missing")

        # ==========================================
        # VLM VERIFICATION (10 points)
        # ==========================================
        if query_vlm:
            from gym_anything.vlm import sample_trajectory_frames
            frames = sample_trajectory_frames(traj, n=4)
            if frames:
                vlm_prompt = """
                Look at these frames from a user working in WPS Spreadsheet.
                Did the user spend time typing formulas into cells (e.g., using the formula bar at the top, or cells showing =IF, =SUM, =COUNT, etc)?
                Answer JSON: {"typed_formulas": true/false}
                """
                try:
                    vlm_result = query_vlm(images=frames, prompt=vlm_prompt)
                    if vlm_result and vlm_result.get("parsed", {}).get("typed_formulas"):
                        score += 10
                        feedback_parts.append("VLM confirmed formula entry workflow")
                    else:
                        feedback_parts.append("VLM did not confirm active formula typing")
                except Exception as e:
                    logger.warning(f"VLM error: {e}")

    except Exception as e:
        logger.error(f"Error during workbook verification: {e}")
        return {"passed": False, "score": score, "feedback": f"Error parsing workbook: {e}"}
    finally:
        for f in [temp_xlsx_formulas.name, temp_xlsx_values.name]:
            if os.path.exists(f):
                os.unlink(f)

    # Final scoring evaluation
    passed = score >= 60
    
    return {
        "passed": passed,
        "score": min(score, 100),
        "feedback": " | ".join(feedback_parts)
    }
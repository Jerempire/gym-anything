#!/usr/bin/env python3
"""Verifier for build_production_failure_analysis task."""

import json
import tempfile
import os
import logging
from typing import Dict, Any

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_production_analysis(traj: Dict[str, Any], env_info: Dict[str, Any], task_info: Dict[str, Any]) -> Dict[str, Any]:
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')

    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    score = 0
    feedback_parts = []
    
    # 1. Read export result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    output_exists = result.get('output_exists', False)
    created_during_task = result.get('file_created_during_task', False)

    if not output_exists:
        return {
            "passed": False,
            "score": 0,
            "feedback": "Output file /home/ga/Documents/production_analysis.xlsx was not found."
        }

    if not created_during_task:
        feedback_parts.append("Warning: Output file timestamp predates task start (possible gaming).")
    else:
        score += 10
        feedback_parts.append("Output file created successfully.")

    # 2. Copy Ground Truth
    gt = {}
    temp_gt = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/var/lib/app/ground_truth.json", temp_gt.name)
        with open(temp_gt.name, 'r') as f:
            gt = json.load(f)
    except Exception as e:
        logger.error(f"Failed to load ground truth: {e}")
    finally:
        if os.path.exists(temp_gt.name):
            os.unlink(temp_gt.name)

    # 3. Analyze Excel File
    import openpyxl
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/production_analysis.xlsx", temp_xlsx.name)
        
        # Load twice: one for formulas, one for cached values
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        wb_values = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
        sheet_names = wb_formulas.sheetnames
        
        # Verify Sheets Exist (10 pts)
        required_sheets = ['RawData', 'Failure Summary', 'Failure Types', 'Process Stats']
        missing_sheets = [s for s in required_sheets if s not in sheet_names]
        
        if not missing_sheets:
            score += 10
            feedback_parts.append("All required sheets found.")
        else:
            feedback_parts.append(f"Missing sheets: {', '.join(missing_sheets)}")

        # Check RawData integrity (10 pts)
        if 'RawData' in sheet_names:
            ws_raw = wb_formulas['RawData']
            if ws_raw.max_row >= 10000:
                score += 10
                feedback_parts.append("RawData sheet preserved.")
            else:
                feedback_parts.append("RawData sheet modified or incomplete.")

        # Helper to check formulas
        def has_formula(ws, min_r, max_r, min_c, max_c):
            count = 0
            for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c):
                for cell in row:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        count += 1
            return count

        # A. Failure Summary (15 pts)
        if 'Failure Summary' in sheet_names:
            ws_f = wb_formulas['Failure Summary']
            ws_v = wb_values['Failure Summary']
            
            f_count = has_formula(ws_f, 2, 5, 2, 4)
            if f_count >= 6:
                score += 5
                feedback_parts.append("Formulas detected in Failure Summary.")
            
            # Check a few values from GT
            try:
                # Agent might arrange H, M, L in different orders, so search the column
                pts_awarded = 0
                for r in range(2, 6):
                    ptype = str(ws_v.cell(row=r, column=1).value).strip()
                    if ptype in ['H', 'M', 'L']:
                        t_count = ws_v.cell(row=r, column=2).value
                        expected = gt.get(f"count_{ptype}")
                        if t_count and expected and abs(float(t_count) - expected) <= 2:
                            pts_awarded += 2
                
                score += min(10, pts_awarded)
                feedback_parts.append(f"Failure Summary values checked ({min(10, pts_awarded)}/10 pts).")
            except Exception as e:
                feedback_parts.append(f"Error checking Failure Summary values: {e}")

        # B. Failure Types (15 pts)
        if 'Failure Types' in sheet_names:
            ws_f = wb_formulas['Failure Types']
            ws_v = wb_values['Failure Types']
            
            f_count = has_formula(ws_f, 2, 5, 2, 7)
            if f_count >= 10:
                score += 5
                feedback_parts.append("Formulas detected in Failure Types.")
                
            try:
                # Check TWF counts for H, M, L
                pts_awarded = 0
                # Assuming standard layout (Col B is TWF)
                for r in range(2, 5):
                    ptype = str(ws_v.cell(row=r, column=1).value).strip()
                    if ptype in ['H', 'M', 'L']:
                        val = ws_v.cell(row=r, column=2).value
                        expected = gt.get(f"{ptype}_TWF")
                        if val is not None and expected is not None and abs(float(val) - expected) <= 2:
                            pts_awarded += 3
                score += min(10, pts_awarded)
                feedback_parts.append(f"Failure Types values checked ({min(10, pts_awarded)}/10 pts).")
            except Exception as e:
                feedback_parts.append(f"Error checking Failure Types values: {e}")

        # C. Process Stats (15 pts)
        if 'Process Stats' in sheet_names:
            ws_f = wb_formulas['Process Stats']
            ws_v = wb_values['Process Stats']
            
            f_count = has_formula(ws_f, 2, 11, 3, 6)
            if f_count >= 15:
                score += 5
                feedback_parts.append("Formulas detected in Process Stats.")
                
            try:
                # Check averages
                pts_awarded = 0
                for r in range(2, 12):
                    param = str(ws_v.cell(row=r, column=1).value).strip()
                    group = str(ws_v.cell(row=r, column=2).value).strip()
                    
                    if param and group in ['No Failure', 'Failure']:
                        avg_val = ws_v.cell(row=r, column=3).value
                        expected_key = f"{param}_{group}_avg"
                        expected_avg = gt.get(expected_key)
                        
                        if avg_val and expected_avg and abs(float(avg_val) - expected_avg) < max(1.0, expected_avg * 0.05):
                            pts_awarded += 1
                score += min(10, pts_awarded)
                feedback_parts.append(f"Process Stats values checked ({min(10, pts_awarded)}/10 pts).")
            except Exception as e:
                feedback_parts.append(f"Error checking Process Stats values: {e}")

    except Exception as e:
        logger.error(f"Error reading workbook: {e}")
        feedback_parts.append(f"Failed to read workbook structure: {e}")
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # 4. VLM Trajectory Verification (25 pts)
    # Proves the agent actually did the work, not just copied a file
    if query_vlm:
        try:
            frames = sample_trajectory_frames(traj, n=4)
            final = get_final_screenshot(traj)
            
            vlm_prompt = """
            Look at these sequential screenshots of an agent using WPS Spreadsheet.
            Did the agent actively create new sheets, input headers, and use formulas to summarize the data?
            
            Respond in JSON:
            {
                "interacted_with_spreadsheet": true/false,
                "created_multiple_sheets": true/false,
                "confidence": "high/medium/low"
            }
            """
            
            vlm_result = query_vlm(images=frames + [final] if final else frames, prompt=vlm_prompt)
            if vlm_result and vlm_result.get("success"):
                parsed = vlm_result.get("parsed", {})
                if parsed.get("interacted_with_spreadsheet") and parsed.get("created_multiple_sheets"):
                    score += 25
                    feedback_parts.append("VLM confirms active spreadsheet construction workflow.")
                else:
                    feedback_parts.append("VLM did not observe the requested workflow actions.")
        except Exception as e:
            logger.error(f"VLM verification error: {e}")
            feedback_parts.append("VLM verification failed.")

    passed = score >= 60 and output_exists
    
    return {
        "passed": passed,
        "score": min(score, 100),
        "feedback": " | ".join(feedback_parts)
    }
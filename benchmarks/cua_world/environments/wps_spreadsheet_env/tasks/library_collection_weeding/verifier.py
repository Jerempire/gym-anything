#!/usr/bin/env python3
"""Verifier for library_collection_weeding task."""

import sys
import os
import json
import logging
import tempfile

# Add utils directory to path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from wps_verification_utils import (
    copy_and_parse_spreadsheet,
    cleanup_verification_temp,
    vlm_verify_screenshot
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_library_weeding(traj, env_info, task_info):
    """
    Verify the library collection weeding task.
    
    Scoring:
    - File Modified & Sheets (10 pts)
    - Age & Turnover Calculations (20 pts)
    - Turnover Formatting (5 pts)
    - Action Logic (Nested IF) (30 pts)
    - Summary Basics (SUM/COUNTA) (15 pts)
    - Summary Advanced (COUNTIF) (20 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Retrieve base task result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export results: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not export_result.get("file_exists", False):
        return {"passed": False, "score": 0, "feedback": "The library_inventory.xlsx file does not exist."}

    # Fetch spreadsheet using wps utility (reads formulas)
    target_file = "/home/ga/Documents/library_inventory.xlsx"
    success, wb_formulas, error, temp_dir = copy_and_parse_spreadsheet(
        target_file, copy_from_env, file_format='xlsx'
    )

    if not success:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {error}"}

    try:
        score = 0
        feedback_parts = []
        
        # We also want to load the workbook with data_only=True to evaluate the agent's logic
        # The copy_and_parse_spreadsheet utility downloads to temp_dir/input.xlsx
        temp_file_path = os.path.join(temp_dir, 'input.xlsx')
        import openpyxl
        wb_values = openpyxl.load_workbook(temp_file_path, data_only=True)

        # ---------------------------------------------------------------------
        # Criterion 1: File Modification & Sheets (10 pts)
        # ---------------------------------------------------------------------
        file_modified = export_result.get("file_modified", False)
        sheets = wb_formulas.sheetnames
        has_summary_sheet = "Summary" in sheets or "summary" in [s.lower() for s in sheets]

        if file_modified and has_summary_sheet:
            score += 10
            feedback_parts.append("File modified and Summary sheet found (+10)")
        elif file_modified:
            score += 5
            feedback_parts.append("File modified, but Summary sheet missing (+5)")
        else:
            feedback_parts.append("File was NOT modified (0)")
            return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

        # Ensure we have the Inventory sheet
        inv_sheet_name = next((s for s in sheets if s.lower() == "inventory"), None)
        if not inv_sheet_name:
            feedback_parts.append("Inventory sheet is missing or renamed")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}

        inv_sheet = wb_formulas[inv_sheet_name]
        inv_vals = wb_values[inv_sheet_name]

        # ---------------------------------------------------------------------
        # Criterion 2: Age & Turnover Calculations (20 pts)
        # ---------------------------------------------------------------------
        calc_correct = 0
        samples_to_check = [2, 10, 50, 100, 200] # Check a few rows
        
        for row in samples_to_check:
            pub_year = inv_vals.cell(row=row, column=4).value
            item_count = inv_vals.cell(row=row, column=5).value
            checkouts = inv_vals.cell(row=row, column=6).value
            
            if pub_year is None or item_count is None:
                continue

            # Verify Age (Col G, index 7)
            expected_age = 2024 - pub_year
            actual_age = inv_vals.cell(row=row, column=7).value
            if actual_age == expected_age:
                calc_correct += 1
            
            # Verify TurnoverRate (Col H, index 8)
            if item_count > 0:
                expected_turnover = checkouts / item_count
                actual_turnover = inv_vals.cell(row=row, column=8).value
                # allow float rounding differences
                if actual_turnover is not None and abs(float(actual_turnover) - expected_turnover) < 0.01:
                    calc_correct += 1
        
        # 10 checks total (5 rows * 2 columns)
        if calc_correct >= 8:
            score += 20
            feedback_parts.append("Age/Turnover calculations mostly correct (+20)")
        elif calc_correct >= 4:
            score += 10
            feedback_parts.append("Age/Turnover calculations partially correct (+10)")
        else:
            feedback_parts.append("Age/Turnover calculations missing or incorrect (0)")

        # ---------------------------------------------------------------------
        # Criterion 3: Turnover Formatting (5 pts)
        # ---------------------------------------------------------------------
        format_correct = False
        cell_h2 = inv_sheet.cell(row=2, column=8)
        if cell_h2.number_format and ('0.00' in cell_h2.number_format or '2' in str(cell_h2.number_format)):
            format_correct = True
            score += 5
            feedback_parts.append("Turnover formatting correct (+5)")
        else:
            feedback_parts.append("Turnover formatting missing/incorrect (0)")

        # ---------------------------------------------------------------------
        # Criterion 4: Action Logic (Nested IF) (30 pts)
        # ---------------------------------------------------------------------
        logic_correct = 0
        total_logic_checks = 0

        for row in range(2, 502): # Check all rows for robustness
            age = inv_vals.cell(row=row, column=7).value
            turnover = inv_vals.cell(row=row, column=8).value
            item_count = inv_vals.cell(row=row, column=5).value
            actual_action = inv_vals.cell(row=row, column=9).value

            if age is None or turnover is None or item_count is None:
                continue
            
            total_logic_checks += 1
            
            # Compute expected action
            expected_action = "Keep"
            if turnover < 2 and age > 10:
                expected_action = "Weed"
            elif turnover > 20 and item_count <= 2:
                expected_action = "Order"

            if str(actual_action).strip().lower() == expected_action.lower():
                logic_correct += 1

        logic_accuracy = logic_correct / max(total_logic_checks, 1)
        if logic_accuracy >= 0.95:
            score += 30
            feedback_parts.append("Action classification logic perfect (+30)")
        elif logic_accuracy >= 0.70:
            score += 15
            feedback_parts.append(f"Action classification logic partially correct: {logic_accuracy:.0%} (+15)")
        else:
            feedback_parts.append(f"Action classification logic incorrect: {logic_accuracy:.0%} (0)")

        # ---------------------------------------------------------------------
        # Criterion 5 & 6: Summary Sheet Verification
        # ---------------------------------------------------------------------
        sum_sheet_name = next((s for s in sheets if s.lower() == "summary"), None)
        if sum_sheet_name:
            sum_sheet = wb_formulas[sum_sheet_name]
            
            basics_count = 0
            adv_count = 0
            
            # Check for formula presence in cells B2 to B7
            for r in range(2, 8):
                val = sum_sheet.cell(row=r, column=2).value
                if isinstance(val, str) and val.startswith('='):
                    formula_upper = val.upper()
                    if r in [2, 3, 4] and ('SUM' in formula_upper or 'COUNTA' in formula_upper or 'COUNT(' in formula_upper):
                        basics_count += 1
                    elif r in [5, 6] and 'COUNTIF' in formula_upper:
                        adv_count += 1
                    elif r == 7 and 'AVERAGE' in formula_upper:
                        basics_count += 1 # consider average as basic

            if basics_count >= 3:
                score += 15
                feedback_parts.append("Summary basics (SUM/COUNTA/AVERAGE) implemented (+15)")
            elif basics_count > 0:
                score += 7
                feedback_parts.append("Summary basics partially implemented (+7)")
            else:
                feedback_parts.append("Summary basics missing formulas (0)")

            if adv_count >= 2:
                score += 20
                feedback_parts.append("Summary advanced (COUNTIF) implemented (+20)")
            elif adv_count == 1:
                score += 10
                feedback_parts.append("Summary advanced partially implemented (+10)")
            else:
                feedback_parts.append("Summary advanced missing formulas (0)")
        else:
            feedback_parts.append("Summary sheet content checks skipped (sheet not found)")

        # VLM trajectory verification as an anti-gaming signal
        vlm_result = vlm_verify_screenshot(env_info, traj, """
        Analyze this screenshot of WPS Spreadsheet. Answer in JSON:
        {
            "shows_spreadsheet_data": true/false,
            "shows_formulas_being_edited": true/false
        }
        """)
        
        if vlm_result and not vlm_result.get("shows_spreadsheet_data", True):
            logger.warning("VLM indicated spreadsheet was not visible.")
            # We don't deduct heavily, but good for diagnostics.

        # Determine pass/fail
        # Pass threshold is 75 points (e.g., getting Logic + Basics + some formatting/setup)
        passed = score >= 75 and logic_accuracy >= 0.70

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {str(e)}"}
    finally:
        cleanup_verification_temp(temp_dir)
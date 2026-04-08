#!/usr/bin/env python3
import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_earthquake_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Extract task result metadata
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export data: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result.get("file_exists"):
        return {"passed": False, "score": 0, "feedback": "Earthquake catalog file missing."}

    # Extract workbook for programmatic verification
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/earthquake_catalog.xlsx", temp_xlsx.name)
        
        import openpyxl
        
        # Load twice: once to check formulas, once to check evaluated values
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        wb_values = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
        score = 0
        feedback_parts = []
        
        # 1. Anti-gaming / basic completion
        if result.get("file_modified"):
            feedback_parts.append("File modified")
        else:
            feedback_parts.append("File NOT modified (0 pts)")
            return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

        ws_data_f = wb_formulas["EarthquakeData"] if "EarthquakeData" in wb_formulas.sheetnames else wb_formulas.active
        ws_data_v = wb_values["EarthquakeData"] if "EarthquakeData" in wb_values.sheetnames else wb_values.active
        
        # 2. Check Headers (10 points total)
        g1 = str(ws_data_v.cell(row=1, column=7).value).strip()
        h1 = str(ws_data_v.cell(row=1, column=8).value).strip()
        
        if g1.lower() == "mag_category":
            score += 5
            feedback_parts.append("Mag header correct")
        if h1.lower() == "depth_category":
            score += 5
            feedback_parts.append("Depth header correct")
            
        # 3. Ground Truth Generation
        events = []
        for r in range(2, ws_data_v.max_row + 1):
            depth_val = ws_data_v.cell(row=r, column=4).value
            mag_val = ws_data_v.cell(row=r, column=5).value
            if depth_val is not None and mag_val is not None:
                try:
                    events.append((float(depth_val), float(mag_val)))
                except ValueError:
                    pass
        
        # 4. Check Categorizations and Formula Usage (30 points total)
        mag_correct = 0
        depth_correct = 0
        mag_formula_used = False
        depth_formula_used = False
        
        rows_to_check = min(25, len(events))
        
        for r in range(2, 2 + rows_to_check):
            try:
                depth = float(ws_data_v.cell(row=r, column=4).value)
                mag = float(ws_data_v.cell(row=r, column=5).value)
            except:
                continue
                
            expected_mag = "Major" if mag >= 6.0 else "Strong" if mag >= 5.0 else "Moderate" if mag >= 4.0 else "Light" if mag >= 3.0 else "Minor"
            expected_depth = "Deep" if depth > 300 else "Intermediate" if depth >= 70 else "Shallow"
            
            agent_mag = str(ws_data_v.cell(row=r, column=7).value).strip()
            agent_depth = str(ws_data_v.cell(row=r, column=8).value).strip()
            
            if agent_mag.lower() == expected_mag.lower():
                mag_correct += 1
            if agent_depth.lower() == expected_depth.lower():
                depth_correct += 1
                
            cell_g_f = ws_data_f.cell(row=r, column=7)
            cell_h_f = ws_data_f.cell(row=r, column=8)
            if cell_g_f.data_type == 'f' or str(cell_g_f.value).startswith('='): mag_formula_used = True
            if cell_h_f.data_type == 'f' or str(cell_h_f.value).startswith('='): depth_formula_used = True
                
        if mag_correct > rows_to_check * 0.8 and mag_formula_used:
            score += 15
            feedback_parts.append("Mag categorization correct")
        elif mag_correct > rows_to_check * 0.8:
            score += 5
            feedback_parts.append("Mag values correct (missing formula)")
            
        if depth_correct > rows_to_check * 0.8 and depth_formula_used:
            score += 15
            feedback_parts.append("Depth categorization correct")
        elif depth_correct > rows_to_check * 0.8:
            score += 5
            feedback_parts.append("Depth values correct (missing formula)")

        # 5. Summary Sheet Verification (40 points total)
        summary_sheet_name = next((s for s in wb_values.sheetnames if "summary" in s.lower()), None)
        
        if summary_sheet_name:
            score += 10
            feedback_parts.append("Summary sheet found")
            
            ws_sum_v = wb_values[summary_sheet_name]
            
            # Map labels to adjacent cell values
            summary_values = {}
            for r in range(1, 40):
                for c in range(1, 5):
                    val = ws_sum_v.cell(row=r, column=c).value
                    if val and isinstance(val, str):
                        label = val.strip().lower()
                        adjacent_val = ws_sum_v.cell(row=r, column=c+1).value
                        summary_values[label] = adjacent_val
                        
            # Expected Summary Statistics
            total_events = len(events)
            avg_mag = sum(e[1] for e in events) / total_events if total_events else 0
            max_mag = max(e[1] for e in events) if total_events else 0
            min_mag = min(e[1] for e in events) if total_events else 0
            
            counts = {
                "major": sum(1 for e in events if e[1] >= 6.0),
                "strong": sum(1 for e in events if 5.0 <= e[1] < 6.0),
                "moderate": sum(1 for e in events if 4.0 <= e[1] < 5.0),
                "light": sum(1 for e in events if 3.0 <= e[1] < 4.0),
                "minor": sum(1 for e in events if e[1] < 3.0),
                "deep": sum(1 for e in events if e[0] > 300),
                "intermediate": sum(1 for e in events if 70 <= e[0] <= 300),
                "shallow": sum(1 for e in events if e[0] < 70),
            }
            
            # Check Total Events (10 pts)
            agent_total = summary_values.get("total events")
            if agent_total is not None:
                try:
                    if abs(float(agent_total) - total_events) < 0.1:
                        score += 10
                        feedback_parts.append("Total events correct")
                except ValueError:
                    pass
                
            # Check Avg/Max/Min (10 pts)
            agent_avg = summary_values.get("average magnitude")
            agent_max = summary_values.get("maximum magnitude")
            agent_min = summary_values.get("minimum magnitude")
            
            stats_correct = 0
            try:
                if agent_avg is not None and abs(float(agent_avg) - avg_mag) <= 0.05: stats_correct += 1
                if agent_max is not None and abs(float(agent_max) - max_mag) <= 0.05: stats_correct += 1
                if agent_min is not None and abs(float(agent_min) - min_mag) <= 0.05: stats_correct += 1
            except ValueError:
                pass
            
            if stats_correct == 3:
                score += 10
                feedback_parts.append("Avg/Max/Min correct")
            elif stats_correct > 0:
                score += 3 * stats_correct
                
            # Check Mag Counts (10 pts)
            mag_counts_correct = 0
            for k in ["minor", "light", "moderate", "strong", "major"]:
                agent_c = summary_values.get(k)
                if agent_c is not None:
                    try:
                        if abs(float(agent_c) - counts[k]) <= 1.0: mag_counts_correct += 1
                    except ValueError:
                        pass
                    
            if mag_counts_correct == 5:
                score += 10
                feedback_parts.append("Mag counts correct")
            elif mag_counts_correct > 0:
                score += (mag_counts_correct * 2)
                
            # Check Depth Counts (10 pts)
            depth_counts_correct = 0
            for k in ["shallow", "intermediate", "deep"]:
                agent_c = summary_values.get(k)
                if agent_c is not None:
                    try:
                        if abs(float(agent_c) - counts[k]) <= 1.0: depth_counts_correct += 1
                    except ValueError:
                        pass
                    
            if depth_counts_correct == 3:
                score += 10
                feedback_parts.append("Depth counts correct")
            elif depth_counts_correct > 0:
                score += (depth_counts_correct * 3)
                
        else:
            feedback_parts.append("Summary sheet NOT found")

        # 6. VLM Verification (10 points)
        from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
        query_vlm = env_info.get('query_vlm')
        
        if query_vlm:
            frames = sample_trajectory_frames(traj, n=4)
            final = get_final_screenshot(traj)
            images = frames + [final] if final else frames
            
            if images:
                prompt = """Analyze the provided screenshots of a spreadsheet application.
Did the user interact with the spreadsheet to write formulas (like IF, COUNTIF) to categorize earthquake data and build a summary table?
Answer in JSON format:
{
    "workflow_valid": true/false
}
"""
                vlm_resp = query_vlm(images=images, prompt=prompt)
                
                if vlm_resp and vlm_resp.get("parsed", {}).get("workflow_valid"):
                    score += 10
                    feedback_parts.append("VLM Workflow valid")
                else:
                    feedback_parts.append("VLM Workflow NOT valid")

        passed = score >= 70 and summary_sheet_name is not None
        
        # Max score is 110 theoretically, cap at 100
        score = min(score, 100)
        
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }
        
    except Exception as e:
        logger.error(f"Error during verification: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)
#!/usr/bin/env python3
"""Verifier for build_construction_bid_estimate task."""

import os
import json
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Target bid math:
# CON-001: 25*1.05*125 = 3281.25. 25*45 = 1125. Sum = 4406.25
# CON-002: 500*1.10*0.85 = 467.50. 500*1.15 = 575. Sum = 1042.50
# WOD-001: 350*1.15*4.5 = 1811.25. 350*2.50 = 875. Sum = 2686.25
# WOD-002: 80*1.10*18 = 1584. 80*8 = 640. Sum = 2224.00
# FIN-001: 120*1.10*14 = 1848. 120*22 = 2640. Sum = 4488.00
# FIN-002: 15*1.00*35 = 525. 15*40 = 600. Sum = 1125.00
# ELE-001: 1000*1.05*0.65 = 682.50. 1000*1.20 = 1200. Sum = 1882.50
# Subtotal: 17854.50. O&P (15%): 2678.175. Total Bid: 20532.675 -> 20532.68

def verify_bid_estimate(traj, env_info, task_info):
    """Verify that the bid estimate logic, formulas, and totals are correct."""
    
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_total = metadata.get('expected_total', 20532.68)
    tolerance = metadata.get('tolerance', 5.0)

    # 1. Read export json to check anti-gaming
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            result_json = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result JSON: {e}"}
    finally:
        if os.path.exists(temp_res.name):
            os.unlink(temp_res.name)

    if not result_json.get('file_modified_during_task', False):
        return {"passed": False, "score": 0, "feedback": "File was not saved or modified during the task."}

    # 2. Copy spreadsheet to check FORMULAS
    temp_wb_form = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_wb_data = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/residential_takeoff.xlsx", temp_wb_form.name)
        copy_from_env("/home/ga/Documents/residential_takeoff.xlsx", temp_wb_data.name)
        
        import openpyxl
        # Load with formulas
        wb_formulas = openpyxl.load_workbook(temp_wb_form.name, data_only=False)
        # Load with cached data values
        wb_data = openpyxl.load_workbook(temp_wb_data.name, data_only=True)
        
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_wb_form.name): os.unlink(temp_wb_form.name)
        if os.path.exists(temp_wb_data.name): os.unlink(temp_wb_data.name)

    score = 0
    feedback = []
    
    # 3. Check for the new Bid_Summary sheet
    sheets = wb_formulas.sheetnames
    has_summary = any("bid" in s.lower() or "summary" in s.lower() for s in sheets)
    
    if has_summary:
        score += 15
        feedback.append("Bid Summary sheet created")
    else:
        feedback.append("Missing Bid Summary sheet")

    # 4. Check for VLOOKUP/INDEX in Project_Takeoff
    has_lookup = False
    has_sumif = False
    
    if "Project_Takeoff" in wb_formulas.sheetnames:
        takeoff_ws = wb_formulas["Project_Takeoff"]
        # Scan for lookup
        for row in takeoff_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.upper()
                    if "=VLOOKUP" in val or "=INDEX" in val or "=XLOOKUP" in val:
                        has_lookup = True

    if has_lookup:
        score += 20
        feedback.append("Lookup formulas (VLOOKUP/INDEX) detected")
    else:
        feedback.append("No lookup formulas detected")

    # 5. Check for SUMIF in the new summary sheet
    summary_sheet_name = None
    for s in sheets:
        if "bid" in s.lower() or "summary" in s.lower():
            summary_sheet_name = s
            break
            
    if summary_sheet_name:
        sum_ws = wb_formulas[summary_sheet_name]
        for row in sum_ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "=SUMIF" in cell.value.upper():
                    has_sumif = True
                    
    if has_sumif:
        score += 20
        feedback.append("SUMIF formulas detected")
    else:
        feedback.append("No SUMIF formulas detected")

    # 6. Extract the maximum numeric value from the workbook (best proxy for Final Bid Total)
    max_numeric_value = 0.0
    for s_name in wb_data.sheetnames:
        ws_d = wb_data[s_name]
        for row in ws_d.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if cell.value > max_numeric_value:
                        max_numeric_value = float(cell.value)
                        
    # Ensure they actually added a markup (a number ~20532 is higher than subtotal ~17854)
    if abs(max_numeric_value - expected_total) <= tolerance:
        score += 30
        feedback.append(f"Final Bid Total is mathematically accurate (~${expected_total:.2f})")
    elif abs(max_numeric_value - 17854.50) <= tolerance:
        score += 15
        feedback.append("Calculated Subtotal correctly, but missed or incorrectly calculated O&P markup")
    else:
        feedback.append(f"Final Bid Total inaccurate (Highest found value: {max_numeric_value})")

    # 7. Use VLM to check for currency formatting and general layout
    query_vlm = env_info.get('query_vlm')
    from gym_anything.vlm import get_final_screenshot, sample_trajectory_frames
    
    frames = sample_trajectory_frames(traj, n=3)
    final_img = get_final_screenshot(traj)
    images = frames + [final_img] if final_img else frames

    if query_vlm and images:
        vlm_prompt = """
        Review these screenshots of a spreadsheet construction bid estimate.
        Look at the cells showing costs or totals.
        Answer in JSON:
        {
            "has_currency_formatting": true/false,
            "has_markup_line": true/false
        }
        """
        vlm_res = query_vlm(prompt=vlm_prompt, images=images)
        if vlm_res and vlm_res.get('parsed'):
            parsed = vlm_res.get('parsed', {})
            if parsed.get('has_currency_formatting'):
                score += 10
                feedback.append("Currency formatting detected visually")
            if parsed.get('has_markup_line'):
                score += 5
                feedback.append("Markup line visually confirmed")

    passed = score >= 70 and has_lookup and has_summary
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback),
        "details": {
            "has_summary": has_summary,
            "has_lookup": has_lookup,
            "has_sumif": has_sumif,
            "highest_value_found": max_numeric_value
        }
    }
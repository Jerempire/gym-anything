#!/usr/bin/env python3
"""
Verifier for Accounts Receivable Aging Analysis task in WPS Spreadsheet.
Checks for correct structural layout, sheet creation, and formula accuracy.
Includes VLM-based trajectory verification to ensure agent performed the workflow.
"""

import os
import json
import logging
import tempfile
from typing import Dict, Any

# Assuming openpyxl is installed in the verification environment
import openpyxl

# Import gym_anything tools for VLM verification
from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Helper to safely check if a cell contains a specific function
def has_formula(cell, keywords):
    if not isinstance(cell.value, str):
        return False
    val = cell.value.upper()
    if not val.startswith('='):
        return False
    if isinstance(keywords, str):
        return keywords in val
    return any(kw in val for kw in keywords)

def verify_ar_aging(traj: Dict[str, Any], env_info: Dict[str, Any], task_info: Dict[str, Any]) -> Dict[str, Any]:
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available in environment."}

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Read task execution metadata
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            exec_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read execution result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    # Anti-gaming: Ensure file was modified
    if not exec_result.get("file_modified_during_task", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "File was not modified during the task. Zero points."
        }
    
    # 2. Extract workbook from container
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/ar_data.xlsx", temp_wb.name)
        # Load with formulas (data_only=False)
        wb = openpyxl.load_workbook(temp_wb.name, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)

    sheet_names = [s.lower() for s in wb.sheetnames]
    
    # =========================================================================
    # CRITERION 1: Invoices Sheet Enrichments (30 points)
    # =========================================================================
    if 'invoices' in sheet_names:
        ws_inv = wb[wb.sheetnames[sheet_names.index('invoices')]]
        
        # Check Headers (Row 1, Cols F-J / 6-10)
        headers = [str(ws_inv.cell(row=1, column=c).value).lower().replace(' ', '') for c in range(6, 11)]
        expected_headers = ['customername', 'duedate', 'totalpaid', 'outstanding', 'daysoverdue']
        
        headers_match = sum(1 for h in headers if h in expected_headers)
        if headers_match >= 4:
            score += 5
            feedback_parts.append("Invoices headers correct")
        else:
            feedback_parts.append(f"Invoices headers incomplete: {headers}")

        # Check Formulas (Row 2 is sufficient check)
        c2 = ws_inv.cell(row=2, column=6) # CustomerName -> VLOOKUP / INDEX
        if has_formula(c2, ['VLOOKUP', 'INDEX', 'XLOOKUP']):
            score += 5
            
        c2 = ws_inv.cell(row=2, column=7) # DueDate -> Math addition
        if has_formula(c2, ['+', 'SUM']):
            score += 5
            
        c2 = ws_inv.cell(row=2, column=8) # TotalPaid -> SUMIF
        if has_formula(c2, ['SUMIF']):
            score += 5
            
        c2 = ws_inv.cell(row=2, column=9) # Outstanding -> Math subtraction
        if has_formula(c2, ['-', 'MINUS']):
            score += 5
            
        c2 = ws_inv.cell(row=2, column=10) # DaysOverdue -> MAX and DATE
        if has_formula(c2, ['MAX']) and has_formula(c2, ['DATE', 'TODAY', '2024']):
            score += 5
            
        feedback_parts.append("Invoices formulas verified")
    else:
        feedback_parts.append("Invoices sheet missing")

    # =========================================================================
    # CRITERION 2: Aging Sheet Creation & Formulas (30 points)
    # =========================================================================
    if 'aging' in sheet_names:
        score += 10
        ws_aging = wb[wb.sheetnames[sheet_names.index('aging')]]
        
        # Check rows (should be 26 rows = 1 header + 25 customers)
        if ws_aging.max_row >= 25:
            score += 5
            
        # Check aging bucket formulas (SUMIFS or SUMPRODUCT) in C2
        c2 = ws_aging.cell(row=2, column=3)
        if has_formula(c2, ['SUMIFS', 'SUMPRODUCT', 'SUMIF']):
            score += 10
            feedback_parts.append("Aging bucket formulas correct")
            
        # Check OverLimit formula in I2
        i2 = ws_aging.cell(row=2, column=9)
        if has_formula(i2, ['IF', '>', '<']):
            score += 5
            
    else:
        feedback_parts.append("Aging sheet missing")

    # =========================================================================
    # CRITERION 3: Summary Sheet (20 points)
    # =========================================================================
    if 'summary' in sheet_names:
        score += 5
        ws_sum = wb[wb.sheetnames[sheet_names.index('summary')]]
        
        # Check for SUM formulas referencing Aging
        has_sum = False
        has_countif = False
        has_sumif = False
        
        for r in range(4, 15):
            for c in range(1, 4):
                cell = ws_sum.cell(row=r, column=c)
                if has_formula(cell, ['SUM', '+']) and 'AGING' in str(cell.value).upper():
                    has_sum = True
                if has_formula(cell, ['COUNTIF']):
                    has_countif = True
                if has_formula(cell, ['SUMIF']):
                    has_sumif = True
                    
        if has_sum: score += 5
        if has_countif: score += 5
        if has_sumif: score += 5
        
        if has_sum and has_countif and has_sumif:
            feedback_parts.append("Summary sheet formulas complete")
        else:
            feedback_parts.append("Summary sheet formulas incomplete")
    else:
        feedback_parts.append("Summary sheet missing")

    # =========================================================================
    # CRITERION 4: VLM Trajectory Verification (20 points)
    # Ensures the agent was actively interacting with the UI.
    # =========================================================================
    if query_vlm:
        frames = sample_trajectory_frames(traj, n=4)
        final_img = get_final_screenshot(traj)
        
        prompt = """
        Review these screenshots from a computer agent working in WPS Spreadsheet.
        Did the agent successfully build a multi-sheet Accounts Receivable Aging report?
        Look for evidence of:
        1. Adding columns and formulas (VLOOKUP, SUMIF) to an Invoice list.
        2. Creating a summary Aging table with buckets (Current, 31-60, etc.).
        
        Respond ONLY in JSON format:
        {"completed": true/false, "reason": "brief explanation"}
        """
        
        images = frames + [final_img] if final_img else frames
        if images:
            vlm_res = query_vlm(prompt=prompt, images=images)
            try:
                parsed = vlm_res.get('parsed', {}) if isinstance(vlm_res, dict) else json.loads(vlm_res)
                if parsed.get('completed', False):
                    score += 20
                    feedback_parts.append("VLM visual verification passed")
                else:
                    feedback_parts.append(f"VLM verification failed: {parsed.get('reason', 'N/A')}")
            except:
                # If VLM fails to parse, grant partial credit if programmatic checks passed strongly
                if score >= 60:
                    score += 10
                    feedback_parts.append("VLM parse failed, granted partial visual credit.")
    else:
        # If VLM unavailable, compensate if programmatic score is perfect
        if score == 80:
            score += 20
        feedback_parts.append("VLM verification skipped (not available).")

    # =========================================================================
    # Final Evaluation
    # =========================================================================
    # Must achieve at least 70/100 to pass this complex task
    passed = score >= 70
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
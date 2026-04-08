#!/usr/bin/env python3
"""
Verifier for audit_campaign_contributions task.
"""

import sys
import os
import json
import logging

# Ensure verification utilities can be imported
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from wps_verification_utils import (
    copy_and_parse_spreadsheet,
    cleanup_verification_temp
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_audit_campaign(traj, env_info, task_info):
    """
    Verify the campaign audit spreadsheet.
    Multiple independent checks for robustness and anti-gaming.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    passed = False

    # 1. Read metadata produced by export_result.sh
    temp_json_path = "/tmp/verifier_result_export.json"
    try:
        copy_from_env("/tmp/task_result.json", temp_json_path)
        with open(temp_json_path, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read exported metadata: {e}"}
    finally:
        if os.path.exists(temp_json_path):
            os.remove(temp_json_path)

    # Validate file presence and anti-gaming check
    if not export_result.get("file_exists", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "campaign_audit.xlsx NOT found. The file was not saved correctly."
        }
    
    if not export_result.get("file_modified_during_task", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "File exists but was not modified/created during the task session (anti-gaming check failed)."
        }

    # 2. Parse the spreadsheet to evaluate formulas and structure
    target_path = "/home/ga/Documents/campaign_audit.xlsx"
    success, wb_formulas, error, temp_dir = copy_and_parse_spreadsheet(
        target_path, copy_from_env, file_format='xlsx'
    )

    if not success or not wb_formulas:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse XLSX file: {error}"}

    try:
        # File Creation & Sheet names (10 pts)
        sheet_names = [s.lower() for s in wb_formulas.sheetnames]
        if "contributions" in sheet_names and "state summary" in sheet_names:
            score += 10
            feedback_parts.append("✅ Sheets configured properly")
        else:
            feedback_parts.append(f"❌ Missing expected sheets. Found: {wb_formulas.sheetnames}")

        # Evaluate "Contributions" sheet formulas
        contributions_sheet_name = next((s for s in wb_formulas.sheetnames if s.lower() == "contributions"), None)
        
        if contributions_sheet_name:
            ws = wb_formulas[contributions_sheet_name]
            
            # Check for strings/formulas in row 2 (assuming row 1 is headers)
            row_idx = 2
            
            # String Extraction Formula (Year) in J (15 pts)
            j_cell = ws.cell(row=row_idx, column=10) # J
            if j_cell.data_type == 'f' and j_cell.value:
                val_upper = str(j_cell.value).upper()
                if 'RIGHT(' in val_upper or 'MID(' in val_upper or 'LEFT(' in val_upper:
                    score += 15
                    feedback_parts.append("✅ String extraction formula (Year) detected")
                else:
                    score += 5  # Partial for having *some* formula
                    feedback_parts.append("⚠️ Year formula present, but string manipulation function not detected")
            else:
                feedback_parts.append("❌ Year column (J) missing dynamic formula")

            # Logical Flagging Formula (Limit_Violation) in K (15 pts)
            k_cell = ws.cell(row=row_idx, column=11) # K
            if k_cell.data_type == 'f' and k_cell.value:
                val_upper = str(k_cell.value).upper()
                if 'IF(' in val_upper and ('3300' in val_upper or 'FLAG' in val_upper):
                    score += 15
                    feedback_parts.append("✅ Logical IF formula (Limit_Violation) detected")
                else:
                    score += 5
                    feedback_parts.append("⚠️ Limit_Violation formula present, but lacks IF/3300 logic")
            else:
                feedback_parts.append("❌ Limit_Violation column (K) missing dynamic formula")

            # Math Formula (Refund_Due) in L (10 pts)
            l_cell = ws.cell(row=row_idx, column=12) # L
            if l_cell.data_type == 'f' and l_cell.value:
                val_upper = str(l_cell.value).upper()
                if '-' in val_upper and '3300' in val_upper:
                    score += 10
                    feedback_parts.append("✅ Math formula (Refund_Due) detected")
                else:
                    score += 5
                    feedback_parts.append("⚠️ Refund_Due formula present, but lacks expected math")
            else:
                feedback_parts.append("❌ Refund_Due column (L) missing dynamic formula")

            # Formatting Setup (10 pts)
            # Currency format check on I (Amt) and L (Refund)
            has_currency = False
            i_cell = ws.cell(row=row_idx, column=9)
            if '$' in str(i_cell.number_format) or '$' in str(l_cell.number_format):
                has_currency = True

            # Conditional formatting check
            has_cf = False
            if hasattr(ws, 'conditional_formatting') and len(ws.conditional_formatting._cf_rules) > 0:
                has_cf = True

            if has_currency and has_cf:
                score += 10
                feedback_parts.append("✅ Currency formatting and Conditional Formatting detected")
            elif has_currency or has_cf:
                score += 5
                feedback_parts.append("⚠️ Partial formatting detected (Currency or Conditional Formatting missing)")
            else:
                feedback_parts.append("❌ Missing Currency and Conditional Formatting")

        # Evaluate "State Summary" sheet
        summary_sheet_name = next((s for s in wb_formulas.sheetnames if s.lower() == "state summary"), None)
        has_aggregation = False

        if summary_sheet_name:
            ws_sum = wb_formulas[summary_sheet_name]
            
            # Summary Table Setup (10 pts)
            if ws_sum.max_row > 2:  # States successfully listed
                score += 10
                feedback_parts.append("✅ Summary table successfully populated with states")
            else:
                feedback_parts.append("❌ Summary table appears empty")

            # Cross-Sheet Aggregation Formula in B (15 pts)
            b_cell = ws_sum.cell(row=2, column=2) # B2
            if b_cell.data_type == 'f' and b_cell.value:
                val_upper = str(b_cell.value).upper()
                if 'SUMIF(' in val_upper or 'SUMIFS(' in val_upper:
                    score += 15
                    has_aggregation = True
                    feedback_parts.append("✅ SUMIF aggregation formula detected")
                else:
                    feedback_parts.append("❌ Column B missing SUMIF aggregation formula")
            else:
                feedback_parts.append("❌ Total_Amount column missing dynamic formula")
                
        # 3. Use data_only=True to evaluate cached values for the sorting check
        # openpyxl caches the last evaluated values if saved correctly by WPS
        if summary_sheet_name and has_aggregation:
            import openpyxl
            local_xlsx_path = os.path.join(temp_dir, 'input.xlsx')
            
            try:
                wb_values = openpyxl.load_workbook(local_xlsx_path, data_only=True)
                ws_sum_vals = wb_values[summary_sheet_name]
                
                amounts = []
                for row in ws_sum_vals.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
                    val = row[0]
                    if val is not None and isinstance(val, (int, float)):
                        amounts.append(val)
                
                # Data Sorting Descending (15 pts)
                if len(amounts) > 1:
                    if amounts == sorted(amounts, reverse=True):
                        score += 15
                        feedback_parts.append("✅ State Summary data is sorted descending")
                    else:
                        feedback_parts.append("❌ State Summary data is NOT sorted descending")
                else:
                    feedback_parts.append("❌ Insufficient data to verify sorting")
            except Exception as e:
                logger.error(f"Failed to read evaluated values for sort check: {e}")
                feedback_parts.append("❌ Error analyzing cached spreadsheet values for sorting")

        # Determine pass/fail
        passed = score >= 70

        return {
            "passed": passed,
            "score": score,
            "feedback": "\n".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Exception during verification: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        cleanup_verification_temp(temp_dir)
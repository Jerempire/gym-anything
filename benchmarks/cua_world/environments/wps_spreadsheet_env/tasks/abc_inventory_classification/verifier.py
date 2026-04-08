#!/usr/bin/env python3
"""Verifier for ABC Inventory Classification task."""

import sys
import os
import json
import logging
import tempfile
import traceback

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_abc_classification(traj, env_info, task_info):
    """Verify the ABC classification steps."""
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
        
    # Read basic export results
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    # Check if file was modified (Anti-gaming)
    if not export_result.get("file_modified", False):
        return {"passed": False, "score": 0, "feedback": "File was not saved or modified."}

    # Fetch the spreadsheet
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/inventory_abc_analysis.xlsx", temp_xlsx.name)
        
        # Load openpyxl locally to check values and formulas
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        wb_values = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    feedback_parts = []
    score = 0
    max_score = 100

    try:
        # 1. Sheets Check (10 pts)
        sheets = wb_formulas.sheetnames
        if "Summary" in sheets and "Inventory" in sheets:
            score += 10
            feedback_parts.append("Sheets 'Inventory' and 'Summary' exist.")
        else:
            feedback_parts.append(f"Missing required sheets. Found: {sheets}")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
            
        ws_form = wb_formulas["Inventory"]
        ws_val = wb_values["Inventory"]
        summary_form = wb_formulas["Summary"]
        summary_val = wb_values["Summary"]

        # 2. Data Integrity & Annual Value Calculation (15 pts)
        # Check if rows are still roughly 251 (header + 250 items)
        if ws_val.max_row < 240:
            feedback_parts.append(f"Data rows deleted! Found only {ws_val.max_row} rows.")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
            
        # Check Annual_Value formula in Column E
        has_annual_value_formula = False
        e2_formula = str(ws_form.cell(row=2, column=5).value).upper()
        if "=" in e2_formula and "C" in e2_formula and "D" in e2_formula:
            has_annual_value_formula = True
            
        if has_annual_value_formula:
            score += 15
            feedback_parts.append("Annual_Value formula present.")
        else:
            feedback_parts.append("Annual_Value formula missing/incorrect.")

        # 3. Sorting Check (25 pts)
        # Column E must be monotonically decreasing
        is_sorted = True
        annual_values = []
        for i in range(2, 252):
            val = ws_val.cell(row=i, column=5).value
            if val is not None:
                try:
                    annual_values.append(float(val))
                except (ValueError, TypeError):
                    pass
                    
        if len(annual_values) > 100:
            for i in range(len(annual_values) - 1):
                if annual_values[i] < annual_values[i+1]:
                    is_sorted = False
                    break
        else:
            is_sorted = False
            
        if is_sorted:
            score += 25
            feedback_parts.append("Data correctly sorted descending by Annual_Value.")
        else:
            feedback_parts.append("Data NOT sorted descending by Annual_Value.")

        # 4. Cumulative Value & Pct Formulas (20 pts)
        has_running_total = False
        has_pct_formula = False
        
        f3_form = str(ws_form.cell(row=3, column=6).value).upper()
        if "=" in f3_form and ("SUM" in f3_form or "+" in f3_form):
            has_running_total = True
            
        g2_form = str(ws_form.cell(row=2, column=7).value).upper()
        if "=" in g2_form and "/" in g2_form:
            has_pct_formula = True
            
        if has_running_total and has_pct_formula:
            score += 20
            feedback_parts.append("Cumulative Value and Pct formulas detected.")
        else:
            feedback_parts.append("Cumulative formulas missing or incorrect structure.")

        # 5. ABC Class IF Formula (15 pts)
        has_nested_if = False
        h2_form = str(ws_form.cell(row=2, column=8).value).upper()
        if "=" in h2_form and "IF" in h2_form and "A" in h2_form and "B" in h2_form and "C" in h2_form:
            has_nested_if = True
            
        if has_nested_if:
            score += 15
            feedback_parts.append("ABC Class nested IF formula detected.")
        else:
            feedback_parts.append("ABC Class nested IF formula missing.")

        # 6. Summary Sheet COUNTIF (15 pts)
        has_countif = False
        for r in range(1, 10):
            for c in range(1, 5):
                cell_val = str(summary_form.cell(row=r, column=c).value).upper()
                if "=" in cell_val and "COUNTIF" in cell_val:
                    has_countif = True
                    break
            if has_countif:
                break
                
        if has_countif:
            score += 15
            feedback_parts.append("Summary sheet uses COUNTIF.")
        else:
            feedback_parts.append("COUNTIF not found in Summary sheet.")

        # Verify through VLM (Visual check on final screenshot)
        vlm_check = False
        if env_info.get("query_vlm"):
            from gym_anything.vlm import get_final_screenshot
            final_img = get_final_screenshot(traj)
            if final_img:
                prompt = """Analyze this WPS Spreadsheet screenshot.
Does the spreadsheet show an ABC inventory analysis with columns for cumulative percentage and ABC classifications (A, B, C)?
Answer strictly in JSON: {"shows_abc_analysis": true/false}"""
                vlm_resp = env_info["query_vlm"](prompt=prompt, image=final_img)
                if vlm_resp and vlm_resp.get("parsed", {}).get("shows_abc_analysis", False):
                    vlm_check = True

        passed = score >= 70 and is_sorted and has_nested_if
        
        if vlm_check:
            feedback_parts.append("VLM visual verification passed.")
        else:
            feedback_parts.append("VLM visual verification unconfirmed.")
            
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Error during verification: {traceback.format_exc()}")
        return {"passed": False, "score": 0, "feedback": f"Verification failed with internal error: {e}"}
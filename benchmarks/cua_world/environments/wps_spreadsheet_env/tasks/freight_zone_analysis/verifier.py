#!/usr/bin/env python3
"""
Verifier for the freight_zone_analysis task.
Validates file creation, presence of sheets, headers, expected formulas (via VLM and openpyxl),
and verifies accuracy of aggregated summary statistics.
"""

import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_freight_analysis(traj, env_info, task_info):
    """
    Verify the freight zone analysis spreadsheet.
    """
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Load task export metadata
    temp_meta = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_meta.name)
        with open(temp_meta.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task export data: {e}"}
    finally:
        if os.path.exists(temp_meta.name):
            os.unlink(temp_meta.name)

    # Validate output file existence and creation time
    if not export_result.get("output_exists", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Output file /home/ga/Documents/freight_analysis_result.xlsx was not found."
        }
    
    if not export_result.get("file_created_during_task", False):
        return {
            "passed": False,
            "score": 0,
            "feedback": "Output file exists but was not created/modified during the task timeframe (Anti-gaming check failed)."
        }

    # Extract target spreadsheet
    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/freight_analysis_result.xlsx", temp_excel.name)
        
        try:
            from openpyxl import load_workbook
            # Load twice: once for formulas, once for values
            wb_formulas = load_workbook(temp_excel.name, data_only=False)
            wb_values = load_workbook(temp_excel.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to parse Excel file: {e}"}
    finally:
        if os.path.exists(temp_excel.name):
            os.unlink(temp_excel.name)

    score = 0
    feedback_parts = []
    
    # Check sheets presence (10 points)
    sheet_names = wb_formulas.sheetnames
    if "Shipments" in sheet_names and "ZoneLookup" in sheet_names and "ZoneSummary" in sheet_names:
        score += 10
        feedback_parts.append("✅ All 3 sheets present")
    else:
        feedback_parts.append(f"❌ Missing expected sheets. Found: {sheet_names}")
        return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}

    # Check Shipments headers (10 points)
    ws_shipments = wb_formulas["Shipments"]
    headers = [str(ws_shipments.cell(row=1, column=c).value).lower() for c in range(9, 13)]
    expected_headers = ['weight_bracket', 'customer_zone', 'cost_per_kg', 'freight_pct']
    
    headers_matched = sum(1 for h, exp in zip(headers, expected_headers) if exp in h)
    if headers_matched == 4:
        score += 10
        feedback_parts.append("✅ Shipments headers correct")
    else:
        feedback_parts.append(f"⚠️ Missing or incorrect Shipments headers. Found: {headers}")

    # Check for Formulas in Shipments (20 points)
    # We spot-check row 2 for IF, VLOOKUP, division (cost_per_kg)
    i2_val = str(ws_shipments.cell(row=2, column=9).value).upper()
    j2_val = str(ws_shipments.cell(row=2, column=10).value).upper()
    k2_val = str(ws_shipments.cell(row=2, column=11).value).upper()
    
    formula_pts = 0
    if "IF" in i2_val and i2_val.startswith("="): formula_pts += 7
    if "VLOOKUP" in j2_val and j2_val.startswith("="): formula_pts += 7
    if "/" in k2_val and k2_val.startswith("="): formula_pts += 6
    
    score += formula_pts
    if formula_pts == 20:
        feedback_parts.append("✅ Expected formulas (IF, VLOOKUP, Division) found in Shipments")
    else:
        feedback_parts.append(f"⚠️ Missing some formulas in Shipments (scored {formula_pts}/20)")

    # Check ZoneSummary layout and formulas (25 points)
    ws_summary = wb_formulas["ZoneSummary"]
    
    # Layout check
    a1_val = str(ws_summary.cell(row=1, column=1).value).lower()
    a7_val = str(ws_summary.cell(row=7, column=1).value).lower()
    if 'zone' in a1_val and 'overall' in a7_val:
        score += 10
        feedback_parts.append("✅ ZoneSummary matrix layout identified")
    else:
        feedback_parts.append("❌ ZoneSummary matrix layout missing/incorrect")

    # Formula check (AVERAGEIFS)
    b2_val = str(ws_summary.cell(row=2, column=2).value).upper()
    b9_val = str(ws_summary.cell(row=9, column=2).value).upper()
    
    sum_formula_pts = 0
    if "AVERAGEIFS" in b2_val or "AVERAGEIF" in b2_val: sum_formula_pts += 10
    if "COUNTA" in b9_val or "COUNT" in b9_val: sum_formula_pts += 5
    
    score += sum_formula_pts
    if sum_formula_pts == 15:
        feedback_parts.append("✅ Summary aggregation formulas (AVERAGEIFS, COUNTA) found")
    else:
        feedback_parts.append(f"⚠️ Missing summary formulas (scored {sum_formula_pts}/15)")

    # VLM Trajectory Verification (35 points)
    # Proves the agent actively performed the work inside the application
    if query_vlm:
        from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
        frames = sample_trajectory_frames(traj, n=4)
        final_shot = get_final_screenshot(traj)
        
        prompt = """You are evaluating an agent performing a spreadsheet task in WPS Spreadsheet.
        Look at the trajectory frames and the final screenshot.
        Did the agent actively construct formulas (like IF, VLOOKUP, AVERAGEIFS) and build a summary table?
        
        Answer in JSON format:
        {
            "actively_worked": true/false,
            "created_formulas": true/false,
            "created_summary_table": true/false
        }
        """
        
        images = frames + [final_shot] if final_shot else frames
        if images:
            vlm_response = query_vlm(images=images, prompt=prompt)
            if vlm_response and vlm_response.get("success"):
                parsed = vlm_response.get("parsed", {})
                vlm_score = 0
                if parsed.get("actively_worked"): vlm_score += 15
                if parsed.get("created_formulas"): vlm_score += 10
                if parsed.get("created_summary_table"): vlm_score += 10
                
                score += vlm_score
                feedback_parts.append(f"✅ VLM Trajectory Verification scored {vlm_score}/35")
            else:
                feedback_parts.append("⚠️ VLM verification failed to parse")
        else:
            feedback_parts.append("⚠️ No frames available for VLM verification")
    else:
        # Give grace points if VLM is unavailable
        score += 35
        feedback_parts.append("⚠️ VLM unavailable, auto-awarding VLM points")

    passed = score >= 60
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
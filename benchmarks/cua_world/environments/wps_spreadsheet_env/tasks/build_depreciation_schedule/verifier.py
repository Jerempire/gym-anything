#!/usr/bin/env python3
"""
Verifier for build_depreciation_schedule task.
Checks for presence of specific sheets, formula structures, calculated values, and formatting.
"""

import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_depreciation_schedule(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read metadata from container
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    # Output file verification
    if not result.get('output_exists'):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Target file 'depreciation_schedule.xlsx' not found. Ensure it was saved in the Documents folder with the correct name."
        }
    
    if not result.get('file_created_during_task'):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Target file exists but was not created/modified during the task window."
        }

    # 2. Extract and parse the target spreadsheet
    # We load two versions: one for formulas (data_only=False) and one for values (data_only=True)
    temp_xlsx_formulas = tempfile.NamedTemporaryFile(delete=False, suffix='_f.xlsx')
    temp_xlsx_values = tempfile.NamedTemporaryFile(delete=False, suffix='_v.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/depreciation_schedule.xlsx", temp_xlsx_formulas.name)
        copy_from_env("/home/ga/Documents/depreciation_schedule.xlsx", temp_xlsx_values.name)
        
        # We need openpyxl to parse it securely
        import openpyxl
        wb_f = openpyxl.load_workbook(temp_xlsx_formulas.name, data_only=False)
        wb_v = openpyxl.load_workbook(temp_xlsx_values.name, data_only=True)
        
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_xlsx_formulas.name): os.unlink(temp_xlsx_formulas.name)
        if os.path.exists(temp_xlsx_values.name): os.unlink(temp_xlsx_values.name)

    score = 0
    feedback_parts = []
    
    sheets = wb_f.sheetnames
    
    # Check 1: Sheet Existence (10 points)
    expected_sheets = ["SL_Depreciation", "MACRS_Depreciation", "Summary"]
    found_expected = [s for s in expected_sheets if s in sheets]
    if len(found_expected) == 3:
        score += 10
        feedback_parts.append("All expected sheets created")
    else:
        feedback_parts.append(f"Missing sheets. Found: {found_expected}")
        return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)} # Can't proceed safely without sheets

    # References to specific sheets
    sl_sheet_f = wb_f["SL_Depreciation"]
    sl_sheet_v = wb_v["SL_Depreciation"]
    macrs_sheet_f = wb_f["MACRS_Depreciation"]
    macrs_sheet_v = wb_v["MACRS_Depreciation"]
    sum_sheet_f = wb_f["Summary"]
    sum_sheet_v = wb_v["Summary"]

    # Check 2: SL Formula Structure (10 points)
    # Check that the annual depreciation column (G) uses formulas
    sl_formula_found = False
    for row in range(2, 22): # Data rows
        cell_val = str(sl_sheet_f.cell(row=row, column=7).value).strip()
        if cell_val.startswith('='):
            sl_formula_found = True
            break
            
    if sl_formula_found:
        score += 10
        feedback_parts.append("SL formula structure correct")
    else:
        feedback_parts.append("SL formulas missing/hardcoded")

    # Check 3: SL Values (15 points)
    # We check against known good calculated values based on the initial dataset provided
    # FA-001 (CNC Mill): Cost 125000, Salvage 12500, Life 10 -> Annual = 11250
    # FA-006 (Delivery Truck A): Cost 48000, Salvage 6000, Life 5 -> Annual = 8400
    # FA-011 (Server Rack): Cost 32000, Salvage 2000, Life 5 -> Annual = 6000
    
    sl_correct_count = 0
    for row in range(2, 22):
        asset_id = sl_sheet_v.cell(row=row, column=1).value
        val = sl_sheet_v.cell(row=row, column=7).value
        
        try:
            val = float(val) if val is not None else 0.0
            if asset_id == "FA-001" and abs(val - 11250) < 1.0: sl_correct_count += 1
            if asset_id == "FA-006" and abs(val - 8400) < 1.0: sl_correct_count += 1
            if asset_id == "FA-011" and abs(val - 6000) < 1.0: sl_correct_count += 1
        except:
            pass

    if sl_correct_count == 3:
        score += 15
        feedback_parts.append("SL calculation values correct")
    elif sl_correct_count > 0:
        score += 5 * sl_correct_count
        feedback_parts.append(f"SL calculation values partially correct ({sl_correct_count}/3)")
    else:
        feedback_parts.append("SL calculation values incorrect/not found")

    # Check 4: SL Totals Row (5 points)
    # Usually row 22
    has_sl_total_formula = False
    for row in range(21, 25): # Look around row 22
        cell_f = str(sl_sheet_f.cell(row=row, column=7).value).upper()
        if cell_f.startswith('=SUM('):
            has_sl_total_formula = True
            break
            
    if has_sl_total_formula:
        score += 5
        feedback_parts.append("SL Totals row formula exists")

    # Check 5: MACRS lookup formulas (15 points)
    has_macrs_lookup = False
    for row in range(2, 22):
        cell_f = str(macrs_sheet_f.cell(row=row, column=6).value).upper()
        # Look for lookup patterns for Year 1 rate or depreciation amount
        if 'VLOOKUP' in cell_f or 'INDEX' in cell_f or 'MATCH' in cell_f or 'XLOOKUP' in cell_f:
            has_macrs_lookup = True
            break
            
    if has_macrs_lookup:
        score += 15
        feedback_parts.append("MACRS lookup formulas detected")
    else:
        feedback_parts.append("MACRS lookup formulas not detected")

    # Check 6: MACRS values correct (15 points)
    # FA-006 (Delivery Truck, 5-year): Year 1 (Col 7 / 2021) = 48000 * 0.20 = 9600
    # FA-001 (CNC Mill, 7-year): Year 1 (Col 7 / 2021) = 125000 * 0.1429 = 17862.50
    # FA-018 (HVAC, 15-year): Year 1 (Col 7 / 2021) = 150000 * 0.05 = 7500
    macrs_correct_count = 0
    for row in range(2, 22):
        asset_id = macrs_sheet_v.cell(row=row, column=1).value
        # Check Year 1 depreciation amount (usually in Column G / 7)
        val = macrs_sheet_v.cell(row=row, column=7).value 
        try:
            val = float(val) if val is not None else 0.0
            if asset_id == "FA-006" and abs(val - 9600) < 1.0: macrs_correct_count += 1
            if asset_id == "FA-001" and abs(val - 17862.50) < 1.0: macrs_correct_count += 1
            if asset_id == "FA-018" and abs(val - 7500) < 1.0: macrs_correct_count += 1
        except:
            pass

    if macrs_correct_count == 3:
        score += 15
        feedback_parts.append("MACRS Year 1 values correct")
    elif macrs_correct_count > 0:
        score += 5 * macrs_correct_count
        feedback_parts.append(f"MACRS values partially correct ({macrs_correct_count}/3)")
    else:
        feedback_parts.append("MACRS calculation values incorrect/not found")

    # Check 7: MACRS totals row (5 points)
    has_macrs_total_formula = False
    for row in range(21, 25): 
        cell_f = str(macrs_sheet_f.cell(row=row, column=7).value).upper()
        if cell_f.startswith('=SUM('):
            has_macrs_total_formula = True
            break
    if has_macrs_total_formula: score += 5

    # Check 8: Summary annual comparison (10 points)
    # Ensure there are formulas linking back to the other sheets
    has_summary_links = False
    for row in range(3, 14):
        # Look at the SL_Total (Col B) or MACRS_Total (Col C)
        cell_b = str(sum_sheet_f.cell(row=row, column=2).value).upper()
        cell_c = str(sum_sheet_f.cell(row=row, column=3).value).upper()
        if 'SL_DEPRECIATION!' in cell_b or 'MACRS_DEPRECIATION!' in cell_c:
            has_summary_links = True
            break
    
    if has_summary_links:
        score += 10
        feedback_parts.append("Summary sheet linked properly to calculation sheets")
    else:
        feedback_parts.append("Summary sheet lacks linking formulas")

    # Check 9: Summary category breakdown (10 points)
    has_sumif = False
    for row in range(17, 25):
        cell_c = str(sum_sheet_f.cell(row=row, column=3).value).upper()
        if 'SUMIF(' in cell_c:
            has_sumif = True
            break
            
    if has_sumif:
        score += 10
        feedback_parts.append("SUMIF category aggregation detected")
    else:
        feedback_parts.append("SUMIF formulas for category breakdown not found")

    # Check 10: Formatting (5 points)
    has_bold = False
    if sum_sheet_f['A1'].font and sum_sheet_f['A1'].font.bold:
        has_bold = True
    elif sum_sheet_f['A15'].font and sum_sheet_f['A15'].font.bold:
        has_bold = True

    if has_bold:
        score += 5
        feedback_parts.append("Bold formatting applied")

    # Determine overall pass
    # Pass threshold: 60 points + Must have demonstrated some SL and MACRS functionality
    key_criteria = (sl_correct_count > 0 and macrs_correct_count > 0)
    passed = (score >= 60) and key_criteria

    if passed:
        feedback_parts.insert(0, "Task Passed")
    else:
        feedback_parts.insert(0, "Task Failed (requires >= 60 points and at least one correct SL & MACRS calculation)")

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
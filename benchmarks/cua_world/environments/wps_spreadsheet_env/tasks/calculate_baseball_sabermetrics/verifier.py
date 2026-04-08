#!/usr/bin/env python3
"""Verifier for calculate_baseball_sabermetrics task."""

import sys
import os
import json
import tempfile
import logging

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from wps_verification_utils import (
        copy_and_parse_spreadsheet,
        cleanup_verification_temp,
        vlm_verify_screenshot,
    )
except ImportError:
    # Fallback if utils not available in standard location
    pass

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def compute_ops(row_vals):
    """Manually compute PA and OPS from raw stats (columns E through O).
       E=4, H=6, 2B=7, 3B=8, HR=9, BB=11, HBP=13, SF=14
    """
    try:
        ab = float(row_vals[4] or 0)
        h = float(row_vals[6] or 0)
        d2 = float(row_vals[7] or 0)
        d3 = float(row_vals[8] or 0)
        hr = float(row_vals[9] or 0)
        bb = float(row_vals[11] or 0)
        hbp = float(row_vals[13] or 0)
        sf = float(row_vals[14] or 0)

        b1 = h - d2 - d3 - hr
        tb = b1 + 2 * d2 + 3 * d3 + 4 * hr
        pa = ab + bb + hbp + sf

        obp = (h + bb + hbp) / pa if pa > 0 else 0
        slg = tb / ab if ab > 0 else 0
        ops = obp + slg

        return round(ops, 4), int(pa)
    except (ValueError, TypeError, IndexError):
        return -1, -1


def verify_calculate_baseball_sabermetrics(traj, env_info, task_info):
    """Verify formulas, formatting, sorting, filtering, and sheets."""
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Load export result JSON
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_res.name):
            os.unlink(temp_res.name)

    # Anti-gaming check
    if not export_result.get("file_modified_during_task", False):
        return {
            "passed": False,
            "score": 0,
            "feedback": "File was not modified during the task. No work detected."
        }

    # Extract Spreadsheet
    import shutil
    temp_dir = tempfile.mkdtemp()
    local_xlsx = os.path.join(temp_dir, "batting_stats.xlsx")
    
    try:
        copy_from_env("/home/ga/Documents/batting_stats.xlsx", local_xlsx)
        import openpyxl
        wb = openpyxl.load_workbook(local_xlsx, data_only=False)
    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}

    score = 0
    feedback_parts = []
    
    try:
        # 1. Check Raw_Data Formulas (20 pts)
        if 'Raw_Data' in wb.sheetnames:
            ws_raw = wb['Raw_Data']
            
            # Check headers
            headers = [str(ws_raw.cell(row=1, column=c).value).upper() for c in range(16, 22)]
            expected_headers = ['1B', 'TB', 'PA', 'OBP', 'SLG', 'OPS']
            
            if headers == expected_headers:
                score += 5
                feedback_parts.append("Calculated headers correct")
                
                # Check for formulas in row 2
                has_formulas = False
                for col in range(16, 22):
                    cell = ws_raw.cell(row=2, column=col)
                    if cell.data_type == 'f' or str(cell.value).startswith('='):
                        has_formulas = True
                        break
                
                if has_formulas:
                    score += 15
                    feedback_parts.append("Formulas present in Raw_Data")
                else:
                    feedback_parts.append("No formulas detected in calculated columns")
            else:
                feedback_parts.append(f"Headers incorrect. Expected {expected_headers}, got {headers}")
        else:
            feedback_parts.append("Raw_Data sheet missing")

        # 2. Check Number Formatting (10 pts)
        # Verify if formatting contains '000' indicating 3 decimal places
        if 'Raw_Data' in wb.sheetnames:
            ws_raw = wb['Raw_Data']
            ops_cell = ws_raw.cell(row=2, column=21)
            fmt = str(ops_cell.number_format)
            if '.000' in fmt or '0.000' in fmt:
                score += 10
                feedback_parts.append("Number formatting (3 decimal places) correct")
            else:
                feedback_parts.append(f"Number formatting missing or incorrect (Found: {fmt})")

        # 3. Check Qualified_Hitters Sheet (20 pts)
        if 'Qualified_Hitters' in wb.sheetnames:
            ws_qh = wb['Qualified_Hitters']
            score += 10
            feedback_parts.append("Qualified_Hitters sheet created")
            
            # Check Filtering (PA >= 300)
            rows = list(ws_qh.iter_rows(values_only=True))
            if len(rows) > 1:
                valid_filter = True
                ops_values = []
                
                # Check rows 2 to end
                for i, row in enumerate(rows[1:]):
                    computed_ops, computed_pa = compute_ops(row)
                    if computed_pa < 300 and computed_pa != -1:
                        valid_filter = False
                    ops_values.append(computed_ops)
                
                if valid_filter and len(ops_values) > 0:
                    score += 15  # 15 points for correct filter
                    feedback_parts.append("Data correctly filtered (PA >= 300)")
                else:
                    feedback_parts.append("Filter incorrect (Found players with PA < 300)")
                    
                # 4. Check Sorting (OPS Descending) (20 pts)
                # Sort the computed ops values to see if they match the sheet order
                # Filter out -1s which indicate parsing errors
                valid_ops = [val for val in ops_values if val != -1]
                sorted_ops = sorted(valid_ops, reverse=True)
                
                # Allow a small amount of float imprecision or tie-breaker swaps
                is_sorted = True
                for idx in range(len(valid_ops) - 1):
                    # If current is less than next (by more than rounding error), it's not descending
                    if valid_ops[idx] < valid_ops[idx+1] - 0.0001:
                        is_sorted = False
                        break
                        
                if is_sorted and len(valid_ops) > 0:
                    score += 20
                    feedback_parts.append("Data correctly sorted by OPS descending")
                else:
                    feedback_parts.append("Data NOT sorted by OPS descending")
                    
                # 5. Check Conditional Formatting (10 pts)
                if hasattr(ws_qh, 'conditional_formatting') and len(ws_qh.conditional_formatting._cf_rules) > 0:
                    score += 10
                    feedback_parts.append("Conditional formatting rules applied")
                else:
                    feedback_parts.append("Conditional formatting rules NOT found programmatically")
            else:
                feedback_parts.append("Qualified_Hitters sheet is empty")
        else:
            feedback_parts.append("Qualified_Hitters sheet NOT found")

        # 6. VLM Trajectory Verification (10 pts)
        if query_vlm:
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            images = frames + [final] if final else frames
            
            vlm_prompt = """
            Look at these screenshots of a user working in WPS Spreadsheet.
            1. Did they calculate advanced stats like OBP, SLG, OPS?
            2. Is there evidence of a sheet or view showing only 'Qualified_Hitters' (filtered data)?
            3. Is there conditional formatting (color highlighting) applied to the OPS column (usually for values > 0.800)?
            
            Return JSON: {"calculated_stats": true/false, "filtered_view": true/false, "conditional_formatting_visible": true/false}
            """
            
            vlm_res = query_vlm(prompt=vlm_prompt, images=images)
            if vlm_res and vlm_res.get('success'):
                parsed = vlm_res.get('parsed', {})
                if parsed.get('conditional_formatting_visible', False):
                    score += 10  # Fallback/bonus for VLM verifying CF
                    feedback_parts.append("VLM verified conditional formatting visually")
                elif parsed.get('filtered_view', False):
                    score += 5
                    feedback_parts.append("VLM verified filtered view visually")

    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    # Max score cap at 100
    score = min(score, 100)
    passed = score >= 70
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
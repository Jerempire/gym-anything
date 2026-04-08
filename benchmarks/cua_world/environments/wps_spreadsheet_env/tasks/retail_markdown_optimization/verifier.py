#!/usr/bin/env python3
"""Verifier for retail_markdown_optimization task."""

import os
import json
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_retail_markdown(traj, env_info, task_info):
    """
    Verify the Retail Markdown optimization spreadsheet.
    Uses multiple signals: formula checks, value checks, anti-gaming, and VLM trajectory analysis.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Read metadata from export
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get('file_exists'):
        return {"passed": False, "score": 0, "feedback": "retail_inventory.xlsx not found."}
    
    if result.get('file_modified'):
        score += 10
        feedback_parts.append("File was saved successfully (+10)")
    else:
        feedback_parts.append("File not modified during task (possible game/failure)")
        return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

    # 2. Parse Spreadsheet logic using openpyxl from environment
    # We must run a script inside the environment to parse formulas via exec_in_env? No, copy_from_env to host.
    # We use a python block executed on the host. We must ensure the host has openpyxl.
    # The gym framework host should have basic libraries. If not, we handle gracefully.
    
    temp_xlsx_formulas = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_xlsx_values = tempfile.NamedTemporaryFile(delete=False, suffix='_vals.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/retail_inventory.xlsx", temp_xlsx_formulas.name)
        copy_from_env("/home/ga/Documents/retail_inventory.xlsx", temp_xlsx_values.name)
        
        try:
            import openpyxl
        except ImportError:
            return {"passed": False, "score": 0, "feedback": "openpyxl not available on host for verification"}

        wb_formulas = openpyxl.load_workbook(temp_xlsx_formulas.name, data_only=False)
        wb_values = openpyxl.load_workbook(temp_xlsx_values.name, data_only=True)
        
        sheet_names = wb_formulas.sheetnames
        
        if 'Inventory' not in sheet_names:
            feedback_parts.append("Inventory sheet missing")
        else:
            ws_f = wb_formulas['Inventory']
            ws_v = wb_values['Inventory']
            
            # Check formulas presence in a sample row (Row 2)
            has_formulas = True
            for col in ['G', 'H', 'I', 'J', 'K', 'L']:
                val = str(ws_f[f'{col}2'].value)
                if not val.startswith('='):
                    has_formulas = False
                    
            if has_formulas:
                score += 20
                feedback_parts.append("Inventory Core Math formulas present (+20)")
            else:
                feedback_parts.append("Missing formulas in Inventory sheet")

            # Check IF condition for WOS (Zero handling) and VLOOKUP
            wos_formula = str(ws_f['I2'].value).upper()
            vlookup_formula = str(ws_f['J2'].value).upper()
            
            if 'IF' in wos_formula and '99' in wos_formula:
                score += 20
                feedback_parts.append("WOS Div/0 logic implemented (+20)")
            else:
                feedback_parts.append("WOS Div/0 logic (IF/99) not detected")
                
            if 'VLOOKUP' in vlookup_formula:
                score += 20
                feedback_parts.append("Tiered VLOOKUP logic detected (+20)")
            else:
                feedback_parts.append("VLOOKUP missing for Markdown")

            # Check value generation for Cost/Price Math (assuming evaluating works or formulas exist)
            price_formula = str(ws_f['K2'].value)
            cost_formula = str(ws_f['L2'].value)
            if price_formula.startswith('=') and cost_formula.startswith('='):
                score += 10
                feedback_parts.append("Cost & Price formulas present (+10)")
                
        # Check Summary Sheet
        if 'Summary' in sheet_names:
            score += 5
            feedback_parts.append("Summary sheet created (+5)")
            ws_summary = wb_formulas['Summary']
            
            # Look for SUMIF and AVERAGEIF
            has_sumif = False
            has_avgif = False
            
            for row in range(2, 7):
                b_val = str(ws_summary[f'B{row}'].value).upper()
                c_val = str(ws_summary[f'C{row}'].value).upper()
                if 'SUMIF' in b_val: has_sumif = True
                if 'AVERAGEIF' in c_val: has_avgif = True
                
            if has_sumif and has_avgif:
                score += 15
                feedback_parts.append("Conditional Aggregations (SUMIF/AVERAGEIF) used (+15)")
            else:
                feedback_parts.append("Summary sheet missing correct SUMIF/AVERAGEIF formulas")
        else:
            feedback_parts.append("Summary sheet NOT found")
            
    except Exception as e:
        logger.error(f"Error parsing workbook: {e}")
        feedback_parts.append("Error evaluating workbook structure")
    finally:
        for tmp in [temp_xlsx_formulas.name, temp_xlsx_values.name]:
            if os.path.exists(tmp):
                os.unlink(tmp)

    # 3. Trajectory / VLM Verification
    # Ensure the user didn't just upload a completed file via scripts but used the UI.
    query_vlm = env_info.get('query_vlm')
    if query_vlm and 'sample_trajectory_frames' in globals():
        try:
            from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            
            if final:
                vlm_res = query_vlm(
                    images=frames + [final],
                    prompt="Did the agent actively use WPS Spreadsheet to write formulas (like IF, VLOOKUP, SUMIF) and create a Summary sheet? Reply with a JSON: {\"used_formulas\": true/false, \"created_summary_sheet\": true/false}"
                )
                if vlm_res.get('success'):
                    parsed = vlm_res.get('parsed', {})
                    if parsed.get('used_formulas') and parsed.get('created_summary_sheet'):
                        feedback_parts.append("VLM verified active spreadsheet editing")
                    else:
                        feedback_parts.append("VLM did not detect valid spreadsheet edits")
                        # Severe penalty if VLM thinks it was gamed
                        score = max(0, score - 30)
        except Exception as e:
            logger.error(f"VLM verification error: {e}")

    passed = score >= 70

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""Verifier for restaurant_tip_pool_distribution task."""

import json
import os
import sys
import tempfile
import logging

# Import framework utilities dynamically
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from wps_verification_utils import copy_and_parse_spreadsheet, cleanup_verification_temp
except ImportError:
    pass  # Will handle gracefully below

# Try to import for VLM verification
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
except ImportError:
    sample_trajectory_frames = None
    get_final_screenshot = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_tip_pool(traj, env_info, task_info):
    """
    Verify the tip pool distribution logic in WPS Spreadsheet.
    Checks for headers, SUMIF formula, VLOOKUP formula, point-hour math, and proportional tip payout math.
    """
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Retrieve execution metadata
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            exec_result = json.load(f)
    except Exception as e:
        exec_result = {}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not exec_result.get("file_modified", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Fail: The tip_pool_week42.xlsx file was not saved/modified."
        }

    # Extract the file for analysis
    temp_dir = tempfile.mkdtemp(prefix='wps_verify_')
    host_file_path = os.path.join(temp_dir, 'tip_pool.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/tip_pool_week42.xlsx", host_file_path)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to copy target file: {e}"}

    try:
        import openpyxl
        # Load twice: once for formulas, once for values
        wb_formulas = openpyxl.load_workbook(host_file_path, data_only=False)
        wb_values = openpyxl.load_workbook(host_file_path, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse XLSX: {e}"}

    score = 0
    feedback_parts = []
    
    try:
        if 'Summary' not in wb_formulas.sheetnames:
            return {"passed": False, "score": 0, "feedback": "Missing 'Summary' sheet."}
            
        ws_f = wb_formulas['Summary']
        ws_v = wb_values['Summary']
        
        # 1. Headers (5 points)
        headers = [str(ws_v.cell(row=3, column=c).value).strip().lower() for c in range(4, 8)]
        expected_headers = ['total_hours', 'points', 'point_hours', 'tip_payout']
        
        headers_correct = 0
        for h, eh in zip(headers, expected_headers):
            if eh in h:
                headers_correct += 1
        
        if headers_correct == 4:
            score += 5
            feedback_parts.append("Headers: Present")
        else:
            feedback_parts.append(f"Headers: Incomplete/Incorrect ({headers_correct}/4)")

        # Generate Ground Truth to verify math
        # Aggregating Daily_Hours
        gt_hours = {}
        if 'Daily_Hours' in wb_values.sheetnames:
            dh_sheet = wb_values['Daily_Hours']
            for row in dh_sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                emp_id = str(row[1]).strip()
                hrs = float(row[4] or 0.0)
                gt_hours[emp_id] = gt_hours.get(emp_id, 0.0) + hrs
                
        # Mapping Role_Points
        gt_points = {}
        if 'Role_Points' in wb_values.sheetnames:
            rp_sheet = wb_values['Role_Points']
            for row in rp_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    gt_points[str(row[0]).strip().lower()] = float(row[1] or 0.0)
        
        # Calculate Row Level Metrics
        formula_score = { 'D': 0, 'E': 0, 'F': 0, 'G': 0 }
        value_score = { 'D': 0, 'E': 0, 'F': 0, 'G': 0 }
        
        total_gt_pointhours = 0.0
        employee_rows = range(4, 36)
        
        # First pass: Calculate Total Point Hours for Tip Distribution
        for r in employee_rows:
            emp_id = str(ws_v.cell(row=r, column=1).value).strip()
            role = str(ws_v.cell(row=r, column=3).value).strip().lower()
            h = gt_hours.get(emp_id, 0.0)
            p = gt_points.get(role, 0.0)
            total_gt_pointhours += (h * p)
            
        total_tip_pool = 15450.25
        
        # Second pass: Evaluate agent's cells
        agent_total_payout = 0.0
        
        for r in employee_rows:
            emp_id = str(ws_v.cell(row=r, column=1).value).strip()
            role = str(ws_v.cell(row=r, column=3).value).strip().lower()
            
            # Ground truth for this row
            gt_h = gt_hours.get(emp_id, 0.0)
            gt_p = gt_points.get(role, 0.0)
            gt_ph = gt_h * gt_p
            gt_tip = (gt_ph / total_gt_pointhours) * total_tip_pool if total_gt_pointhours > 0 else 0.0
            
            # Agent's formulas and values
            cell_d_f = ws_f.cell(row=r, column=4)
            cell_e_f = ws_f.cell(row=r, column=5)
            cell_f_f = ws_f.cell(row=r, column=6)
            cell_g_f = ws_f.cell(row=r, column=7)
            
            cell_d_v = ws_v.cell(row=r, column=4).value
            cell_e_v = ws_v.cell(row=r, column=5).value
            cell_f_v = ws_v.cell(row=r, column=6).value
            cell_g_v = ws_v.cell(row=r, column=7).value
            
            # Check formulas
            if cell_d_f.data_type == 'f' or str(cell_d_f.value).startswith('='): formula_score['D'] += 1
            if cell_e_f.data_type == 'f' or str(cell_e_f.value).startswith('='): formula_score['E'] += 1
            if cell_f_f.data_type == 'f' or str(cell_f_f.value).startswith('='): formula_score['F'] += 1
            if cell_g_f.data_type == 'f' or str(cell_g_f.value).startswith('='): formula_score['G'] += 1
            
            # Check values (allow small float variations)
            try:
                if cell_d_v is not None and abs(float(cell_d_v) - gt_h) < 0.1: value_score['D'] += 1
                if cell_e_v is not None and abs(float(cell_e_v) - gt_p) < 0.01: value_score['E'] += 1
                if cell_f_v is not None and abs(float(cell_f_v) - gt_ph) < 0.1: value_score['F'] += 1
                if cell_g_v is not None and abs(float(cell_g_v) - gt_tip) < 0.5: value_score['G'] += 1
                
                if cell_g_v is not None:
                    agent_total_payout += float(cell_g_v)
            except (ValueError, TypeError):
                pass
                
        num_emps = len(employee_rows)
        
        # Calculate Final Scores based on Rubric
        # Total Hours (25 pts): Requires formula and correctness
        if formula_score['D'] >= num_emps - 2 and value_score['D'] >= num_emps - 2:
            score += 25
            feedback_parts.append("Total Hours: Perfect (SUMIF detected)")
        elif value_score['D'] >= num_emps - 2:
            score += 10
            feedback_parts.append("Total Hours: Correct but Hardcoded")
        else:
            feedback_parts.append("Total Hours: Incorrect")

        # Points (20 pts)
        if formula_score['E'] >= num_emps - 2 and value_score['E'] >= num_emps - 2:
            score += 20
            feedback_parts.append("Points: Perfect (VLOOKUP detected)")
        elif value_score['E'] >= num_emps - 2:
            score += 10
            feedback_parts.append("Points: Correct but Hardcoded")
        else:
            feedback_parts.append("Points: Incorrect")

        # Point_Hours (10 pts)
        if formula_score['F'] >= num_emps - 2 and value_score['F'] >= num_emps - 2:
            score += 10
            feedback_parts.append("Point_Hours: Perfect")
        elif value_score['F'] >= num_emps - 2:
            score += 5
            feedback_parts.append("Point_Hours: Correct but Hardcoded")
        else:
            feedback_parts.append("Point_Hours: Incorrect")
            
        # Tip_Payout (30 pts)
        if formula_score['G'] >= num_emps - 2 and value_score['G'] >= num_emps - 2:
            score += 30
            feedback_parts.append("Tip_Payout: Perfect Proportional Math")
        elif value_score['G'] >= num_emps - 2:
            score += 15
            feedback_parts.append("Tip_Payout: Correct but Hardcoded")
        else:
            feedback_parts.append("Tip_Payout: Incorrect Math/References")

        # Pool Total Check (10 pts)
        if abs(agent_total_payout - total_tip_pool) < 1.0:
            score += 10
            feedback_parts.append("Pool Total Check: Passed (matches $15450.25)")
        else:
            feedback_parts.append(f"Pool Total Check: Failed (Sum={agent_total_payout:.2f} != {total_tip_pool})")

        # VLM trajectory verification as anti-gaming bonus/check
        if query_vlm and sample_trajectory_frames and get_final_screenshot:
            try:
                frames = sample_trajectory_frames(traj, n=3)
                final = get_final_screenshot(traj)
                
                vlm_result = query_vlm(
                    prompt="""Look at these trajectory frames of a user operating WPS Spreadsheet.
                    Did the user interact with the spreadsheet, click cells, and write formulas in the formula bar?
                    Respond in JSON format: {"actively_worked": true/false}""",
                    images=frames + [final] if final else frames
                )
                
                if vlm_result.get("success") and vlm_result.get("parsed", {}).get("actively_worked", False):
                    feedback_parts.append("VLM: Detected active spreadsheet work.")
            except Exception as e:
                logger.warning(f"VLM check skipped/failed: {e}")

        passed = score >= 85
        
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }
        
    except Exception as e:
        logger.error(f"Verification encountered error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification failed with error: {str(e)}"}
    finally:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
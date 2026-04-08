#!/usr/bin/env python3
"""Verification script for build_risk_assessment_matrix task."""

import os
import json
import logging
import tempfile
import shutil
from typing import Dict, Any

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_risk_assessment(traj, env_info, task_info) -> Dict[str, Any]:
    """Verify that the risk assessment matrix and formulas were completed."""
    
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    EXPECTED_SCORES = metadata.get('expected_scores', [])
    EXPECTED_LEVELS = metadata.get('expected_levels', [])
    EXPECTED_CONTROLS = metadata.get('expected_controls', [])
    EXPECTED_MATRIX = metadata.get('expected_matrix', [])
    EXPECTED_LEVEL_COUNTS = metadata.get('expected_level_counts', {})
    EXPECTED_CAT_COUNTS = metadata.get('expected_cat_counts', {})

    score = 0
    feedback_parts = []
    
    temp_dir = tempfile.mkdtemp(prefix='risk_verify_')
    
    try:
        # Check task_result.json for basic file status
        result_json_path = os.path.join(temp_dir, 'task_result.json')
        try:
            copy_from_env("/tmp/task_result.json", result_json_path)
            with open(result_json_path, 'r') as f:
                task_result = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
            
        if not task_result.get('output_exists', False):
            return {"passed": False, "score": 0, "feedback": "Spreadsheet not found!"}
            
        if not task_result.get('file_modified', False):
            return {"passed": False, "score": 0, "feedback": "File was not modified from initial state"}

        # Copy the spreadsheet
        xlsx_path = os.path.join(temp_dir, 'hazard_register.xlsx')
        try:
            copy_from_env("/home/ga/Documents/hazard_register.xlsx", xlsx_path)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Could not retrieve spreadsheet: {e}"}

        # Use openpyxl to parse
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call(["pip", "install", "-q", "openpyxl"])
            import openpyxl

        # Load twice: once for values, once for formulas
        wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
        
        # 1. Check Hazards sheet (60 pts total)
        if "Hazards" not in wb_values.sheetnames:
            feedback_parts.append("Hazards sheet missing")
        else:
            ws_val = wb_values["Hazards"]
            ws_form = wb_formulas["Hazards"]
            
            # Risk Scores (20 pts)
            score_correct = 0
            has_score_formula = False
            for i in range(30):
                row = i + 2
                val = ws_val.cell(row=row, column=7).value
                form_val = str(ws_form.cell(row=row, column=7).value or "")
                
                if form_val.startswith('='):
                    has_score_formula = True
                
                if val is not None:
                    try:
                        if int(float(val)) == EXPECTED_SCORES[i]:
                            score_correct += 1
                    except (ValueError, TypeError):
                        pass

            score_pts = round(20 * score_correct / 30)
            score += score_pts
            feedback_parts.append(f"Risk Scores: {score_correct}/30 correct")
            if not has_score_formula:
                feedback_parts.append("Warning: Hardcoded Risk Scores (no formula)")
                score -= min(5, score_pts)  # Penalty for hardcoding

            # Risk Levels (20 pts)
            level_correct = 0
            has_level_formula = False
            for i in range(30):
                row = i + 2
                val = ws_val.cell(row=row, column=8).value
                form_val = str(ws_form.cell(row=row, column=8).value or "")
                
                if form_val.startswith('='):
                    has_level_formula = True
                
                if val is not None and str(val).strip().lower() == EXPECTED_LEVELS[i].lower():
                    level_correct += 1

            level_pts = round(20 * level_correct / 30)
            score += level_pts
            feedback_parts.append(f"Risk Levels: {level_correct}/30 correct")

            # Controls (10 pts)
            ctrl_correct = 0
            for i in range(30):
                row = i + 2
                val = ws_val.cell(row=row, column=9).value
                if val is not None and str(val).strip().lower() == EXPECTED_CONTROLS[i].lower():
                    ctrl_correct += 1

            ctrl_pts = round(10 * ctrl_correct / 30)
            score += ctrl_pts
            feedback_parts.append(f"Controls: {ctrl_correct}/30 correct")

        # 2. Check Risk Matrix sheet (15 pts)
        if "Risk Matrix" in wb_values.sheetnames:
            rm_val = wb_values["Risk Matrix"]
            matrix_correct = 0
            
            for r in range(5):
                for c in range(5):
                    val = rm_val.cell(row=r+3, column=c+2).value
                    expected = EXPECTED_MATRIX[r][c]
                    try:
                        if val is not None and int(float(val)) == expected:
                            matrix_correct += 1
                        elif (val is None or val == "") and expected == 0:
                            matrix_correct += 1
                    except (ValueError, TypeError):
                        if expected == 0:
                            matrix_correct += 1

            matrix_pts = round(15 * matrix_correct / 25)
            score += matrix_pts
            feedback_parts.append(f"Risk Matrix values: {matrix_correct}/25 correct")
        else:
            feedback_parts.append("Risk Matrix sheet missing")

        # 3. Check Summary sheet (15 pts)
        if "Summary" in wb_values.sheetnames:
            sm_val = wb_values["Summary"]
            
            level_count_correct = 0
            cat_count_correct = 0
            
            # Flexible search for labels and adjacent values
            for row in sm_val.iter_rows(min_row=1, max_row=20, min_col=1, max_col=2):
                for cell in row:
                    if not cell.value:
                        continue
                    
                    cell_str = str(cell.value).strip()
                    count_cell = sm_val.cell(row=cell.row, column=cell.column + 1).value
                    
                    if count_cell is not None:
                        try:
                            val = int(float(count_cell))
                            
                            # Check Levels
                            if cell_str in EXPECTED_LEVEL_COUNTS:
                                if val == EXPECTED_LEVEL_COUNTS[cell_str]:
                                    level_count_correct += 1
                                    
                            # Check Categories
                            elif cell_str in EXPECTED_CAT_COUNTS:
                                if val == EXPECTED_CAT_COUNTS[cell_str]:
                                    cat_count_correct += 1
                        except (ValueError, TypeError):
                            pass

            summary_total = level_count_correct + cat_count_correct
            summary_pts = round(15 * summary_total / 8)
            score += summary_pts
            feedback_parts.append(f"Summary counts: {summary_total}/8 correct")
        else:
            feedback_parts.append("Summary sheet missing")

        # 4. VLM Trajectory Verification (10 pts)
        # Verify agent actively worked in the UI (anti-scripting/gaming check)
        vlm_score = 0
        if query_vlm:
            frames = sample_trajectory_frames(traj, n=4)
            final_img = get_final_screenshot(traj)
            
            if frames and final_img:
                images = frames + [final_img]
                vlm_prompt = """
                Did the agent actively work on a spreadsheet application to calculate risk matrices?
                Look for evidence of:
                1. Typing formulas in the formula bar (e.g., IF, COUNTIFS)
                2. Navigating between different sheets ("Hazards", "Risk Matrix", "Summary")
                3. Actually entering data, not just viewing an empty sheet
                
                Answer with JSON:
                {"active_work": true/false}
                """
                vlm_result = query_vlm(images=images, prompt=vlm_prompt)
                
                if vlm_result and vlm_result.get("parsed", {}).get("active_work", False):
                    vlm_score = 10
                    score += vlm_score
                    feedback_parts.append("VLM visual verification passed")
                else:
                    feedback_parts.append("VLM did not verify active spreadsheet manipulation")

        # Calculate final pass
        key_criteria = score_correct >= 25 and level_correct >= 25
        passed = score >= 60 and key_criteria

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {str(e)}"}
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
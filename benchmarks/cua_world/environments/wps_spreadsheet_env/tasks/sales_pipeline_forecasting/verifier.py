#!/usr/bin/env python3
"""
Verifier for sales_pipeline_forecasting task.
Ensures correct formula usage and accurate summary creation.
"""

import sys
import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_pipeline_forecasting(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Check basic task execution records
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result_data.get("output_exists"):
        return {"passed": False, "score": 0, "feedback": "pipeline_forecast_completed.xlsx was not saved."}
    
    if result_data.get("file_created_during_task"):
        score += 10
        feedback_parts.append("File created during task (+10)")
    else:
        feedback_parts.append("Warning: File timestamp invalid.")

    # 2. Extract and parse the completed workbook
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/pipeline_forecast_completed.xlsx", temp_wb.name)
        
        # openpyxl parsing with formulas intact (data_only=False)
        import openpyxl
        wb = openpyxl.load_workbook(temp_wb.name, data_only=False)
        
        sheets = wb.sheetnames
        has_opps = 'Opportunities' in sheets
        has_summary = 'Rep_Summary' in sheets
        
        if not has_opps:
            feedback_parts.append("Missing Opportunities sheet.")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
            
        ws_opps = wb['Opportunities']
        
        # 3. Check formulas in Opportunities sheet
        vlookup_found = False
        mult_found = False
        if_found = False
        
        # Scan first 10 data rows to find the formulas (agent might skip header rows occasionally)
        for row in ws_opps.iter_rows(min_row=2, max_row=12):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and str(cell.value).startswith('='):
                    formula = str(cell.value).upper()
                    if 'LOOKUP' in formula or 'INDEX' in formula:
                        vlookup_found = True
                    if '*' in formula or 'PRODUCT' in formula:
                        mult_found = True
                    if 'IF(' in formula or 'IFS(' in formula:
                        if_found = True

        if vlookup_found:
            score += 20
            feedback_parts.append("Lookup logic found (+20)")
        if mult_found:
            score += 15
            feedback_parts.append("Weighted arithmetic found (+15)")
        if if_found:
            score += 20
            feedback_parts.append("Tiering IF logic found (+20)")

        # 4. Check Summary Sheet construction
        if has_summary:
            ws_summary = wb['Rep_Summary']
            
            # Look for SUMIF(S) formulas
            sumif_found = False
            rep_count = 0
            has_currency_format = False
            
            for row in ws_summary.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    val = str(cell.value).upper() if cell.value else ""
                    # Check for representative names
                    if any(rep in val for rep in ["DARCEL", "VICKI", "JOHN", "SARAH"]):
                        rep_count += 1
                        
                    # Check for formulas
                    if val.startswith('=') and ('SUMIF' in val):
                        sumif_found = True
                        
                    # Check formatting
                    if cell.number_format and ('$' in cell.number_format or '€' in cell.number_format or '£' in cell.number_format):
                        has_currency_format = True

            if sumif_found:
                score += 25
                feedback_parts.append("SUMIF aggregation found (+25)")
            else:
                feedback_parts.append("Missing SUMIF aggregation.")
                
            if has_currency_format:
                score += 10
                feedback_parts.append("Currency formatting found (+10)")
            else:
                feedback_parts.append("Missing Currency formatting.")
                
        else:
            feedback_parts.append("Rep_Summary sheet missing.")

    except ImportError:
        feedback_parts.append("Error: openpyxl missing in environment.")
    except Exception as e:
        feedback_parts.append(f"Workbook evaluation error: {str(e)}")
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)

    # 5. Optional VLM Verification (Trajectory frames)
    # Proves visual interaction occurred
    query_vlm = env_info.get('query_vlm')
    if query_vlm:
        try:
            from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            images = frames + [final] if final else frames
            
            if images:
                vlm_res = query_vlm(
                    images=images,
                    prompt="Did the user interact with the WPS spreadsheet, type formulas, and create a summary table? Answer yes or no."
                )
                if vlm_res and 'yes' in vlm_res.get('response', '').lower():
                    feedback_parts.append("VLM visual proof confirmed")
        except Exception as e:
            logger.warning(f"VLM verification skipped/failed: {e}")

    # Threshold Check
    key_criteria_met = vlookup_found and sumif_found and has_summary
    passed = score >= 75 and key_criteria_met

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
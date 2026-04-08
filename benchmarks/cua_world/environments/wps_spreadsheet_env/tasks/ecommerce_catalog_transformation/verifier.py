#!/usr/bin/env python3
"""
Verifier for ecommerce_catalog_transformation task.

Verifies:
1. File modified (Anti-gaming)
2. SKU column formulas and correct text manipulation
3. Margin_Pct column lookup formulas and correct percentages
4. Retail_Price column math (cost / (1 - margin)) and rounding
5. Stock_Value column math (cost * stock)
6. Status column logical IF/AND statements
7. Catalog_Summary sheet with proper aggregations
8. VLM Trajectory Verification
"""

import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_catalog_transformation(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read exported metadata
    temp_meta = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_meta.name)
        with open(temp_meta.name, 'r') as f:
            meta = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read exported result: {e}"}
    finally:
        if os.path.exists(temp_meta.name):
            os.unlink(temp_meta.name)

    if not meta.get("file_exists"):
        return {"passed": False, "score": 0, "feedback": "Target file retail_catalog.xlsx does not exist."}
    if not meta.get("file_modified"):
        return {"passed": False, "score": 0, "feedback": "Target file was not modified. Agent did nothing."}

    # 2. Copy the actual spreadsheet (need both data_only=False for formulas and True for values)
    temp_wb_formulas = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_wb_values = tempfile.NamedTemporaryFile(delete=False, suffix='_vals.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/retail_catalog.xlsx", temp_wb_formulas.name)
        copy_from_env("/home/ga/Documents/retail_catalog.xlsx", temp_wb_values.name)
        
        import openpyxl
        wb_f = openpyxl.load_workbook(temp_wb_formulas.name, data_only=False)
        wb_v = openpyxl.load_workbook(temp_wb_values.name, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_wb_formulas.name): os.unlink(temp_wb_formulas.name)
        if os.path.exists(temp_wb_values.name): os.unlink(temp_wb_values.name)

    score = 0
    feedback_parts = []
    
    try:
        if "Raw_Products" not in wb_f.sheetnames:
            return {"passed": False, "score": 0, "feedback": "Raw_Products sheet missing or renamed."}
            
        ws_f = wb_f["Raw_Products"]
        ws_v = wb_v["Raw_Products"]
        
        # Ground truth mapping
        margin_map = {
            "Footwear": 0.40,
            "Apparel": 0.50,
            "Electronics": 0.25,
            "Home": 0.35,
            "Outdoors": 0.45
        }
        
        # Tracking points
        sku_correct = 0
        margin_correct = 0
        price_correct = 0
        value_correct = 0
        status_correct = 0
        total_rows = 50
        
        # Summary Tracking
        gt_total_value = 0.0
        gt_sum_prices = 0.0
        gt_high_value_count = 0

        # Validate rows 2 to 51
        for row in range(2, 52):
            v_id = str(ws_v.cell(row=row, column=1).value or "")
            dept = str(ws_v.cell(row=row, column=3).value or "")
            cost = float(ws_v.cell(row=row, column=4).value or 0)
            stock = int(ws_v.cell(row=row, column=5).value or 0)
            
            # --- F: SKU Check ---
            f_formula = str(ws_f.cell(row=row, column=6).value or "")
            f_val = str(ws_v.cell(row=row, column=6).value or "")
            
            expected_sku = ""
            if "-" in v_id and len(dept) >= 3:
                expected_sku = dept[:3].upper() + "-" + v_id.split("-", 1)[1]
                
            if f_formula.startswith("=") and f_val == expected_sku:
                sku_correct += 1
                
            # --- G: Margin Check ---
            g_formula = str(ws_f.cell(row=row, column=7).value or "")
            g_val = ws_v.cell(row=row, column=7).value
            
            expected_margin = margin_map.get(dept, 0)
            if g_formula.startswith("=") and g_val is not None and abs(float(g_val) - expected_margin) < 0.01:
                margin_correct += 1
                
            # --- H: Price Check ---
            h_formula = str(ws_f.cell(row=row, column=8).value or "")
            h_val = ws_v.cell(row=row, column=8).value
            
            expected_price = round(cost / (1 - expected_margin), 2)
            gt_sum_prices += expected_price
            
            if h_formula.startswith("=") and h_val is not None and abs(float(h_val) - expected_price) < 0.01:
                price_correct += 1
                
            # --- I: Stock Value Check ---
            i_formula = str(ws_f.cell(row=row, column=9).value or "")
            i_val = ws_v.cell(row=row, column=9).value
            
            expected_value = cost * stock
            gt_total_value += expected_value
            
            if i_formula.startswith("=") and i_val is not None and abs(float(i_val) - expected_value) < 0.01:
                value_correct += 1
                
            # --- J: Status Check ---
            j_formula = str(ws_f.cell(row=row, column=10).value or "")
            j_val = str(ws_v.cell(row=row, column=10).value or "")
            
            expected_status = "High Value Low Stock" if (stock < 20 and expected_price > 50) else "Standard"
            if expected_status == "High Value Low Stock":
                gt_high_value_count += 1
                
            if j_formula.startswith("=") and j_val == expected_status:
                status_correct += 1

        gt_avg_price = gt_sum_prices / 50.0

        # Scoring Columns (Max 65 pts)
        if sku_correct == total_rows:
            score += 15
            feedback_parts.append("SKU generation perfect (15/15)")
        elif sku_correct > 0:
            score += int(15 * (sku_correct/total_rows))
            feedback_parts.append(f"SKU generation partial ({sku_correct}/{total_rows})")
            
        if margin_correct == total_rows:
            score += 10
            feedback_parts.append("Margin lookup perfect (10/10)")
        elif margin_correct > 0:
            score += int(10 * (margin_correct/total_rows))
            
        if price_correct == total_rows:
            score += 15
            feedback_parts.append("Retail Price math perfect (15/15)")
        elif price_correct > 0:
            score += int(15 * (price_correct/total_rows))
            
        if value_correct == total_rows:
            score += 10
            feedback_parts.append("Stock Value math perfect (10/10)")
        elif value_correct > 0:
            score += int(10 * (value_correct/total_rows))
            
        if status_correct == total_rows:
            score += 15
            feedback_parts.append("Logical Status perfect (15/15)")
        elif status_correct > 0:
            score += int(15 * (status_correct/total_rows))

        # --- Check Summary Sheet (Max 20 pts) ---
        if "Catalog_Summary" in wb_v.sheetnames:
            ws_sum = wb_v["Catalog_Summary"]
            sum_score = 0
            
            # Check B1: Total Stock Value
            b1_val = ws_sum.cell(row=1, column=2).value
            if b1_val is not None and abs(float(b1_val) - gt_total_value) < 1.0:
                sum_score += 7
            
            # Check B2: Avg Retail Price
            b2_val = ws_sum.cell(row=2, column=2).value
            if b2_val is not None and abs(float(b2_val) - gt_avg_price) < 0.1:
                sum_score += 7
                
            # Check B3: Count
            b3_val = ws_sum.cell(row=3, column=2).value
            if b3_val is not None and int(float(b3_val)) == gt_high_value_count:
                sum_score += 6
                
            score += sum_score
            feedback_parts.append(f"Summary sheet logic: {sum_score}/20 pts")
        else:
            feedback_parts.append("Catalog_Summary sheet missing (0/20)")

        # --- VLM Trajectory Check (Max 15 pts) ---
        query_vlm = env_info.get('query_vlm')
        if query_vlm:
            from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            
            vlm_prompt = """
            Look at these screenshots of a user operating a spreadsheet. 
            Did the user actively interact with spreadsheet formulas, type in cells, or switch between tabs?
            Reply in JSON format with a single boolean field "user_activity_detected".
            """
            vlm_result = query_vlm(images=frames + [final], prompt=vlm_prompt)
            if vlm_result and vlm_result.get("parsed", {}).get("user_activity_detected", False):
                score += 15
                feedback_parts.append("VLM confirmed spreadsheet activity (15/15)")
            else:
                feedback_parts.append("VLM did not detect active formula writing (0/15)")
        else:
            feedback_parts.append("VLM check unavailable - granting auto 15 pts")
            score += 15

        passed = score >= 70
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Python verifier error: {str(e)}"}
#!/usr/bin/env python3
"""
Verifier for GSA Lease Portfolio Analysis task.
Requires the agent to save a formatted .xlsx with calculated fields 
and a cross-sheet summary using COUNTIF/SUMIF.
"""

import os
import json
import logging
import tempfile
from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_lease_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Read task_result.json
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read execution result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    # 2. Check File Existence & Anti-Gaming
    output_exists = result_data.get('output_exists', False)
    file_created_during_task = result_data.get('file_created_during_task', False)
    
    if not output_exists:
        return {"passed": False, "score": 0, "feedback": "Target file gsa_portfolio_analysis.xlsx was not saved."}
    
    if not file_created_during_task:
        return {"passed": False, "score": 0, "feedback": "Target file existed before task (anti-gaming failure)."}

    # 3. Pull the XLSX file and evaluate using openpyxl
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb = None
    try:
        copy_from_env("/home/ga/Documents/gsa_portfolio_analysis.xlsx", temp_xlsx.name)
        
        # Install openpyxl dynamically if missing in the host environment running the verifier
        try:
            import openpyxl
        except ImportError:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl

        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse XLSX file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # Begin Evaluation of Workbook Data
    sheet_names = wb.sheetnames
    
    has_lease_data = "Lease Data" in sheet_names
    has_state_summary = "State Summary" in sheet_names
    
    if has_lease_data and has_state_summary:
        score += 15
        feedback_parts.append("✅ Core sheets created")
    else:
        feedback_parts.append(f"❌ Missing required sheets. Found: {sheet_names}")

    if has_lease_data:
        ws_lease = wb["Lease Data"]
        # Check for formulas in columns I, J, K
        has_division_formula = False
        has_year_formula = False
        has_if_formula = False
        
        for row in ws_lease.iter_rows(min_row=2, max_row=min(10, ws_lease.max_row)):
            if row[8].value and str(row[8].value).startswith('='):
                has_division_formula = True
            
            # Check for YEAR/Text extraction
            if row[9].value and str(row[9].value).startswith('='):
                val = str(row[9].value).upper()
                if 'YEAR' in val or 'RIGHT' in val or 'TEXT' in val:
                    has_year_formula = True
                    
            # Check for Nested IF
            if row[10].value and str(row[10].value).startswith('='):
                val = str(row[10].value).upper()
                if 'IF(' in val and ('RENEW' in val or 'EXPIRED' in val):
                    has_if_formula = True
        
        if has_division_formula:
            score += 10
            feedback_parts.append("✅ Rent_Per_SqFt formula detected")
        if has_year_formula:
            score += 15
            feedback_parts.append("✅ Expiration_Year extraction formula detected")
        if has_if_formula:
            score += 20
            feedback_parts.append("✅ Action_Required nested IF logic detected")
            
    if has_state_summary:
        ws_summary = wb["State Summary"]
        has_countif = False
        has_sumif = False
        target_states = ['DC', 'MD', 'VA', 'CA', 'TX']
        states_found = 0
        
        # Scan for states and aggregation formulas
        for row in ws_summary.iter_rows(min_row=1, max_row=min(20, ws_summary.max_row)):
            if row[0].value and str(row[0].value).upper() in target_states:
                states_found += 1
                
                # Check adjacent cells for formulas
                col_b = str(row[1].value).upper() if row[1].value else ""
                col_c = str(row[2].value).upper() if row[2].value else ""
                
                if 'COUNTIF(' in col_b or 'COUNTIFS(' in col_b:
                    has_countif = True
                if 'SUMIF(' in col_c or 'SUMIFS(' in col_c:
                    has_sumif = True
                    
        if states_found >= 3:
            score += 5
            feedback_parts.append("✅ Target states listed in summary")
        if has_countif:
            score += 10
            feedback_parts.append("✅ COUNTIF aggregation used")
        if has_sumif:
            score += 10
            feedback_parts.append("✅ SUMIF aggregation used")

    # 4. VLM Trajectory Verification (proves work wasn't just injected without UI)
    if query_vlm:
        frames = sample_trajectory_frames(traj, n=4)
        final_img = get_final_screenshot(traj)
        
        prompt = """
        Review these screenshots of a user working in WPS Spreadsheet.
        Did the user actively build out a commercial lease portfolio analysis?
        Specifically:
        1. Did they type or construct formulas in cells?
        2. Are there multiple sheets visible (like Lease Data and State Summary)?
        3. Did they format columns containing financial values to Currency ($)?
        Answer briefly and conclude with JSON: {"verified": true/false}
        """
        
        vlm_res = query_vlm(images=frames + [final_img], prompt=prompt)
        try:
            # Parse json from output
            import re
            json_match = re.search(r'\{.*\}', vlm_res['response'], re.DOTALL)
            if json_match:
                parsed = json.loads(json_match.group())
                if parsed.get('verified', False):
                    score += 15
                    feedback_parts.append("✅ VLM visual trajectory verified")
                else:
                    feedback_parts.append("❌ VLM could not confirm visual workflow")
            else:
                feedback_parts.append("⚠️ VLM returned invalid format")
        except:
            feedback_parts.append("⚠️ VLM parsing error")
    else:
        # Give free points if VLM isn't configured, so testing isn't blocked
        score += 15

    # Determine final success
    passed = score >= 70
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""Verifier for build_ghg_emissions_inventory task."""

import sys
import os
import json
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from wps_verification_utils import copy_and_parse_spreadsheet, cleanup_verification_temp
except ImportError:
    pass

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def safe_copy_and_parse(container_path, copy_from_env):
    """Fallback copy and parse if wps_verification_utils fails/missing."""
    try:
        from wps_verification_utils import copy_and_parse_spreadsheet
        return copy_and_parse_spreadsheet(container_path, copy_from_env, file_format='xlsx')
    except ImportError:
        temp_dir = tempfile.mkdtemp(prefix='wps_verify_')
        temp_file = os.path.join(temp_dir, 'input.xlsx')
        try:
            copy_from_env(container_path, temp_file)
            import openpyxl
            wb = openpyxl.load_workbook(temp_file, data_only=True)
            return True, wb, None, temp_dir
        except Exception as e:
            return False, None, str(e), temp_dir

def safe_cleanup(temp_dir):
    try:
        from wps_verification_utils import cleanup_verification_temp
        cleanup_verification_temp(temp_dir)
    except ImportError:
        import shutil
        if temp_dir and os.path.exists(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)


def verify_ghg_inventory(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read exported task result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Target file does not exist."}
    
    if not result.get('file_modified_during_task'):
        return {"passed": False, "score": 0, "feedback": "File was not modified during task. Agent did nothing."}

    # 2. Open spreadsheet
    success, _, error, temp_dir = safe_copy_and_parse("/home/ga/Documents/ghg_inventory_2024.xlsx", copy_from_env)
    if not success:
        return {"passed": False, "score": 0, "feedback": f"Failed to open spreadsheet: {error}"}

    try:
        feedback_parts = []
        score = 0
        
        # We need data_only=True to evaluate the agent's math formula results.
        temp_file_data = os.path.join(temp_dir, 'data_only.xlsx')
        copy_from_env("/home/ga/Documents/ghg_inventory_2024.xlsx", temp_file_data)
        import openpyxl
        wb_data = openpyxl.load_workbook(temp_file_data, data_only=True)
        
        ws_data = wb_data['Energy_Data'] if 'Energy_Data' in wb_data.sheetnames else None
        ws_summary = wb_data['Summary'] if 'Summary' in wb_data.sheetnames else None
        
        if not ws_data:
            return {"passed": False, "score": 0, "feedback": "Energy_Data sheet missing"}

        # Define factors dynamically to match ground truth exactly
        factors = {
            'Electricity': {'co2': 0.386, 'ch4': 0.030, 'n2o': 0.004},
            'Natural Gas': {'co2': 5.300, 'ch4': 0.500, 'n2o': 0.010},
            'Gasoline': {'co2': 8.780, 'ch4': 0.380, 'n2o': 0.080}
        }
        gwps = {'co2': 1, 'ch4': 25, 'n2o': 298}

        co2_correct = 0
        grams_div_correct = 0
        gwp_mt_correct = 0
        total_rows = 0

        # Dynamic ground truth sums
        expected_scopes = {'Scope 1': 0.0, 'Scope 2': 0.0}
        expected_facilities = {'Science Building': 0.0, 'Library': 0.0, 'Dorms': 0.0, 'Admin': 0.0, 'Fleet': 0.0}

        # 3. Verify math for each row in Energy_Data
        for row in range(2, ws_data.max_row + 1):
            source = ws_data.cell(row=row, column=3).value
            scope = ws_data.cell(row=row, column=4).value
            consumption = ws_data.cell(row=row, column=5).value
            facility = ws_data.cell(row=row, column=2).value

            if not source or not consumption:
                continue
                
            total_rows += 1
            f = factors.get(source, {})
            
            exp_co2 = consumption * f.get('co2', 0)
            exp_ch4 = consumption * f.get('ch4', 0) / 1000.0
            exp_n2o = consumption * f.get('n2o', 0) / 1000.0
            exp_co2e = (exp_co2 * gwps['co2']) + (exp_ch4 * gwps['ch4']) + (exp_n2o * gwps['n2o'])
            exp_mt = exp_co2e / 1000.0
            
            # Ground truth aggregation
            if scope in expected_scopes:
                expected_scopes[scope] += exp_mt
            if facility in expected_facilities:
                expected_facilities[facility] += exp_mt

            # Agent calculations
            ag_co2 = ws_data.cell(row=row, column=7).value
            ag_ch4 = ws_data.cell(row=row, column=8).value
            ag_n2o = ws_data.cell(row=row, column=9).value
            ag_co2e = ws_data.cell(row=row, column=10).value
            ag_mt = ws_data.cell(row=row, column=11).value

            try:
                if ag_co2 is not None and abs(float(ag_co2) - exp_co2) < 0.1:
                    co2_correct += 1
                if ag_ch4 is not None and abs(float(ag_ch4) - exp_ch4) < 0.001 and \
                   ag_n2o is not None and abs(float(ag_n2o) - exp_n2o) < 0.001:
                    grams_div_correct += 1
                if ag_co2e is not None and abs(float(ag_co2e) - exp_co2e) < 0.1 and \
                   ag_mt is not None and abs(float(ag_mt) - exp_mt) < 0.1:
                    gwp_mt_correct += 1
            except (ValueError, TypeError):
                pass

        if total_rows > 0:
            if co2_correct / total_rows > 0.9:
                score += 15
                feedback_parts.append("CO2 calc correct (15/15)")
            else:
                feedback_parts.append(f"CO2 calc incorrect ({co2_correct}/{total_rows})")

            if grams_div_correct / total_rows > 0.9:
                score += 20
                feedback_parts.append("Grams to kg division correct (20/20)")
            else:
                feedback_parts.append("Grams to kg division incorrect (missed /1000 step?)")

            if gwp_mt_correct / total_rows > 0.9:
                score += 25
                feedback_parts.append("GWP and MT conversion correct (25/25)")
            else:
                feedback_parts.append("GWP or MT conversion incorrect")

        # 4. Check SUMIFs in Summary sheet
        if ws_summary:
            found_scope1 = False
            found_scope2 = False
            fac_matches = 0
            
            # Since layout isn't strictly enforced, scan for the values near the labels
            for row in range(1, ws_summary.max_row + 1):
                for col in range(1, ws_summary.max_column + 1):
                    val = ws_summary.cell(row=row, column=col).value
                    if isinstance(val, str):
                        # Scope checks
                        if 'Scope 1' in val:
                            adj = ws_summary.cell(row=row, column=col+1).value
                            if adj is not None and isinstance(adj, (int, float)) and abs(float(adj) - expected_scopes['Scope 1']) < 1.0:
                                found_scope1 = True
                        elif 'Scope 2' in val:
                            adj = ws_summary.cell(row=row, column=col+1).value
                            if adj is not None and isinstance(adj, (int, float)) and abs(float(adj) - expected_scopes['Scope 2']) < 1.0:
                                found_scope2 = True
                        
                        # Facility checks
                        for fac_key in expected_facilities.keys():
                            if fac_key in val:
                                adj = ws_summary.cell(row=row, column=col+1).value
                                if adj is not None and isinstance(adj, (int, float)) and abs(float(adj) - expected_facilities[fac_key]) < 1.0:
                                    fac_matches += 1

            if found_scope1 and found_scope2:
                score += 15
                feedback_parts.append("Scope summary correct (15/15)")
            else:
                feedback_parts.append("Scope summary missing or incorrect")

            if fac_matches >= 4:  # Allow 1 minor miss due to formatting
                score += 15
                feedback_parts.append(f"Facility summary correct (15/15)")
            else:
                feedback_parts.append(f"Facility summary incorrect ({fac_matches}/5 matched)")
        else:
            feedback_parts.append("Summary sheet NOT found")

        # 5. VLM Chart Check (Pie or Doughnut Chart)
        query_vlm = env_info.get('query_vlm')
        if query_vlm:
            try:
                frames = sample_trajectory_frames(traj, n=3)
                final = get_final_screenshot(traj)
                images = frames + [final] if final else frames
                
                vlm_res = query_vlm(
                    images=images,
                    prompt="""Look at these screenshots of a spreadsheet. 
Did the user insert a Pie Chart or Doughnut Chart summarizing the data (e.g., Scope 1 vs Scope 2 emissions)? 
Respond strictly in JSON format:
{"has_chart": true} or {"has_chart": false}"""
                )
                
                if vlm_res and vlm_res.get('parsed', {}).get('has_chart', False):
                    score += 10
                    feedback_parts.append("Visual chart detected (10/10)")
                else:
                    feedback_parts.append("Visual chart NOT detected")
            except Exception as e:
                feedback_parts.append(f"VLM check failed: {e}")
        else:
            feedback_parts.append("VLM not available")

        # Final pass requires math thresholds (Critical fail if missing gram conversion)
        passed = score >= 80 and (grams_div_correct / max(total_rows, 1) > 0.9)
        if not passed and score >= 80:
            feedback_parts.append("CRITICAL: Grams to kg division mandatory step failed.")

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {str(e)}"}
    finally:
        safe_cleanup(temp_dir)
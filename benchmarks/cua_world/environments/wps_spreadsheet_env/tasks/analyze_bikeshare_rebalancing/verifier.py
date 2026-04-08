#!/usr/bin/env python3
"""Verifier for analyze_bikeshare_rebalancing task."""

import json
import os
import sys
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_bikeshare_rebalancing(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Read export metadata
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_meta = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result_meta.get("file_modified", False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "File was not modified during the task."
        }

    # Fetch and parse the spreadsheet using data_only=True to evaluate formulas natively
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/bikeshare_q3_data.xlsx", temp_xlsx.name)
        
        try:
            from openpyxl import load_workbook
            wb_data = load_workbook(temp_xlsx.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
            
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    score = 0
    feedback_parts = []
    
    try:
        ws_trips = wb_data["Trips"]
        ws_stations = wb_data["Station_Summary"]
        ws_riders = wb_data["Rider_Summary"]
    except KeyError as e:
        return {"passed": False, "score": 0, "feedback": f"Missing required sheet: {e}"}

    # --- 1. Extract Ground Truth directly from the agent's edited Trips sheet ---
    trips_data = []
    duration_passed = False
    durations_calculated = 0

    for row in ws_trips.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        start_time = row[1]
        end_time = row[2]
        st_station = row[3]
        en_station = row[4]
        rtype = row[5]
        
        trips_data.append({
            'start_time': start_time,
            'end_time': end_time,
            'start_station': st_station,
            'end_station': en_station,
            'rider_type': rtype
        })
        
        # Check duration (Column G is index 6)
        if len(row) > 6 and row[6] is not None:
            if isinstance(row[6], (int, float)):
                durations_calculated += 1

    if durations_calculated >= len(trips_data) * 0.9: # Allow small margin
        duration_passed = True
        score += 15
        feedback_parts.append("Duration calculated correctly")
    else:
        feedback_parts.append(f"Duration calculation missing or incomplete ({durations_calculated}/{len(trips_data)})")

    # Re-calculate accurate ground truth
    gt_stations = {}
    for t in trips_data:
        st = t['start_station']
        en = t['end_station']
        if st not in gt_stations: gt_stations[st] = {'starts': 0, 'ends': 0}
        if en not in gt_stations: gt_stations[en] = {'starts': 0, 'ends': 0}
        gt_stations[st]['starts'] += 1
        gt_stations[en]['ends'] += 1

    for st, vals in gt_stations.items():
        vals['net'] = vals['ends'] - vals['starts']

    gt_riders = {'Casual': {'trips': 0, 'durations': []}, 'Member': {'trips': 0, 'durations': []}}
    for t in trips_data:
        rtype = t['rider_type']
        if rtype in gt_riders:
            gt_riders[rtype]['trips'] += 1
            if t['end_time'] and t['start_time']:
                try:
                    diff = t['end_time'] - t['start_time']
                    mins = diff.total_seconds() / 60.0
                    gt_riders[rtype]['durations'].append(mins)
                except:
                    pass

    for rtype in gt_riders:
        if gt_riders[rtype]['trips'] > 0:
            gt_riders[rtype]['avg_duration'] = sum(gt_riders[rtype]['durations']) / gt_riders[rtype]['trips']
        else:
            gt_riders[rtype]['avg_duration'] = 0

    # --- 2. Verify Station Aggregation & Action Logic & Sorting ---
    station_aggregation_passed = True
    action_correct = True
    sorted_correctly = True
    prev_net = -999999
    valid_station_rows = 0

    for row in ws_stations.iter_rows(min_row=2, values_only=True):
        st = row[0]
        if not st: continue
        starts = row[1]
        ends = row[2]
        net = row[3]
        action = row[4]
        
        valid_station_rows += 1
        
        if st in gt_stations:
            if starts != gt_stations[st]['starts']: station_aggregation_passed = False
            if ends != gt_stations[st]['ends']: station_aggregation_passed = False
            if net != gt_stations[st]['net']: station_aggregation_passed = False
        
        if net is not None and isinstance(net, (int, float)):
            if net < prev_net:
                sorted_correctly = False
            prev_net = net
            
            expected_action = "Monitor"
            if net <= -10: expected_action = "Add Bikes"
            elif net >= 10: expected_action = "Remove Bikes"
            
            if not action or str(action).strip().lower() != expected_action.lower():
                action_correct = False
        else:
            sorted_correctly = False
            action_correct = False

    if valid_station_rows == 0:
        station_aggregation_passed = False
        action_correct = False
        sorted_correctly = False

    if station_aggregation_passed:
        score += 25
        feedback_parts.append("Station aggregation accurate")
    else:
        feedback_parts.append("Station aggregation incorrect")
        
    if action_correct:
        score += 20
        feedback_parts.append("Rebalancing logic correct")
    else:
        feedback_parts.append("Rebalancing IF logic incorrect")

    if sorted_correctly and valid_station_rows > 1:
        score += 15
        feedback_parts.append("Stations sorted correctly")
    else:
        feedback_parts.append("Stations not sorted correctly by Net_Flow")

    # --- 3. Verify Rider Metrics ---
    rider_metrics_passed = True
    valid_rider_rows = 0
    
    for row in ws_riders.iter_rows(min_row=2, values_only=True):
        rtype = row[0]
        if rtype in gt_riders:
            valid_rider_rows += 1
            trips = row[1]
            avg_dur = row[2]
            
            if trips != gt_riders[rtype]['trips']:
                rider_metrics_passed = False
            
            if avg_dur is not None and isinstance(avg_dur, (int, float)):
                gt_dur = gt_riders[rtype]['avg_duration']
                if gt_dur > 0 and abs(avg_dur - gt_dur) > gt_dur * 0.05: # Allow 5% tolerance
                    rider_metrics_passed = False
            else:
                rider_metrics_passed = False

    if valid_rider_rows == 2 and rider_metrics_passed:
        score += 15
        feedback_parts.append("Rider metrics accurate")
    else:
        feedback_parts.append("Rider metrics incorrect")

    # --- 4. VLM Verification ---
    try:
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
        from wps_verification_utils import vlm_verify_screenshot
        
        vlm_result = vlm_verify_screenshot(env_info, traj, """
Analyze this WPS Spreadsheet screenshot. Answer in JSON:
{
    "shows_formulas_or_calculated_data": true/false,
    "shows_multiple_sheets_tabs": true/false
}
Does the spreadsheet show:
1. Calculated columns, sorting actions, or formula bars?
2. The sheet tabs (Trips, Station_Summary, Rider_Summary) visible at the bottom?
""")
        if vlm_result is not None:
            if vlm_result.get("shows_formulas_or_calculated_data"): score += 5
            if vlm_result.get("shows_multiple_sheets_tabs"): score += 5
            feedback_parts.append("VLM visual verification: completed")
        else:
            feedback_parts.append("VLM visual verification: skipped")
    except Exception as e:
        logger.warning(f"VLM verification error: {e}")
        feedback_parts.append("VLM visual verification: failed")

    # Ensure critical step passed
    key_criteria = station_aggregation_passed
    passed = score >= 60 and key_criteria
    
    if not key_criteria and score >= 60:
        feedback_parts.append("FAIL: Core requirement (Station Aggregation) not met")
        passed = False

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
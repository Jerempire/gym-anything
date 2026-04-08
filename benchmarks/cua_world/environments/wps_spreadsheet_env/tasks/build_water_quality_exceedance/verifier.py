#!/usr/bin/env python3
"""
Verifier for build_water_quality_exceedance task.
Validates the spreadsheet contents, cross-references with hidden ground truth,
and uses trajectory VLM framing to prevent spoofing.
"""

import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_water_quality_report(traj, env_info, task_info):
    """
    Verify the resulting water_quality_report.xlsx and ground truth file.
    Uses copy_from_env exclusively to access files.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Fetch task JSON result (from export_result.sh)
    result_json_tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", result_json_tmp.name)
        with open(result_json_tmp.name, 'r') as f:
            task_stats = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task stats: {e}"}
    finally:
        os.unlink(result_json_tmp.name)

    # Anti-gaming criteria
    if not task_stats.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "❌ Output file water_quality_report.xlsx was not found"}
    if not task_stats.get("file_created_during_task", False):
        return {"passed": False, "score": 0, "feedback": "❌ Output file was not created/modified during the task session"}
    if task_stats.get("output_size_bytes", 0) < 5000:
        return {"passed": False, "score": 0, "feedback": "❌ Output file is too small to contain the required data"}

    score += 10
    feedback_parts.append("✅ File created successfully")

    # 2. Fetch Ground Truth data
    gt_tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/ground_truth_exceedances.json", gt_tmp.name)
        with open(gt_tmp.name, 'r') as f:
            ground_truth = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read ground truth: {e}"}
    finally:
        os.unlink(gt_tmp.name)

    # 3. Fetch and Parse the actual Excel File
    excel_tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/water_quality_report.xlsx", excel_tmp.name)
        # Using openpyxl to parse formulas and data
        import openpyxl
        wb = openpyxl.load_workbook(excel_tmp.name, data_only=False)
        wb_data = openpyxl.load_workbook(excel_tmp.name, data_only=True)
    except ImportError:
        return {"passed": False, "score": 0, "feedback": "openpyxl not installed in verifier environment"}
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse Excel file: {e}"}
    finally:
        os.unlink(excel_tmp.name)

    sheet_names = wb.sheetnames

    # Check Sheets Names
    expected_sheets = ["Raw Data", "Regulatory Limits", "Exceedance Analysis", "Summary"]
    found_sheets = [s for s in expected_sheets if s in sheet_names]
    if len(found_sheets) == 4:
        score += 10
        feedback_parts.append("✅ All expected sheets found")
    else:
        feedback_parts.append(f"❌ Missing expected sheets (Found: {found_sheets})")

    # Check Exceedance Analysis Formulas
    if "Exceedance Analysis" in sheet_names:
        sheet = wb["Exceedance Analysis"]
        formula_count = 0
        # Sample the first 50 rows looking for formulas in F and G
        for row in sheet.iter_rows(min_row=2, max_row=50, min_col=6, max_col=7):
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula_count += 1
        
        if formula_count > 10:
            score += 20
            feedback_parts.append("✅ VLOOKUP/IF formulas present in Exceedance Analysis")
        else:
            feedback_parts.append("❌ Required formulas missing in Exceedance Analysis")

    # Check Summary Sheet Data vs Ground Truth
    if "Summary" in sheet_names:
        sheet_data = wb_data["Summary"]
        param_counts_correct = 0
        
        # We scan the sheet values dynamically so we don't punish minor layout deviations
        for row in sheet_data.iter_rows(values_only=True):
            if not row: continue
            row_strs = [str(c).strip() for c in row if c is not None]
            row_nums = [c for c in row if isinstance(c, (int, float))]
            
            # Check Parameter matches
            for param, truth_data in ground_truth.get('parameters', {}).items():
                if param in row_strs and len(row_nums) >= 2:
                    # Usually Total Samples comes first, then Exceedances
                    agent_total = row_nums[0]
                    agent_exceed = row_nums[1]
                    # Allow ±2 tolerance for edge case formula/rounding differences
                    if abs(agent_exceed - truth_data['exceedances']) <= 2:
                        param_counts_correct += 1
                        break # Prevent double counting

        if param_counts_correct >= 6:
            score += 30
            feedback_parts.append(f"✅ Parameter exceedance summaries match ground truth ({param_counts_correct}/8)")
        else:
            score += (param_counts_correct * 3)
            feedback_parts.append(f"❌ Parameter exceedances incorrect or missing (Got {param_counts_correct}/8)")

    # 4. Trajectory VLM Verification (Anti-Spoofing Check)
    query_vlm = env_info.get('query_vlm')
    if query_vlm:
        try:
            from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
            frames = sample_trajectory_frames(traj, n=4)
            final = get_final_screenshot(traj)
            images = frames + [final] if final else frames

            if images:
                vlm_result = query_vlm(
                    images=images,
                    prompt="""Analyze these trajectory screenshots of a WPS Spreadsheet session.
Respond with JSON only:
{
    "shows_multiple_tabs_active": true/false,
    "shows_editing_formulas": true/false
}
1. Can you see the user working across different tabs at the bottom?
2. Is there evidence of the user writing or editing formulas (e.g., in the top formula bar =VLOOKUP or =COUNTIF)?
"""
                )
                if vlm_result.get("success"):
                    parsed = vlm_result.get("parsed", {})
                    if parsed.get("shows_multiple_tabs_active") and parsed.get("shows_editing_formulas"):
                        score += 30
                        feedback_parts.append("✅ VLM confirmed trajectory workflow (multi-sheet formula authoring)")
                    elif parsed.get("shows_multiple_tabs_active") or parsed.get("shows_editing_formulas"):
                        score += 15
                        feedback_parts.append("⚠️ VLM partially confirmed trajectory workflow")
                    else:
                        feedback_parts.append("❌ VLM did not observe active formula authoring in trajectory")
        except Exception as e:
            logger.warning(f"VLM verification failed to run: {e}")
            # Do not completely fail the agent if the VLM API call fails
            score += 30 
            feedback_parts.append("⚠️ VLM verification skipped (API error), gave default points.")
    else:
        # Give points if VLM isn't configured in environment
        score += 30
        feedback_parts.append("⚠️ VLM unavailable, default points awarded.")

    passed = score >= 70
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
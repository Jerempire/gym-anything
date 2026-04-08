#!/usr/bin/env python3
"""
Verifier for nyc_restaurant_inspection_audit task.
Checks for cross-sheet formula writing (VLOOKUP, SUMIFS, IF logic) and verifies computed values.
"""

import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_nyc_audit(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Fetch metadata result (detect file modification)
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result JSON: {e}"}
    finally:
        if os.path.exists(temp_res.name):
            os.unlink(temp_res.name)

    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Target spreadsheet file missing."}
    
    if not result.get('file_modified'):
        return {"passed": False, "score": 0, "feedback": "File not modified since task started. Did you save?"}

    # 2. Extract workbook and parse
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/nyc_inspections_audit.xlsx", temp_xlsx.name)
        
        try:
            from openpyxl import load_workbook
            # Load twice: one for retrieving raw formula text, one for verifying cached evaluated values
            wb_formulas = load_workbook(temp_xlsx.name, data_only=False)
            wb_values = load_workbook(temp_xlsx.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to parse Excel file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    score = 0
    feedback = []

    try:
        ws_log_f = wb_formulas['Violations_Log']
        ws_log_v = wb_values['Violations_Log']
        ws_roll_f = wb_formulas['Inspection_Rollup']
        ws_roll_v = wb_values['Inspection_Rollup']
    except KeyError as e:
        return {"passed": False, "score": 0, "feedback": f"Missing expected sheet: {e}"}

    def check_formula(sheet, col, any_keywords=None, all_keywords=None):
        """Scans the first few data rows to detect required formula footprints."""
        for row in range(2, min(10, sheet.max_row + 1)):
            val = str(sheet.cell(row=row, column=col).value).upper().strip()
            if not val.startswith('='):
                continue
                
            any_pass = any(kw.upper() in val for kw in any_keywords) if any_keywords else True
            all_pass = all(kw.upper() in val for kw in all_keywords) if all_keywords else True
                
            if any_pass and all_pass:
                return True
        return False

    # Scoring Criteria 1: Base Fine (VLOOKUP) - 15 points
    if check_formula(ws_log_f, 7, any_keywords=['VLOOKUP', 'XLOOKUP', 'INDEX']):
        score += 15
        feedback.append("Base Fine formula OK")
    else:
        feedback.append("Base Fine formula missing or incorrect")

    # Scoring Criteria 2: Adjusted Fine (IF multiplier) - 15 points
    if check_formula(ws_log_f, 8, all_keywords=['IF'], any_keywords=['1.5', '150%', '1.50']):
        score += 15
        feedback.append("Adjusted Fine logic OK")
    else:
        feedback.append("Adjusted Fine formula missing or incorrect")

    # Scoring Criteria 3: Total Points (SUMIFS) - 20 points
    if check_formula(ws_roll_f, 4, any_keywords=['SUMIFS', 'SUMPRODUCT']):
        score += 20
        feedback.append("Total Points aggregation OK")
    else:
        feedback.append("Total Points aggregation missing or incorrect")

    # Scoring Criteria 4: Total Fines (SUMIFS) - 15 points
    if check_formula(ws_roll_f, 5, any_keywords=['SUMIFS', 'SUMPRODUCT']):
        score += 15
        feedback.append("Total Fines aggregation OK")
    else:
        feedback.append("Total Fines aggregation missing or incorrect")

    # Scoring Criteria 5: Projected Grade (IF) - 15 points
    if check_formula(ws_roll_f, 6, any_keywords=['IF', 'IFS']):
        score += 15
        feedback.append("Projected Grade logic OK")
    else:
        feedback.append("Projected Grade formula missing")

    # Scoring Criteria 6: Audit Action (IF & OR/>500) - 10 points
    if check_formula(ws_roll_f, 7, all_keywords=['IF'], any_keywords=['>500', '500']):
        score += 10
        feedback.append("Audit Action logic OK")
    else:
        feedback.append("Audit Action formula missing")

    # Scoring Criteria 7: Formatting, Correctness & Saving - 10 points
    # Ground truth for row 2 (RIVIERA CATERERS): Expected 14 Pts, $900 Total Fines
    try:
        pts = ws_roll_v.cell(row=2, column=4).value
        fines = ws_roll_v.cell(row=2, column=5).value
        grade = str(ws_roll_v.cell(row=2, column=6).value).upper().strip()
        action = str(ws_roll_v.cell(row=2, column=7).value).upper().strip()

        if pts == 14 and fines == 900:
            score += 5
            feedback.append("Computed numerical values verified")
        else:
            feedback.append(f"Value mismatch: Expected 14 pts, 900 fines; got {pts}, {fines}")
            
        if grade == "B" and action == "URGENT REVIEW":
            score += 5
            feedback.append("Computed classification matches expected")
        else:
            feedback.append(f"Classification mismatch: Expected B, Urgent Review; got {grade}, {action}")
            
    except Exception as e:
        feedback.append("Could not verify computed values (was the file saved in WPS?)")

    # VLM Anti-gaming Verification
    query_vlm = env_info.get('query_vlm')
    if query_vlm:
        from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
        frames = sample_trajectory_frames(traj, n=3)
        final = get_final_screenshot(traj)
        images = frames + [final] if final else frames
        
        prompt = """
        Review these screenshots from a WPS Spreadsheet session.
        Did the user actively use the application and edit spreadsheet cells or formulas (e.g., VLOOKUP, SUMIFS, IF)?
        Answer in JSON format: {"used_spreadsheet": true/false}
        """
        vlm_res = query_vlm(images=images, prompt=prompt)
        if vlm_res and vlm_res.get('success'):
            if vlm_res.get('parsed', {}).get('used_spreadsheet'):
                feedback.append("VLM: Confirmed formula authoring")
            else:
                feedback.append("VLM: Did not clearly detect spreadsheet usage")
                score = max(0, score - 20)  # Heavy penalization for suspected spoofing

    passed = score >= 70
    
    return {
        "passed": passed,
        "score": min(100, score),
        "feedback": " | ".join(feedback)
    }
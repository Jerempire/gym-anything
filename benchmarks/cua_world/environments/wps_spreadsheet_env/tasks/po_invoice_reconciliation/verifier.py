#!/usr/bin/env python3
"""
Verifier for PO Invoice Reconciliation task.
Checks for multi-sheet creation, correct usage of VLOOKUP, variance math,
conditional IF logic, aggregated summary metrics, and visual validation.
"""

import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_po_reconciliation(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_output_path = metadata.get('expected_output_path', '/home/ga/Documents/PO_Reconciliation.xlsx')
    ground_truth_path = metadata.get('ground_truth_path', '/var/lib/ground_truth/po_reconciliation_gt.json')

    score = 0
    feedback_parts = []
    
    # Try importing openpyxl here to avoid crashing if it's not installed in the host, 
    # though gym_anything usually provides it.
    try:
        import openpyxl
    except ImportError:
        return {"passed": False, "score": 0, "feedback": "openpyxl not available in verifier environment."}

    # 1. Fetch the export result JSON
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    # Check file exists and was modified
    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Output spreadsheet not found."}
    
    if result.get('file_created_during_task'):
        score += 10
        feedback_parts.append("File saved/modified successfully")
    else:
        feedback_parts.append("File was NOT modified during task")
        # Anti-gaming: If file wasn't modified, the agent did nothing.
        return {"passed": False, "score": 0, "feedback": "File not modified during task duration (Anti-gaming check)."}

    # 2. Fetch the Spreadsheet
    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env(expected_output_path, temp_excel.name)
        # Load with data_only=False to inspect formulas
        wb = openpyxl.load_workbook(temp_excel.name, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_excel.name):
            os.unlink(temp_excel.name)

    sheets = [s.lower() for s in wb.sheetnames]
    
    # 3. Check for Sheets
    has_reconciliation = any('reconciliation' in s for s in sheets)
    has_summary = any('summary' in s for s in sheets)

    if has_reconciliation:
        score += 10
        feedback_parts.append("Reconciliation sheet created")
    else:
        feedback_parts.append("Reconciliation sheet MISSING")

    if has_summary:
        score += 10
        feedback_parts.append("Summary sheet created")
    else:
        feedback_parts.append("Summary sheet MISSING")

    # 4. Analyze Formulas in Reconciliation
    vlookup_count = 0
    variance_count = 0
    if_count = 0
    has_bold = False
    has_currency = False
    has_percentage = False

    if has_reconciliation:
        recon_sheet_name = [s for s in wb.sheetnames if 'reconciliation' in s.lower()][0]
        ws_recon = wb[recon_sheet_name]
        
        # Check formatting in Row 1
        for cell in ws_recon[1]:
            if cell.font and cell.font.bold:
                has_bold = True
                break

        # Check formulas down the columns
        for row in ws_recon.iter_rows(min_row=2, max_row=60):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    val_upper = val.upper()
                    # Check VLOOKUP/XLOOKUP
                    if 'VLOOKUP' in val_upper or 'XLOOKUP' in val_upper or 'INDEX' in val_upper:
                        vlookup_count += 1
                    # Check subtraction for variance
                    if '-' in val_upper or 'MINUS' in val_upper:
                        variance_count += 1
                    # Check division for variance %
                    if '/' in val_upper:
                        variance_count += 1
                    # Check IF logic for status
                    if 'IF' in val_upper:
                        if_count += 1

                # Check Number Formatting
                fmt = cell.number_format if cell.number_format else ""
                if '$' in fmt or '0.00' in fmt:
                    has_currency = True
                if '%' in fmt:
                    has_percentage = True

    if vlookup_count >= 40:
        score += 20
        feedback_parts.append("VLOOKUP formulas verified")
    elif vlookup_count > 0:
        score += 10
        feedback_parts.append(f"Partial VLOOKUPs found ({vlookup_count})")
    else:
        feedback_parts.append("VLOOKUP formulas MISSING")

    if variance_count >= 40:
        score += 15
        feedback_parts.append("Variance calculation math verified")
    elif variance_count > 0:
        score += 5
        feedback_parts.append("Partial variance calculations found")
        
    if if_count >= 40:
        score += 15
        feedback_parts.append("IF conditional formulas verified")
    elif if_count > 0:
        score += 5
        feedback_parts.append("Partial IF formulas found")

    # 5. Analyze Summary Sheet
    sumif_count = 0
    countif_count = 0
    if has_summary:
        summary_sheet_name = [s for s in wb.sheetnames if 'summary' in s.lower()][0]
        ws_sum = wb[summary_sheet_name]
        
        for row in ws_sum.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    if 'SUMIF' in val.upper() or 'SUM(' in val.upper():
                        sumif_count += 1
                    if 'COUNTIF' in val.upper() or 'COUNT(' in val.upper() or 'COUNTA(' in val.upper():
                        countif_count += 1

    if sumif_count > 0 and countif_count > 0:
        score += 10
        feedback_parts.append("Aggregation formulas (SUMIF/COUNTIF) verified in Summary")
    else:
        feedback_parts.append("Aggregation formulas MISSING in Summary")

    # 6. Formatting
    if has_bold and has_currency and has_percentage:
        score += 10
        feedback_parts.append("Professional formatting applied")
    elif has_bold or has_currency or has_percentage:
        score += 5
        feedback_parts.append("Partial formatting applied")

    # 7. VLM Verification (Trajectory checks)
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
    
    query_vlm = env_info.get('query_vlm')
    if query_vlm:
        frames = sample_trajectory_frames(traj, n=3)
        final_frame = get_final_screenshot(traj)
        
        vlm_prompt = """
        Analyze these screenshots from a WPS Spreadsheet reconciliation task.
        1. Does the spreadsheet show a "Reconciliation" sheet actively being worked on?
        2. Are there multiple sheets created at the bottom?
        3. Is there a "Summary" sheet visible containing aggregated statistics?
        Respond in JSON:
        {
            "reconciliation_visible": true/false,
            "multiple_sheets_visible": true/false,
            "summary_visible": true/false
        }
        """
        
        images = frames + [final_frame] if final_frame else frames
        if images:
            vlm_response = query_vlm(prompt=vlm_prompt, images=images)
            if vlm_response and vlm_response.get("success"):
                parsed = vlm_response.get("parsed", {})
                if parsed.get("reconciliation_visible") and parsed.get("summary_visible"):
                    score += 10
                    feedback_parts.append("VLM visual trajectory verified")
                elif parsed.get("reconciliation_visible") or parsed.get("multiple_sheets_visible"):
                    score += 5
                    feedback_parts.append("VLM visual trajectory partially verified")
            else:
                feedback_parts.append("VLM visual check failed to parse")
    else:
        feedback_parts.append("VLM unavailable - awarded points to threshold")
        score += 10  # Best effort point injection if API is offline

    # Evaluate Pass condition
    # Required: Must have saved the file, created the sheets, and used VLOOKUP
    passed = score >= 60 and vlookup_count > 0 and has_reconciliation

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
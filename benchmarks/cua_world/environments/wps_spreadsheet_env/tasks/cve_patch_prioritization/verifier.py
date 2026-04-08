#!/usr/bin/env python3
import sys
import os
import json
import logging
import tempfile
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_cve_patch_prioritization(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read export result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result JSON: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Output file does not exist."}
    if not result.get('file_modified_during_task'):
        return {"passed": False, "score": 0, "feedback": "File was not modified during the task."}

    # 2. Read the Excel file
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/vuln_scan_results.xlsx", temp_xlsx.name)
        import openpyxl
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        wb_data = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read Excel file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    score = 0
    feedback_parts = []
    
    try:
        ws_asset = wb_data["Asset_Inventory"]
        ws_cve = wb_data["NVD_CVE_Info"]
        ws_scan_f = wb["Scan_Data"]
        ws_scan_d = wb_data["Scan_Data"]
    except KeyError as e:
        return {"passed": False, "score": 0, "feedback": f"Missing required sheet: {e}"}
        
    # Build dictionary reference data
    assets = {}
    for row in ws_asset.iter_rows(min_row=2, values_only=True):
        if row[0]:
            assets[row[0]] = {"Environment": row[2], "Impact": row[3], "Team_Owner": row[4]}
            
    cves = {}
    for row in ws_cve.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cves[row[0]] = {"Base_Score": float(row[2]) if row[2] is not None else 0.0, "Exploit_Maturity": row[4]}

    total_rows = 0
    correct_lookups = 0
    correct_math = 0
    correct_sla = 0
    has_formulas = False
    
    team_sla3_counts = {"Web": 0, "Database": 0, "Infrastructure": 0, "HR_IT": 0, "Finance": 0}

    # Evaluate Scan_Data rows
    for row_idx, (row_f, row_d) in enumerate(zip(ws_scan_f.iter_rows(min_row=2), ws_scan_d.iter_rows(min_row=2, values_only=True))):
        if len(row_d) < 3:
            continue
        srv_id = row_d[1]
        cve_id = row_d[2]
        
        if not srv_id or not cve_id:
            continue
            
        total_rows += 1
        
        env = row_d[3] if len(row_d) > 3 else None
        impact = row_d[4] if len(row_d) > 4 else None
        b_score = row_d[5] if len(row_d) > 5 else None
        maturity = row_d[6] if len(row_d) > 6 else None
        team = row_d[7] if len(row_d) > 7 else None
        adj_score = row_d[8] if len(row_d) > 8 else None
        sla_days = row_d[9] if len(row_d) > 9 else None
        
        # Check formulas to prevent hardcoded solutions
        for cell in row_f[3:10]:
            if cell.data_type == 'f':
                has_formulas = True
        
        expected_env = assets.get(srv_id, {}).get("Environment")
        expected_impact = assets.get(srv_id, {}).get("Impact")
        expected_team = assets.get(srv_id, {}).get("Team_Owner")
        expected_bscore = cves.get(cve_id, {}).get("Base_Score")
        expected_maturity = cves.get(cve_id, {}).get("Exploit_Maturity")
        
        if (env == expected_env and impact == expected_impact and team == expected_team and
            maturity == expected_maturity):
            if b_score is not None and expected_bscore is not None and math.isclose(float(b_score), float(expected_bscore), abs_tol=0.1):
                correct_lookups += 1
                
        env_mult = 1.2 if expected_env == "Prod" else 0.8 if expected_env == "Dev" else 1.0
        imp_mult = 1.5 if expected_impact == "High" else 1.0 if expected_impact == "Medium" else 0.5 if expected_impact == "Low" else 1.0
        
        expected_adj = min(10.0, float(expected_bscore or 0.0) * env_mult * imp_mult)
        
        if adj_score is not None:
            try:
                if math.isclose(float(adj_score), expected_adj, abs_tol=0.1):
                    correct_math += 1
            except ValueError:
                pass
            
        if expected_adj >= 9.0 or expected_maturity in ["Functional", "High"]:
            expected_sla = 3
        elif expected_adj >= 7.0:
            expected_sla = 14
        else:
            expected_sla = 30
            
        if sla_days is not None:
            try:
                if int(float(sla_days)) == expected_sla:
                    correct_sla += 1
            except ValueError:
                pass

        if expected_sla == 3 and expected_team in team_sla3_counts:
            team_sla3_counts[expected_team] += 1

    lookup_ratio = correct_lookups / total_rows if total_rows > 0 else 0
    math_ratio = correct_math / total_rows if total_rows > 0 else 0
    sla_ratio = correct_sla / total_rows if total_rows > 0 else 0
    
    if lookup_ratio > 0.9:
        score += 25
        feedback_parts.append("Lookups correct")
    else:
        feedback_parts.append(f"Lookups missing/incorrect ({lookup_ratio:.1%} correct)")

    if math_ratio > 0.9:
        score += 30
        feedback_parts.append("Adjusted_Score math correct (including caps)")
    else:
        feedback_parts.append(f"Adjusted_Score math missing/incorrect ({math_ratio:.1%} correct)")
        
    if sla_ratio > 0.9:
        score += 20
        feedback_parts.append("SLA_Days logic correct")
    else:
        feedback_parts.append(f"SLA_Days logic missing/incorrect ({sla_ratio:.1%} correct)")
        
    if has_formulas:
        score += 10
        feedback_parts.append("Used formulas")
    else:
        feedback_parts.append("Formulas not detected (hardcoded values?)")
        
    # Evaluate Summary Sheet
    summary_correct = False
    if "Remediation_Summary" in wb.sheetnames:
        ws_sum = wb_data["Remediation_Summary"]
        teams_found = 0
        correct_counts = 0
        
        for row in ws_sum.iter_rows(values_only=True):
            if row[0] in team_sla3_counts:
                teams_found += 1
                try:
                    if int(float(row[1])) == team_sla3_counts[row[0]]:
                        correct_counts += 1
                except:
                    pass
                    
        if teams_found >= 5 and correct_counts >= 5:
            summary_correct = True
            
    if summary_correct:
        score += 15
        feedback_parts.append("Remediation_Summary correct")
    else:
        feedback_parts.append("Remediation_Summary missing/incorrect")
        
    passed = score >= 70
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
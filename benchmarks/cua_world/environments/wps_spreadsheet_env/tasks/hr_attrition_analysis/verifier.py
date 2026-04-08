#!/usr/bin/env python3
"""
Verifier for hr_attrition_analysis task.
Checks for correct multi-sheet formulas, aggregations, conditional formatting, and charts.
"""

import sys
import os
import json
import logging
import tempfile
import math

# Add utilities path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from wps_verification_utils import copy_and_parse_spreadsheet, cleanup_verification_temp
from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def evaluate_ground_truth(wb):
    """Calculates the expected aggregation values directly from the actual EmployeeData sheet."""
    gt = {
        'roles': {},
        'satisfaction': {1: {'total': 0, 'attrited': 0}, 2: {'total': 0, 'attrited': 0}, 
                         3: {'total': 0, 'attrited': 0}, 4: {'total': 0, 'attrited': 0}}
    }
    
    if 'EmployeeData' not in wb.sheetnames:
        return gt
        
    ws = wb['EmployeeData']
    
    # Map column headers to indices
    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
    if 'JobRole' not in headers:
        return gt
        
    idx_role = headers['JobRole']
    idx_attr = headers['Attrition']
    idx_inc = headers['MonthlyIncome']
    idx_sat = headers['JobSatisfaction']
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
            
        role = str(row[idx_role]).strip()
        attr = str(row[idx_attr]).strip()
        
        try:
            inc = float(row[idx_inc])
        except (ValueError, TypeError):
            inc = 0.0
            
        try:
            sat = int(row[idx_sat])
        except (ValueError, TypeError):
            sat = 0
            
        # Tally Role
        if role not in gt['roles']:
            gt['roles'][role] = {'total': 0, 'attrited': 0, 'retained': 0, 'inc_attrited': [], 'inc_retained': []}
            
        gt['roles'][role]['total'] += 1
        if attr.lower() == 'yes':
            gt['roles'][role]['attrited'] += 1
            gt['roles'][role]['inc_attrited'].append(inc)
        else:
            gt['roles'][role]['retained'] += 1
            gt['roles'][role]['inc_retained'].append(inc)
            
        # Tally Satisfaction
        if sat in gt['satisfaction']:
            gt['satisfaction'][sat]['total'] += 1
            if attr.lower() == 'yes':
                gt['satisfaction'][sat]['attrited'] += 1

    return gt

def safe_float(val):
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def verify_attrition_dashboard(traj, env_info, task_info):
    """
    Main verification function.
    """
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read task metadata
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export JSON: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    file_modified = export_result.get('file_modified_during_task', False)
    
    # 2. Extract Spreadsheet
    # We load TWO versions: data_only=True (to read computed math) and data_only=False (for charts/CF)
    try:
        # Load with data_only=True
        temp_xlsx_path = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
        copy_from_env("/home/ga/Documents/hr_attrition_data.xlsx", temp_xlsx_path)
        
        import openpyxl
        wb_data = openpyxl.load_workbook(temp_xlsx_path, data_only=True)
        wb_structure = openpyxl.load_workbook(temp_xlsx_path, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if 'temp_xlsx_path' in locals() and os.path.exists(temp_xlsx_path):
            os.unlink(temp_xlsx_path)

    feedback = []
    score = 0
    max_score = 100

    if not file_modified:
        feedback.append("Anti-gaming check failed: File was not modified.")
        return {"passed": False, "score": 0, "feedback": " | ".join(feedback)}

    if 'Attrition_Summary' not in wb_data.sheetnames:
        feedback.append("Sheet 'Attrition_Summary' not found.")
        return {"passed": False, "score": 0, "feedback": " | ".join(feedback)}

    score += 10
    feedback.append("Found Attrition_Summary sheet")

    ws_data = wb_data['Attrition_Summary']
    ws_struct = wb_structure['Attrition_Summary']
    
    gt = evaluate_ground_truth(wb_data)

    # --- Verify Table 1 (Roles) ---
    expected_roles = [
        "Healthcare Representative", "Human Resources", "Laboratory Technician", 
        "Manager", "Manufacturing Director", "Research Director", 
        "Research Scientist", "Sales Executive", "Sales Representative"
    ]
    
    t1_math_correct = True
    t1_income_correct = True
    risk_flag_correct = True
    
    for row_idx, role in enumerate(expected_roles, start=2):
        # Read user values
        cell_role = str(ws_data.cell(row=row_idx, column=1).value).strip()
        u_total = safe_float(ws_data.cell(row=row_idx, column=2).value)
        u_attr = safe_float(ws_data.cell(row=row_idx, column=3).value)
        u_ret = safe_float(ws_data.cell(row=row_idx, column=4).value)
        u_rate = safe_float(ws_data.cell(row=row_idx, column=5).value)
        u_inc_a = safe_float(ws_data.cell(row=row_idx, column=6).value)
        u_inc_r = safe_float(ws_data.cell(row=row_idx, column=7).value)
        u_risk = str(ws_data.cell(row=row_idx, column=8).value).strip().lower()

        if cell_role != role:
            t1_math_correct = False
            continue
            
        g = gt['roles'].get(role, {})
        g_total = g.get('total', 0)
        g_attr = g.get('attrited', 0)
        g_ret = g.get('retained', 0)
        g_rate = g_attr / g_total if g_total > 0 else 0
        
        inc_a_list = g.get('inc_attrited', [])
        inc_r_list = g.get('inc_retained', [])
        g_inc_a = sum(inc_a_list)/len(inc_a_list) if inc_a_list else 0
        g_inc_r = sum(inc_r_list)/len(inc_r_list) if inc_r_list else 0

        # Math checks
        if u_total != g_total or u_attr != g_attr or u_ret != g_ret:
            t1_math_correct = False
        
        if u_rate is None or not math.isclose(u_rate, g_rate, abs_tol=0.01):
            t1_math_correct = False

        # Income checks
        if u_inc_r is None or not math.isclose(u_inc_r, g_inc_r, abs_tol=5.0):
            t1_income_correct = False
        # Allow blank/error/0 for empty attrited incomes
        if inc_a_list and (u_inc_a is None or not math.isclose(u_inc_a, g_inc_a, abs_tol=5.0)):
            t1_income_correct = False

        # IF Logic checks
        expected_risk = "high risk" if g_rate > 0.15 else "normal"
        if u_risk != expected_risk:
            risk_flag_correct = False

    if t1_math_correct:
        score += 20
        feedback.append("Table 1 Role Counts Correct")
    else:
        feedback.append("Table 1 Role Counts Incorrect")

    if t1_income_correct:
        score += 15
        feedback.append("Table 1 Avg Incomes Correct")
    else:
        feedback.append("Table 1 Avg Incomes Incorrect")

    if risk_flag_correct:
        score += 10
        feedback.append("Risk Flag Logic Correct")
    else:
        feedback.append("Risk Flag Logic Incorrect")

    # --- Verify Table 2 (Satisfaction) ---
    t2_math_correct = True
    for row_idx, sat in enumerate([1, 2, 3, 4], start=14):
        cell_sat = safe_float(ws_data.cell(row=row_idx, column=1).value)
        u_total = safe_float(ws_data.cell(row=row_idx, column=2).value)
        u_attr = safe_float(ws_data.cell(row=row_idx, column=3).value)
        u_rate = safe_float(ws_data.cell(row=row_idx, column=4).value)
        
        if cell_sat != sat:
            t2_math_correct = False
            continue
            
        g = gt['satisfaction'].get(sat, {})
        g_total = g.get('total', 0)
        g_attr = g.get('attrited', 0)
        g_rate = g_attr / g_total if g_total > 0 else 0
        
        if u_total != g_total or u_attr != g_attr:
            t2_math_correct = False
        if u_rate is None or not math.isclose(u_rate, g_rate, abs_tol=0.01):
            t2_math_correct = False

    if t2_math_correct:
        score += 15
        feedback.append("Table 2 Satisfaction Correct")
    else:
        feedback.append("Table 2 Satisfaction Incorrect")

    # --- Verify Conditional Formatting (Structure) ---
    has_cf = False
    if hasattr(ws_struct, 'conditional_formatting') and ws_struct.conditional_formatting:
        if ws_struct.conditional_formatting._cf_rules:
            has_cf = True
    
    if has_cf:
        score += 10
        feedback.append("Conditional Formatting Applied")
    else:
        feedback.append("No Conditional Formatting detected programmatically")

    # --- Verify Charts & VLM Check ---
    has_chart_obj = hasattr(ws_struct, '_charts') and len(ws_struct._charts) > 0
    if has_chart_obj:
        score += 10
        feedback.append("Chart Object detected in file")
    else:
        feedback.append("No Chart Object detected in file")

    # VLM verification to confirm visual presence of chart and conditional formatting
    vlm_score = 0
    if query_vlm:
        frames = sample_trajectory_frames(traj, n=3)
        final_frame = get_final_screenshot(traj)
        if final_frame:
            frames.append(final_frame)
            
        vlm_prompt = """
        Review these screenshots of WPS Spreadsheet. 
        The agent was tasked with building an HR Attrition Dashboard.
        1. Do you see a column chart visualizing Attrition Rate by Job Role?
        2. Do you see colored cell backgrounds (conditional formatting) applied to the Attrition Rate numbers?
        
        Respond in JSON:
        {
            "has_chart": true/false,
            "has_colored_cells": true/false
        }
        """
        
        vlm_resp = query_vlm(prompt=vlm_prompt, images=frames)
        if vlm_resp and vlm_resp.get("success"):
            parsed = vlm_resp.get("parsed", {})
            if parsed.get("has_chart", False):
                vlm_score += 5
                feedback.append("VLM confirmed Chart visibility")
            if parsed.get("has_colored_cells", False):
                vlm_score += 5
                if not has_cf: # Compensate if openpyxl failed to read CF
                    score += 10
                feedback.append("VLM confirmed Conditional Formatting visibility")
    
    score += vlm_score

    # Passing Threshold: Must have Table 1 math mostly correct and at least 65 points total
    passed = score >= 65 and t1_math_correct

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback)
    }
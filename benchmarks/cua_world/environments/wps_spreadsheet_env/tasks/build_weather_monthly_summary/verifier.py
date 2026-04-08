#!/usr/bin/env python3
"""Verify the weather monthly summary task."""

import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

def verify_weather_summary(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    score = 0
    feedback_parts = []
    
    # 1. Check Task Result output (Anti-gaming check)
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            task_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        os.unlink(temp_result.name)

    if not task_result.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "Output file not found"}

    if not task_result.get("file_modified_during_task", False):
        feedback_parts.append("WARNING: File was not modified during task execution")

    # 2. Read calculated ground truth
    temp_gt = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/ground_truth.json", temp_gt.name)
        with open(temp_gt.name, 'r') as f:
            gt = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Could not load ground truth: {e}"}
    finally:
        os.unlink(temp_gt.name)

    # 3. Read agent spreadsheet
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/central_park_weather_2023.xlsx", temp_xlsx.name)
        
        from openpyxl import load_workbook
        try:
            wb_formula = load_workbook(temp_xlsx.name, data_only=False)
            wb_values = load_workbook(temp_xlsx.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Invalid spreadsheet file: {e}"}
            
        if 'Monthly Summary' in wb_formula.sheetnames:
            score += 5
            feedback_parts.append("'Monthly Summary' sheet exists (+5)")
        else:
            return {"passed": False, "score": 0, "feedback": "'Monthly Summary' sheet not found"}

        ws_f = wb_formula['Monthly Summary']
        ws_v = wb_values['Monthly Summary']

        # Verify Headers
        expected_headers = ['Month', 'Avg High', 'Avg Low', 'Total Precip', 'Max Temp', 'Min Temp', 'Precip Days']
        actual_headers = [str(ws_v.cell(row=1, column=c).value or '') for c in range(1, 8)]
        header_matches = sum(1 for exp, act in zip(expected_headers, actual_headers) if exp.lower() in act.lower())
        if header_matches >= 5:
            score += 5
            feedback_parts.append(f"Headers correct (+5)")
        elif header_matches >= 3:
            score += 2
            feedback_parts.append(f"Headers partially correct (+2)")

        # Verify Month Columns
        months_found = 0
        for row in range(2, 14):
            cell_val = str(ws_v.cell(row=row, column=1).value or '')
            if any(m.lower() in cell_val.lower() for m in MONTH_NAMES) or cell_val.isdigit():
                months_found += 1
        if months_found >= 12:
            score += 5
            feedback_parts.append("All 12 months present (+5)")
        elif months_found >= 6:
            score += 2
            
        # Evaluation Helper
        def has_formula(row, col):
            val = ws_f.cell(row=row, column=col).value
            return isinstance(val, str) and val.startswith('=')

        def check_column(col_idx, gt_key, pts_formula, pts_value, tol=0.5):
            formula_count = 0
            value_matches = 0
            for month_num in range(1, 13):
                row = month_num + 1
                if has_formula(row, col_idx):
                    formula_count += 1
                val = ws_v.cell(row=row, column=col_idx).value
                expected = gt['monthly'].get(str(month_num), {}).get(gt_key)
                if val is not None and expected is not None:
                    try:
                        if abs(float(val) - expected) <= tol:
                            value_matches += 1
                    except (ValueError, TypeError):
                        pass
            
            pts = 0
            if formula_count >= 10: pts += pts_formula
            elif formula_count >= 6: pts += pts_formula // 2
            
            if value_matches >= 10: pts += pts_value
            elif value_matches >= 6: pts += pts_value // 2
            
            return pts, formula_count, value_matches

        # Column Verifications
        pts_high, f_high, v_high = check_column(2, 'avg_high', 5, 5, 0.5)
        score += pts_high
        feedback_parts.append(f"Avg High: {f_high} forms, {v_high} vals (+{pts_high})")

        pts_low, f_low, v_low = check_column(3, 'avg_low', 5, 5, 0.5)
        score += pts_low
        feedback_parts.append(f"Avg Low: {f_low} forms, {v_low} vals (+{pts_low})")

        pts_prcp, f_prcp, v_prcp = check_column(4, 'total_prcp', 5, 5, 0.05)
        score += pts_prcp
        feedback_parts.append(f"Total Precip: {f_prcp} forms, {v_prcp} vals (+{pts_prcp})")

        pts_max, f_max, v_max = check_column(5, 'max_temp', 3, 4, 0.5)
        score += pts_max
        feedback_parts.append(f"Max Temp: {f_max} forms, {v_max} vals (+{pts_max})")

        pts_min, f_min, v_min = check_column(6, 'min_temp', 3, 4, 0.5)
        score += pts_min
        feedback_parts.append(f"Min Temp: {f_min} forms, {v_min} vals (+{pts_min})")

        pts_pdays, f_pdays, v_pdays = check_column(7, 'prcp_days', 4, 4, 0.1)
        score += pts_pdays
        feedback_parts.append(f"Precip Days: {f_pdays} forms, {v_pdays} vals (+{pts_pdays})")

        # Annual Summary Assessment
        annual_pts = 0
        annual_found = False
        annual_start_row = None
        for r in range(14, 26):
            cell_val = str(ws_v.cell(row=r, column=1).value or '')
            if 'annual' in cell_val.lower() or 'summary' in cell_val.lower():
                annual_found = True
                annual_start_row = r
                break

        if annual_found:
            annual_pts += 2
            annual_values = {}
            for r in range(annual_start_row, annual_start_row + 8):
                label = str(ws_v.cell(row=r, column=1).value or '').lower()
                val = ws_v.cell(row=r, column=2).value
                if 'avg high' in label or 'average high' in label:
                    annual_values['avg_high'] = val
                elif 'avg low' in label or 'average low' in label:
                    annual_values['avg_low'] = val
                elif 'total' in label and 'precip' in label and 'day' not in label:
                    annual_values['total_prcp'] = val
                elif 'precip' in label and 'day' in label:
                    annual_values['total_prcp_days'] = val

            for key, expected in gt['annual'].items():
                if key in annual_values and annual_values[key] is not None:
                    try:
                        actual = float(annual_values[key])
                        tol = 0.5 if 'temp' in key or 'high' in key or 'low' in key else 0.05 if 'prcp' == key[-4:] else 0
                        if key == 'total_prcp_days' and int(actual) == expected:
                            annual_pts += 2
                        elif abs(actual - expected) <= tol:
                            annual_pts += 2
                    except (ValueError, TypeError):
                        pass

        score += annual_pts
        feedback_parts.append(f"Annual Summary (+{annual_pts})")

        # Cell Formatting Verifications
        format_pts = 0
        temp_formatted = 0
        prcp_formatted = 0
        for row in range(2, 14):
            for col in [2, 3, 5, 6]:
                fmt = ws_v.cell(row=row, column=col).number_format
                if fmt and ('0.0' in str(fmt) or '#.0' in str(fmt)):
                    temp_formatted += 1
            for col in [4]:
                fmt = ws_v.cell(row=row, column=col).number_format
                if fmt and ('0.00' in str(fmt) or '#.00' in str(fmt)):
                    prcp_formatted += 1

        if temp_formatted >= 24: format_pts += 3
        elif temp_formatted >= 12: format_pts += 1
        if prcp_formatted >= 6: format_pts += 2
        elif prcp_formatted >= 3: format_pts += 1
        
        score += format_pts
        feedback_parts.append(f"Formatting (+{format_pts})")

        # Overall Structure & Formula usage Ratio Verification
        score += 3
        
        total_data_cells = 0
        formula_cells = 0
        for row in range(2, 14):
            for col in range(2, 8):
                val = ws_f.cell(row=row, column=col).value
                if val is not None:
                    total_data_cells += 1
                    if isinstance(val, str) and val.startswith('='):
                        formula_cells += 1
        
        if total_data_cells > 0 and (formula_cells / total_data_cells) >= 0.7:
            score += 5
            feedback_parts.append("Formulas used predominantly (+5)")

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {e}"}
    finally:
        os.unlink(temp_xlsx.name)

    # 4. Trajectory Pattern VLM Verification (Anti-gaming & Progress Tracking)
    vlm_score = 0
    query_vlm = env_info.get('query_vlm')
    if query_vlm:
        from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
        frames = sample_trajectory_frames(traj, n=4)
        final = get_final_screenshot(traj)
        if frames and final:
            vlm_res = query_vlm(
                prompt="""Analyze these screenshots of a user working in WPS Spreadsheet.
Did the user actively create the 'Monthly Summary' sheet and type formulas (like AVERAGEIF, SUMIF) to calculate the monthly and annual statistics?
Look for evidence of workflow progression (e.g., empty sheet -> typing headers -> entering formulas -> formatting -> final result).
Reply with ONLY a JSON object: {"workflow_observed": true/false}""",
                images=frames + [final]
            )
            if vlm_res and vlm_res.get('parsed', {}).get('workflow_observed', False):
                vlm_score = 10
                feedback_parts.append("VLM: Workflow progression observed (+10)")
            else:
                feedback_parts.append("VLM: Workflow progression NOT observed")
    
    score += vlm_score
    passed = score >= 60

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
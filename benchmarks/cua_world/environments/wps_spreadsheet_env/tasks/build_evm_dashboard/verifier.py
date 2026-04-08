#!/usr/bin/env python3
"""
Verifier for build_evm_dashboard task.
Evaluates both the spreadsheet calculations and VLM visual output.
"""

import os
import json
import logging
import tempfile
import re

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def parse_number(val):
    """Helper to extract numbers from strings like '$2,400,000.00'."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        clean = re.sub(r'[^\d\.-]', '', val)
        if clean and clean != '-' and clean != '.':
            try:
                return float(clean)
            except ValueError:
                pass
    return None

def has_target_value(numbers, target, is_ratio=False):
    """Check if the target value exists in the extracted numbers array within tolerance."""
    if is_ratio:
        tol = 0.05  # Generous tolerance for index rounding (SPI/CPI)
    else:
        tol = max(abs(target * 0.03), 10.0)  # 3% tolerance for dollar amounts
        
    for n in numbers:
        if abs(n - target) <= tol:
            return True
    return False

def verify_evm_dashboard(traj, env_info, task_info):
    """Verify EVM Dashboard implementation."""
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available."}

    # Extract metadata metrics
    expected_metrics = task_info.get('metadata', {}).get('expected_metrics', {})
    
    # 1. Read JSON export metadata
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_meta = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    # 2. Copy the actual spreadsheet (twice, one for formulas, one for data)
    temp_xlsx_data = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_xlsx_formulas = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env("/home/ga/Documents/project_evm_data.xlsx", temp_xlsx_data.name)
        copy_from_env("/home/ga/Documents/project_evm_data.xlsx", temp_xlsx_formulas.name)
        
        try:
            import openpyxl
        except ImportError:
            return {"passed": False, "score": 0, "feedback": "openpyxl not available."}
            
        # Load workbook
        wb_data = openpyxl.load_workbook(temp_xlsx_data.name, data_only=True)
        wb_formulas = openpyxl.load_workbook(temp_xlsx_formulas.name, data_only=False)
        
        sheet_names = [s.lower() for s in wb_data.sheetnames]
        
        dashboard_sheet_name = None
        for name in wb_data.sheetnames:
            if 'evm_dashboard' in name.lower() or 'dashboard' in name.lower():
                dashboard_sheet_name = name
                break
                
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_xlsx_data.name):
            os.unlink(temp_xlsx_data.name)
        if os.path.exists(temp_xlsx_formulas.name):
            os.unlink(temp_xlsx_formulas.name)

    score = 0
    feedback_parts = []
    
    # Anti-gaming: Check if file was actually modified
    if not result_meta.get('file_modified_during_task', False):
        return {"passed": False, "score": 0, "feedback": "Anti-gaming: Spreadsheet was not modified during the task."}

    if not dashboard_sheet_name:
        return {
            "passed": False,
            "score": 0,
            "feedback": "Failed to find a sheet named 'EVM_Dashboard'."
        }

    score += 5
    feedback_parts.append("EVM_Dashboard sheet found")

    ws_data = wb_data[dashboard_sheet_name]
    ws_formulas = wb_formulas[dashboard_sheet_name]
    
    # Extract all numbers from the dashboard
    dashboard_numbers = []
    for row in ws_data.iter_rows(values_only=True):
        for cell in row:
            num = parse_number(cell)
            if num is not None:
                dashboard_numbers.append(num)

    # Validate Formulas vs Hardcoding
    formula_count = 0
    for row in ws_formulas.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                
    if formula_count < 3:
        feedback_parts.append("WARNING: Few formulas detected (possible hardcoding)")
        # Cap score heavily if they just typed the answers without referencing
        max_possible_score = 40
    else:
        max_possible_score = 100

    # Verification Weights
    metric_weights = {
        "EV": 15, "AC": 10, "PV": 10, "BAC": 5, "SV": 5, "CV": 5, 
        "SPI": 8, "CPI": 8, "EAC": 7, "ETC": 5, "VAC": 5, "TCPI": 5
    }
    
    # Check Math Output
    metrics_passed = 0
    for key, weight in metric_weights.items():
        target = expected_metrics.get(key)
        is_ratio = key in ["SPI", "CPI", "TCPI"]
        
        if target is not None and has_target_value(dashboard_numbers, target, is_ratio):
            score += weight
            metrics_passed += 1
        else:
            feedback_parts.append(f"Missing/Incorrect {key}")

    if metrics_passed == len(metric_weights):
        feedback_parts.append("All EVM calculations correct")
    else:
        feedback_parts.append(f"{metrics_passed}/{len(metric_weights)} calculations correct")

    # Chart Verification (OpenPyxl programmatic + VLM fallback)
    has_chart = len(ws_data._charts) > 0
    chart_verified = False

    if has_chart:
        chart_verified = True
        feedback_parts.append("Chart object found programmatically")
    elif query_vlm:
        # Fallback to VLM if openpyxl misses a chart object
        frames = sample_trajectory_frames(traj, n=3)
        final_img = get_final_screenshot(traj)
        if final_img:
            frames.append(final_img)
            
        vlm_prompt = """Look at these screenshots of a WPS Spreadsheet task.
        Is there a visual line chart (S-Curve) plotted on the spreadsheet showing multiple intersecting lines over time (like EV, PV, AC)?
        Respond with valid JSON only: {"chart_visible": true/false}"""
        
        vlm_result = query_vlm(images=frames, prompt=vlm_prompt)
        if vlm_result and vlm_result.get('parsed', {}).get('chart_visible', False):
            chart_verified = True
            feedback_parts.append("Chart found via VLM visual check")
            
    if chart_verified:
        score += 7
    else:
        feedback_parts.append("S-Curve chart not detected")
        
    # Enforce anti-cheat cap
    if score > max_possible_score:
        score = max_possible_score
        feedback_parts.append(f"Score capped at {max_possible_score} due to lack of formulas")

    # Evaluate Pass condition
    passed = score >= 60 and metrics_passed >= 4

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
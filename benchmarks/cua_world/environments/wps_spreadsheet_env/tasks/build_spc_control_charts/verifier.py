#!/usr/bin/env python3
"""Verifier for build_spc_control_charts task."""

import sys
import os
import json
import logging
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from wps_verification_utils import copy_and_parse_spreadsheet, cleanup_verification_temp
from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_numbers(sheet):
    """Extract all numeric values from a sheet to check for calculated stats."""
    nums = []
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, (int, float)):
                nums.append(float(cell))
            elif isinstance(cell, str):
                try:
                    nums.append(float(cell.strip()))
                except ValueError:
                    pass
    return nums

def has_value(nums, target, tol=0.005):
    """Check if target value exists in list within tolerance."""
    return any(abs(n - target) <= tol for n in nums)

def verify_spc_charts(traj, env_info, task_info):
    """
    Verify SPC charts generation logic.
    Uses multi-criteria scoring to evaluate spreadsheet formulas, values, charts, and visual progression.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    
    # Check modification status
    try:
        temp_res = os.path.join(tempfile.gettempdir(), 'task_result.json')
        copy_from_env("/tmp/task_result.json", temp_res)
        with open(temp_res, 'r') as f:
            res_data = json.load(f)
        if res_data.get('file_modified'):
            score += 5
            feedback_parts.append("File correctly saved/modified")
        else:
            feedback_parts.append("File was NOT saved during task")
        os.unlink(temp_res)
    except Exception as e:
        logger.warning(f"Could not read task_result.json: {e}")

    # Copy spreadsheet from container
    success, wb, error, temp_dir = copy_and_parse_spreadsheet(
        "/home/ga/Documents/piston_ring_measurements.xlsx", copy_from_env, file_format='xlsx'
    )

    if not success:
        return {"passed": False, "score": score, "feedback": f"Failed to open spreadsheet: {error}"}

    try:
        # Re-load with data_only=True to evaluate cached formula outputs (WPS saves them natively)
        import openpyxl
        temp_file = os.path.join(temp_dir, 'input.xlsx')
        wb_values = openpyxl.load_workbook(temp_file, data_only=True)
        sheet_names = wb_values.sheetnames

        # Verify Raw Data Calculations
        if "Raw Data" not in sheet_names:
            feedback_parts.append("Missing 'Raw Data' sheet")
        else:
            ws_raw = wb_values["Raw Data"]
            exp_xbars, exp_rs = [], []
            
            # Dynamically calculate ground truth from actual sheet data (robust against minor starting data edits)
            for row in ws_raw.iter_rows(min_row=2, max_row=26, min_col=2, max_col=6, values_only=True):
                vals = [v for v in row if isinstance(v, (int, float))]
                if len(vals) == 5:
                    exp_xbars.append(sum(vals)/5.0)
                    exp_rs.append(max(vals) - min(vals))
            
            agent_xbars, agent_rs = [], []
            for row in ws_raw.iter_rows(min_row=2, max_row=26, min_col=7, max_col=8, values_only=True):
                agent_xbars.append(row[0] if isinstance(row[0], (int, float)) else None)
                agent_rs.append(row[1] if isinstance(row[1], (int, float)) else None)
            
            # Score Xbar column (G)
            if len(agent_xbars) >= 25 and all(a is not None and abs(a-e) < 0.005 for a, e in zip(agent_xbars[:25], exp_xbars)):
                score += 15
                feedback_parts.append("X-bar values correct")
            else:
                feedback_parts.append("X-bar column incorrect")
                
            # Score R column (H)
            if len(agent_rs) >= 25 and all(a is not None and abs(a-e) < 0.005 for a, e in zip(agent_rs[:25], exp_rs)):
                score += 15
                feedback_parts.append("R values correct")
            else:
                feedback_parts.append("R column incorrect")

            # Verify Limit Sheets and Summaries using Ground Truth
            if len(exp_xbars) > 0:
                x_dbl_bar = sum(exp_xbars) / len(exp_xbars)
                r_bar = sum(exp_rs) / len(exp_rs)
                ucl_xbar = x_dbl_bar + 0.577 * r_bar
                lcl_xbar = x_dbl_bar - 0.577 * r_bar
                ucl_r = 2.114 * r_bar
                
                sigma_hat = r_bar / 2.326
                cp = (74.035 - 73.965) / (6 * sigma_hat)
                cpk = min((74.035 - x_dbl_bar)/(3*sigma_hat), (x_dbl_bar - 73.965)/(3*sigma_hat))
                
                # Check Control Limits sheet
                if "Control Limits" in sheet_names:
                    cl_nums = extract_numbers(wb_values["Control Limits"])
                    targets = [x_dbl_bar, r_bar, ucl_xbar, lcl_xbar, ucl_r]
                    found = sum(1 for t in targets if has_value(cl_nums, t, 0.005))
                    if found == 5:
                        score += 15
                        feedback_parts.append("Control limits formulas correct")
                    else:
                        score += found * 3
                        feedback_parts.append(f"Control limits partial ({found}/5)")
                else:
                    feedback_parts.append("Missing 'Control Limits' sheet")
                    
                # Check Summary sheet
                if "Summary" in sheet_names:
                    sum_nums = extract_numbers(wb_values["Summary"])
                    c_found = 0
                    if has_value(sum_nums, cp, 0.05): c_found += 1
                    if has_value(sum_nums, cpk, 0.05): c_found += 1
                    
                    if c_found == 2:
                        score += 15
                        feedback_parts.append("Cp and Cpk correct")
                    else:
                        score += c_found * 7
                        feedback_parts.append(f"Cp/Cpk partial ({c_found}/2)")
                else:
                    feedback_parts.append("Missing 'Summary' sheet")

        # Check for presence of Chart objects
        charts_found = 0
        if "X-bar Chart" in sheet_names and len(wb_values["X-bar Chart"]._charts) > 0:
            charts_found += 1
        if "R Chart" in sheet_names and len(wb_values["R Chart"]._charts) > 0:
            charts_found += 1
            
        if charts_found == 2:
            score += 15
            feedback_parts.append("Dedicated Chart sheets created")
        elif charts_found == 1:
            score += 7
            feedback_parts.append("One chart sheet missing/empty")
        else:
            # Fallback if placed on main data sheet
            if "Raw Data" in sheet_names and len(wb_values["Raw Data"]._charts) >= 2:
                score += 10
                feedback_parts.append("Charts found on Raw Data instead of dedicated sheets")
            else:
                feedback_parts.append("No charts found")

        # VLM Trajectory Verification
        query_vlm = env_info.get('query_vlm')
        if query_vlm and traj:
            try:
                frames = sample_trajectory_frames(traj, n=3)
                final = get_final_screenshot(traj)
                images = frames + [final] if final else frames
                
                prompt = """
Analyze these screenshots of a spreadsheet task workflow.
Respond in JSON format:
{
    "shows_charts": true/false,
    "shows_control_limit_lines": true/false,
    "shows_statistical_workflow": true/false
}
Does the visual workflow indicate:
1. Line charts being actively created or viewed?
2. Line charts that include distinct horizontal reference lines (Control Limits)?
3. The user interacting with process capability or statistical calculations?
"""
                vlm_response = query_vlm(prompt=prompt, images=images)
                
                parsed = {}
                if vlm_response and vlm_response.get("success"):
                    if vlm_response.get("parsed"):
                        parsed = vlm_response["parsed"]
                    elif vlm_response.get("text"):
                        match = re.search(r'\{.*\}', vlm_response["text"], re.DOTALL)
                        if match:
                            parsed = json.loads(match.group(0))
                
                vlm_pts = 0
                if parsed.get("shows_charts"): vlm_pts += 5
                if parsed.get("shows_control_limit_lines"): vlm_pts += 10
                if parsed.get("shows_statistical_workflow"): vlm_pts += 5
                
                score += vlm_pts
                feedback_parts.append(f"VLM Visual check: +{vlm_pts}/20")
            except Exception as e:
                feedback_parts.append(f"VLM error: {e}")
        else:
            feedback_parts.append("VLM verification skipped")

        passed = score >= 60

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": score, "feedback": f"Error during verification: {str(e)}"}
    finally:
        cleanup_verification_temp(temp_dir)
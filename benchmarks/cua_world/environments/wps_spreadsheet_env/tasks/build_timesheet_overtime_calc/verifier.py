#!/usr/bin/env python3
"""Verifier for build_timesheet_overtime_calc task."""

import os
import sys
import json
import logging
import tempfile
import shutil

# Ensure we can import wps_verification_utils if it exists in the environment structure
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from wps_verification_utils import (
        copy_and_parse_spreadsheet,
        cleanup_verification_temp
    )
    WPS_UTILS_AVAILABLE = True
except ImportError:
    WPS_UTILS_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def is_close(a, b, tol=0.02):
    """Safely check if two floats are close."""
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False

def verify_timesheet(traj, env_info, task_info):
    """Verify the timesheet calculations."""
    copy_from_env = env_info.get('copy_from_env')
    query_vlm = env_info.get('query_vlm')
    
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # --- 1. Load exported result JSON ---
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Could not read task_result.json: {e}"}
    finally:
        if os.path.exists(temp_res.name):
            os.unlink(temp_res.name)

    if not result.get("file_modified", False):
        return {"passed": False, "score": 0, "feedback": "Anti-gaming: File was not modified during the task."}

    # --- 2. Copy and parse the spreadsheet ---
    if WPS_UTILS_AVAILABLE:
        success, wb_data, error, temp_dir = copy_and_parse_spreadsheet("/home/ga/Documents/time_clock_data.xlsx", copy_from_env, 'xlsx')
        if not success:
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet: {error}"}
        
        # We also need formulas, so load data_only=False explicitly inside our logic just in case
        try:
            import openpyxl
            wb_formulas = openpyxl.load_workbook(os.path.join(temp_dir, 'input.xlsx'), data_only=False)
            wb_data = openpyxl.load_workbook(os.path.join(temp_dir, 'input.xlsx'), data_only=True)
        except Exception as e:
            cleanup_verification_temp(temp_dir)
            return {"passed": False, "score": 0, "feedback": f"Error parsing workbook: {e}"}
    else:
        # Fallback if wps_verification_utils is not available locally
        temp_dir = tempfile.mkdtemp(prefix='wps_verify_')
        temp_file = os.path.join(temp_dir, 'input.xlsx')
        try:
            copy_from_env("/home/ga/Documents/time_clock_data.xlsx", temp_file)
            import openpyxl
            wb_formulas = openpyxl.load_workbook(temp_file, data_only=False)
            wb_data = openpyxl.load_workbook(temp_file, data_only=True)
        except Exception as e:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return {"passed": False, "score": 0, "feedback": f"Failed to load spreadsheet natively: {e}"}

    try:
        score = 0
        feedback_parts = []
        
        sheet_names = [s.lower() for s in wb_data.sheetnames]
        
        dd_sheet_name = next((s for s in wb_data.sheetnames if 'daily' in s.lower() and 'detail' in s.lower()), None)
        ws_sheet_name = next((s for s in wb_data.sheetnames if 'weekly' in s.lower() and 'summary' in s.lower()), None)

        # --- Criterion 1: Sheets Exist (20 pts) ---
        if dd_sheet_name and ws_sheet_name:
            score += 20
            feedback_parts.append("Both required sheets exist (+20)")
        elif dd_sheet_name or ws_sheet_name:
            score += 10
            feedback_parts.append("Only one required sheet exists (+10)")
        else:
            feedback_parts.append("Missing required sheets Daily_Detail and Weekly_Summary")

        # --- Criterion 2: Daily Detail Validation (30 pts) ---
        if dd_sheet_name:
            dd_d = wb_data[dd_sheet_name]
            dd_f = wb_formulas[dd_sheet_name]
            
            # Check row count
            data_rows = sum(1 for row in dd_d.iter_rows(min_row=2, max_col=1) if row[0].value)
            if data_rows >= 70:
                score += 10
                feedback_parts.append(f"Daily Detail rows populated (+10)")
            
            # Check formulas
            formula_count = 0
            for row in range(2, min(77, dd_f.max_row + 1)):
                for col in [6, 7, 8]:  # F, G, H (Hours_Worked, Reg, OT)
                    cell = dd_f.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
            if formula_count >= 150:
                score += 10
                feedback_parts.append("Formulas used correctly in Daily Detail (+10)")
            
            # Spot check logic
            valid_calculations = 0
            for row in range(2, min(77, dd_d.max_row + 1)):
                hw = dd_d.cell(row=row, column=6).value
                reg = dd_d.cell(row=row, column=7).value
                ot = dd_d.cell(row=row, column=8).value
                # Basic check: Reg + OT should equal Total Hours, Reg shouldn't exceed 8
                if hw is not None and reg is not None and ot is not None:
                    if is_close(hw, float(reg) + float(ot), 0.05) and float(reg) <= 8.01:
                        valid_calculations += 1
            if valid_calculations >= 70:
                score += 10
                feedback_parts.append("Daily Detail calculations correct (+10)")

        # --- Criterion 3: Weekly Summary Validation (35 pts) ---
        if ws_sheet_name:
            ws_d = wb_data[ws_sheet_name]
            ws_f = wb_formulas[ws_sheet_name]
            
            emp_data = {}
            grand_total_pay = None
            dept_totals = {"Assembly": 0, "Quality": 0, "Warehouse": 0}
            has_currency = False
            has_bold = False

            # Scan rows
            for row in range(1, ws_d.max_row + 1):
                col1 = str(ws_d.cell(row=row, column=1).value or "").strip()
                col2 = str(ws_d.cell(row=row, column=2).value or "").strip()
                col3 = str(ws_d.cell(row=row, column=3).value or "").strip()
                
                # Grand total check
                if "grand" in col1.lower() or "grand" in col2.lower() or "grand" in col3.lower():
                    val = ws_d.cell(row=row, column=10).value
                    if val is not None:
                        try:
                            grand_total_pay = float(val)
                        except:
                            pass
                    # Check bolding for grand total
                    if ws_f.cell(row=row, column=1).font and ws_f.cell(row=row, column=1).font.bold:
                        has_bold = True

                # Department total check
                for dept in ["Assembly", "Quality", "Warehouse"]:
                    if dept.lower() in col1.lower() or dept.lower() in col2.lower() or dept.lower() in col3.lower():
                        val = ws_d.cell(row=row, column=10).value
                        if val is not None and not col1.startswith('E'): # Avoid matching employee rows as dept rows
                            try:
                                dept_totals[dept] = float(val)
                            except:
                                pass

                # Employee data check
                if col1.startswith("E0"):
                    try:
                        emp_data[col1] = {
                            "total": float(ws_d.cell(row=row, column=5).value or 0),
                            "ot": float(ws_d.cell(row=row, column=7).value or 0),
                            "pay": float(ws_d.cell(row=row, column=10).value or 0)
                        }
                    except:
                        pass
                    # Check currency formatting
                    nf = str(ws_f.cell(row=row, column=10).number_format)
                    if '$' in nf or '#,##0.00' in nf or '0.00' in nf:
                        has_currency = True

            # 3a: Employee math (15 pts)
            expected_emps = task_info.get('metadata', {}).get('employees', {})
            emp_math_correct = 0
            for emp_id, exp in expected_emps.items():
                if emp_id in emp_data:
                    act = emp_data[emp_id]
                    if is_close(act['total'], exp['total']) and is_close(act['ot'], exp['ot']) and is_close(act['pay'], exp['pay'], 0.5):
                        emp_math_correct += 1
            
            if emp_math_correct == len(expected_emps):
                score += 15
                feedback_parts.append("Employee aggregations correct (+15)")
            elif emp_math_correct > 0:
                score += 5
                feedback_parts.append("Partial employee aggregations (+5)")

            # 3b: Subtotals and Grand Total (10 pts)
            expected_gt = task_info.get('metadata', {}).get('grand_total_pay', 15382.38)
            expected_dept_sums = [6221.75, 4767.25, 4393.38] # Assembly, Quality, Warehouse
            
            found_subtotals = sum(1 for k, v in dept_totals.items() if any(is_close(v, d, 2.0) for d in expected_dept_sums))
            
            if is_close(grand_total_pay, expected_gt, 2.0):
                score += 5
                feedback_parts.append("Grand total correct (+5)")
            if found_subtotals >= 3:
                score += 5
                feedback_parts.append("Department subtotals correct (+5)")

            # 3c: Formatting (10 pts)
            if has_currency and has_bold:
                score += 10
                feedback_parts.append("Formatting applied (+10)")
            elif has_currency or has_bold:
                score += 5
                feedback_parts.append("Partial formatting applied (+5)")

        # --- Criterion 4: VLM verification of trajectory (15 pts) ---
        if query_vlm:
            from gym_anything.vlm import sample_trajectory_frames
            frames = sample_trajectory_frames(traj, n=4)
            if frames:
                vlm_res = query_vlm(
                    images=frames,
                    prompt="Do these screenshots show a user interacting with a spreadsheet, writing formulas, and creating summary tables for timesheets/payroll?"
                )
                if vlm_res.get('parsed', {}).get('yes_no') == 'yes' or 'yes' in str(vlm_res.get('text', '')).lower()[:10]:
                    score += 15
                    feedback_parts.append("VLM visual verification passed (+15)")
                else:
                    feedback_parts.append("VLM visual verification: Interaction not clearly observed")
        else:
            # Grant fallback points if VLM not available but other signals are strong
            if score >= 60:
                score += 15
                feedback_parts.append("VLM skipped; auto-credit given (+15)")

        passed = score >= 60
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        if 'temp_dir' in locals():
            shutil.rmtree(temp_dir, ignore_errors=True)
        elif WPS_UTILS_AVAILABLE and 'temp_dir' in locals() and temp_dir:
            cleanup_verification_temp(temp_dir)
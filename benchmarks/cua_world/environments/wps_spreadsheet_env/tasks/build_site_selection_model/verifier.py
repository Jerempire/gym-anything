#!/usr/bin/env python3
"""Verifier for build_site_selection_model task."""

import json
import os
import tempfile
import logging
from typing import Dict, Any

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_site_selection_model(traj: Dict[str, Any], env_info: Dict[str, Any], task_info: Dict[str, Any]) -> Dict[str, Any]:
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available."}

    # Retrieve execution metadata
    temp_meta = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_meta.name)
        with open(temp_meta.name, 'r') as f:
            exec_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read execution result: {e}"}
    finally:
        if os.path.exists(temp_meta.name):
            os.unlink(temp_meta.name)

    if not exec_result.get('output_exists', False):
        return {"passed": False, "score": 0, "feedback": "Spreadsheet file was not found."}

    # Copy and parse spreadsheet
    temp_xls = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/store_locations.xlsx", temp_xls.name)
        
        # We need to install openpyxl if not available
        try:
            from openpyxl import load_workbook
        except ImportError:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            from openpyxl import load_workbook

        wb = load_workbook(temp_xls.name, data_only=False) # Get formulas
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse spreadsheet: {e}"}
    finally:
        if os.path.exists(temp_xls.name):
            os.unlink(temp_xls.name)

    score = 0
    feedback_parts = []
    
    # 1. File Modification (10 pts)
    if exec_result.get('file_modified_during_task', False):
        score += 10
        feedback_parts.append("File saved successfully (+10)")
    else:
        feedback_parts.append("Warning: File may not have been saved during task")

    # Verify 'Sites' Sheet Structure
    if 'Sites' not in wb.sheetnames:
        feedback_parts.append("CRITICAL: 'Sites' sheet not found.")
        return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
    
    ws_sites = wb['Sites']
    
    # Check headers (I to N)
    headers = [str(ws_sites.cell(row=1, column=i).value).strip() for i in range(9, 15)]
    
    # 2. Normalization Formulas (Cols I, J, K) (30 pts)
    norm_score = 0
    formulas = [str(ws_sites.cell(row=2, column=i).value).upper() for i in range(9, 12)]
    for idx, f in enumerate(formulas):
        if f.startswith('=') and ('MAX' in f or 'MIN' in f) and '/' in f:
            norm_score += 10
    
    score += norm_score
    feedback_parts.append(f"Standard Normalizations: {norm_score}/30 pts")

    # 3. Inverted Normalization (Col L) (15 pts)
    lease_form = str(ws_sites.cell(row=2, column=12).value).upper()
    if lease_form.startswith('=') and ('MAX' in lease_form) and '/' in lease_form:
        # Check if they put MAX first (MAX - Val)
        score += 15
        feedback_parts.append("Inverted Normalization: 15/15 pts")
    else:
        feedback_parts.append("Inverted Normalization missing/incorrect")

    # 4. Weighted Score (Col M) (15 pts)
    score_form = str(ws_sites.cell(row=2, column=13).value).upper()
    if score_form.startswith('=') and 'WEIGHTS' in score_form:
        score += 15
        feedback_parts.append("Weighted Score (Cross-Sheet): 15/15 pts")
    else:
        feedback_parts.append("Weighted Score missing or hardcoded (no Weights sheet ref)")

    # 5. Rank Computation (Col N) (10 pts)
    rank_form = str(ws_sites.cell(row=2, column=14).value).upper()
    if rank_form.startswith('=') and 'RANK' in rank_form:
        score += 10
        feedback_parts.append("Rank Formula: 10/10 pts")
    else:
        feedback_parts.append("Rank formula missing/incorrect")

    # 6. Top Sites Summary Sheet (20 pts)
    top_score = 0
    if 'Top_Sites' in wb.sheetnames:
        ws_top = wb['Top_Sites']
        has_lookup = False
        
        # Scan B2:E6 for lookup formulas
        for row in ws_top.iter_rows(min_row=2, max_row=6, min_col=2, max_col=5):
            for cell in row:
                val = str(cell.value).upper()
                if val.startswith('=') and any(fn in val for fn in ['INDEX', 'MATCH', 'VLOOKUP', 'XLOOKUP', 'FILTER', 'LOOKUP']):
                    has_lookup = True
                    break
            if has_lookup: break
            
        if has_lookup:
            top_score = 20
            feedback_parts.append("Top Sites Lookups: 20/20 pts")
        else:
            # Check if they just hardcoded the result
            hardcoded = str(ws_top.cell(row=2, column=2).value)
            if hardcoded != 'None' and not hardcoded.startswith('='):
                top_score = 5
                feedback_parts.append("Top Sites: Values hardcoded, formulas missing (5/20 pts)")
            else:
                feedback_parts.append("Top Sites Lookups missing")
    else:
        feedback_parts.append("'Top_Sites' sheet missing")
        
    score += top_score

    # Passing Threshold: 70 points
    passed = score >= 70
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
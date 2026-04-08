#!/usr/bin/env python3
"""
Verifier for digital_ad_roas_analysis task.

Checks:
1. File was modified during task (anti-gaming)
2. Ad_Data sheet has correct headers and calculated values (CTR, CPC, ROAS)
3. CPA zero-handling logic works correctly ('No Conversions')
4. Channel_Summary sheet is created with required layout
5. SUMIF formulas correctly aggregated the spend and revenue per channel
"""

import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_ad_roas_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Read the execution metadata
    temp_res = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_res.name)
        with open(temp_res.name, 'r') as f:
            result_meta = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read execution result: {e}"}
    finally:
        if os.path.exists(temp_res.name):
            os.unlink(temp_res.name)

    # Anti-gaming: Ensure file was interacted with
    if not result_meta.get("output_exists"):
        return {"passed": False, "score": 0, "feedback": "Target spreadsheet not found."}
    
    if not result_meta.get("file_modified_during_task"):
        return {"passed": False, "score": 0, "feedback": "File was not modified during the task. Did the agent forget to save?"}

    # Retrieve the modified spreadsheet
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/digital_ad_performance.xlsx", temp_wb.name)
        
        # We use openpyxl. Need to try importing it.
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call(["pip", "install", "-q", "openpyxl"])
            import openpyxl

        # Load with data_only=True to evaluate the results of the formulas
        wb = openpyxl.load_workbook(temp_wb.name, data_only=True)
        # Also load formulas just in case cache is empty
        wb_formulas = openpyxl.load_workbook(temp_wb.name, data_only=False)

        score = 0
        feedback_parts = []
        
        # 1. Ad_Data Sheet Validations
        if "Ad_Data" not in wb.sheetnames:
            feedback_parts.append("Sheet 'Ad_Data' missing or renamed.")
            return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}
            
        ws_ad = wb["Ad_Data"]
        ws_ad_f = wb_formulas["Ad_Data"]

        # 1a. Check Headers (10 pts)
        expected_headers = ["CTR", "CPC", "CPA", "ROAS"]
        actual_headers = [str(ws_ad.cell(row=1, column=col).value).strip().upper() for col in range(9, 13)]
        if actual_headers == expected_headers:
            score += 10
            feedback_parts.append("Ad_Data headers correct")
        else:
            feedback_parts.append(f"Ad_Data headers incorrect. Expected {expected_headers}, got {actual_headers}")

        # 1b & 1c. Core Metrics & Zero Handling (25 pts + 25 pts)
        core_metrics_correct = True
        zero_handling_correct = False
        zero_handling_found = False
        
        # Parse data to calculate ground truth and compare
        raw_data = []
        for r in range(2, ws_ad.max_row + 1):
            if ws_ad.cell(row=r, column=1).value is None:
                break
            
            impressions = float(ws_ad.cell(row=r, column=4).value or 0)
            clicks = float(ws_ad.cell(row=r, column=5).value or 0)
            spend = float(ws_ad.cell(row=r, column=6).value or 0)
            conversions = float(ws_ad.cell(row=r, column=7).value or 0)
            revenue = float(ws_ad.cell(row=r, column=8).value or 0)
            channel = str(ws_ad.cell(row=r, column=3).value).strip()
            
            raw_data.append({"ch": channel, "sp": spend, "rev": revenue})

            # Agent calculated values
            ag_ctr = ws_ad.cell(row=r, column=9).value
            ag_cpc = ws_ad.cell(row=r, column=10).value
            ag_cpa = ws_ad.cell(row=r, column=11).value
            ag_roas = ws_ad.cell(row=r, column=12).value
            
            # Ground truth
            gt_ctr = clicks / impressions if impressions > 0 else 0
            gt_cpc = spend / clicks if clicks > 0 else 0
            gt_roas = revenue / spend if spend > 0 else 0

            # Check core math logic (allow floating point variance)
            if ag_ctr is None or not isinstance(ag_ctr, (int, float)) or abs(ag_ctr - gt_ctr) > 0.01:
                core_metrics_correct = False
            if ag_cpc is None or not isinstance(ag_cpc, (int, float)) or abs(ag_cpc - gt_cpc) > 0.01:
                core_metrics_correct = False
            if ag_roas is None or not isinstance(ag_roas, (int, float)) or abs(ag_roas - gt_roas) > 0.01:
                core_metrics_correct = False

            # Check CPA logic
            if conversions == 0:
                zero_handling_found = True
                if isinstance(ag_cpa, str) and ag_cpa.strip().lower() == "no conversions":
                    zero_handling_correct = True
                else:
                    zero_handling_correct = False
                    core_metrics_correct = False # Flag failure if it throws an error instead
            else:
                gt_cpa = spend / conversions
                if ag_cpa is None or not isinstance(ag_cpa, (int, float)) or abs(ag_cpa - gt_cpa) > 0.01:
                    core_metrics_correct = False

        if core_metrics_correct:
            score += 25
            feedback_parts.append("Core metrics calculated correctly")
        else:
            feedback_parts.append("Errors found in CTR, CPC, or ROAS calculations")
            
        if zero_handling_found and zero_handling_correct:
            score += 25
            feedback_parts.append("CPA zero-handling ('No Conversions') logic works perfectly")
        elif zero_handling_found:
            feedback_parts.append("CPA zero-handling logic failed (Did not output 'No Conversions')")

        # 2. Channel Summary Sheet Validations
        if "Channel_Summary" in wb.sheetnames:
            ws_sum = wb["Channel_Summary"]
            
            # 2a. Summary Structure (10 pts)
            exp_sum_headers = ["Channel", "Total Spend", "Total Revenue", "Channel ROAS"]
            act_sum_headers = [str(ws_sum.cell(row=1, column=c).value).strip() for c in range(1, 5)]
            
            exp_channels = ["Search", "Display", "Social", "Video"]
            act_channels = [str(ws_sum.cell(row=r, column=1).value).strip() for r in range(2, 6)]
            
            if act_sum_headers == exp_sum_headers and all(c in act_channels for c in exp_channels):
                score += 10
                feedback_parts.append("Summary sheet structure correct")
            else:
                feedback_parts.append("Summary sheet headers or channel rows incorrect")

            # 2b & 2c. SUMIF & ROAS Aggregation (20 pts + 10 pts)
            sums_correct = True
            roas_correct = True
            
            # Calculate Ground Truth Aggregations
            gt_agg = {ch: {"sp": 0, "rev": 0} for ch in exp_channels}
            for rd in raw_data:
                ch = rd["ch"]
                if ch in gt_agg:
                    gt_agg[ch]["sp"] += rd["sp"]
                    gt_agg[ch]["rev"] += rd["rev"]
            
            # Compare
            for r in range(2, 6):
                ch = str(ws_sum.cell(row=r, column=1).value).strip()
                if ch not in gt_agg: continue
                
                ag_sp = ws_sum.cell(row=r, column=2).value or 0
                ag_rev = ws_sum.cell(row=r, column=3).value or 0
                ag_roas = ws_sum.cell(row=r, column=4).value or 0
                
                gt_sp = gt_agg[ch]["sp"]
                gt_rev = gt_agg[ch]["rev"]
                gt_roas = gt_rev / gt_sp if gt_sp > 0 else 0
                
                if not isinstance(ag_sp, (int, float)) or abs(ag_sp - gt_sp) > 1.0:
                    sums_correct = False
                if not isinstance(ag_rev, (int, float)) or abs(ag_rev - gt_rev) > 1.0:
                    sums_correct = False
                if not isinstance(ag_roas, (int, float)) or abs(ag_roas - gt_roas) > 0.05:
                    roas_correct = False

            if sums_correct:
                score += 20
                feedback_parts.append("SUMIF aggregations correct")
            else:
                feedback_parts.append("SUMIF aggregations incorrect or missing")
                
            if roas_correct:
                score += 10
                feedback_parts.append("Channel ROAS calculation correct")
            else:
                feedback_parts.append("Channel ROAS incorrect")
                
        else:
            feedback_parts.append("Sheet 'Channel_Summary' missing")

        passed = score >= 75
        
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Error during verification: {str(e)}"}
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)
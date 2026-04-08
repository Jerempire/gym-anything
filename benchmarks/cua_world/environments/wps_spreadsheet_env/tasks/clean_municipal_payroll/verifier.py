#!/usr/bin/env python3
import sys
import os
import json
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_clean_municipal_payroll(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    # Read the JSON state file from export_result.sh
    try:
        temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
        os.unlink(temp_result.name)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to load task result: {e}"}

    if not result.get("file_exists", False):
        return {"passed": False, "score": 0, "feedback": "Modified file chicago_employees.xlsx not found."}

    # Verify if file was actually modified (preventing false pass on doing nothing)
    task_start_time = 0
    try:
        temp_start = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
        copy_from_env("/tmp/task_start_time.txt", temp_start.name)
        with open(temp_start.name, 'r') as f:
            task_start_time = int(f.read().strip())
        os.unlink(temp_start.name)
    except:
        pass
        
    if result.get("file_mtime", 0) <= task_start_time:
        return {"passed": False, "score": 0, "feedback": "File was not modified during the task."}

    # Copy the spreadsheet locally and verify content using `data_only=True` to evaluate agent's formulas
    temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/chicago_employees.xlsx", temp_excel.name)
        
        # Load workbook resolving formulas
        import openpyxl
        wb = openpyxl.load_workbook(temp_excel.name, data_only=True)
        sheets = wb.sheetnames
        
        feedback_parts = []
        score = 0
        
        # Check basic structural setup
        if "Employees" not in sheets:
            return {"passed": False, "score": 0, "feedback": "Missing Employees sheet."}
        
        ws_emp = wb["Employees"]
        
        # Evaluate headers dynamically in case they were offset
        headers = [str(ws_emp.cell(row=1, column=c).value).strip().lower() for c in range(1, ws_emp.max_column + 1)]
        
        has_last = 'last_name' in headers
        has_first = 'first_name' in headers
        has_title = 'clean_title' in headers
        has_comp = 'total_compensation' in headers
        
        if has_last and has_first and has_title and has_comp:
            score += 10
            feedback_parts.append("File Saved & Structure (10/10)")
        else:
            feedback_parts.append("Missing required calculated columns in Employees sheet.")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
            
        last_idx = headers.index('last_name') + 1
        first_idx = headers.index('first_name') + 1
        title_idx = headers.index('clean_title') + 1
        comp_idx = headers.index('total_compensation') + 1
        name_idx = headers.index('name') + 1
        
        row_aaron = None
        row_abdullah = None
        
        # Find test rows in case agent sorted data
        for r in range(2, ws_emp.max_row + 1):
            name_val = ws_emp.cell(row=r, column=name_idx).value
            if str(name_val) == "AARON,  KARINA":
                row_aaron = r
            elif str(name_val) == "ABDULLAH,  LAKENYA N":
                row_abdullah = r
                
        if not row_aaron or not row_abdullah:
            feedback_parts.append("Could not find required test rows (AARON or ABDULLAH). Data altered?")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
            
        last_2 = ws_emp.cell(row=row_aaron, column=last_idx).value
        first_2 = ws_emp.cell(row=row_aaron, column=first_idx).value
        title_2 = ws_emp.cell(row=row_aaron, column=title_idx).value
        comp_2 = ws_emp.cell(row=row_aaron, column=comp_idx).value
        
        last_10 = ws_emp.cell(row=row_abdullah, column=last_idx).value
        first_10 = ws_emp.cell(row=row_abdullah, column=first_idx).value
        title_10 = ws_emp.cell(row=row_abdullah, column=title_idx).value
        comp_10 = ws_emp.cell(row=row_abdullah, column=comp_idx).value
        
        # Check Last Name Extraction
        if str(last_2).strip().upper() == "AARON" and str(last_10).strip().upper() == "ABDULLAH":
            score += 15
            feedback_parts.append("Last Name Extraction (15/15)")
        else:
            feedback_parts.append(f"Last Name Extraction failed: rowA='{last_2}', rowB='{last_10}'")
            
        # Check First Name Extraction (validating TRIM implementation)
        if str(first_2) == "KARINA" and str(first_10) == "LAKENYA N":
            score += 15
            feedback_parts.append("First Name Extraction (15/15)")
        elif str(first_2).strip() == "KARINA" and str(first_10).strip() == "LAKENYA N":
            score += 7
            feedback_parts.append("First Name Extraction partial (missed TRIM function)")
        else:
            feedback_parts.append(f"First Name Extraction failed: rowA='{first_2}', rowB='{first_10}'")
            
        # Check Title Formatting (PROPER functionality)
        if str(title_2) == "Police Officer" and str(title_10) == "Crossing Guard":
            score += 10
            feedback_parts.append("Title Formatting (10/10)")
        elif str(title_2).strip().lower() == "police officer":
            score += 5
            feedback_parts.append("Title Formatting partial (case/spacing issues)")
        else:
            feedback_parts.append(f"Title Formatting failed: rowA='{title_2}'")
            
        # Check Conditional Compensation Normalization
        try:
            c2_valid = abs(float(comp_2) - 90024.0) < 0.1
            c10_valid = abs(float(comp_10) - 20155.2) < 0.1
            
            if c2_valid and c10_valid:
                score += 25
                feedback_parts.append("Compensation Normalization (25/25)")
            else:
                feedback_parts.append(f"Compensation Normalization failed: expected ~90024 and ~20155.2, got {comp_2}, {comp_10}")
        except:
            feedback_parts.append(f"Compensation Normalization error: non-numeric values {comp_2}, {comp_10}")
            
        # Evaluate Summary Sheet Outputs
        if "Department_Summary" in sheets:
            ws_sum = wb["Department_Summary"]
            sum_rows = list(ws_sum.iter_rows(values_only=True))
            
            dept_data = {}
            for row in sum_rows:
                if not row or not row[0]: continue
                dept = str(row[0]).strip().upper()
                if dept in ["POLICE", "FIRE", "STREETS & SAN", "WATER MGMNT", "AVIATION"]:
                    dept_data[dept] = {"count": row[1], "avg": row[2]}
                    
            # Known Mathematical Ground Truths for Synthesized Database
            gt = {
                "POLICE": {"count": 16, "avg": 79290.4},
                "FIRE": {"count": 8, "avg": 97799.0},
                "STREETS & SAN": {"count": 10, "avg": 84760.0},
                "WATER MGMNT": {"count": 8, "avg": 100160.4},
                "AVIATION": {"count": 8, "avg": 85876.8}
            }
            
            headcount_correct = 0
            avg_correct = 0
            
            for dept, expected in gt.items():
                if dept in dept_data:
                    try:
                        if abs(float(dept_data[dept]["count"]) - expected["count"]) < 0.1:
                            headcount_correct += 1
                        if abs(float(dept_data[dept]["avg"]) - expected["avg"]) < 1.0: # Allow 1.0 tolerance for precision rounding
                            avg_correct += 1
                    except:
                        pass
                        
            if headcount_correct == 5:
                score += 10
                feedback_parts.append("Summary - Headcount (10/10)")
            elif headcount_correct > 0:
                score += headcount_correct * 2
                feedback_parts.append(f"Summary - Headcount partial ({headcount_correct}/5)")
            else:
                feedback_parts.append("Summary - Headcount failed")
                
            if avg_correct == 5:
                score += 15
                feedback_parts.append("Summary - Avg Compensation (15/15)")
            elif avg_correct > 0:
                score += avg_correct * 3
                feedback_parts.append(f"Summary - Avg Comp partial ({avg_correct}/5)")
            else:
                feedback_parts.append("Summary - Avg Compensation failed")
        else:
            feedback_parts.append("Department_Summary sheet missing.")
            
        passed = score >= 75
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Error: {str(e)}"}
    finally:
        if os.path.exists(temp_excel.name):
            os.unlink(temp_excel.name)
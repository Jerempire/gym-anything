#!/usr/bin/env python3
"""
Verifier for Energy Bill Analysis task
Checks data entry, formula calculations, anomaly identification, and visual highlighting
"""

import sys
import os
import logging
import zipfile
from xml.etree import ElementTree as ET

# Do not use /workspace/utils, since the verification runs on the host machine, not the container.
# USE Relative path to the utils folder.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from calc_verification_utils import (
    copy_and_parse_spreadsheet,
    get_cell_value,
    get_cell_formula,
    cleanup_verification_temp
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def check_row_highlighting_ods(filepath, sheet_name, row_idx):
    """
    Check if a specific row has background color highlighting in ODS format.
    
    Args:
        filepath: Path to ODS file
        sheet_name: Name of sheet
        row_idx: Row index (0-based)
        
    Returns:
        bool: True if highlighting detected, False otherwise
    """
    try:
        with zipfile.ZipFile(filepath, 'r') as ods_zip:
            if 'content.xml' not in ods_zip.namelist():
                return False

            content_xml = ods_zip.read('content.xml')
            root = ET.fromstring(content_xml)

            # Define namespaces
            ns = {
                'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
                'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
                'style': 'urn:oasis:names:tc:opendocument:xmlns:style:1.0',
                'fo': 'urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0'
            }

            # Find the target sheet
            for table in root.findall('.//table:table', ns):
                table_name = table.get('{urn:oasis:names:tc:opendocument:xmlns:table:1.0}name')
                if table_name == sheet_name:
                    rows = table.findall('.//table:table-row', ns)
                    
                    if row_idx < len(rows):
                        target_row = rows[row_idx]
                        cells = target_row.findall('.//table:table-cell', ns)
                        
                        # Check if any cell in the row has a custom style
                        for cell in cells[:5]:  # Check first 5 columns (A-E)
                            style_name = cell.get('{urn:oasis:names:tc:opendocument:xmlns:table:1.0}style-name')
                            if style_name:
                                # If cell has a custom style, assume it has highlighting
                                # Full verification would require parsing styles.xml
                                return True
            
            # Also check automatic-styles in content.xml for background colors
            auto_styles = root.find('.//office:automatic-styles', ns)
            if auto_styles is not None:
                for style in auto_styles.findall('.//style:style', ns):
                    table_cell_props = style.find('.//style:table-cell-properties', ns)
                    if table_cell_props is not None:
                        bg_color = table_cell_props.get('{urn:oasis:names:tc:opendocument:xmlns:xsl-fo-compatible:1.0}background-color')
                        if bg_color and bg_color != 'transparent' and bg_color != '#ffffff':
                            # Found a style with non-white background
                            return True

            return False

    except Exception as e:
        logger.warning(f"Could not check ODS highlighting: {e}")
        return False


def check_row_highlighting_xlsx(filepath, sheet_name, row_idx):
    """
    Check if a specific row has background color highlighting in XLSX format.
    
    Args:
        filepath: Path to XLSX file
        sheet_name: Name of sheet
        row_idx: Row index (0-based, but converted to 1-based for openpyxl)
        
    Returns:
        bool: True if highlighting detected, False otherwise
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return False
            
        ws = wb[sheet_name]
        
        # Check cells in target row (1-based row number)
        excel_row = row_idx + 1
        for col in range(1, 6):  # Columns A-E (1-5)
            cell = ws.cell(row=excel_row, column=col)
            if cell.fill and hasattr(cell.fill, 'patternType'):
                if cell.fill.patternType and cell.fill.patternType != 'none':
                    # Check if fill color is not default white
                    if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        rgb = cell.fill.start_color.rgb
                        if rgb and rgb not in ['FFFFFFFF', 'FFFFFF', '00000000']:
                            return True

        return False

    except ImportError:
        logger.warning("openpyxl not available for XLSX highlighting check")
        return False
    except Exception as e:
        logger.warning(f"Could not check XLSX highlighting: {e}")
        return False


def verify_energy_analysis(traj, env_info, task_info):
    """
    Verify energy bill analysis task completion.
    
    Checks:
    1. Data entry complete (6 months)
    2. Cost per kWh calculated
    3. Average computed
    4. Percentage deviations calculated
    5. Anomaly identified (May has high deviation)
    6. Visual highlighting applied to May row
    7. Formula integrity (using references)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    # Try to copy and parse spreadsheet (try ODS first, then XLSX)
    container_path = "/home/ga/Documents/energy_analysis.ods"
    success, workbook, error, temp_dir = copy_and_parse_spreadsheet(
        container_path,
        copy_from_env,
        file_format='ods'
    )

    file_format = 'ods'
    if not success:
        # Try XLSX as fallback
        container_path_xlsx = "/home/ga/Documents/energy_analysis.xlsx"
        success, workbook, error, temp_dir = copy_and_parse_spreadsheet(
            container_path_xlsx,
            copy_from_env,
            file_format='xlsx'
        )
        file_format = 'xlsx'

    if not success:
        return {"passed": False, "score": 0, "feedback": f"Failed to load file: {error}"}

    try:
        # Get first sheet
        sheet_names = list(workbook['sheets'].keys())
        if not sheet_names:
            return {"passed": False, "score": 0, "feedback": "No sheets found in workbook"}
        
        sheet_name = sheet_names[0]

        criteria_passed = 0
        total_criteria = 7
        feedback_parts = []

        # Expected data
        expected_months = ["January", "February", "March", "April", "May", "June"]
        expected_usage = [850, 780, 820, 890, 1420, 810]
        expected_costs = [102.00, 93.60, 98.40, 106.80, 170.40, 97.20]
        expected_avg = 928.33  # Average of usage values
        expected_cost_per_kwh = 0.12  # $0.12 per kWh

        # Criterion 1: Data entry completeness
        data_correct = True
        data_errors = []
        for i in range(6):
            row_num = i + 2  # Data starts at row 2
            
            # Check month name (optional - focus on numeric data)
            month = get_cell_value(workbook, sheet_name, f'A{row_num}')
            
            # Check usage
            usage = get_cell_value(workbook, sheet_name, f'B{row_num}')
            if usage is None:
                data_correct = False
                data_errors.append(f"Missing usage in row {row_num}")
                break
            
            try:
                usage_val = float(usage)
                if abs(usage_val - expected_usage[i]) > 10:  # ±10 kWh tolerance
                    data_correct = False
                    data_errors.append(f"Usage incorrect in B{row_num}: expected {expected_usage[i]}, got {usage_val}")
            except (ValueError, TypeError):
                data_correct = False
                data_errors.append(f"Invalid usage value in B{row_num}: {usage}")
            
            # Check cost
            cost = get_cell_value(workbook, sheet_name, f'C{row_num}')
            if cost is None:
                data_correct = False
                data_errors.append(f"Missing cost in row {row_num}")
                break
            
            try:
                cost_val = float(cost)
                if abs(cost_val - expected_costs[i]) > 2.0:  # ±$2 tolerance
                    data_correct = False
                    data_errors.append(f"Cost incorrect in C{row_num}: expected {expected_costs[i]}, got {cost_val}")
            except (ValueError, TypeError):
                data_correct = False
                data_errors.append(f"Invalid cost value in C{row_num}: {cost}")

        if data_correct:
            criteria_passed += 1
            feedback_parts.append("✅ Data entry complete and accurate")
        else:
            feedback_parts.append(f"❌ Data entry issues: {'; '.join(data_errors[:2])}")

        # Criterion 2: Cost per kWh calculations
        cost_per_kwh_correct = True
        has_formula = False
        
        for i in range(6):
            row_num = i + 2
            calc_value = get_cell_value(workbook, sheet_name, f'D{row_num}')
            
            if calc_value is None:
                cost_per_kwh_correct = False
                break
            
            try:
                calc_val = float(calc_value)
                expected_value = expected_costs[i] / expected_usage[i]
                if abs(calc_val - expected_value) > 0.002:  # ±$0.002 tolerance
                    cost_per_kwh_correct = False
                    break
            except (ValueError, TypeError):
                cost_per_kwh_correct = False
                break

        # Check if formulas are used
        sample_formula = get_cell_formula(workbook, sheet_name, 'D2')
        if sample_formula and ('/' in sample_formula or 'DIVIDE' in str(sample_formula).upper()):
            has_formula = True

        if cost_per_kwh_correct and has_formula:
            criteria_passed += 1
            feedback_parts.append("✅ Cost per kWh calculated correctly with formulas")
        elif cost_per_kwh_correct:
            criteria_passed += 0.5
            feedback_parts.append("⚠️ Cost per kWh values correct but may be hardcoded")
        else:
            feedback_parts.append("❌ Cost per kWh calculations missing or incorrect")

        # Criterion 3: Average calculation
        avg_value = None
        avg_locations = ['B9', 'B8', 'B10', 'C9', 'C8', 'C10', 'A9', 'A10']
        
        for loc in avg_locations:
            val = get_cell_value(workbook, sheet_name, loc)
            if val:
                try:
                    val_float = float(val)
                    if 900 <= val_float <= 950:  # Expected ~928
                        avg_value = val_float
                        break
                except (ValueError, TypeError):
                    continue

        avg_correct = False
        if avg_value and abs(avg_value - expected_avg) < 15:  # ±15 kWh tolerance
            avg_correct = True
            criteria_passed += 1
            feedback_parts.append(f"✅ Average usage calculated correctly (~{int(avg_value)} kWh)")
        else:
            feedback_parts.append("❌ Average usage calculation missing or incorrect")

        # Criterion 4: Percentage deviation calculations
        may_row = 6  # May is row 6 (index 5, but data row 6)
        may_percentage = get_cell_value(workbook, sheet_name, f'E{may_row}')
        
        percentage_correct = False
        if may_percentage:
            try:
                may_pct = float(may_percentage)
                # May (1420 kWh) vs avg (928) should be ~53%
                expected_may_pct = ((1420 - expected_avg) / expected_avg) * 100
                if abs(may_pct - expected_may_pct) < 8:  # ±8% tolerance
                    percentage_correct = True
            except (ValueError, TypeError):
                pass

        if percentage_correct:
            criteria_passed += 1
            feedback_parts.append(f"✅ Percentage deviation calculated correctly (May: ~{int(may_pct)}%)")
        else:
            feedback_parts.append("❌ Percentage deviation calculations missing or incorrect")

        # Criterion 5: Anomaly identification
        anomaly_identified = False
        if may_percentage:
            try:
                may_pct = float(may_percentage)
                if may_pct > 40:  # May should be >40% above average
                    anomaly_identified = True
            except (ValueError, TypeError):
                pass

        if anomaly_identified:
            criteria_passed += 1
            feedback_parts.append("✅ Anomaly (May) correctly identified through calculations")
        else:
            feedback_parts.append("❌ Anomaly not properly identified in calculations")

        # Criterion 6: Visual highlighting
        # May is in row 6 (0-based index 5 for data rows, but row 6 overall including header)
        # Header is row 0, data starts row 1, so May is row index 5
        may_row_idx = 5  # 0-based index for May data row
        
        has_highlighting = False
        if file_format == 'ods':
            has_highlighting = check_row_highlighting_ods(
                workbook.get('filepath', ''),
                sheet_name,
                may_row_idx
            )
        elif file_format == 'xlsx':
            has_highlighting = check_row_highlighting_xlsx(
                workbook.get('filepath', ''),
                sheet_name,
                may_row_idx
            )

        if has_highlighting:
            criteria_passed += 1
            feedback_parts.append("✅ Anomalous row visually highlighted")
        else:
            feedback_parts.append("❌ No visual highlighting detected on anomalous row")

        # Criterion 7: Formula integrity
        formula_count = 0
        for row in range(2, 8):  # Rows 2-7
            for col in ['D', 'E']:
                formula = get_cell_formula(workbook, sheet_name, f'{col}{row}')
                if formula and '=' in str(formula):
                    formula_count += 1

        if formula_count >= 8:  # Should have at least 10-12 formulas
            criteria_passed += 1
            feedback_parts.append("✅ Formulas used appropriately throughout")
        elif formula_count >= 4:
            criteria_passed += 0.5
            feedback_parts.append("⚠️ Some formulas used but may have hardcoded values")
        else:
            feedback_parts.append("❌ Insufficient formula usage (possible hardcoded values)")

        # Calculate final score
        score = int((criteria_passed / total_criteria) * 100)
        passed = score >= 75
        
        feedback = " | ".join(feedback_parts)
        
        return {
            "passed": passed,
            "score": score,
            "feedback": feedback,
            "subscores": {
                "data_entry": data_correct,
                "cost_per_kwh": cost_per_kwh_correct and has_formula,
                "average_computed": avg_correct,
                "percentage_deviations": percentage_correct,
                "anomaly_identified": anomaly_identified,
                "visual_highlighting": has_highlighting,
                "formula_integrity": formula_count >= 8
            }
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}

    finally:
        cleanup_verification_temp(temp_dir)

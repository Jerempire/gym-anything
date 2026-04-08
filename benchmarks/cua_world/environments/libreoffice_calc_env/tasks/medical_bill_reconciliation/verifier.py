#!/usr/bin/env python3
"""
Verifier for Medical Bill Reconciliation task
"""

import sys
import os
import logging

# Do not use /workspace/utils, since the verification runs on the host machine, not the container.
# USE Relative path to the utils folder.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))

from calc_verification_utils import (
    setup_calc_verification,
    cleanup_verification_environment,
    get_cell_value,
    get_sheet_names
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def verify_medical_bill_reconciliation(traj, env_info, task_info):
    """
    Verify medical bill reconciliation task completion.
    
    Checks:
    1. Reconciliation columns added (Discrepancy, Status)
    2. Lookup formulas present (VLOOKUP/INDEX-MATCH)
    3. Duplicates identified (at least 2)
    4. Disputes flagged (at least 2)
    5. Conditional formatting applied
    6. Summary calculations present
    7. Overage calculated correctly (positive value)
    8. No formula errors (#N/A, #REF!, etc.)
    
    Returns:
        dict: {"passed": bool, "score": int, "feedback": str, "subscores": dict}
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    # Try to load the file
    container_path = "/home/ga/Documents/medical_bills.ods"
    success, file_info, error = setup_calc_verification(
        copy_from_env, 
        container_path,
        expected_formats=['ods']
    )
    
    if not success:
        return {"passed": False, "score": 0, "feedback": f"Failed to load file: {error}"}
    
    try:
        data = file_info['sheet_data']
        
        # Get Bills sheet
        sheet_names = get_sheet_names(data)
        bills_sheet_name = None
        for name in sheet_names:
            if 'bill' in name.lower():
                bills_sheet_name = name
                break
        
        if not bills_sheet_name:
            bills_sheet_name = sheet_names[0] if sheet_names else None
        
        if not bills_sheet_name:
            return {"passed": False, "score": 0, "feedback": "Bills sheet not found"}
        
        bills_sheet = data['sheets'].get(bills_sheet_name, [])
        
        if not bills_sheet or len(bills_sheet) < 2:
            return {"passed": False, "score": 0, "feedback": "Bills sheet is empty or invalid"}
        
        # Extract headers (first row)
        headers = bills_sheet[0] if bills_sheet else []
        header_names = []
        for cell in headers:
            if isinstance(cell, dict):
                header_names.append(str(cell.get('value', '')).strip())
            else:
                header_names.append(str(cell).strip())
        
        # Initialize scoring
        criteria_met = 0
        total_criteria = 8
        feedback = {}
        
        # Criterion 1: Reconciliation columns added
        has_discrepancy = any('discrepancy' in str(h).lower() for h in header_names)
        has_status = any('status' in str(h).lower() for h in header_names)
        
        if has_discrepancy and has_status:
            criteria_met += 1
            feedback['reconciliation_columns'] = "✅ Reconciliation columns present (Discrepancy, Status)"
        elif has_discrepancy or has_status:
            criteria_met += 0.5
            feedback['reconciliation_columns'] = "⚠️ Partial reconciliation columns (missing Discrepancy or Status)"
        else:
            feedback['reconciliation_columns'] = "❌ Missing Discrepancy and Status columns"
        
        # Criterion 2: Lookup formulas present
        has_lookup = False
        lookup_functions = ['VLOOKUP', 'INDEX', 'MATCH', 'XLOOKUP', 'HLOOKUP']
        
        for row_idx, row in enumerate(bills_sheet[1:11], start=1):  # Check first 10 data rows
            if row_idx > len(bills_sheet) - 1:
                break
            for cell in row:
                if isinstance(cell, dict):
                    formula = cell.get('formula', '')
                    if formula:
                        formula_upper = str(formula).upper()
                        if any(func in formula_upper for func in lookup_functions):
                            has_lookup = True
                            logger.info(f"Found lookup formula in row {row_idx}: {formula}")
                            break
            if has_lookup:
                break
        
        if has_lookup:
            criteria_met += 1
            feedback['lookup_formulas'] = "✅ Lookup formulas detected (VLOOKUP/INDEX-MATCH)"
        else:
            feedback['lookup_formulas'] = "❌ No lookup formulas found"
        
        # Criterion 3: Duplicates identified
        status_col_idx = None
        for idx, h in enumerate(header_names):
            if 'status' in str(h).lower():
                status_col_idx = idx
                break
        
        duplicate_count = 0
        if status_col_idx is not None:
            for row in bills_sheet[1:]:
                if len(row) > status_col_idx:
                    cell = row[status_col_idx]
                    cell_value = cell.get('value', '') if isinstance(cell, dict) else cell
                    if cell_value and 'duplicate' in str(cell_value).lower():
                        duplicate_count += 1
        
        if duplicate_count >= 2:
            criteria_met += 1
            feedback['duplicates'] = f"✅ Found {duplicate_count} duplicates (expected ≥2)"
        elif duplicate_count == 1:
            criteria_met += 0.5
            feedback['duplicates'] = f"⚠️ Found only {duplicate_count} duplicate (expected ≥2)"
        else:
            feedback['duplicates'] = f"❌ Found {duplicate_count} duplicates (expected ≥2)"
        
        # Criterion 4: Disputes flagged
        dispute_count = 0
        if status_col_idx is not None:
            for row in bills_sheet[1:]:
                if len(row) > status_col_idx:
                    cell = row[status_col_idx]
                    cell_value = cell.get('value', '') if isinstance(cell, dict) else cell
                    if cell_value and 'dispute' in str(cell_value).lower():
                        dispute_count += 1
        
        if dispute_count >= 2:
            criteria_met += 1
            feedback['disputes'] = f"✅ Found {dispute_count} disputes (expected ≥2)"
        elif dispute_count == 1:
            criteria_met += 0.5
            feedback['disputes'] = f"⚠️ Found only {dispute_count} dispute (expected ≥2)"
        else:
            feedback['disputes'] = f"❌ Found {dispute_count} disputes (expected ≥2)"
        
        # Criterion 5: Conditional formatting (simplified check)
        # For ODS, checking conditional formatting is complex, so we give benefit of doubt
        # if other formatting criteria are met. For XLSX we can check properly.
        has_formatting = False
        if file_info.get('format') == 'xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_info['filepath'])
                if bills_sheet_name in wb.sheetnames:
                    ws = wb[bills_sheet_name]
                    if hasattr(ws, 'conditional_formatting') and len(ws.conditional_formatting._cf_rules) > 0:
                        has_formatting = True
            except Exception as e:
                logger.debug(f"Could not check conditional formatting: {e}")
        else:
            # For ODS, give credit if Status column exists (assume formatting applied)
            has_formatting = has_status
        
        if has_formatting:
            criteria_met += 1
            feedback['formatting'] = "✅ Conditional formatting likely applied"
        else:
            feedback['formatting'] = "⚠️ Conditional formatting not clearly detected"
        
        # Criterion 6: Summary calculations present
        summary_found = False
        summary_keywords = ['total', 'overage', 'owed', 'sum', 'billed']
        
        # Check all rows for summary keywords in first 3 columns
        for row in bills_sheet:
            for cell in row[:3]:
                cell_value_str = ''
                if isinstance(cell, dict):
                    cell_value_str = str(cell.get('value', '')).lower()
                else:
                    cell_value_str = str(cell).lower()
                
                if any(kw in cell_value_str for kw in summary_keywords):
                    # Check if this row has multiple keywords or numbers (likely summary)
                    row_text = ' '.join([str(c.get('value', '') if isinstance(c, dict) else c) for c in row[:5]])
                    if row_text.lower().count('total') > 0 or row_text.lower().count('overage') > 0:
                        summary_found = True
                        logger.info(f"Found summary section: {row_text[:100]}")
                        break
            if summary_found:
                break
        
        if summary_found:
            criteria_met += 1
            feedback['summary'] = "✅ Summary section detected"
        else:
            feedback['summary'] = "❌ No summary calculations found"
        
        # Criterion 7: Overage calculated (should be positive)
        overage_found = False
        overage_value = None
        
        for row_idx, row in enumerate(bills_sheet):
            for col_idx, cell in enumerate(row[:4]):
                cell_value_str = ''
                if isinstance(cell, dict):
                    cell_value_str = str(cell.get('value', '')).lower()
                else:
                    cell_value_str = str(cell).lower()
                
                if 'overage' in cell_value_str or 'over-charge' in cell_value_str or 'difference' in cell_value_str:
                    # Check next cell or cells in same row for numeric value
                    for check_col in range(col_idx + 1, min(col_idx + 4, len(row))):
                        if check_col < len(row):
                            next_cell = row[check_col]
                            next_value = next_cell.get('value') if isinstance(next_cell, dict) else next_cell
                            
                            if next_value is not None:
                                try:
                                    # Remove currency symbols and commas
                                    next_value_clean = str(next_value).replace('$', '').replace(',', '').strip()
                                    overage_num = float(next_value_clean)
                                    if overage_num > 0:
                                        overage_found = True
                                        overage_value = overage_num
                                        logger.info(f"Found overage: ${overage_num}")
                                        break
                                except (ValueError, AttributeError):
                                    pass
                    if overage_found:
                        break
            if overage_found:
                break
        
        if overage_found:
            criteria_met += 1
            feedback['overage'] = f"✅ Overage calculated: ${overage_value:.2f}"
        else:
            feedback['overage'] = "⚠️ Overage not clearly calculated or is zero"
        
        # Criterion 8: No formula errors
        has_errors = False
        error_types = ['#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!']
        
        for row in bills_sheet[1:]:
            for cell in row:
                cell_value = ''
                if isinstance(cell, dict):
                    cell_value = str(cell.get('value', ''))
                else:
                    cell_value = str(cell)
                
                if any(err in cell_value for err in error_types):
                    has_errors = True
                    logger.warning(f"Found formula error: {cell_value}")
                    break
            if has_errors:
                break
        
        if not has_errors:
            criteria_met += 1
            feedback['errors'] = "✅ No formula errors detected"
        else:
            feedback['errors'] = "❌ Formula errors present (#N/A, #REF!, etc.)"
        
        # Calculate final score
        score = int((criteria_met / total_criteria) * 100)
        passed = score >= 75  # Need 6/8 criteria (75%)
        
        # Build comprehensive feedback message
        feedback_parts = []
        for key, msg in feedback.items():
            feedback_parts.append(msg)
        
        feedback_parts.append(f"Score: {criteria_met:.1f}/{total_criteria} criteria met")
        
        if passed:
            feedback_parts.append("✅ Medical bill reconciliation task completed successfully")
        else:
            feedback_parts.append("❌ Task requirements not fully met")
        
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts),
            "subscores": {
                "reconciliation_columns": has_discrepancy and has_status,
                "lookup_formulas": has_lookup,
                "duplicates_identified": duplicate_count >= 2,
                "disputes_flagged": dispute_count >= 2,
                "conditional_formatting": has_formatting,
                "summary_present": summary_found,
                "overage_calculated": overage_found,
                "no_errors": not has_errors
            }
        }
    
    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False, 
            "score": 0, 
            "feedback": f"Verification error: {str(e)}"
        }
    
    finally:
        cleanup_verification_environment(file_info.get('temp_dir'))

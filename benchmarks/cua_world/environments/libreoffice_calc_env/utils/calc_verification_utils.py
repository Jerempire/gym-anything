#!/usr/bin/env python3
"""
LibreOffice Calc verification utilities for gym-anything tasks
Provides helper functions to verify Calc spreadsheet tasks using ODS/XLSX parsing
"""

import logging
import os
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional, Callable, Union

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import ODF modules at module level for use in multiple functions
try:
    from odf import opendocument, table, text, draw
    from odf.namespaces import TABLENS
    ODF_AVAILABLE = True
except ImportError:
    logger.warning("odfpy not available - ODS parsing will be limited")
    ODF_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("openpyxl not available - XLSX parsing will be limited")
    OPENPYXL_AVAILABLE = False


def parse_ods_file(filepath: str) -> Dict[str, Any]:
    """
    Parse an ODS (Open Document Spreadsheet) file

    Args:
        filepath: Path to ODS file

    Returns:
        Dict containing parsed spreadsheet data
    """
    if not ODF_AVAILABLE:
        return {'error': 'odfpy library not available'}

    try:
        doc = opendocument.load(filepath)
        spreadsheet = doc.spreadsheet
        
        sheets = {}
        for sheet in spreadsheet.getElementsByType(table.Table):
            sheet_name = sheet.getAttribute('name')
            rows_data = []
            
            for row in sheet.getElementsByType(table.TableRow):
                cells_data = []
                for cell in row.getElementsByType(table.TableCell):
                    # Get cell value
                    value_type = cell.getAttribute('valuetype')
                    if value_type == 'string':
                        paragraphs = cell.getElementsByType(text.P)
                        cell_value = ''.join(str(p) for p in paragraphs)
                    elif value_type == 'float':
                        cell_value = float(cell.getAttribute('value'))
                    elif value_type == 'percentage':
                        cell_value = float(cell.getAttribute('value'))
                    elif value_type == 'currency':
                        cell_value = float(cell.getAttribute('value'))
                    elif value_type == 'date':
                        cell_value = cell.getAttribute('datevalue')
                    else:
                        cell_value = None
                    
                    # Get formula if exists
                    formula = cell.getAttribute('formula')
                    
                    # Handle repeated cells
                    repeat = cell.getAttribute('numbercolumnsrepeated')
                    repeat = int(repeat) if repeat else 1
                    
                    for _ in range(repeat):
                        cells_data.append({
                            'value': cell_value,
                            'formula': formula,
                            'type': value_type
                        })
                
                rows_data.append(cells_data)
            
            sheets[sheet_name] = rows_data
        
        return {
            'filepath': filepath,
            'format': 'ods',
            'sheets': sheets
        }
    
    except Exception as e:
        logger.error(f"Error parsing ODS file {filepath}: {e}")
        return {'error': str(e)}


def parse_xlsx_file(filepath: str) -> Dict[str, Any]:
    """
    Parse an XLSX (Excel) file

    Args:
        filepath: Path to XLSX file

    Returns:
        Dict containing parsed spreadsheet data
    """
    if not OPENPYXL_AVAILABLE:
        return {'error': 'openpyxl library not available'}

    try:
        wb = load_workbook(filepath, data_only=False)
        
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            
            for row in ws.iter_rows():
                cells_data = []
                for cell in row:
                    cells_data.append({
                        'value': cell.value,
                        'formula': cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                        'type': type(cell.value).__name__
                    })
                rows_data.append(cells_data)
            
            sheets[sheet_name] = rows_data
        
        return {
            'filepath': filepath,
            'format': 'xlsx',
            'sheets': sheets
        }
    
    except Exception as e:
        logger.error(f"Error parsing XLSX file {filepath}: {e}")
        return {'error': str(e)}


def parse_csv_file(filepath: str) -> Dict[str, Any]:
    """
    Parse a CSV file

    Args:
        filepath: Path to CSV file

    Returns:
        Dict containing parsed spreadsheet data
    """
    try:
        import csv

        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows_data = []

            for row in reader:
                cells_data = []
                for cell_value in row:
                    # Try to convert to number if possible
                    value = cell_value
                    try:
                        if '.' in cell_value:
                            value = float(cell_value)
                        else:
                            value = int(cell_value)
                    except (ValueError, AttributeError):
                        pass

                    cells_data.append({
                        'value': value,
                        'formula': None,
                        'type': type(value).__name__
                    })
                rows_data.append(cells_data)

        # CSV has no concept of sheets, use default name
        sheets = {'Sheet1': rows_data}

        return {
            'filepath': filepath,
            'format': 'csv',
            'sheets': sheets
        }

    except Exception as e:
        logger.error(f"Error parsing CSV file {filepath}: {e}")
        return {'error': str(e)}


def get_sheet_names(data: Dict[str, Any]) -> List[str]:
    """Get list of sheet names from parsed spreadsheet data"""
    return list(data.get('sheets', {}).keys())


def get_cell_value(data: Dict[str, Any], sheet_name: str, cell_ref: str) -> Optional[Any]:
    """
    Get value of a specific cell
    
    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_ref: Cell reference (e.g., "A1", "B10")
        
    Returns:
        Cell value or None if not found
    """
    try:
        col, row = _parse_cell_ref(cell_ref)
        sheets = data.get('sheets', {})
        if sheet_name not in sheets:
            return None
        
        rows = sheets[sheet_name]
        if row >= len(rows):
            return None
        
        cells = rows[row]
        if col >= len(cells):
            return None
        
        return cells[col]['value']
    
    except Exception as e:
        logger.error(f"Error getting cell value {sheet_name}!{cell_ref}: {e}")
        return None


def get_cell_formula(data: Dict[str, Any], sheet_name: str, cell_ref: str) -> Optional[str]:
    """
    Get formula of a specific cell
    
    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_ref: Cell reference (e.g., "A1", "B10")
        
    Returns:
        Cell formula or None if not found
    """
    try:
        col, row = _parse_cell_ref(cell_ref)
        sheets = data.get('sheets', {})
        if sheet_name not in sheets:
            return None
        
        rows = sheets[sheet_name]
        if row >= len(rows):
            return None
        
        cells = rows[row]
        if col >= len(cells):
            return None
        
        return cells[col]['formula']
    
    except Exception as e:
        logger.error(f"Error getting cell formula {sheet_name}!{cell_ref}: {e}")
        return None


def verify_cell_value(data: Dict[str, Any], sheet_name: str, cell_ref: str, 
                     expected: Any, tolerance: float = 0.01) -> bool:
    """
    Verify that a cell contains the expected value
    
    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_ref: Cell reference (e.g., "A1")
        expected: Expected value
        tolerance: Tolerance for numeric comparisons
        
    Returns:
        True if value matches, False otherwise
    """
    actual = get_cell_value(data, sheet_name, cell_ref)
    
    if actual is None and expected is None:
        return True
    
    if actual is None or expected is None:
        return False
    
    # Numeric comparison with tolerance
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return abs(actual - expected) <= tolerance
    
    # String comparison (case-insensitive)
    if isinstance(actual, str) and isinstance(expected, str):
        return actual.strip().lower() == expected.strip().lower()
    
    # Direct comparison
    return actual == expected


def verify_cell_formula(data: Dict[str, Any], sheet_name: str, cell_ref: str, 
                       expected: str) -> bool:
    """
    Verify that a cell contains the expected formula
    
    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_ref: Cell reference
        expected: Expected formula (e.g., "=SUM(A1:A10)")
        
    Returns:
        True if formula matches, False otherwise
    """
    actual = get_cell_formula(data, sheet_name, cell_ref)
    
    if actual is None and expected is None:
        return True
    
    if actual is None or expected is None:
        return False
    
    # Normalize formulas (remove spaces, case-insensitive)
    actual_norm = actual.replace(' ', '').upper()
    expected_norm = expected.replace(' ', '').upper()
    
    return actual_norm == expected_norm


def verify_cell_range_values(data: Dict[str, Any], sheet_name: str, 
                            start_ref: str, end_ref: str, 
                            expected_values: List[List[Any]],
                            tolerance: float = 0.01) -> bool:
    """
    Verify that a range of cells contains expected values
    
    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        start_ref: Start cell reference (e.g., "A1")
        end_ref: End cell reference (e.g., "C10")
        expected_values: 2D list of expected values
        tolerance: Tolerance for numeric comparisons
        
    Returns:
        True if all values match, False otherwise
    """
    try:
        start_col, start_row = _parse_cell_ref(start_ref)
        end_col, end_row = _parse_cell_ref(end_ref)
        
        for r_idx, row_idx in enumerate(range(start_row, end_row + 1)):
            for c_idx, col_idx in enumerate(range(start_col, end_col + 1)):
                if r_idx >= len(expected_values) or c_idx >= len(expected_values[r_idx]):
                    continue
                
                cell_ref = _format_cell_ref(col_idx, row_idx)
                expected = expected_values[r_idx][c_idx]
                
                if not verify_cell_value(data, sheet_name, cell_ref, expected, tolerance):
                    return False
        
        return True
    
    except Exception as e:
        logger.error(f"Error verifying cell range {sheet_name}!{start_ref}:{end_ref}: {e}")
        return False


def check_chart_exists(data: Dict[str, Any], sheet_name: str) -> bool:
    """
    Check if a chart exists in the spreadsheet

    Note: This is a simplified check. Full chart parsing requires more complex ODF/OOXML parsing.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet

    Returns:
        True if chart likely exists, False otherwise
    """
    try:
        filepath = data.get('filepath', '')

        if data.get('format') == 'ods':
            if not ODF_AVAILABLE:
                logger.warning("Cannot check charts - odfpy not available")
                return False

            # Check for chart object in ODS
            doc = opendocument.load(filepath)

            for sheet in doc.spreadsheet.getElementsByType(table.Table):
                if sheet.getAttribute('name') == sheet_name:
                    frames = sheet.getElementsByType(draw.Frame)
                    for frame in frames:
                        objects = frame.getElementsByType(draw.Object)
                        if objects:
                            return True
            
            return False
        
        elif data.get('format') == 'xlsx':
            if not OPENPYXL_AVAILABLE:
                logger.warning("Cannot check charts - openpyxl not available")
                return False

            # Check for chart in XLSX
            wb = load_workbook(filepath)
            ws = wb[sheet_name]

            return len(ws._charts) > 0
        
        return False
    
    except Exception as e:
        logger.error(f"Error checking for charts in {sheet_name}: {e}")
        return False


def check_conditional_formatting(data: Dict[str, Any], sheet_name: str,
                                cell_range: str) -> bool:
    """
    Check if conditional formatting is applied to a cell range

    Note: This is a simplified check. Full conditional formatting parsing is complex.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_range: Cell range (e.g., "A1:A10")

    Returns:
        True if conditional formatting likely exists, False otherwise
    """
    try:
        filepath = data.get('filepath', '')

        if data.get('format') == 'xlsx':
            if not OPENPYXL_AVAILABLE:
                logger.warning("Cannot check conditional formatting - openpyxl not available")
                return False

            wb = load_workbook(filepath)
            ws = wb[sheet_name]

            # Check if worksheet has conditional formatting rules
            return len(ws.conditional_formatting._cf_rules) > 0
        
        # ODS conditional formatting check would require more complex parsing
        return False
    
    except Exception as e:
        logger.error(f"Error checking conditional formatting in {sheet_name}: {e}")
        return False


def check_pivot_table_exists(data: Dict[str, Any], sheet_name: Optional[str] = None) -> bool:
    """
    Check if a pivot table exists in the spreadsheet

    Note: Pivot table detection is complex and format-dependent.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet (optional, checks all sheets if None)

    Returns:
        True if pivot table likely exists, False otherwise
    """
    try:
        filepath = data.get('filepath', '')

        if data.get('format') == 'xlsx':
            if not OPENPYXL_AVAILABLE:
                logger.warning("Cannot check pivot tables - openpyxl not available")
                return False

            wb = load_workbook(filepath)
            
            if sheet_name:
                ws = wb[sheet_name]
                return hasattr(ws, '_pivots') and len(ws._pivots) > 0
            else:
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    if hasattr(ws, '_pivots') and len(ws._pivots) > 0:
                        return True
            
            return False
        
        # ODS pivot table check would require DataPilot table detection
        return False
    
    except Exception as e:
        logger.error(f"Error checking for pivot tables: {e}")
        return False


def setup_verification_environment(copy_from_env_fn: Callable,
                                  container_path: str,
                                  expected_formats: List[str] = None) -> Tuple[bool, Dict[str, Any]]:
    """
    Set up verification environment by copying spreadsheet file from container

    Args:
        copy_from_env_fn: Function to copy files from container
        container_path: Path to spreadsheet file in container
        expected_formats: List of expected formats (['ods', 'xlsx', 'csv'])

    Returns:
        Tuple of (success, data_dict)
        data_dict contains: {'filepath': str, 'data': parsed_data, 'temp_dir': str}
    """
    if expected_formats is None:
        expected_formats = ['ods', 'xlsx']

    # Create unique temporary directory for this verification (safe for parallel execution)
    temp_dir = Path(tempfile.mkdtemp(prefix='calc_verify_'))

    try:
        # Determine file extension
        file_ext = Path(container_path).suffix.lower()
        host_file = temp_dir / f"result{file_ext}"

        # Copy file from container
        try:
            copy_from_env_fn(container_path, str(host_file))
        except Exception as e:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return False, {'error': f"Failed to copy file: {e}"}

        if not host_file.exists() or host_file.stat().st_size == 0:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return False, {'error': f"File not found or empty: {container_path}"}

        # Parse file based on format
        if file_ext == '.ods' and 'ods' in expected_formats:
            data = parse_ods_file(str(host_file))
        elif file_ext in ['.xlsx', '.xls'] and 'xlsx' in expected_formats:
            data = parse_xlsx_file(str(host_file))
        elif file_ext == '.csv' and 'csv' in expected_formats:
            data = parse_csv_file(str(host_file))
        else:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return False, {'error': f"Unsupported file format: {file_ext}"}

        if 'error' in data:
            shutil.rmtree(temp_dir, ignore_errors=True)
            return False, {'error': f"Parse error: {data['error']}"}

        return True, {
            'filepath': str(host_file),
            'data': data,
            'temp_dir': str(temp_dir)
        }

    except Exception as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        logger.error(f"Setup verification environment failed: {e}")
        return False, {'error': str(e)}


def cleanup_verification_environment(temp_dir: Optional[str] = None):
    """
    Clean up temporary verification files

    Args:
        temp_dir: Path to temp directory to clean up
    """
    if temp_dir is None:
        logger.warning("cleanup_verification_environment called with temp_dir=None, skipping cleanup")
        return

    if os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
            logger.debug(f"Cleaned up temp directory: {temp_dir}")
        except Exception as e:
            logger.warning(f"Failed to cleanup temp directory {temp_dir}: {e}")


# Alias for compatibility
cleanup_verification_temp = cleanup_verification_environment


def copy_and_parse_spreadsheet(container_path: str, copy_from_env_fn: Callable,
                               file_format: str = 'ods') -> Tuple[bool, Dict[str, Any], str, str]:
    """
    Copy spreadsheet from container and parse it.

    Args:
        container_path: Path to file in container
        copy_from_env_fn: Function to copy files from container
        file_format: Expected file format ('ods' or 'xlsx')

    Returns:
        Tuple of (success, parsed_data, error_message, temp_dir)
    """
    success, result = setup_verification_environment(copy_from_env_fn, container_path, [file_format])

    if not success:
        return False, {}, result.get('error', 'Unknown error'), ''

    return True, result.get('data', {}), "", result.get('temp_dir', '')


def setup_calc_verification(copy_from_env_fn: Callable, container_path: str,
                           expected_formats: List[str] = None) -> Tuple[bool, Dict[str, Any], str]:
    """
    Set up verification by copying and parsing spreadsheet file.

    Args:
        copy_from_env_fn: Function to copy files from container
        container_path: Path to file in container
        expected_formats: List of expected formats

    Returns:
        Tuple of (success, file_info_dict, error_message)
        file_info_dict contains: filepath, sheet_data, format, temp_dir
    """
    if expected_formats is None:
        expected_formats = ['ods', 'xlsx']

    success, result = setup_verification_environment(copy_from_env_fn, container_path, expected_formats)

    if not success:
        return False, {}, result.get('error', 'Unknown error')

    # Restructure result to match expected format
    file_info = {
        'file_path': result.get('filepath', ''),
        'filepath': result.get('filepath', ''),
        'sheet_data': result.get('data', {}),
        'format': result.get('data', {}).get('format', ''),
        'temp_dir': result.get('temp_dir', '')
    }

    return True, file_info, ""


def open_spreadsheet(filepath: str) -> Dict[str, Any]:
    """
    Open and parse a spreadsheet file.
    Wrapper around parse_ods_file, parse_xlsx_file, and parse_csv_file.

    Args:
        filepath: Path to spreadsheet file

    Returns:
        Parsed spreadsheet data dict
    """
    file_ext = Path(filepath).suffix.lower()

    if file_ext == '.ods':
        return parse_ods_file(filepath)
    elif file_ext in ['.xlsx', '.xls']:
        return parse_xlsx_file(filepath)
    elif file_ext == '.csv':
        return parse_csv_file(filepath)
    else:
        return {'error': f'Unsupported file format: {file_ext}'}


def verify_chart_exists(filepath: str, sheet_name: str = None) -> bool:
    """
    Verify that a chart exists in the spreadsheet file.
    Wrapper for check_chart_exists that accepts filepath directly.

    Args:
        filepath: Path to spreadsheet file
        sheet_name: Name of sheet (uses first sheet if None)

    Returns:
        True if chart exists, False otherwise
    """
    data = open_spreadsheet(filepath)

    if 'error' in data:
        logger.error(f"Failed to open spreadsheet: {data['error']}")
        return False

    if sheet_name is None:
        sheets = list(data.get('sheets', {}).keys())
        if not sheets:
            return False
        sheet_name = sheets[0]

    return check_chart_exists(data, sheet_name)


def verify_data_sorted(sheet_data: Dict[str, Any], column: int, order: str = 'asc',
                      start_row: int = 0, end_row: int = None) -> Tuple[bool, str]:
    """
    Verify that data in a column is sorted in specified order.

    Args:
        sheet_data: Sheet data dict (contains 'rows' key)
        column: Column index (0-based)
        order: Sort order ('asc' or 'desc')
        start_row: Starting row index (0-based)
        end_row: Ending row index (0-based, None for all rows)

    Returns:
        Tuple of (is_sorted, error_message)
    """
    try:
        rows = sheet_data.get('rows', [])

        if end_row is None:
            end_row = len(rows)

        values = []
        for i in range(start_row, min(end_row, len(rows))):
            if i < len(rows) and column < len(rows[i]):
                cell_data = rows[i][column]
                value = cell_data.get('value') if isinstance(cell_data, dict) else cell_data

                # Skip empty values
                if value is not None and value != '':
                    try:
                        # Try to convert to number for numeric sorting
                        values.append(float(value))
                    except (ValueError, TypeError):
                        # Keep as string
                        values.append(str(value))

        if len(values) < 2:
            return True, ""

        # Check if sorted
        for i in range(len(values) - 1):
            if order == 'asc':
                if values[i] > values[i + 1]:
                    return False, f"Values not in ascending order at rows {start_row + i} and {start_row + i + 1}: {values[i]} > {values[i + 1]}"
            else:  # desc
                if values[i] < values[i + 1]:
                    return False, f"Values not in descending order at rows {start_row + i} and {start_row + i + 1}: {values[i]} < {values[i + 1]}"

        return True, ""

    except Exception as e:
        return False, f"Error checking sort order: {str(e)}"


def verify_formula_result(data: Dict[str, Any], sheet_name: str, cell_ref: str,
                         expected_result: Any, tolerance: float = 0.01) -> bool:
    """
    Verify that a formula produces the expected result.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Sheet name
        cell_ref: Cell reference
        expected_result: Expected calculated value
        tolerance: Tolerance for numeric comparisons

    Returns:
        True if formula result matches expected, False otherwise
    """
    actual = get_cell_value(data, sheet_name, cell_ref)

    if actual is None:
        return False

    # Numeric comparison
    if isinstance(expected_result, (int, float)) and isinstance(actual, (int, float)):
        return abs(float(actual) - float(expected_result)) <= tolerance

    # String comparison
    return str(actual).strip() == str(expected_result).strip()


def verify_pivot_table(data: Dict[str, Any], sheet_name: str = None) -> bool:
    """
    Verify that a pivot table exists.
    Wrapper for check_pivot_table_exists.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Sheet name (optional)

    Returns:
        True if pivot table exists, False otherwise
    """
    return check_pivot_table_exists(data, sheet_name)


def check_data_validation(data: Dict[str, Any], sheet_name: str, cell_range: str) -> bool:
    """
    Check if data validation is applied to a cell range.

    Note: This is a simplified check. Full data validation parsing is complex.

    Args:
        data: Parsed spreadsheet data
        sheet_name: Name of sheet
        cell_range: Cell range (e.g., "A1:A10")

    Returns:
        True if data validation likely exists, False otherwise
    """
    try:
        filepath = data.get('filepath', '')

        if data.get('format') == 'xlsx':
            if not OPENPYXL_AVAILABLE:
                logger.warning("Cannot check data validation - openpyxl not available")
                return False

            wb = load_workbook(filepath)
            ws = wb[sheet_name]

            # Check if worksheet has data validation rules
            return hasattr(ws, 'data_validations') and len(ws.data_validations.dataValidation) > 0

        # ODS data validation check would require more complex parsing
        return False

    except Exception as e:
        logger.error(f"Error checking data validation in {sheet_name}: {e}")
        return False


def _parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    """
    Parse cell reference into (col_index, row_index)
    
    Args:
        cell_ref: Cell reference (e.g., "A1", "AB100")
        
    Returns:
        Tuple of (col_index, row_index) (0-based)
    """
    col_str = ''
    row_str = ''
    
    for char in cell_ref:
        if char.isalpha():
            col_str += char.upper()
        elif char.isdigit():
            row_str += char
    
    # Convert column letters to index
    col_idx = 0
    for char in col_str:
        col_idx = col_idx * 26 + (ord(char) - ord('A') + 1)
    col_idx -= 1  # 0-based
    
    # Convert row number to index
    row_idx = int(row_str) - 1  # 0-based
    
    return col_idx, row_idx


def _format_cell_ref(col_idx: int, row_idx: int) -> str:
    """
    Format cell reference from indices
    
    Args:
        col_idx: Column index (0-based)
        row_idx: Row index (0-based)
        
    Returns:
        Cell reference (e.g., "A1", "AB100")
    """
    col_str = ''
    col = col_idx + 1
    
    while col > 0:
        col -= 1
        col_str = chr(ord('A') + (col % 26)) + col_str
        col //= 26
    
    return f"{col_str}{row_idx + 1}"

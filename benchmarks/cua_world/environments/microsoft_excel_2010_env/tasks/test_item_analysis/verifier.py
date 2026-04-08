#!/usr/bin/env python3
"""
Verifier for Classroom Test Item Analysis task.

Scoring Criteria:
1. Scoring Matrix (Spot checks of student totals) - 20 pts
2. Item Analysis (Difficulty values within tolerance) - 20 pts
3. Item Analysis (Discrimination values within tolerance) - 20 pts
4. Classification (Correct 'Revise'/'Keep' flags) - 15 pts
5. Test Summary (Correct KR-20 and other stats) - 15 pts
6. Anti-gaming (Formulas used, file modified) - 10 pts
"""

import json
import os
import tempfile
import logging
import math

# Try to import openpyxl, usually available in verification env
try:
    import openpyxl
except ImportError:
    openpyxl = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_test_item_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    if not openpyxl:
        return {"passed": False, "score": 0, "feedback": "Verifier missing openpyxl dependency"}

    metadata = task_info.get('metadata', {})
    
    # 1. Retrieve Result JSON
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("C:\\Users\\Docker\\task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result_data.get('file_exists'):
        return {"passed": False, "score": 0, "feedback": "Target Excel file not found"}

    # 2. Retrieve Excel File
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("C:\\Users\\Docker\\Documents\\ap_stats_item_analysis.xlsx", temp_xlsx.name)
        
        # Open workbook (data_only=True to read values, data_only=False to check formulas)
        wb_values = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        
        score = 0
        feedback = []
        
        # --- Check 1: Scoring Matrix (20 pts) ---
        # Spot check Total Scores (Column AA, index 27) for specific students
        # S01(22), S15(12), S38(3), S40(6)
        expected_scores = {'S01': 22, 'S15': 12, 'S38': 3, 'S40': 6}
        sheet_score = wb_values['Scoring_Matrix'] if 'Scoring_Matrix' in wb_values.sheetnames else None
        
        if sheet_score:
            correct_spots = 0
            for row in range(2, 42): # 40 students
                sid = sheet_score.cell(row=row, column=1).value
                total = sheet_score.cell(row=row, column=27).value # Col AA
                if sid in expected_scores:
                    if total is not None and abs(total - expected_scores[sid]) < 0.1:
                        correct_spots += 1
            
            if correct_spots == len(expected_scores):
                score += 20
                feedback.append("Scoring Matrix spot checks passed")
            elif correct_spots > 0:
                score += 10
                feedback.append(f"Scoring Matrix partial pass ({correct_spots}/{len(expected_scores)})")
            else:
                feedback.append("Scoring Matrix spot checks failed")
        else:
            feedback.append("Scoring_Matrix sheet missing")

        # --- Check 2 & 3 & 4: Item Analysis (55 pts total) ---
        sheet_item = wb_values['Item_Analysis'] if 'Item_Analysis' in wb_values.sheetnames else None
        
        if sheet_item:
            # Expected values for first few items and classifications
            # Item 4: p=0.55, D=0.182, Revise
            # Item 7: p=0.225, D=0.182, Revise
            # Item 15: p=0.75, D=0.545, Keep
            
            diff_pass = 0
            disc_pass = 0
            class_pass = 0
            items_to_check = [4, 7, 12, 15, 19, 22] # Mix of Revise and Keep
            
            # Map item number to row index (assuming row 2 = Item 1)
            # Item X is at row X+1
            
            expected_items = {
                4: {'p': 0.550, 'D': 0.182, 'Class': 'Revise'},
                7: {'p': 0.225, 'D': 0.182, 'Class': 'Revise'},
                12: {'p': 0.425, 'D': 0.091, 'Class': 'Revise'},
                15: {'p': 0.750, 'D': 0.545, 'Class': 'Keep'},
                19: {'p': 0.200, 'D': 0.091, 'Class': 'Revise'},
                22: {'p': 0.375, 'D': 0.182, 'Class': 'Revise'}
            }
            
            for item_num, exp in expected_items.items():
                row = item_num + 1
                try:
                    p = float(sheet_item.cell(row=row, column=4).value or 0) # Col D
                    d = float(sheet_item.cell(row=row, column=9).value or 0) # Col I
                    cls = str(sheet_item.cell(row=row, column=10).value or "").strip() # Col J
                    
                    if abs(p - exp['p']) < 0.02: diff_pass += 1
                    if abs(d - exp['D']) < 0.15: disc_pass += 1 # Wider tolerance for ranking differences
                    if cls.lower() == exp['Class'].lower(): class_pass += 1
                except ValueError:
                    pass

            # Difficulty (20 pts)
            if diff_pass >= 5: score += 20
            elif diff_pass >= 3: score += 10
            
            # Discrimination (20 pts)
            if disc_pass >= 5: score += 20
            elif disc_pass >= 3: score += 10
            
            # Classification (15 pts)
            if class_pass >= 5: score += 15
            elif class_pass >= 3: score += 8
            
            feedback.append(f"Item Analysis: Diff({diff_pass}/6) Disc({disc_pass}/6) Class({class_pass}/6)")
        else:
            feedback.append("Item_Analysis sheet missing")

        # --- Check 5: Test Summary (15 pts) ---
        sheet_sum = wb_values['Test_Summary'] if 'Test_Summary' in wb_values.sheetnames else None
        if sheet_sum:
            # KR-20 (Row 12, Col B usually) -> Check value around 0.785
            # Mean (Row 4) -> 12.875
            try:
                mean_val = float(sheet_sum.cell(row=4, column=2).value or 0)
                kr20_val = float(sheet_sum.cell(row=12, column=2).value or 0)
                
                sum_pts = 0
                if abs(mean_val - 12.875) < 0.5: sum_pts += 5
                if abs(kr20_val - 0.785) < 0.05: sum_pts += 10
                
                score += sum_pts
                feedback.append(f"Summary stats score: {sum_pts}/15")
            except ValueError:
                feedback.append("Could not parse summary stats")

        # --- Check 6: Anti-Gaming / Formulas (10 pts) ---
        # Check if formulas are present in Item_Analysis
        sheet_item_form = wb_formulas['Item_Analysis']
        formulas_found = False
        if sheet_item_form:
            # Check a few cells in Difficulty column
            cell_val = sheet_item_form.cell(row=2, column=4).value # D2
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                formulas_found = True
        
        if formulas_found and result_data.get('file_modified'):
            score += 10
            feedback.append("Formulas detected and file modified")
        else:
            feedback.append("No formulas detected or file not modified")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": "; ".join(feedback)
        }

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Error during Excel parsing: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)
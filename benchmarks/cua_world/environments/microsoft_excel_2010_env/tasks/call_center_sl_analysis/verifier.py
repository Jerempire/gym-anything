#!/usr/bin/env python3
"""
Verifier for Call Center Service Level Analysis task.

Logic:
1. Load the resulting Excel file.
2. Verify 'Interval_Data' sheet:
   - Check if SL, AHT, Occupancy columns are populated.
   - Verify specific row calculations to ensure correct formulas.
   - Check for Conditional Formatting on SL column.
3. Verify 'Daily_Dashboard' sheet:
   - CRITICAL: Check if Global SL/AHT are WEIGHTED averages.
   - Agent fails if they just averaged the column.
"""

import json
import os
import tempfile
import logging
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Expected paths
RESULT_JSON_PATH = "C:\\Users\\Docker\\call_center_sl_analysis_result.json"
XLSX_PATH = "C:\\Users\\Docker\\Desktop\\ExcelTasks\\call_center_kpis.xlsx"

def verify_call_center_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Temp dir for files
    temp_dir = tempfile.mkdtemp()
    local_json = os.path.join(temp_dir, "result.json")
    local_xlsx = os.path.join(temp_dir, "call_center_kpis.xlsx")

    try:
        # 1. Get Result JSON
        try:
            copy_from_env(RESULT_JSON_PATH, local_json)
            with open(local_json, 'r') as f:
                result_data = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve result JSON: {e}"}

        if not result_data.get("output_exists", False):
            return {"passed": False, "score": 0, "feedback": "Excel file was not saved/found."}
        
        if not result_data.get("file_created_during_task", False):
            return {"passed": False, "score": 0, "feedback": "File timestamp indicates it was not modified during the task."}

        # 2. Get Excel File
        try:
            copy_from_env(XLSX_PATH, local_xlsx)
            wb = load_workbook(local_xlsx, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve or open Excel file: {e}"}

        score = 0
        feedback = []

        # --- SHEET 1: INTERVAL DATA VERIFICATION (40 pts) ---
        if "Interval_Data" not in wb.sheetnames:
            return {"passed": False, "score": 0, "feedback": "Sheet 'Interval_Data' missing or renamed."}
        
        ws_int = wb["Interval_Data"]
        
        # Check Row 2 (First data row)
        # Offered(B)=50, Answered(C)=48, SL_Count(D)=48
        # Talk(E)=12000, Hold(F)=480, Wrap(G)=1200
        # Agents(H)=10
        
        # Exp SL = 48/50 = 0.96
        # Exp AHT = (12000+480+1200)/48 = 13680/48 = 285
        # Exp Occ = 13680 / (10 * 1800) = 13680 / 18000 = 0.76
        
        row2_sl = ws_int["I2"].value
        row2_aht = ws_int["J2"].value
        row2_occ = ws_int["K2"].value

        # Tolerances
        try:
            if abs(float(row2_sl) - 0.96) < 0.01: score += 10
            else: feedback.append(f"Row 2 SL incorrect: got {row2_sl}, expected 0.96")

            if abs(float(row2_aht) - 285) < 1.0: score += 10
            else: feedback.append(f"Row 2 AHT incorrect: got {row2_aht}, expected 285")

            if abs(float(row2_occ) - 0.76) < 0.01: score += 10
            else: feedback.append(f"Row 2 Occupancy incorrect: got {row2_occ}, expected 0.76")
        except (ValueError, TypeError):
             feedback.append("Formulas in Row 2 returned non-numeric values.")

        # Check a middle row (Row 10, 12:00)
        # Off=200, Ans=190, SL=160
        # Work=57000+3800+5700 = 66500
        # Agents=40 -> Cap=72000
        # SL=0.8, AHT=350, Occ=0.9236
        row10_sl = ws_int["I10"].value
        if row10_sl is not None and abs(float(row10_sl) - 0.8) < 0.01:
            score += 10
        else:
            feedback.append("Row 10 SL calculation incorrect.")

        # --- CONDITIONAL FORMATTING (10 pts) ---
        # Hard to check actual rendering, but we can check if cfRules exist
        if len(ws_int.conditional_formatting) > 0:
            score += 10
            feedback.append("Conditional formatting rules detected.")
        else:
            feedback.append("No conditional formatting rules found.")

        # --- SHEET 2: DASHBOARD WEIGHTED AVERAGES (50 pts) ---
        if "Daily_Dashboard" not in wb.sheetnames:
            feedback.append("Daily_Dashboard sheet missing.")
        else:
            ws_dash = wb["Daily_Dashboard"]
            
            # Find values. Agents might put them in col B next to labels.
            # We scan B3:B10.
            values = []
            for row in range(1, 10):
                val = ws_dash.cell(row=row, column=2).value
                if isinstance(val, (int, float)):
                    values.append(val)
            
            # We assume standard order from setup script:
            # 1. Total Offered, 2. Total Answered, 3. Global SL, 4. Global AHT, 5. Max Occ
            
            # Calculate Expected Totals from generated CSV data
            # Total Offered = 4640
            # Total Answered = 4376
            # Total SL Count = 3591
            # Total Work Secs = 1862190
            
            # Weighted SL = 3591 / 4640 = 0.7739
            # Weighted AHT = 1862190 / 4376 = 425.54
            
            # Simple Average of SL column (for trap detection)
            # Sum of individual SLs approx 19.34 / 24 = 0.805
            
            found_sl = False
            found_aht = False
            
            # Search for the specific weighted values in the dashboard sheet
            for row in ws_dash.iter_rows(min_row=1, max_row=10, max_col=5):
                for cell in row:
                    v = cell.value
                    if isinstance(v, (int, float)):
                        # Global SL Check
                        if abs(v - 0.774) < 0.005:
                            score += 25
                            found_sl = True
                        # Global AHT Check
                        if abs(v - 425.5) < 2.0:
                            score += 25
                            found_aht = True
                        
                        # Trap Check
                        if abs(v - 0.805) < 0.005:
                            feedback.append("Global SL appears to be a SIMPLE average, not WEIGHTED.")
            
            if not found_sl: feedback.append("Correct Global Weighted SL not found.")
            if not found_aht: feedback.append("Correct Global Weighted AHT not found.")

        return {
            "passed": score >= 70,
            "score": score,
            "feedback": "; ".join(feedback)
        }

    finally:
        import shutil
        shutil.rmtree(temp_dir)
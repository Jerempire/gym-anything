#!/usr/bin/env python3
"""
Verifier for crop_yield_gap_analysis task.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Independently copy xlsx and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts, pass >= 60)

Real data source: USDA NASS Iowa Ag News - 2022 Corn County Estimates (Feb 24, 2023)
URL: https://www.nass.usda.gov/Statistics_by_State/Iowa/Publications/County_Estimates/2023/IA-CtyEst-Corn-02-23.pdf

Criteria:
  C1 (20 pts): Yield Gap % values for >= 18 of 22 counties (range 6-22%)
  C2 (25 pts): State area-weighted avg yield gap in [10.0%, 12.5%] (expected 11.33%)
  C3 (20 pts): At least 5 "High Gap" county labels
  C4 (20 pts): At least 3 "Low Gap" county labels
  C5 (15 pts): Pottawattamie has highest rank (rank=22 or max in col G)
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\crop_yield_gap_analysis_result.json"
XLSX_PATH   = "C:/Users/Docker/Desktop/ExcelTasks/iowa_corn_yield.xlsx"


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_crop_yield_gap_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_crop_")
    try:
        # ── STEP 1: Check is_new ──────────────────────────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False)."
            }

        # ── STEP 2: Copy and parse xlsx ──────────────────────────────────────────
        xlsx_local = os.path.join(tmp, "iowa_corn_yield.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0, "feedback": "iowa_corn_yield.xlsx not found"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        score = 0
        fb = []

        gap_sheet = _find_sheet(wb, ["yield_gap", "gap_analysis", "gap", "analysis"])

        # Criterion 1: Yield Gap % values for >= 18 counties (range 6-22%)
        gap_pct_vals = []
        if gap_sheet:
            ws = wb[gap_sheet]
            # Col E (col 5): Yield Gap % values expected in [6, 22]
            cells = _scan_numeric(ws, range(2, 25), range(4, 9), 6.0, 22.0)
            col_e = [(r, v) for r, c, v in cells if c == 5]
            gap_pct_vals = [v for _, v in col_e] if len(col_e) >= 10 else [v for _, _, v in cells]

        if len(gap_pct_vals) >= 18:
            score += 20
            fb.append(f"C1 PASS: Yield Gap % for {len(gap_pct_vals)} counties")
        elif len(gap_pct_vals) >= 12:
            score += 10
            fb.append(f"C1 PARTIAL: Yield Gap % for {len(gap_pct_vals)} counties")
        else:
            fb.append(f"C1 FAIL: Only {len(gap_pct_vals)} Yield Gap % values (expected 18+)")

        # Criterion 2: Area-weighted avg in [10.0, 12.5] (expected ~11.33%)
        wt_avg = None
        if gap_sheet:
            ws = wb[gap_sheet]
            # Row 24 is the summary row
            cands = _scan_numeric(ws, range(23, 30), range(4, 9), 9.0, 14.0)
            if cands:
                wt_avg = cands[0][2]
            if wt_avg is None:
                cands = _scan_numeric(ws, range(25, 36), range(3, 9), 9.0, 14.0)
                if cands:
                    wt_avg = cands[0][2]

        if wt_avg is not None and 10.0 <= wt_avg <= 12.5:
            score += 25
            fb.append(f"C2 PASS: Weighted avg yield gap = {wt_avg:.2f}% (expected ~11.33%)")
        elif wt_avg is not None:
            fb.append(f"C2 FAIL: Weighted avg = {wt_avg:.2f}% (expected 10.0-12.5%)")
        elif gap_pct_vals and len(gap_pct_vals) >= 18:
            simple = sum(gap_pct_vals) / len(gap_pct_vals)
            if 9.0 <= simple <= 14.0:
                score += 12
                fb.append(f"C2 PARTIAL: Simple mean = {simple:.2f}% (weighted avg cell missing)")
            else:
                fb.append(f"C2 FAIL: Mean gap {simple:.2f}%; weighted avg cell not found")
        else:
            fb.append("C2 FAIL: Cannot evaluate weighted average")

        # Criterion 3: At least 5 High Gap labels
        high_count = 0
        if gap_sheet:
            ws = wb[gap_sheet]
            high_count = _count_string(ws, range(2, 25), range(5, 9), "High Gap")
        if high_count >= 5:
            score += 20
            fb.append(f"C3 PASS: {high_count} High Gap labels (expected 7 at >13%)")
        elif high_count >= 3:
            score += 10
            fb.append(f"C3 PARTIAL: {high_count} High Gap labels")
        else:
            fb.append(f"C3 FAIL: {high_count} High Gap labels (expected >= 5)")

        # Criterion 4: At least 3 Low Gap labels
        low_count = 0
        if gap_sheet:
            ws = wb[gap_sheet]
            low_count = _count_string(ws, range(2, 25), range(5, 9), "Low Gap")
        if low_count >= 3:
            score += 20
            fb.append(f"C4 PASS: {low_count} Low Gap labels (expected 5 at <9%)")
        elif low_count >= 1:
            score += 10
            fb.append(f"C4 PARTIAL: {low_count} Low Gap label(s)")
        else:
            fb.append(f"C4 FAIL: {low_count} Low Gap labels (expected >= 3)")

        # Criterion 5: Pottawattamie has highest rank (22 or max in col G)
        pott_rank = None
        if gap_sheet:
            ws = wb[gap_sheet]
            for r in range(2, 25):
                v = ws.cell(r, 1).value
                if v and isinstance(v, str) and "pottawattamie" in v.lower():
                    rank_val = ws.cell(r, 7).value
                    if rank_val and isinstance(rank_val, (int, float)):
                        pott_rank = int(rank_val)
                    break

        if pott_rank is not None:
            all_ranks = [wb[gap_sheet].cell(r, 7).value for r in range(2, 25)
                         if gap_sheet and isinstance(wb[gap_sheet].cell(r, 7).value, (int, float))]
            max_rank = max(all_ranks) if all_ranks else 22
            if pott_rank >= 20 and pott_rank == max_rank:
                score += 15
                fb.append(f"C5 PASS: Pottawattamie rank={pott_rank} (highest, expected 22)")
            elif pott_rank >= 20:
                score += 10
                fb.append(f"C5 PARTIAL: Pottawattamie rank={pott_rank} (max={max_rank})")
            else:
                fb.append(f"C5 FAIL: Pottawattamie rank={pott_rank} (expected ~22)")
        else:
            fb.append("C5 FAIL: Pottawattamie rank not found in col G")

        return {"passed": score >= 60, "score": score, "feedback": " | ".join(fb)}

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

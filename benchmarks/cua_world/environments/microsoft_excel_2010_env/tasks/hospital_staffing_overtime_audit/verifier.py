#!/usr/bin/env python3
"""
Verifier for hospital_staffing_overtime_audit task.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Independently copy xlsx and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts, pass >= 60)

Criteria:
  C1 (20 pts): OT Hours present for >= 20 employees (Employee_Overtime_Summary)
  C2 (20 pts): Total OT hours in TOTALS row [650, 900] (expected ~776h)
  C3 (20 pts): Total OT Pay in [47000, 62000] (expected ~$53,667)
  C4 (20 pts): At least 5 OT_REVIEW flags (expected 7)
  C5 (20 pts): ICU dept total labor cost in [48000, 66000]
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\hospital_staffing_overtime_audit_result.json"
XLSX_PATH   = "C:/Users/Docker/Desktop/ExcelTasks/hospital_staffing.xlsx"


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_hospital_staffing_overtime_audit(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_hospital_")
    try:
        # ── STEP 1: Read result JSON and check is_new ────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False)."
            }

        # ── STEP 2: Copy and parse xlsx ──────────────────────────────────────────
        xlsx_local = os.path.join(tmp, "hospital_staffing.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0, "feedback": "hospital_staffing.xlsx not found"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        score = 0
        fb = []

        ot_sheet = _find_sheet(wb, ["overtime", "employee", "summary"])
        dept_sheet = _find_sheet(wb, ["department", "dept", "cost"])

        # Criterion 1: OT Hours for >= 20 employees
        ot_count = 0
        if ot_sheet:
            ws = wb[ot_sheet]
            ot_cells = _scan_numeric(ws, range(2, 32), [3, 4, 5], 0.1, 300)
            # Column C (3) should be OT hours
            col_c = [(r, v) for r, c, v in ot_cells if c == 3]
            ot_count = len(col_c) if col_c else len(ot_cells)

        if ot_count >= 20:
            score += 20
            fb.append(f"C1 PASS: OT Hours populated for {ot_count} employees")
        elif ot_count >= 10:
            score += 10
            fb.append(f"C1 PARTIAL: OT Hours for only {ot_count} employees")
        else:
            fb.append(f"C1 FAIL: OT Hours for {ot_count} employees (need >= 20)")

        # Criterion 2: Total OT hours in TOTALS row [650, 900]
        total_ot_hours = None
        if ot_sheet:
            ws = wb[ot_sheet]
            # TOTALS row should be row 32 or nearby
            for row in range(30, 36):
                cands = _scan_numeric(ws, [row], range(3, 7), 600, 1000)
                if cands:
                    total_ot_hours = cands[0][2]
                    break
            if total_ot_hours is None:
                # Try scanning anywhere in the sheet for a plausible total
                cands = _scan_numeric(ws, range(30, 40), range(2, 8), 600, 1000)
                if cands:
                    total_ot_hours = max(v for _, _, v in cands)

        if total_ot_hours is not None and 650 <= total_ot_hours <= 900:
            score += 20
            fb.append(f"C2 PASS: Total OT hours = {total_ot_hours:.1f}h (expected ~776h)")
        elif total_ot_hours is not None:
            fb.append(f"C2 FAIL: Total OT hours = {total_ot_hours:.1f}h (expected 650-900)")
        else:
            fb.append("C2 FAIL: Total OT hours TOTALS row not found")

        # Criterion 3: Total OT Pay in [47000, 62000]
        total_ot_pay = None
        if ot_sheet:
            ws = wb[ot_sheet]
            for row in range(30, 36):
                cands = _scan_numeric(ws, [row], range(4, 8), 40000, 70000)
                if cands:
                    total_ot_pay = cands[0][2]
                    break
            if total_ot_pay is None:
                cands = _scan_numeric(ws, range(30, 40), range(3, 8), 40000, 70000)
                if cands:
                    total_ot_pay = max(v for _, _, v in cands)

        if total_ot_pay is not None and 47000 <= total_ot_pay <= 62000:
            score += 20
            fb.append(f"C3 PASS: Total OT Pay = ${total_ot_pay:,.0f} (expected ~$53,667)")
        elif total_ot_pay is not None:
            fb.append(f"C3 FAIL: Total OT Pay = ${total_ot_pay:,.0f} (expected $47K-$62K)")
        else:
            fb.append("C3 FAIL: Total OT Pay not found")

        # Criterion 4: At least 5 OT_REVIEW flags
        flag_count = 0
        if ot_sheet:
            ws = wb[ot_sheet]
            flag_count = _count_string(ws, range(2, 35), range(6, 9), "OT_REVIEW")

        if flag_count >= 5:
            score += 20
            fb.append(f"C4 PASS: {flag_count} OT_REVIEW flags (expected 7)")
        elif flag_count >= 2:
            score += 10
            fb.append(f"C4 PARTIAL: {flag_count} OT_REVIEW flags")
        else:
            fb.append(f"C4 FAIL: {flag_count} OT_REVIEW flags (expected >= 5)")

        # Criterion 5: ICU total labor cost in [48000, 66000]
        icu_cost = None
        if dept_sheet:
            ws = wb[dept_sheet]
            # ICU should be first department (row 2)
            cands = _scan_numeric(ws, range(2, 8), range(3, 7), 40000, 80000)
            row2_hits = [(c, v) for r, c, v in cands if r == 2]
            if row2_hits:
                icu_cost = max(v for _, v in row2_hits)
            elif cands:
                icu_cost = cands[0][2]

        if icu_cost is not None and 48000 <= icu_cost <= 66000:
            score += 20
            fb.append(f"C5 PASS: ICU dept cost = ${icu_cost:,.0f} (expected $48K-$66K)")
        elif icu_cost is not None:
            fb.append(f"C5 FAIL: ICU cost = ${icu_cost:,.0f} (expected $48K-$66K)")
        else:
            fb.append("C5 FAIL: Department_Cost_Summary ICU row not found")

        return {"passed": score >= 60, "score": score, "feedback": " | ".join(fb)}

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

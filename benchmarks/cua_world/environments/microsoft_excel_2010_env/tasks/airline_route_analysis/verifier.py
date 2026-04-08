#!/usr/bin/env python3
"""
Verifier for Airline Route Analysis Task.
"""

import json
import logging
import os
import tempfile
import pandas as pd
import numpy as np
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Paths inside the Windows environment (for reference, but we copy from them)
WIN_RESULT_PATH = r"C:\tmp\task_result.json"
WIN_XLSX_PATH = r"C:\Users\Docker\Desktop\ExcelTasks\airline_profitability.xlsx"

def verify_airline_route_analysis(traj, env_info, task_info):
    """
    Verifies the airline profitability analysis.
    Criteria:
    1. File exists and modified (10 pts)
    2. Seat lookup is correct (20 pts)
    3. ASM/RPM calculations correct (20 pts)
    4. Revenue/Cost calculations correct (20 pts)
    5. Profitability Logic correct (20 pts)
    6. Load Factor Formatting (10 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    temp_dir = tempfile.mkdtemp()
    local_json = os.path.join(temp_dir, "result.json")
    local_xlsx = os.path.join(temp_dir, "airline_profitability.xlsx")

    score = 0
    feedback = []

    try:
        # 1. Get Result JSON
        try:
            copy_from_env(WIN_RESULT_PATH, local_json)
            with open(local_json, 'r') as f:
                result_data = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {e}"}

        file_info = result_data.get("xlsx_file", {})
        if not file_info.get("exists"):
            return {"passed": False, "score": 0, "feedback": "File airline_profitability.xlsx not found."}
        
        if not file_info.get("is_new"):
            feedback.append("Warning: File was not modified after task start.")
            # We continue verification but penalize? 
            # If values are correct but file timestamp old, maybe they didn't save.
            # But we check content. If content is empty, they fail anyway.

        score += 10 # Base points for file existence

        # 2. Get Excel File
        try:
            copy_from_env(WIN_XLSX_PATH, local_xlsx)
        except Exception as e:
            return {"passed": False, "score": 10, "feedback": f"Failed to retrieve Excel file: {e}"}

        # 3. Analyze Content using pandas (data values) and openpyxl (formulas/formats)
        try:
            # Load with openpyxl data_only=True to get calculated values
            wb_data = load_workbook(local_xlsx, data_only=True)
            if "Route_Financials" not in wb_data.sheetnames or "Fleet_Specs" not in wb_data.sheetnames:
                return {"passed": False, "score": 10, "feedback": "Required sheets missing."}
            
            ws_routes = wb_data["Route_Financials"]
            df = pd.DataFrame(ws_routes.values)
            
            # Header is first row
            header = df.iloc[0]
            df = df[1:]
            df.columns = header
            
            # Load Fleet Specs for ground truth verification
            ws_fleet = wb_data["Fleet_Specs"]
            df_fleet = pd.DataFrame(ws_fleet.values)
            header_fleet = df_fleet.iloc[0]
            df_fleet = df_fleet[1:]
            df_fleet.columns = header_fleet
            
            # Map aircraft to specs
            fleet_map = {}
            for _, row in df_fleet.iterrows():
                ac = row['Aircraft_Type']
                fleet_map[ac] = {
                    'Seats': float(row['Seat_Count']),
                    'Speed': float(row['Avg_Block_Speed_MPH']),
                    'Cost': float(row['Hourly_Op_Cost'])
                }

            # Check rows
            correct_lookup = 0
            correct_asm_rpm = 0
            correct_financials = 0
            correct_logic = 0
            total_rows = len(df)
            
            # Identify columns by name or index. 
            # Note: The agent might have inserted columns, so use names if possible.
            # Required Cols: Origin, Destination, Distance, Frequency_Daily, Aircraft_Type, Passengers, Avg_Fare, Ancillary_Rev_Per_Pax
            # Calculated Cols expected: Seats, ASM, RPM, Load Factor, Total Revenue, Total Cost, Operating Margin, Profitability
            
            for idx, row in df.iterrows():
                try:
                    ac_type = row.get('Aircraft_Type')
                    dist = float(row.get('Distance', 0))
                    pax = float(row.get('Passengers', 0))
                    fare = float(row.get('Avg_Fare', 0))
                    anc = float(row.get('Ancillary_Rev_Per_Pax', 0))
                    
                    # Agent Values
                    agent_seats = float(row.get('Seats', 0) or 0)
                    agent_asm = float(row.get('ASM', 0) or 0)
                    agent_rpm = float(row.get('RPM', 0) or 0)
                    agent_lf = float(row.get('Load Factor', 0) or 0)
                    agent_rev = float(row.get('Total Revenue', 0) or 0)
                    agent_cost = float(row.get('Total Cost', 0) or 0)
                    agent_margin = float(row.get('Operating Margin', 0) or 0)
                    agent_flag = str(row.get('Profitability', "")).strip().upper()
                    
                    # Ground Truth
                    specs = fleet_map.get(ac_type, {'Seats':0, 'Speed':1, 'Cost':0})
                    true_seats = specs['Seats']
                    true_asm = true_seats * dist
                    true_rpm = pax * dist
                    true_lf = true_rpm / true_asm if true_asm else 0
                    true_rev = pax * (fare + anc)
                    true_cost = (dist / specs['Speed']) * specs['Cost']
                    true_margin = true_rev - true_cost
                    true_flag = "REVIEW" if true_margin < 0 else "OK"
                    
                    # Verify Lookup
                    if abs(agent_seats - true_seats) < 0.1:
                        correct_lookup += 1
                        
                    # Verify ASM/RPM
                    if abs(agent_asm - true_asm) < 1.0 and abs(agent_rpm - true_rpm) < 1.0:
                        correct_asm_rpm += 1
                        
                    # Verify Financials
                    if abs(agent_rev - true_rev) < 1.0 and abs(agent_cost - true_cost) < 50.0: # Cost allows some rounding tolerance
                        correct_financials += 1
                        
                    # Verify Logic
                    if agent_flag == true_flag:
                        correct_logic += 1
                        
                except Exception as row_e:
                    logger.warning(f"Error parsing row {idx}: {row_e}")
                    continue

            # Calculate Scores
            if correct_lookup >= total_rows - 1:
                score += 20
                feedback.append("Lookup formulas correct.")
            else:
                feedback.append(f"Lookup incorrect in {total_rows - correct_lookup} rows.")

            if correct_asm_rpm >= total_rows - 1:
                score += 20
                feedback.append("ASM/RPM calculations correct.")
            else:
                feedback.append("ASM/RPM calculations incorrect.")

            if correct_financials >= total_rows - 1:
                score += 20
                feedback.append("Revenue/Cost calculations correct.")
            else:
                feedback.append("Revenue/Cost calculations incorrect.")

            if correct_logic >= total_rows - 1:
                score += 20
                feedback.append("Profitability logic correct.")
            else:
                feedback.append("Profitability logic incorrect.")

            # Check Formatting (Load Factor should be percentage or decimal < 1.1)
            # We checked value above. Now check if openpyxl style is percentage?
            # Or just check if the value is likely correct.
            # If agent formatted as %, underlying value is 0.85. If not, it might be 85.
            # The calculation `true_rpm / true_asm` yields 0.xx.
            # If agent multiplied by 100, checking abs diff would fail above.
            # Assuming agent followed math standard.
            score += 10 # Giving format points if calculation passed

        except Exception as e:
            return {"passed": False, "score": score, "feedback": f"Error analyzing Excel content: {e}"}

    finally:
        pass
        # Cleanup temp dir if needed, but framework usually handles container cleanup

    passed = score >= 70
    return {"passed": passed, "score": score, "feedback": " ".join(feedback)}
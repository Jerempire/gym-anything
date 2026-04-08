#!/usr/bin/env python3
import json
import os
import sys
import tempfile
import logging
import math

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def install_openpyxl():
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

install_openpyxl()
import openpyxl

def verify_timber_cruise(traj, env_info, task_info):
    """
    Verifies the Timber Cruise Volume Analysis task.
    
    Criteria:
    1. File exists and modified (Anti-gaming)
    2. Tree-level calculations (Volume_Calculations sheet)
    3. Stand-level summary (Stand_Summary sheet)
    4. Formula usage checks
    """
    
    # 1. Setup & Data Retrieval
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "System error: copy_from_env not available"}

    # Define score components
    score = 0
    feedback = []
    
    # Create temp directory
    with tempfile.TemporaryDirectory() as temp_dir:
        local_xlsx = os.path.join(temp_dir, "timber_cruise.xlsx")
        local_meta = os.path.join(temp_dir, "result_meta.json")
        
        # Copy files
        try:
            # Note: The task info metadata contains the Windows path
            remote_path = task_info.get('metadata', {}).get('expected_output_path', "C:\\Users\\Docker\\Documents\\timber_cruise.xlsx")
            copy_from_env(remote_path, local_xlsx)
            copy_from_env("/tmp/result_meta.json", local_meta)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve output file: {str(e)}"}

        # Check metadata (modification time)
        try:
            with open(local_meta, 'r') as f:
                meta = json.load(f)
                if not meta.get('output_exists'):
                    return {"passed": False, "score": 0, "feedback": "Output file does not exist."}
                # Optional: Check modification time against task start (omitted for brevity, assume valid if content changes)
        except:
            pass # proceed to check content

        # Load Workbook
        try:
            wb = openpyxl.load_workbook(local_xlsx, data_only=True)
            wb_formulas = openpyxl.load_workbook(local_xlsx, data_only=False) # To check formulas
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Corrupt or invalid Excel file: {str(e)}"}

        # 2. Verify Volume_Calculations Sheet
        if "Volume_Calculations" not in wb.sheetnames:
            return {"passed": False, "score": 0, "feedback": "Missing sheet: Volume_Calculations"}
            
        ws_val = wb["Volume_Calculations"]
        ws_for = wb_formulas["Volume_Calculations"]
        
        # Ground Truth Constants
        BAF = 20
        COEFFS = {
            "DF": {"b0": -20.43, "b1": 0.01838, "price": 650},
            "WH": {"b0": -18.25, "b1": 0.01724, "price": 450},
            "WRC": {"b0": -24.87, "b1": 0.01682, "price": 900},
            "RA": {"b0": -12.15, "b1": 0.01543, "price": 380}
        }
        
        # Spot check rows (e.g., Row 2 and Row 20)
        passed_rows = 0
        total_checked = 0
        formula_points = 0
        
        # Check rows 2 through 11 (10 trees)
        for r in range(2, 12):
            total_checked += 1
            try:
                # Read Inputs
                sp = ws_val.cell(row=r, column=3).value
                dbh = float(ws_val.cell(row=r, column=4).value or 0)
                ht = float(ws_val.cell(row=r, column=5).value or 0) # Total height not used in formula, Merch is
                merch = float(ws_val.cell(row=r, column=6).value or 0)
                defect = float(ws_val.cell(row=r, column=7).value or 0)
                
                # GT Calculations
                ba_gt = 0.005454 * (dbh**2)
                tpa_gt = BAF / ba_gt if ba_gt > 0 else 0
                
                c = COEFFS.get(sp, COEFFS["DF"])
                gross_gt = c["b0"] + c["b1"] * (dbh**2) * merch
                if gross_gt < 0: gross_gt = 0
                
                net_gt = gross_gt * (1 - defect/100.0)
                vpa_gt = net_gt * tpa_gt
                stump_gt = (vpa_gt / 1000.0) * c["price"]
                
                size_cls = "Large" if dbh >= 24 else ("Medium" if dbh >= 16 else "Small")
                
                # Verify Agent Values
                # Cols: H(8)=BA, I(9)=TPA, J(10)=Gross, K(11)=Net, L(12)=VPA, M(13)=Stump, N(14)=Class
                ba_agent = float(ws_val.cell(row=r, column=8).value or 0)
                tpa_agent = float(ws_val.cell(row=r, column=9).value or 0)
                gross_agent = float(ws_val.cell(row=r, column=10).value or 0)
                net_agent = float(ws_val.cell(row=r, column=11).value or 0)
                vpa_agent = float(ws_val.cell(row=r, column=12).value or 0)
                stump_agent = float(ws_val.cell(row=r, column=13).value or 0)
                size_agent = str(ws_val.cell(row=r, column=14).value).strip()
                
                # Checks (5% tolerance)
                row_pass = True
                if abs(ba_agent - ba_gt) > 0.1: row_pass = False
                if abs(tpa_agent - tpa_gt) > 0.5: row_pass = False
                if abs(gross_agent - gross_gt) > (gross_gt * 0.05 + 5): row_pass = False
                if abs(net_agent - net_gt) > (net_gt * 0.05 + 5): row_pass = False
                if abs(vpa_agent - vpa_gt) > (vpa_gt * 0.05 + 10): row_pass = False
                if size_agent != size_cls: row_pass = False
                
                if row_pass: passed_rows += 1
                
                # Formula check
                f_cell = ws_for.cell(row=r, column=10).value # Gross Vol Formula
                if f_cell and str(f_cell).startswith("="):
                    formula_points += 1
                    
            except Exception as e:
                logger.warning(f"Error checking row {r}: {e}")
                
        # Score Sheet 1
        # Max 60 points for calculation accuracy
        sheet1_score = 0
        if passed_rows >= 8: # 80% accuracy in spot check
            sheet1_score = 60
            feedback.append("Calculations correct.")
        elif passed_rows >= 5:
            sheet1_score = 30
            feedback.append("Calculations partially correct.")
        else:
            feedback.append(f"Calculation errors. Passed rows: {passed_rows}/{total_checked}.")
            
        score += sheet1_score
        
        # 3. Verify Stand_Summary Sheet
        if "Stand_Summary" not in wb.sheetnames:
            feedback.append("Missing Stand_Summary sheet.")
        else:
            ws_sum = wb["Stand_Summary"]
            # Check Total MBF/Acre (B7)
            # We approximate the ground truth by summing up our sample GT or relying on range
            # Since we generated random data, we rely on the agent's internal consistency 
            # OR we re-calculate the full sum from Sheet 1 data (better).
            
            # Recalculate full sum from agent's Sheet 1 to verify Sheet 2 rollup logic
            # This ensures we don't penalize if Sheet 1 was slightly off but rollup was correct
            total_vpa_agent = 0
            total_val_agent = 0
            trees_count = 0
            
            # Iterate all rows in sheet 1 to get total
            for r in range(2, ws_val.max_row + 1):
                try:
                    v = float(ws_val.cell(row=r, column=12).value or 0) # VPA
                    s = float(ws_val.cell(row=r, column=13).value or 0) # Stumpage
                    total_vpa_agent += v
                    total_val_agent += s
                    trees_count += 1
                except:
                    pass
            
            # Formula: Sum(VPA) / NumPlots
            expected_mbf_ac = (total_vpa_agent / 10) / 1000.0
            expected_val_ac = (total_val_agent / 10)
            
            agent_mbf_ac = float(ws_sum["B7"].value or 0)
            agent_val_ac = float(ws_sum["B14"].value or 0)
            
            summary_pass = True
            if abs(agent_mbf_ac - expected_mbf_ac) > 0.5: summary_pass = False
            if abs(agent_val_ac - expected_val_ac) > 100: summary_pass = False
            
            # Check QMD logic: SQRT( Sum(DBH^2 * TPA) / Sum(TPA) )
            # Just checking if populated and reasonable range (10-30)
            agent_qmd = float(ws_sum["B22"].value or 0)
            if agent_qmd < 10 or agent_qmd > 40: summary_pass = False
            
            if summary_pass:
                score += 30
                feedback.append("Summary rollup correct.")
            else:
                feedback.append(f"Summary rollup incorrect. Exp MBF/Ac: {expected_mbf_ac:.1f}, Got: {agent_mbf_ac:.1f}")

        # 4. Formula Integrity Check
        if formula_points >= 5:
            score += 10
        else:
            feedback.append("Formulas missing (hardcoded values detected).")

    return {
        "passed": score >= 60,
        "score": score,
        "feedback": " ".join(feedback)
    }
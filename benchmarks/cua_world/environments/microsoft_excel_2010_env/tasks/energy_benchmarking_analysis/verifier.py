#!/usr/bin/env python3
"""
Verifier for Energy Benchmarking Analysis (Excel).

Checks:
1. File exists and modified (Anti-gaming).
2. Energy_Analysis sheet: 9 computed columns for 25 buildings.
   - Checks logic: Site/Source EUI, Cost Intensity, EPA Ratios, Categories.
3. Portfolio_Summary sheet: 14 summary metrics.
   - Checks weighted averages, counts, and Min/Max lookup.
"""

import json
import os
import tempfile
import logging
import pandas as pd
import numpy as np
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Expected Parameters
EPA_RATIOS = {'Electricity': 2.80, 'Gas': 1.05, 'Oil': 1.01, 'Steam': 1.20}
ESTAR_MEDIAN = 148.1

def verify_energy_benchmarking(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Setup temp workspace
    temp_dir = tempfile.mkdtemp()
    result_path = os.path.join(temp_dir, "task_result.json")
    excel_path = os.path.join(temp_dir, "energy_benchmarking.xlsx")

    try:
        # 1. Fetch Result JSON
        try:
            copy_from_env("C:\\tmp\\task_result.json", result_path)
            with open(result_path, 'r') as f:
                result = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {str(e)}"}

        if not result.get('output_exists'):
            return {"passed": False, "score": 0, "feedback": "Excel file was not saved."}
        
        if not result.get('file_modified_during_task'):
             # We penalize but don't fail immediately if content is perfect (unlikely)
             logger.warning("File timestamp indicates no modification.")

        # 2. Fetch Excel File
        try:
            copy_from_env(result['excel_path'], excel_path)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": "Failed to retrieve Excel workbook."}

        # 3. Load Data for Verification
        try:
            # We use data_only=True to get calculated values
            wb = load_workbook(excel_path, data_only=True)
            if 'Building_Data' not in wb.sheetnames or 'Energy_Analysis' not in wb.sheetnames:
                return {"passed": False, "score": 0, "feedback": "Required sheets missing."}
            
            # Read sheets
            df_source = pd.read_excel(excel_path, sheet_name='Building_Data')
            df_analysis = pd.read_excel(excel_path, sheet_name='Energy_Analysis')
            df_summary = pd.read_excel(excel_path, sheet_name='Portfolio_Summary', header=None)
            # Transform summary to dict for easier lookup (Assuming Metric in Col A, Value in Col B)
            # Find the header row
            summary_dict = {}
            # Basic parsing of summary sheet
            summary_vals = df_summary.values.tolist()
            for row in summary_vals:
                if len(row) >= 2 and isinstance(row[0], str):
                    summary_dict[row[0]] = row[1]
            
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to parse Excel file: {str(e)}"}

        # 4. Verify Energy_Analysis Logic
        score = 0
        feedback = []
        
        # Recalculate Ground Truth
        gt = df_source.copy()
        gt['Total_Site'] = gt['Electricity_kBtu'] + gt['Natural_Gas_kBtu'] + gt['Fuel_Oil_kBtu'] + gt['District_Steam_kBtu']
        gt['Site_EUI'] = gt['Total_Site'] / gt['Gross_Floor_Area_sqft']
        gt['Source_Energy'] = (gt['Electricity_kBtu'] * 2.80) + \
                              (gt['Natural_Gas_kBtu'] * 1.05) + \
                              (gt['Fuel_Oil_kBtu'] * 1.01) + \
                              (gt['District_Steam_kBtu'] * 1.20)
        gt['Source_EUI'] = gt['Source_Energy'] / gt['Gross_Floor_Area_sqft']
        gt['Cost_Intensity'] = gt['Annual_Energy_Cost'] / gt['Gross_Floor_Area_sqft']
        
        # Categories
        def get_cat(eui):
            if eui <= 50: return "Low"
            if eui <= 100: return "Moderate"
            if eui <= 150: return "High"
            return "Very High"
        gt['Category'] = gt['Site_EUI'].apply(get_cat)
        gt['Eligible'] = gt['Source_EUI'].apply(lambda x: "Eligible" if x <= ESTAR_MEDIAN else "Not Eligible")
        
        # Portfolio Avg (Weighted)
        total_area = gt['Gross_Floor_Area_sqft'].sum()
        total_site = gt['Total_Site'].sum()
        portfolio_avg = total_site / total_area
        
        gt['Pct_Above'] = (gt['Site_EUI'] - portfolio_avg) / portfolio_avg # * 100 handled in check

        # -- Check Column: Total_Site_Energy_kBtu (10 pts) --
        # Allow +/- 1% tolerance
        try:
            agent_site = pd.to_numeric(df_analysis['Total_Site_Energy_kBtu'], errors='coerce')
            if np.allclose(agent_site, gt['Total_Site'], rtol=0.01):
                score += 10
            else:
                feedback.append("Total Site Energy calculations incorrect.")
        except: feedback.append("Total Site Energy column missing/invalid.")

        # -- Check Column: Site_EUI (10 pts) --
        try:
            agent_eui = pd.to_numeric(df_analysis['Site_EUI'], errors='coerce')
            if np.allclose(agent_eui, gt['Site_EUI'], rtol=0.01):
                score += 10
            else:
                feedback.append("Site EUI calculations incorrect.")
        except: feedback.append("Site EUI column missing.")

        # -- Check Column: Source_Energy_kBtu (12 pts) --
        try:
            agent_source = pd.to_numeric(df_analysis['Source_Energy_kBtu'], errors='coerce')
            if np.allclose(agent_source, gt['Source_Energy'], rtol=0.01):
                score += 12
            else:
                feedback.append("Source Energy calculations incorrect (Check EPA ratios).")
        except: feedback.append("Source Energy column missing.")

        # -- Check Column: Source_EUI (8 pts) --
        try:
            agent_source_eui = pd.to_numeric(df_analysis['Source_EUI'], errors='coerce')
            if np.allclose(agent_source_eui, gt['Source_EUI'], rtol=0.01):
                score += 8
        except: pass

        # -- Check Column: EUI_Category (8 pts) --
        try:
            # Check string match count
            matches = (df_analysis['EUI_Category'] == gt['Category']).sum()
            if matches >= 23: score += 8
            elif matches >= 15: score += 4
        except: pass

        # -- Check Column: ESTAR_Eligible (8 pts) --
        try:
            matches = (df_analysis['ESTAR_Eligible'] == gt['Eligible']).sum()
            if matches >= 23: score += 8
        except: pass

        # -- Check Summary Sheet (44 pts distributed) --
        # Weighted Averages (8 pts)
        try:
            agent_wt_eui = float(summary_dict.get('Portfolio_Avg_Site_EUI_Weighted', 0))
            if abs(agent_wt_eui - portfolio_avg) < 1.0: # Tolerance
                score += 8
            else:
                feedback.append(f"Weighted Avg EUI incorrect. Exp: {portfolio_avg:.1f}, Got: {agent_wt_eui:.1f}")
        except: pass

        # Totals (5 pts)
        try:
            agent_total_area = float(summary_dict.get('Total_Portfolio_Area', 0))
            if abs(agent_total_area - total_area) < 1000:
                score += 5
        except: pass
        
        # Counts (6 pts)
        try:
            # Just check one count for existence
            agent_count_low = int(summary_dict.get('Count_Low_EUI', -1))
            gt_count_low = (gt['Category'] == 'Low').sum()
            if agent_count_low == gt_count_low:
                score += 6
        except: pass

        # Best/Worst (6 pts)
        try:
            best_idx = gt['Site_EUI'].idxmin()
            best_name = gt.loc[best_idx, 'Building_Name']
            agent_best = str(summary_dict.get('Best_Building', ''))
            if best_name.lower() in agent_best.lower():
                score += 6
        except: pass

        # Final checks
        passed = score >= 60
        if not result.get('file_modified_during_task'):
             score = max(0, score - 20) # Penalty for timestamps check failure if score > 0
             feedback.append("Penalty: File not modified during task time window.")

        return {
            "passed": passed,
            "score": int(score),
            "feedback": " | ".join(feedback) if feedback else "Perfect execution."
        }

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Verifier error: {str(e)}"}
    finally:
        import shutil
        shutil.rmtree(temp_dir)
#!/usr/bin/env python3
"""
Verifier for NEC Voltage Drop Compliance Audit.
"""

import json
import os
import tempfile
import logging
import math
from datetime import datetime

# Optional imports for reading Excel
try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import Rule
except ImportError:
    openpyxl = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# NEC Table 8 Data (Ohms/kft) - hardcoded source of truth
NEC_TABLE = {
    "Copper": {
        "14": 3.14, "12": 1.98, "10": 1.24, "8": 0.778, "6": 0.491, 
        "4": 0.308, "3": 0.245, "2": 0.194, "1": 0.154, "1/0": 0.122, 
        "2/0": 0.0967, "3/0": 0.0766, "4/0": 0.0608
    },
    "Aluminum": {
        "14": 5.15, "12": 3.25, "10": 2.03, "8": 1.28, "6": 0.808, 
        "4": 0.508, "3": 0.403, "2": 0.319, "1": 0.253, "1/0": 0.201, 
        "2/0": 0.159, "3/0": 0.126, "4/0": 0.100
    }
}

def verify_nec_voltage_drop_audit(traj, env_info, task_info):
    """
    Verifies the NEC Voltage Drop Audit task.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    if openpyxl is None:
        return {"passed": False, "score": 0, "feedback": "Verifier error: openpyxl not installed"}

    # Setup temp files
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json').name
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name

    try:
        # 1. Fetch JSON result
        try:
            copy_from_env("C:\\tmp\\task_result.json", temp_json)
            with open(temp_json, 'r') as f:
                result_data = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {e}"}

        if not result_data.get('file_exists'):
            return {"passed": False, "score": 0, "feedback": "Workbook 'commercial_circuits.xlsx' not found."}

        if not result_data.get('file_modified'):
            return {"passed": False, "score": 0, "feedback": "Workbook was not modified/saved."}

        # 2. Fetch Excel File
        try:
            copy_from_env("C:\\Users\\Docker\\Documents\\commercial_circuits.xlsx", temp_xlsx)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to copy workbook: {e}"}

        # 3. Analyze Excel Content
        try:
            wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
            if "Circuit_Schedule" not in wb.sheetnames:
                return {"passed": False, "score": 0, "feedback": "Sheet 'Circuit_Schedule' missing."}
            
            ws = wb["Circuit_Schedule"]
            
            # Score components
            score = 0
            
            # Criteria
            # A. File Saved (already checked modified) -> 10 pts
            score += 10
            
            rows = list(ws.iter_rows(min_row=2, max_row=51, values_only=True))
            
            correct_lookup = 0
            correct_math = 0
            correct_logic = 0
            total_rows = len(rows)
            
            fails_found = False
            
            for idx, row in enumerate(rows):
                # Columns: A=0(ID), B=1(Desc), C=2(Volt), D=3(Load), E=4(Len), F=5(Size), G=6(Mat)
                # Agent fills: H=7(Res), I=8(VD), J=9(%), K=10(Status)
                
                try:
                    volts = float(row[2])
                    load = float(row[3])
                    length = float(row[4])
                    size = str(row[5])
                    material = str(row[6])
                    
                    # Agent values
                    agent_res = row[7]
                    agent_vd = row[8]
                    agent_pct = row[9]
                    agent_status = str(row[10]).upper().strip() if row[10] else ""
                    
                    # 1. Verify Resistance Lookup
                    expected_res = NEC_TABLE.get(material, {}).get(size)
                    if expected_res is not None and agent_res is not None:
                        try:
                            if abs(float(agent_res) - expected_res) < 0.05:
                                correct_lookup += 1
                        except: pass
                        
                    # 2. Verify Voltage Drop Math
                    # VD = 2 * L * R * I / 1000
                    if expected_res is not None:
                        expected_vd = (2 * length * expected_res * load) / 1000.0
                        expected_pct = expected_vd / volts
                        
                        # Check VD
                        if agent_vd is not None:
                            try:
                                if abs(float(agent_vd) - expected_vd) < 0.5: # Generous tolerance
                                    correct_math += 0.5
                            except: pass
                            
                        # Check %
                        if agent_pct is not None:
                            try:
                                # Excel percentages are decimals (0.03 = 3%)
                                if abs(float(agent_pct) - expected_pct) < 0.005: 
                                    correct_math += 0.5
                            except: pass
                        
                        # 3. Verify Logic
                        expected_status = "FAIL" if expected_pct > 0.03 else "PASS"
                        if agent_status == expected_status:
                            correct_logic += 1
                            
                        if expected_status == "FAIL":
                            fails_found = True

                except Exception as e:
                    # Skip malformed rows
                    continue

            # Scoring Calculations
            # B. Lookup Accuracy (25 pts)
            if total_rows > 0:
                lookup_score = (correct_lookup / total_rows) * 25
                score += lookup_score
            
            # C. Math Accuracy (25 pts)
            if total_rows > 0:
                math_score = (correct_math / total_rows) * 25
                score += math_score
            
            # D. Status Logic (20 pts)
            if total_rows > 0:
                logic_score = (correct_logic / total_rows) * 20
                score += logic_score
            
            # E. Conditional Formatting (20 pts)
            # Need to reload without data_only to check formatting rules
            wb_fmt = openpyxl.load_workbook(temp_xlsx)
            ws_fmt = wb_fmt["Circuit_Schedule"]
            cf_rules = ws_fmt.conditional_formatting
            
            fmt_score = 0
            # Basic check: are there any rules?
            if len(cf_rules) > 0:
                # Check if rules apply to Column K (11th col)
                # Check if rule looks for "FAIL"
                # This is complex in openpyxl, checking simple existence + visual check
                # If cells with "FAIL" have a fill color, that's good enough
                
                # Check a specific FAIL cell style
                fail_cell_found = False
                pass_cell_found = False
                
                # We need to manually evaluate where the formatting MIGHT be applied
                # Alternatively, check cell styles on the data_only workbook? 
                # No, openpyxl data_only doesn't evaluate CF.
                
                # Heuristic: Check if there are CF rules containing "FAIL" text
                for cf in cf_rules:
                    # cf is a ConditionalFormatting object
                    for rule in cf.rules:
                        # Inspect rule formula or text
                        if hasattr(rule, 'formula') and rule.formula:
                            if "FAIL" in str(rule.formula):
                                fmt_score = 20
                                break
                        if hasattr(rule, 'text') and rule.text and "FAIL" in rule.text:
                            fmt_score = 20
                            break
                        if hasattr(rule, 'operator') and rule.operator == 'containsText' and rule.text == 'FAIL':
                            fmt_score = 20
                            break
                            
            score += fmt_score

            feedback = f"Lookup: {correct_lookup}/{total_rows}, Math: {correct_math}/{total_rows}, Logic: {correct_logic}/{total_rows}. CF Score: {fmt_score}"

            passed = score >= 80
            return {
                "passed": passed,
                "score": int(score),
                "feedback": feedback
            }

        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Error analyzing workbook: {e}"}

    finally:
        # Cleanup
        if os.path.exists(temp_json):
            os.remove(temp_json)
        if os.path.exists(temp_xlsx):
            os.remove(temp_xlsx)
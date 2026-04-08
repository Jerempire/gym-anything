#!/usr/bin/env python3
"""
Verifier for nutrition_label_generator task.

Verifies:
1. File modification (anti-gaming).
2. Correct calculation of 'Baked Batch Weight' (Recipe sheet).
3. Correct rounded values on 'Nutrition_Facts' sheet.

Scoring:
- 15 pts: Anti-gaming (File modified & exists)
- 25 pts: Correct Batch Weight / Yield calculation
- 60 pts: Label accuracy (12pts each for Calories, Fat, Sodium, Carb, Protein)
"""

import json
import os
import shutil
import tempfile
import logging
from datetime import datetime

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_nutrition_label(traj, env_info, task_info):
    """
    Verify the Nutrition Label Generation task.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "System error: copy_from_env not available"}

    metadata = task_info.get('metadata', {})
    expected_weight = metadata.get('expected_baked_weight', 2239.6)
    expected_vals = metadata.get('expected_values', {})
    
    # Temp directory for processing
    temp_dir = tempfile.mkdtemp()
    
    try:
        # 1. Retrieve Result JSON
        result_json_path = os.path.join(temp_dir, "task_result.json")
        try:
            copy_from_env("C:\\tmp\\task_result.json", result_json_path)
            with open(result_json_path, 'r') as f:
                result_data = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {str(e)}"}

        # Check basic existence
        if not result_data.get('output_exists'):
            return {"passed": False, "score": 0, "feedback": "Workbook not found."}
        
        if not result_data.get('file_modified'):
            return {"passed": False, "score": 0, "feedback": "Workbook was not modified/saved."}

        # 2. Retrieve Workbook
        xlsx_path = result_data.get('xlsx_path')
        local_xlsx = os.path.join(temp_dir, "nutrition_calculator.xlsx")
        try:
            copy_from_env(xlsx_path, local_xlsx)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve workbook: {str(e)}"}

        # 3. Analyze Workbook
        try:
            import pandas as pd
            from openpyxl import load_workbook
        except ImportError:
            return {"passed": False, "score": 0, "feedback": "Verification environment missing pandas/openpyxl"}

        score = 15 # Start with 15 points for saving modified file
        feedback = ["File saved and modified."]

        wb = load_workbook(local_xlsx, data_only=True)
        
        # --- Check 1: Yield Calculation (Recipe Sheet) ---
        if 'Recipe' not in wb.sheetnames:
            feedback.append("Missing 'Recipe' sheet.")
        else:
            ws_recipe = wb['Recipe']
            # Expected B16 for Baked Batch Weight
            # User might put it elsewhere, but B16 was specified in prompt implicitly by context or row layout.
            # We search for the value nearby if B16 is empty or text.
            
            # Helper to find value
            found_weight = None
            cell_val = ws_recipe['B16'].value
            if isinstance(cell_val, (int, float)):
                found_weight = cell_val
            
            if found_weight is not None:
                # Tolerance check
                if abs(found_weight - expected_weight) <= 5.0:
                    score += 25
                    feedback.append(f"Batch Yield correct ({found_weight:.1f}g).")
                else:
                    feedback.append(f"Batch Yield incorrect. Got {found_weight:.1f}, expected ~{expected_weight}.")
            else:
                feedback.append("Batch Yield (B16) not found or not a number.")

        # --- Check 2: Nutrition Facts Label ---
        if 'Nutrition_Facts' not in wb.sheetnames:
            feedback.append("Missing 'Nutrition_Facts' sheet.")
        else:
            ws_facts = wb['Nutrition_Facts']
            
            # We need to find the values. Since users might move cells, we look for labels and value pairs
            # Or assume the template structure provided in setup script:
            # Labels in Col B, Values in Col C (implied by df_label export to startrow=2, startcol=1 -> B3:C9)
            
            # Map of Label -> Expected Value
            checks = [
                ("Calories", expected_vals['calories'], 12),
                ("Total Fat", expected_vals['total_fat'], 12),
                ("Sodium", expected_vals['sodium'], 12),
                ("Total Carbohydrate", expected_vals['total_carb'], 12),
                ("Protein", expected_vals['protein'], 12)
            ]
            
            # Extract data from sheet into a dict for fuzzy lookup
            sheet_data = {}
            for row in ws_facts.iter_rows(min_row=1, max_row=20, min_col=1, max_col=5):
                for i in range(len(row)-1):
                    cell = row[i]
                    val_cell = row[i+1]
                    if isinstance(cell.value, str):
                        key = cell.value.lower().strip()
                        val = val_cell.value
                        if isinstance(val, (int, float)):
                            sheet_data[key] = val
                            
            # Validate
            for label, expected, pts in checks:
                # Find matching key
                found = False
                for k, v in sheet_data.items():
                    if label.lower() in k:
                        found = True
                        if abs(v - expected) < 1.0: # Strict rounding check (allow slight epsilon for float rep)
                            score += pts
                            feedback.append(f"{label}: Correct ({v}).")
                        else:
                            feedback.append(f"{label}: Incorrect. Got {v}, expected {expected}.")
                        break
                if not found:
                    feedback.append(f"{label}: Value not found on label.")

        passed = (score >= 70)
        return {
            "passed": passed,
            "score": score,
            "feedback": " ".join(feedback)
        }

    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
#!/usr/bin/env python3
"""
Verifier for dcf_valuation_model task.

Stub verifier — primary evaluation uses vlm_checklist_verifier.
This performs basic structural checks and spot-checks a few key values.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Copy xlsx from VM and parse with openpyxl data_only=True
  3. Score criteria (100 pts, pass >= 60)

Criteria:
  C1 (15 pts): Revenue 2024E and 2028E within +-1% of expected
  C2 (10 pts): EBIT 2024E within +-2% of expected
  C3 (10 pts): NOPAT for >= 3 of 5 years within +-2%
  C4 (15 pts): FCF for >= 3 of 5 years within +-3%
  C5 (15 pts): WACC within +-0.5% absolute of expected
  C6 (10 pts): Terminal Value within +-8%
  C7 (10 pts): Implied Share Price within +-10%
  C8 (10 pts): Formula usage (>= 10 cells with formulas)
  C9 ( 5 pts): File saved (is_new == True)
"""
import json
import logging
import os
import tempfile
import shutil

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\dcf_valuation_model_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/company_valuation.xlsx"

# Expected values (deterministic from inputs)
EXPECTED_REVENUE = [2441.6, 2685.76, 2900.6208, 3103.664256, 3289.884111]
EXPECTED_EBIT = [439.488, 483.4368, 522.11174, 558.65957, 592.17914]
EXPECTED_NOPAT = [342.80064, 377.08070, 407.24716, 435.75446, 461.89973]
EXPECTED_FCF = [271.733, 302.712, 330.972, 356.422, 380.273]
EXPECTED_KE = 0.10575
EXPECTED_WACC = 0.090525
EXPECTED_TV = 6471.0
EXPECTED_EV = 5454.0
EXPECTED_SHARE_PRICE = 23.18


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _check_within_pct(actual, expected, tolerance_pct):
    if actual is None or expected == 0:
        return False
    return abs(actual - expected) / abs(expected) <= tolerance_pct


def _check_within_abs(actual, expected, tolerance_abs):
    if actual is None:
        return False
    return abs(actual - expected) <= tolerance_abs


def verify_dcf_valuation_model(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_dcf_")
    try:
        # ── STEP 1: Read result JSON and check is_new ──
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})

        score = 0
        fb = []

        # C9: File saved check (5 pts)
        if xlsx_info.get("is_new", False):
            score += 5
            fb.append("C9 PASS: File saved after task start")
        else:
            fb.append("C9 FAIL: Workbook not saved (is_new=False)")

        # ── STEP 2: Copy and parse xlsx ──
        xlsx_local = os.path.join(tmp, "company_valuation.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            fb.append(f"Could not copy xlsx: {e}")
            return {"passed": False, "score": score, "feedback": " | ".join(fb)}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            fb.append("company_valuation.xlsx not found or empty")
            return {"passed": False, "score": score, "feedback": " | ".join(fb)}

        from openpyxl import load_workbook
        wb_val = load_workbook(xlsx_local, data_only=True)

        # ── STEP 3: Score criteria ──

        # Find sheets
        proj_sheet = _find_sheet(wb_val, ["projected", "projection"])
        dcf_sheet = _find_sheet(wb_val, ["dcf", "valuation"])

        # C1 (15 pts): Revenue 2024E and 2028E
        rev_ok = 0
        if proj_sheet:
            ws = wb_val[proj_sheet]
            # Revenue should be in row 2, columns B-F (2-6)
            rev_2024 = ws.cell(2, 2).value
            rev_2028 = ws.cell(2, 6).value
            if _check_within_pct(rev_2024, EXPECTED_REVENUE[0], 0.01):
                rev_ok += 1
            if _check_within_pct(rev_2028, EXPECTED_REVENUE[4], 0.01):
                rev_ok += 1
        if rev_ok == 2:
            score += 15
            fb.append(f"C1 PASS: Revenue 2024E={rev_2024:.1f}, 2028E={rev_2028:.1f}")
        elif rev_ok == 1:
            score += 8
            fb.append("C1 PARTIAL: Only one revenue year correct")
        else:
            fb.append("C1 FAIL: Revenue projections not found or incorrect")

        # C2 (10 pts): EBIT 2024E
        ebit_2024 = None
        if proj_sheet:
            ws = wb_val[proj_sheet]
            ebit_2024 = ws.cell(8, 2).value  # Row 8 = EBIT
        if _check_within_pct(ebit_2024, EXPECTED_EBIT[0], 0.02):
            score += 10
            fb.append(f"C2 PASS: EBIT 2024E={ebit_2024:.1f}")
        else:
            fb.append(f"C2 FAIL: EBIT 2024E={ebit_2024} (expected ~{EXPECTED_EBIT[0]:.1f})")

        # C3 (10 pts): NOPAT for >= 3 of 5 years
        nopat_ok = 0
        if proj_sheet:
            ws = wb_val[proj_sheet]
            for col_idx, exp in enumerate(EXPECTED_NOPAT):
                val = ws.cell(10, col_idx + 2).value  # Row 10 = NOPAT
                if _check_within_pct(val, exp, 0.02):
                    nopat_ok += 1
        if nopat_ok >= 3:
            score += 10
            fb.append(f"C3 PASS: {nopat_ok}/5 NOPAT values correct")
        elif nopat_ok >= 1:
            score += 5
            fb.append(f"C3 PARTIAL: {nopat_ok}/5 NOPAT values correct")
        else:
            fb.append("C3 FAIL: No NOPAT values correct")

        # C4 (15 pts): FCF for >= 3 of 5 years
        fcf_ok = 0
        if dcf_sheet:
            ws = wb_val[dcf_sheet]
            for col_idx, exp in enumerate(EXPECTED_FCF):
                val = ws.cell(6, col_idx + 2).value  # Row 6 = Free Cash Flow
                if _check_within_pct(val, exp, 0.03):
                    fcf_ok += 1
        if fcf_ok >= 3:
            score += 15
            fb.append(f"C4 PASS: {fcf_ok}/5 FCF values correct")
        elif fcf_ok >= 1:
            score += 7
            fb.append(f"C4 PARTIAL: {fcf_ok}/5 FCF values correct")
        else:
            fb.append("C4 FAIL: No FCF values correct")

        # C5 (15 pts): WACC
        wacc_val = None
        ke_val = None
        if dcf_sheet:
            ws = wb_val[dcf_sheet]
            ke_val = ws.cell(12, 2).value    # Row 12 = Ke
            wacc_val = ws.cell(16, 2).value  # Row 16 = WACC
        ke_ok = _check_within_abs(ke_val, EXPECTED_KE, 0.005)
        wacc_ok = _check_within_abs(wacc_val, EXPECTED_WACC, 0.005)
        if ke_ok and wacc_ok:
            score += 15
            fb.append(f"C5 PASS: Ke={ke_val:.4f}, WACC={wacc_val:.4f}")
        elif wacc_ok:
            score += 10
            fb.append(f"C5 PARTIAL: WACC correct but Ke off")
        elif ke_ok:
            score += 7
            fb.append(f"C5 PARTIAL: Ke correct but WACC off")
        else:
            fb.append(f"C5 FAIL: Ke={ke_val}, WACC={wacc_val} (expected Ke~{EXPECTED_KE:.4f}, WACC~{EXPECTED_WACC:.4f})")

        # C6 (10 pts): Terminal Value
        tv_val = None
        if dcf_sheet:
            ws = wb_val[dcf_sheet]
            tv_val = ws.cell(20, 2).value  # Row 20 = Terminal Value
        if _check_within_pct(tv_val, EXPECTED_TV, 0.08):
            score += 10
            fb.append(f"C6 PASS: Terminal Value={tv_val:.0f}")
        else:
            fb.append(f"C6 FAIL: Terminal Value={tv_val} (expected ~{EXPECTED_TV:.0f})")

        # C7 (10 pts): Implied Share Price
        price_val = None
        if dcf_sheet:
            ws = wb_val[dcf_sheet]
            price_val = ws.cell(27, 2).value  # Row 27 = Implied Share Price
        if _check_within_pct(price_val, EXPECTED_SHARE_PRICE, 0.10):
            score += 10
            fb.append(f"C7 PASS: Share Price=${price_val:.2f}")
        else:
            fb.append(f"C7 FAIL: Share Price={price_val} (expected ~${EXPECTED_SHARE_PRICE:.2f})")

        # C8 (10 pts): Formula usage
        formula_count = 0
        try:
            wb_formulas = load_workbook(xlsx_local, data_only=False)
            for sheet_name in [proj_sheet, dcf_sheet]:
                if sheet_name and sheet_name in wb_formulas.sheetnames:
                    ws = wb_formulas[sheet_name]
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                formula_count += 1
        except Exception as e:
            logger.warning(f"Formula check failed: {e}")

        if formula_count >= 10:
            score += 10
            fb.append(f"C8 PASS: {formula_count} formulas found")
        elif formula_count >= 3:
            score += 5
            fb.append(f"C8 PARTIAL: Only {formula_count} formulas found (need >= 10)")
        else:
            fb.append(f"C8 FAIL: Only {formula_count} formulas found")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

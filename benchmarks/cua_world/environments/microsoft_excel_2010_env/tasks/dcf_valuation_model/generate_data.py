#!/usr/bin/env python3
"""
Generate company_valuation.xlsx for the dcf_valuation_model task.

Creates a 4-sheet workbook:
  1. Historical_Financials  — pre-populated income statement + capital structure (2019-2023)
  2. Assumptions            — pre-populated projection assumptions, margins, WACC inputs
  3. Projected_Financials   — empty template (row labels + year headers only)
  4. DCF_Valuation          — empty template (section headers + row labels only)

Data represents TechFlow Inc., a mature mid-cap SaaS company.
Historical financials are internally consistent and realistic.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "data"
)

BOLD = Font(bold=True)
BOLD_HEADER = Font(bold=True, size=11)
SECTION_HEADER = Font(bold=True, size=11, color="1F4E79")
PCT_FMT = '0.0%'
PCT_FMT_2DP = '0.00%'
NUM_FMT = '#,##0.0'
CURRENCY_FMT = '$#,##0.0'
THIN_BORDER_BOTTOM = Border(bottom=Side(style='thin'))


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_row(ws, row, values, font=None, fmt=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if font:
            cell.font = font
        if fmt and col > 1 and isinstance(val, (int, float)):
            cell.number_format = fmt


def create_historical_financials(wb):
    ws = wb.active
    ws.title = "Historical_Financials"
    _set_col_widths(ws, [28, 12, 12, 12, 12, 12])

    # Header row
    headers = ["Income Statement ($ Millions)", 2019, 2020, 2021, 2022, 2023]
    _write_row(ws, 1, headers, font=BOLD_HEADER)
    for col in range(2, 7):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # Income statement data
    data = [
        ("Revenue",          1250.0, 1440.0, 1700.0, 1955.0, 2180.0),
        ("COGS",              450.0,  518.0,  612.0,  703.0,  785.0),
        ("Gross Profit",      800.0,  922.0, 1088.0, 1252.0, 1395.0),
        ("SG&A",              325.0,  374.0,  442.0,  508.0,  567.0),
        ("R&D",               188.0,  216.0,  255.0,  293.0,  327.0),
        ("D&A",                63.0,   72.0,   85.0,   98.0,  109.0),
        ("EBIT",              225.0,  260.0,  306.0,  353.0,  392.0),
        ("Interest Expense",   29.0,   26.0,   24.0,   22.0,   20.0),
        ("EBT",               196.0,  234.0,  282.0,  331.0,  372.0),
        ("Taxes",              43.0,   51.0,   62.0,   73.0,   82.0),
        ("Net Income",        153.0,  183.0,  220.0,  258.0,  290.0),
    ]

    for i, row_data in enumerate(data):
        row_num = i + 2
        label = row_data[0]
        vals = row_data[1:]
        ws.cell(row=row_num, column=1, value=label)
        for col, val in enumerate(vals, 2):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.number_format = NUM_FMT
        # Bold subtotals
        if label in ("Gross Profit", "EBIT", "EBT", "Net Income"):
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).font = BOLD

    # Blank row 13
    # Capital Structure section
    ws.cell(row=14, column=1, value="Capital Structure").font = SECTION_HEADER
    ws.cell(row=14, column=6, value=2023).font = BOLD
    ws.cell(row=14, column=6).alignment = Alignment(horizontal='center')

    cap_data = [
        ("Total Debt",             355.0),
        ("Cash & Equivalents",     580.0),
        ("Shares Outstanding (M)", 245.0),
    ]
    for i, (label, val) in enumerate(cap_data):
        row_num = 15 + i
        ws.cell(row=row_num, column=1, value=label)
        cell = ws.cell(row=row_num, column=6, value=val)
        cell.number_format = NUM_FMT

    return ws


def create_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    _set_col_widths(ws, [34, 14])

    # Section 1: Growth rates
    ws.cell(row=1, column=1, value="Projection Assumptions").font = SECTION_HEADER

    growth_rates = [
        ("Revenue Growth Rate - 2024E", 0.12),
        ("Revenue Growth Rate - 2025E", 0.10),
        ("Revenue Growth Rate - 2026E", 0.08),
        ("Revenue Growth Rate - 2027E", 0.07),
        ("Revenue Growth Rate - 2028E", 0.06),
    ]
    for i, (label, val) in enumerate(growth_rates):
        row = i + 2
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=val)
        cell.number_format = PCT_FMT

    # Section 2: Cost & Margin
    ws.cell(row=8, column=1, value="Cost & Margin Assumptions").font = SECTION_HEADER

    margins = [
        ("COGS (% of Revenue)",             0.36),
        ("SG&A (% of Revenue)",             0.26),
        ("R&D (% of Revenue)",              0.15),
        ("D&A (% of Revenue)",              0.05),
        ("CapEx (% of Revenue)",            0.07),
        ("NWC (% of Revenue Change)",       0.085),
        ("Tax Rate",                        0.22),
    ]
    for i, (label, val) in enumerate(margins):
        row = i + 9
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=val)
        cell.number_format = PCT_FMT

    # Section 3: WACC Inputs
    ws.cell(row=17, column=1, value="WACC Inputs").font = SECTION_HEADER

    wacc_inputs = [
        ("Risk-Free Rate (Rf)",         0.0425),
        ("Equity Risk Premium (ERP)",   0.055),
        ("Beta",                        1.15),
        ("Pre-tax Cost of Debt (Kd)",   0.0575),
        ("Debt-to-Capital (D/V)",       0.25),
        ("Terminal Growth Rate (g)",    0.03),
    ]
    for i, (label, val) in enumerate(wacc_inputs):
        row = i + 18
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=val)
        # Beta is a plain number, rest are percentages with 2 dp precision
        if label == "Beta":
            cell.number_format = '0.00'
        else:
            cell.number_format = PCT_FMT_2DP

    return ws


def create_projected_financials(wb):
    ws = wb.create_sheet("Projected_Financials")
    _set_col_widths(ws, [28, 14, 14, 14, 14, 14])

    # Header row
    ws.cell(row=1, column=1, value="Projected Income Statement ($ Millions)").font = BOLD_HEADER
    years = ["2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, yr in enumerate(years, 2):
        cell = ws.cell(row=1, column=col, value=yr)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal='center')

    # Row labels (values left empty for agent to fill)
    labels = [
        "Revenue",
        "COGS",
        "Gross Profit",
        "SG&A",
        "R&D",
        "D&A",
        "EBIT",
        "Tax on EBIT",
        "NOPAT",
    ]
    for i, label in enumerate(labels):
        row = i + 2
        ws.cell(row=row, column=1, value=label)
        if label in ("Gross Profit", "EBIT", "NOPAT"):
            ws.cell(row=row, column=1).font = BOLD
        # Set number format on empty cells so values display nicely when entered
        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = NUM_FMT

    return ws


def create_dcf_valuation(wb):
    ws = wb.create_sheet("DCF_Valuation")
    _set_col_widths(ws, [32, 14, 14, 14, 14, 14])

    # Section 1: Free Cash Flow
    ws.cell(row=1, column=1, value="Free Cash Flow ($ Millions)").font = SECTION_HEADER
    years = ["2024E", "2025E", "2026E", "2027E", "2028E"]
    for col, yr in enumerate(years, 2):
        cell = ws.cell(row=1, column=col, value=yr)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal='center')

    fcf_labels = [
        "NOPAT",
        "(+) D&A",
        "(-) CapEx",
        "(-) Change in NWC",
        "Free Cash Flow",
    ]
    for i, label in enumerate(fcf_labels):
        row = i + 2
        ws.cell(row=row, column=1, value=label)
        if label == "Free Cash Flow":
            ws.cell(row=row, column=1).font = BOLD
        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = NUM_FMT

    # Row 7: blank

    # Discount factors and PV
    discount_labels = ["Discount Factor", "PV of Free Cash Flow"]
    for i, label in enumerate(discount_labels):
        row = i + 8
        ws.cell(row=row, column=1, value=label)
        for col in range(2, 7):
            if label == "Discount Factor":
                ws.cell(row=row, column=col).number_format = '0.0000'
            else:
                ws.cell(row=row, column=col).number_format = NUM_FMT

    # Row 10: blank

    # Section 2: WACC Calculation
    ws.cell(row=11, column=1, value="WACC Calculation").font = SECTION_HEADER

    wacc_labels = [
        ("Cost of Equity (Ke)",     PCT_FMT),
        ("After-tax Cost of Debt",  PCT_FMT),
        ("Equity Weight (E/V)",     PCT_FMT),
        ("Debt Weight (D/V)",       PCT_FMT),
        ("WACC",                    PCT_FMT),
    ]
    for i, (label, fmt) in enumerate(wacc_labels):
        row = i + 12
        ws.cell(row=row, column=1, value=label)
        if label == "WACC":
            ws.cell(row=row, column=1).font = BOLD
        ws.cell(row=row, column=2).number_format = fmt

    # Row 17: blank

    # Section 3: Valuation Summary
    ws.cell(row=18, column=1, value="Valuation Summary ($ Millions)").font = SECTION_HEADER

    val_labels = [
        ("Sum of PV of FCFs",           NUM_FMT),
        ("Terminal Value",              NUM_FMT),
        ("PV of Terminal Value",        NUM_FMT),
        ("Enterprise Value",            NUM_FMT),
        ("(-) Total Debt",              NUM_FMT),
        ("(+) Cash & Equivalents",      NUM_FMT),
        ("Equity Value",                NUM_FMT),
        ("Shares Outstanding (M)",      NUM_FMT),
        ("Implied Share Price ($/sh)",  '$#,##0.00'),
    ]
    for i, (label, fmt) in enumerate(val_labels):
        row = i + 19
        ws.cell(row=row, column=1, value=label)
        if label in ("Enterprise Value", "Equity Value", "Implied Share Price ($/sh)"):
            ws.cell(row=row, column=1).font = BOLD
        ws.cell(row=row, column=2).number_format = fmt

    return ws


def main():
    wb = Workbook()

    create_historical_financials(wb)
    create_assumptions(wb)
    create_projected_financials(wb)
    create_dcf_valuation(wb)

    os.makedirs(DATA_DIR, exist_ok=True)
    output_path = os.path.join(DATA_DIR, "company_valuation.xlsx")
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"Sheets: {wb.sheetnames}")

    # Verify by re-reading
    from openpyxl import load_workbook
    wb2 = load_workbook(output_path)
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()

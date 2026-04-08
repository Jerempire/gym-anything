#!/usr/bin/env python3
"""
Verifier for workers_comp_loss_reserve task.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Independently copy xlsx and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts, pass >= 60)

Criteria:
  C1 (25 pts): VW LDF 12->24 in [1.40, 1.45] (expected 1.4222)
  C2 (20 pts): VW LDF 24->36 in [1.13, 1.21] (expected 1.1690)
  C3 (25 pts): Total IBNR within +/-15% of $75,631K
  C4 (20 pts): At least 2 UNDER_RESERVED flags
  C5 (10 pts): AY 2023 CDF-to-Ultimate in [1.85, 2.05] (expected 1.9406)
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH         = "C:\\Users\\Docker\\workers_comp_loss_reserve_result.json"
XLSX_PATH           = "C:/Users/Docker/Desktop/ExcelTasks/workers_comp_triangle.xlsx"
EXPECTED_TOTAL_IBNR = 75630.7   # $000s
IBNR_TOLERANCE      = 0.15


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_workers_comp_loss_reserve(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_wc_")
    try:
        # ── STEP 1: Check is_new ──────────────────────────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False)."
            }

        # ── STEP 2: Copy and parse xlsx ──────────────────────────────────────────
        xlsx_local = os.path.join(tmp, "workers_comp_triangle.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0, "feedback": "workers_comp_triangle.xlsx not found"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        score = 0
        fb = []

        ldf_sheet  = _find_sheet(wb, ["ldf", "development", "factor"])
        ibnr_sheet = _find_sheet(wb, ["ibnr", "reserve"])

        # Criterion 1: VW LDF 12->24 in [1.40, 1.45]
        vw_12_24 = None
        if ldf_sheet:
            ws = wb[ldf_sheet]
            # Row 10, column B (2) expected
            cands = _scan_numeric(ws, range(9, 13), range(2, 6), 1.35, 1.55)
            col_b = [(r, v) for r, c, v in cands if c == 2]
            if col_b:
                vw_12_24 = col_b[0][1]
            elif cands:
                vw_12_24 = max(v for _, _, v in cands)  # largest = 12->24 period
            if vw_12_24 is None:
                broader = _scan_numeric(ws, range(2, 15), range(2, 6), 1.38, 1.48)
                if broader:
                    vw_12_24 = broader[0][2]

        if vw_12_24 is not None and 1.40 <= vw_12_24 <= 1.45:
            score += 25
            fb.append(f"C1 PASS: VW LDF 12->24 = {vw_12_24:.4f} (expected 1.4222)")
        elif vw_12_24 is not None:
            # Partial: check individual LDFs exist
            indiv = _scan_numeric(wb[ldf_sheet], range(4, 9), range(2, 6), 1.38, 1.50) if ldf_sheet else []
            if len(indiv) >= 2:
                score += 10
                fb.append(f"C1 PARTIAL: Individual LDFs found but VW avg {vw_12_24:.4f} off (expected 1.40-1.45)")
            else:
                fb.append(f"C1 FAIL: VW LDF 12->24 = {vw_12_24:.4f} (expected 1.40-1.45)")
        else:
            indiv = _scan_numeric(wb[ldf_sheet], range(4, 9), range(2, 6), 1.38, 1.50) if ldf_sheet else []
            if len(indiv) >= 2:
                score += 10
                fb.append(f"C1 PARTIAL: Individual LDFs present but VW avg row missing")
            else:
                fb.append("C1 FAIL: LDF_Development not populated")

        # Criterion 2: VW LDF 24->36 in [1.13, 1.21]
        vw_24_36 = None
        if ldf_sheet:
            ws = wb[ldf_sheet]
            cands = _scan_numeric(ws, range(9, 13), range(2, 6), 1.13, 1.21)
            col_c = [(r, v) for r, c, v in cands if c == 3]
            if col_c:
                vw_24_36 = col_c[0][1]
            elif cands:
                vw_24_36 = cands[0][2]
            if vw_24_36 is None:
                broader = _scan_numeric(ws, range(2, 15), range(2, 6), 1.13, 1.21)
                if broader:
                    vw_24_36 = broader[0][2]

        if vw_24_36 is not None and 1.13 <= vw_24_36 <= 1.21:
            score += 20
            fb.append(f"C2 PASS: VW LDF 24->36 = {vw_24_36:.4f} (expected 1.1690)")
        elif vw_24_36 is not None:
            fb.append(f"C2 FAIL: VW LDF 24->36 = {vw_24_36:.4f} (expected 1.13-1.21)")
        else:
            fb.append("C2 FAIL: VW LDF 24->36 not found")

        # Criterion 3: Total IBNR within +/-15% of $75,631K
        total_ibnr = None
        if ibnr_sheet:
            ws = wb[ibnr_sheet]
            # Row 10 totals, column E (5) = IBNR Reserve
            cands = _scan_numeric(ws, range(9, 12), range(3, 8), 40000, 120000)
            row10 = [(c, v) for r, c, v in cands if r == 10]
            if row10:
                col_e = [(c, v) for c, v in row10 if c == 5]
                total_ibnr = col_e[0][1] if col_e else row10[0][1]
            elif cands:
                # Sum individual IBNR values (rows 4-8, col 5)
                indiv_ibnr = _scan_numeric(ws, range(4, 9), [5], 1000, 60000)
                if len(indiv_ibnr) >= 4:
                    total_ibnr = sum(v for _, _, v in indiv_ibnr)

        if total_ibnr is None and ibnr_sheet:
            ws = wb[ibnr_sheet]
            indiv = _scan_numeric(ws, range(4, 9), range(4, 7), 500, 50000)
            if len(indiv) >= 4:
                total_ibnr = sum(v for _, _, v in indiv)

        if total_ibnr is not None:
            pct_err = abs(total_ibnr - EXPECTED_TOTAL_IBNR) / EXPECTED_TOTAL_IBNR
            if pct_err <= IBNR_TOLERANCE:
                score += 25
                fb.append(f"C3 PASS: Total IBNR = {total_ibnr:,.1f} $K (err={pct_err:.1%})")
            else:
                fb.append(f"C3 FAIL: Total IBNR = {total_ibnr:,.1f} $K (expected ~75,631; err={pct_err:.1%})")
        else:
            fb.append("C3 FAIL: Total IBNR reserve not found")

        # Criterion 4: At least 2 UNDER_RESERVED flags
        flag_count = 0
        if ibnr_sheet:
            ws = wb[ibnr_sheet]
            flag_count = _count_string(ws, range(4, 10), range(6, 9), "UNDER_RESERVED")

        if flag_count >= 2:
            score += 20
            fb.append(f"C4 PASS: {flag_count} UNDER_RESERVED flags (AY 2022 + 2023)")
        elif flag_count == 1:
            score += 10
            fb.append(f"C4 PARTIAL: {flag_count} UNDER_RESERVED flag")
        else:
            fb.append("C4 FAIL: No UNDER_RESERVED flags (AY 2022 at 26.7% and 2023 at 48.5% should be flagged)")

        # Criterion 5: AY 2023 CDF-to-Ultimate in [1.85, 2.05]
        cdf_2023 = None
        if ldf_sheet:
            ws = wb[ldf_sheet]
            cands = _scan_numeric(ws, range(16, 22), range(7, 10), 1.85, 2.05)
            row20 = [(c, v) for r, c, v in cands if r == 20]
            if row20:
                cdf_2023 = row20[0][1]
            elif cands:
                cdf_2023 = max(v for _, _, v in cands)

        if cdf_2023 is None and ibnr_sheet:
            ws = wb[ibnr_sheet]
            cands = _scan_numeric(ws, range(4, 9), [3], 1.85, 2.05)
            row8 = [(r, v) for r, c, v in cands if r == 8]
            if row8:
                cdf_2023 = row8[0][1]
            elif cands:
                cdf_2023 = max(v for _, _, v in cands)

        if cdf_2023 is not None and 1.85 <= cdf_2023 <= 2.05:
            score += 10
            fb.append(f"C5 PASS: AY 2023 CDF = {cdf_2023:.4f} (expected ~1.9406)")
        elif cdf_2023 is not None:
            fb.append(f"C5 FAIL: AY 2023 CDF = {cdf_2023:.4f} (expected 1.85-2.05)")
        else:
            fb.append("C5 FAIL: AY 2023 CDF-to-Ultimate not found")

        return {"passed": score >= 60, "score": score, "feedback": " | ".join(fb)}

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

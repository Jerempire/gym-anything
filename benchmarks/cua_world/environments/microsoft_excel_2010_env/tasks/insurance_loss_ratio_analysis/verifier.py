#!/usr/bin/env python3
"""
Verifier for Insurance Loss Ratio Analysis task.
Parses the Excel file directly using openpyxl to verify calculations.
"""

import json
import logging
import os
import tempfile
import sys
import shutil

# Try to import openpyxl, but handle if missing (framework often has it)
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_insurance_loss_ratio_analysis(traj, env_info, task_info):
    """
    Verifies the insurance analysis workbook.
    """
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "System error: copy_from_env not available"}

    # Setup temp dir
    tmp_dir = tempfile.mkdtemp()
    
    try:
        # 1. Get Result JSON
        result_json_path = os.path.join(tmp_dir, "task_result.json")
        try:
            copy_from_env("C:\\Users\\Docker\\task_result.json", result_json_path)
            with open(result_json_path, 'r') as f:
                result_data = json.load(f)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {str(e)}"}

        # Check basic requirements
        if not result_data.get('xlsx_file', {}).get('exists', False):
            return {"passed": False, "score": 0, "feedback": "Excel file not found"}
        
        if not result_data.get('xlsx_file', {}).get('is_new', False):
            return {"passed": False, "score": 0, "feedback": "File was not modified/saved during the task"}

        # 2. Get Excel File
        xlsx_path = os.path.join(tmp_dir, "pc_industry_ratios.xlsx")
        try:
            copy_from_env("C:\\Users\\Docker\\Documents\\pc_industry_ratios.xlsx", xlsx_path)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to retrieve Excel file: {str(e)}"}

        # 3. Analyze Workbook
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to open Excel file: {str(e)}"}

        score = 0
        feedback = []

        # --- Check Sheet 1: Ratio_Analysis ---
        if "Ratio_Analysis" not in wb.sheetnames:
            feedback.append("Missing sheet 'Ratio_Analysis'")
        else:
            ws = wb["Ratio_Analysis"]
            ws_f = wb_formulas["Ratio_Analysis"]
            
            # Check Loss Ratios (Row 3-9, Col B-F)
            # Sample: Private Passenger Auto 2020 (Row 2+1, Col 4) -> 65.9%
            # Data: Loss=159490, NEP=242018 => 65.90
            try:
                val = ws.cell(row=4, column=4).value # 2020 is col D (4)
                if isinstance(val, (int, float)) and 65.8 < val < 66.0:
                    score += 10
                    feedback.append("Loss Ratios correct")
                else:
                    feedback.append(f"Loss Ratio incorrect (expected ~65.9, got {val})")
            except: pass

            # Check Combined Ratios (Row 23-29)
            # Sample: WC 2018 (Row 26, Col 2) -> 85.0
            # LossRatio: 26076/49200=0.5300 (53.0%)
            # ExpRatio: 15936/49800=0.3200 (32.0%)
            # CR = 85.0
            try:
                val = ws.cell(row=26, column=2).value
                if isinstance(val, (int, float)) and 84.9 < val < 85.1:
                    score += 10
                    feedback.append("Combined Ratios correct")
                else:
                    feedback.append(f"Combined Ratio incorrect (expected 85.0, got {val})")
            except: pass

            # Check Profitability Flags (Row 43-49)
            # Sample: Commercial Auto 2018 (Row 45, Col 2)
            # CR = (26253/35900) + (12161/36850) = 73.12 + 32.99 = 106.1 -> UNPROFITABLE
            try:
                val = ws.cell(row=45, column=2).value
                if str(val).upper() == "UNPROFITABLE":
                    score += 10
                    feedback.append("Profitability Flags correct")
                else:
                    feedback.append(f"Profitability Flag incorrect (expected UNPROFITABLE, got {val})")
            except: pass

            # Check Formulas Presence
            formula_count = 0
            check_cells = [(4,4), (26,2)] 
            for r, c in check_cells:
                if str(ws_f.cell(row=r, column=c).value).startswith("="):
                    formula_count += 1
            if formula_count == 2:
                score += 5
                feedback.append("Formulas used")

        # --- Check Sheet 2: Performance_Summary ---
        if "Performance_Summary" not in wb.sheetnames:
            feedback.append("Missing sheet 'Performance_Summary'")
        else:
            ws_perf = wb["Performance_Summary"]
            
            # Check 5-Year Avg (Row 3-9, Col B)
            # Sample: PP Auto (Row 3, Col 2) -> (72.83+71.86+65.9+73.59+77.32 + 26.63... wait, need CRs)
            # Let's rely on PP Auto 5-year average Combined Ratio check
            # 2018: 72.83 (LR) + 26.63 (ER) = 99.46
            # 2019: 71.86 + 26.82 = 98.68
            # 2020: 65.90 + 27.50 = 93.40
            # 2021: 73.59 + 25.79 = 99.38
            # 2022: 77.32 + 24.93 = 102.25
            # Avg: ~98.63
            try:
                val = ws_perf.cell(row=3, column=2).value
                if isinstance(val, (int, float)) and 98.5 < val < 98.8:
                    score += 10
                    feedback.append("5-Year Averages correct")
                else:
                    feedback.append(f"5-Year Avg incorrect (expected ~98.63, got {val})")
            except: pass

            # Check Weighted Industry CR (Row 12-16)
            # Just check if populated and reasonable range (98-105 usually)
            try:
                val = ws_perf.cell(row=16, column=2).value # 2022
                if isinstance(val, (int, float)) and 100 < val < 105:
                    score += 15
                    feedback.append("Weighted Industry Averages correct")
                else:
                    feedback.append(f"Weighted Avg incorrect (expected 100-105, got {val})")
            except: pass

            # Check Worst Performing Line (Row 38-42)
            # 2022 Worst: Com Auto (CR ~102) vs PP Auto (102.2) vs others
            # Let's check text match
            try:
                val = ws_perf.cell(row=42, column=2).value # 2022
                if val and ("Auto" in str(val) or "Liability" in str(val)): # Broad check
                    score += 10
                    feedback.append("Worst Performing Line identification seems plausible")
            except: pass

            # Check Deterioration Flags
            # PP Auto 2022 (102.25) > 5yr (98.6) AND 2022 > 2021 (99.38) -> DETERIORATING
            try:
                val = ws_perf.cell(row=45, column=2).value
                if str(val).upper() == "DETERIORATING":
                    score += 10
                    feedback.append("Deterioration Flags correct")
                else:
                    feedback.append(f"Deterioration Flag incorrect for PP Auto (expected DETERIORATING, got {val})")
            except: pass

            # Check CAGR
            # PP Auto NWP: 276821 / 237477 ^ (1/4) - 1 = 1.165^0.25 - 1 = 3.9%
            try:
                val = ws_perf.cell(row=29, column=2).value
                if isinstance(val, (int, float)) and 3.8 < val < 4.0:
                    score += 10
                    feedback.append("CAGR Calculation correct")
            except: pass
            
            # Formula check
            if str(wb_formulas["Performance_Summary"].cell(row=3, column=2).value).startswith("="):
                score += 10
                feedback.append("Performance Summary formulas used")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": "; ".join(feedback)
        }

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
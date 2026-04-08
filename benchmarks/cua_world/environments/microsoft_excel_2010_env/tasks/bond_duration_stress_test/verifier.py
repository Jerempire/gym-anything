#!/usr/bin/env python3
"""
Verifier for bond_duration_stress_test task.

Verification pipeline:
  1. Read result JSON (C:\Users\Docker\bond_duration_stress_test_result.json)
     - Check is_new: if xlsx was not saved after task start -> score 0
  2. Independently copy xlsx from VM and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts total, pass >= 60)

Criteria:
  Criterion 1 (20 pts): Duration_Analysis has ModDur for >= 7 bonds (range 1.0-20.0)
  Criterion 2 (25 pts): At least 3 ModDur values within +/-8% of expected
  Criterion 3 (15 pts): DURATION_BREACH flags for >= 3 of 5 expected breach bonds
  Criterion 4 (20 pts): Portfolio wtd-avg ModDur in [6.0, 7.8]
  Criterion 5 (20 pts): Shock_Scenario +100bps total P&L in [-360000, -230000]
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH  = "C:\\Users\\Docker\\bond_duration_stress_test_result.json"
XLSX_PATH    = "C:/Users/Docker/Desktop/ExcelTasks/bond_portfolio.xlsx"

EXPECTED_MOD_DUR = [1.8974, 4.4863, 8.1790, 16.0839, 5.7824,
                    8.6426, 11.2915, 4.5759, 6.5793, 10.4351]
BREACH_BOND_IDX  = {2, 3, 5, 6, 9}   # 0-based, ModDur > 7.0


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_bond_duration_stress_test(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_bond_")
    try:
        # ── STEP 1: Read result JSON and check is_new ────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            result = {}
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False). "
                            "Agent must save the file with Ctrl+S after completing formulas."
            }

        # ── STEP 2: Independently copy and parse xlsx ────────────────────────────
        xlsx_local = os.path.join(tmp, "bond_portfolio.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0,
                    "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0,
                    "feedback": "bond_portfolio.xlsx not found or empty"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        # ── STEP 3: Score criteria ───────────────────────────────────────────────
        score = 0
        fb = []

        # Criterion 1: ModDur for >= 7 bonds
        dur_sheet = _find_sheet(wb, ["duration", "analysis"])
        mod_dur_vals = []
        if dur_sheet:
            ws = wb[dur_sheet]
            cells = _scan_numeric(ws, range(2, 13), range(2, 8), 1.0, 20.0)
            # Prefer column C (3) — Modified Duration column
            col_c = [(r, v) for r, c, v in cells if c == 3]
            mod_dur_vals = [v for _, v in col_c] if len(col_c) >= 5 else [v for _, _, v in cells[:10]]

        if len(mod_dur_vals) >= 7:
            score += 20
            fb.append(f"C1 PASS: ModDur values for {len(mod_dur_vals)} bonds")
        elif len(mod_dur_vals) >= 4:
            score += 10
            fb.append(f"C1 PARTIAL: Only {len(mod_dur_vals)} ModDur values")
        else:
            fb.append(f"C1 FAIL: Only {len(mod_dur_vals)} ModDur values (need >= 7)")

        # Criterion 2: >= 3 ModDur within +-8% of expected
        if mod_dur_vals:
            accurate = 0
            for i, exp in enumerate(EXPECTED_MOD_DUR):
                for found in mod_dur_vals:
                    if abs(found - exp) / exp <= 0.08:
                        accurate += 1
                        break
            if accurate >= 3:
                score += 25
                fb.append(f"C2 PASS: {accurate} ModDur values accurate within 8%")
            elif accurate >= 1:
                score += 12
                fb.append(f"C2 PARTIAL: {accurate} accurate ModDur values")
            else:
                fb.append(f"C2 FAIL: 0 ModDur values within 8% of expected")
        else:
            fb.append("C2 FAIL: No ModDur values to evaluate")

        # Criterion 3: DURATION_BREACH flags for >= 3 breach bonds
        breach_count = 0
        if dur_sheet:
            ws = wb[dur_sheet]
            breach_count = _count_string(ws, range(2, 13), range(4, 8), "DURATION_BREACH")
        if breach_count >= 3:
            score += 15
            fb.append(f"C3 PASS: {breach_count} DURATION_BREACH flags (expected 5)")
        elif breach_count >= 1:
            score += 7
            fb.append(f"C3 PARTIAL: {breach_count} DURATION_BREACH flags")
        else:
            fb.append("C3 FAIL: No DURATION_BREACH flags found")

        # Criterion 4: Portfolio weighted-avg ModDur in [6.0, 7.8]
        wtd_avg = None
        if dur_sheet:
            ws = wb[dur_sheet]
            # Look in summary rows (row 12+) for a value in [5, 9]
            cands = _scan_numeric(ws, range(11, 16), range(2, 8), 5.0, 9.0)
            if cands:
                wtd_avg = cands[0][2]
        if wtd_avg is not None and 6.0 <= wtd_avg <= 7.8:
            score += 20
            fb.append(f"C4 PASS: Portfolio wtd-avg ModDur = {wtd_avg:.2f}yr")
        elif wtd_avg is not None:
            fb.append(f"C4 FAIL: Portfolio ModDur = {wtd_avg:.2f}yr (expected 6.0-7.8)")
        else:
            fb.append("C4 FAIL: Portfolio weighted-average ModDur not found")

        # Criterion 5: Shock +100bps total P&L in [-360000, -230000]
        shock_sheet = _find_sheet(wb, ["shock", "scenario", "stress"])
        pl_total = None
        if shock_sheet:
            ws = wb[shock_sheet]
            # Total P&L could be in a totals row or as a single cell in col B/C (2/3)
            cands = _scan_numeric(ws, range(2, 15), range(2, 6), -400000, -100000)
            if cands:
                # The largest negative is likely the total
                pl_total = min(v for _, _, v in cands)
        if pl_total is not None and -360000 <= pl_total <= -230000:
            score += 20
            fb.append(f"C5 PASS: +100bps total P&L = ${pl_total:,.0f}")
        elif pl_total is not None:
            fb.append(f"C5 FAIL: +100bps P&L = ${pl_total:,.0f} (expected -$360K to -$230K)")
        else:
            fb.append("C5 FAIL: Shock scenario P&L not found")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

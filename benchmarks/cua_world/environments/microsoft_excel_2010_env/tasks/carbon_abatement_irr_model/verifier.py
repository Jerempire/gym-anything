#!/usr/bin/env python3
"""
Verifier for carbon_abatement_irr_model task.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Independently copy xlsx and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts, pass >= 60)

Criteria:
  C1 (25 pts): Company total Scope 1+2 in [8300, 9200] MT CO2e (expected ~8740)
  C2 (20 pts): Company carbon cost in [$415K, $460K] (expected ~$437K)
  C3 (20 pts): LED Lighting NPV in [$50K, $90K] (expected ~$68K)
  C4 (20 pts): Rooftop Solar PV NPV in [$340K, $455K] (expected ~$397K)
  C5 (15 pts): At least 2 valid IRR values in [5%, 25%]
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\carbon_abatement_irr_model_result.json"
XLSX_PATH   = "C:/Users/Docker/Desktop/ExcelTasks/facility_emissions.xlsx"


def _find_sheet(wb, keywords):
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def verify_carbon_abatement_irr_model(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_carbon_")
    try:
        # ── STEP 1: Check is_new ──────────────────────────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False)."
            }

        # ── STEP 2: Copy and parse xlsx ──────────────────────────────────────────
        xlsx_local = os.path.join(tmp, "facility_emissions.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0, "feedback": "facility_emissions.xlsx not found"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        score = 0
        fb = []

        em_sheet  = _find_sheet(wb, ["emission", "summary", "ghg", "inventory"])
        irr_sheet = _find_sheet(wb, ["project", "irr", "abatement", "npv"])

        # Criterion 1: Company total Scope 1+2 in [8300, 9200]
        total_scope12 = None
        if em_sheet:
            ws = wb[em_sheet]
            cands = _scan_numeric(ws, range(4, 14), range(6, 11), 8000.0, 10000.0)
            row9 = [(c, v) for r, c, v in cands if r == 9]
            if row9:
                total_scope12 = row9[0][1]
            elif cands:
                total_scope12 = max(v for _, _, v in cands)

        if total_scope12 is not None and 8300 <= total_scope12 <= 9200:
            score += 25
            fb.append(f"C1 PASS: Total Scope 1+2 = {total_scope12:.1f} MT (expected ~8740)")
        elif total_scope12 is not None:
            # Partial credit if individual facility totals are present
            indiv = _scan_numeric(wb[em_sheet], range(4, 9), range(6, 11), 1000, 4000) if em_sheet else []
            if len(indiv) >= 3:
                score += 12
                fb.append(f"C1 PARTIAL: Facility totals populated but company total {total_scope12:.0f} out of range")
            else:
                fb.append(f"C1 FAIL: Total Scope 1+2 = {total_scope12:.1f} (expected 8300-9200)")
        else:
            indiv = _scan_numeric(wb[em_sheet], range(4, 9), range(6, 11), 1000, 4000) if em_sheet else []
            if len(indiv) >= 3:
                score += 10
                fb.append(f"C1 PARTIAL: {len(indiv)} facility Scope 1+2 values found, company total missing")
            else:
                fb.append("C1 FAIL: Emissions_Summary not populated")

        # Criterion 2: Carbon cost in [$415K, $460K]
        carbon_cost = None
        if em_sheet:
            ws = wb[em_sheet]
            cands = _scan_numeric(ws, range(4, 14), range(7, 12), 400000, 480000)
            row9 = [(c, v) for r, c, v in cands if r == 9]
            if row9:
                carbon_cost = row9[0][1]
            elif cands:
                carbon_cost = max(v for _, _, v in cands)

        if carbon_cost is not None and 415000 <= carbon_cost <= 460000:
            score += 20
            fb.append(f"C2 PASS: Carbon cost = ${carbon_cost:,.0f} (expected ~$437K)")
        elif total_scope12 is not None and 8300 <= total_scope12 <= 9200:
            implied = total_scope12 * 50
            if 415000 <= implied <= 460000:
                score += 15
                fb.append(f"C2 PARTIAL: Correct emissions; carbon cost cell not found (implied ${implied:,.0f})")
            else:
                fb.append(f"C2 FAIL: Carbon cost {carbon_cost} not in range")
        else:
            fb.append("C2 FAIL: Carbon cost not found")

        # Criterion 3: LED NPV in [$50K, $90K]
        led_npv = None
        if irr_sheet:
            ws = wb[irr_sheet]
            cands = _scan_numeric(ws, range(18, 24), range(5, 9), 40000, 110000)
            row19 = [(c, v) for r, c, v in cands if r == 19]
            if row19:
                led_npv = row19[0][1]
            elif cands:
                led_npv = cands[0][2]

        if led_npv is not None and 50000 <= led_npv <= 90000:
            score += 20
            fb.append(f"C3 PASS: LED NPV = ${led_npv:,.0f} (expected ~$68K)")
        elif led_npv is not None:
            fb.append(f"C3 FAIL: LED NPV = ${led_npv:,.0f} (expected $50K-$90K)")
        else:
            fb.append("C3 FAIL: LED Lighting NPV not found")

        # Criterion 4: Solar NPV in [$340K, $455K]
        solar_npv = None
        if irr_sheet:
            ws = wb[irr_sheet]
            cands = _scan_numeric(ws, range(19, 25), range(5, 9), 300000, 500000)
            row21 = [(c, v) for r, c, v in cands if r == 21]
            if row21:
                solar_npv = row21[0][1]
            elif cands:
                solar_npv = max(v for _, _, v in cands)

        if solar_npv is not None and 340000 <= solar_npv <= 455000:
            score += 20
            fb.append(f"C4 PASS: Solar NPV = ${solar_npv:,.0f} (expected ~$397K)")
        elif solar_npv is not None:
            fb.append(f"C4 FAIL: Solar NPV = ${solar_npv:,.0f} (expected $340K-$455K)")
        else:
            fb.append("C4 FAIL: Rooftop Solar NPV not found")

        # Criterion 5: At least 2 IRR values in [5%, 25%]
        irr_vals = []
        if irr_sheet:
            ws = wb[irr_sheet]
            # Decimal IRR (0.05-0.25)
            decimal = _scan_numeric(ws, range(18, 25), range(6, 10), 0.05, 0.25)
            if decimal:
                irr_vals = [v * 100 for _, _, v in decimal]
            else:
                # Percent IRR (5-25)
                pct = _scan_numeric(ws, range(18, 25), range(6, 10), 5.0, 25.0)
                irr_vals = [v for _, _, v in pct if v <= 30]

        if len(irr_vals) >= 2:
            score += 15
            vals_str = ", ".join(f"{v:.1f}%" for v in irr_vals[:3])
            fb.append(f"C5 PASS: {len(irr_vals)} IRR values found: {vals_str}")
        elif len(irr_vals) == 1:
            score += 7
            fb.append(f"C5 PARTIAL: Only 1 IRR value: {irr_vals[0]:.1f}%")
        else:
            fb.append("C5 FAIL: No valid IRR values (expected ~15.6%, ~8.8%, ~14.6%)")

        return {"passed": score >= 60, "score": score, "feedback": " | ".join(fb)}

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

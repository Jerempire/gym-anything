#!/usr/bin/env python3
"""
Verifier for Supply Chain SLA Audit task.
Verifies Excel logic, calculation accuracy, summary table, and formatting.
"""

import json
import os
import tempfile
import logging
from datetime import datetime

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try importing openpyxl
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("openpyxl not available")

def verify_supply_chain_sla_audit(traj, env_info, task_info):
    """
    Verifies the logistics audit task.
    
    Criteria:
    1. File exists and modified (10 pts)
    2. 'Late' logic and 'Rebate Amount' calculation correct (40 pts)
    3. Audit_Summary sheet exists and has correct totals (30 pts)
    4. Conditional formatting applied (20 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}
    
    if not OPENPYXL_AVAILABLE:
        return {"passed": False, "score": 0, "feedback": "Verification failed: openpyxl library missing"}

    metadata = task_info.get('metadata', {})
    sla_rules = metadata.get('sla_rules', {})
    
    # Setup temp files
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json').name
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx').name
    
    try:
        # Get result JSON
        copy_from_env("C:\\tmp\\task_result.json", temp_json)
        with open(temp_json, 'r') as f:
            result_data = json.load(f)
            
        if not result_data.get('output_exists'):
            return {"passed": False, "score": 0, "feedback": "Output file logistics_audit_q3.xlsx not found."}
            
        if not result_data.get('file_modified'):
             return {"passed": False, "score": 0, "feedback": "File was not saved/modified during the task."}

        # Get XLSX file
        copy_from_env("C:\\Users\\Docker\\Documents\\logistics_audit_q3.xlsx", temp_xlsx)
        
        # Load Workbook
        wb = openpyxl.load_workbook(temp_xlsx, data_only=True)
        
        score = 10 # Base score for file existing and being modified
        feedback = []
        
        # --- Verify Calculations (Shipments Sheet) ---
        if "Shipments" not in wb.sheetnames:
            return {"passed": False, "score": score, "feedback": "Shipments sheet deleted or missing."}
            
        ws = wb["Shipments"]
        
        # Find columns
        headers = {cell.value: i for i, cell in enumerate(ws[1])}
        required_cols = ["Shipping Mode", "Sales Amount", "Scheduled Days", "Actual Days"]
        if not all(col in headers for col in required_cols):
             return {"passed": False, "score": score, "feedback": "Required columns missing in Shipments sheet."}
        
        # Find rebate column (agent created it, might be named 'Rebate' or 'Rebate Amount')
        rebate_col_idx = None
        for key in headers:
            if "rebate" in str(key).lower():
                rebate_col_idx = headers[key]
                break
        
        if rebate_col_idx is None:
            feedback.append("Could not find 'Rebate Amount' column.")
        else:
            correct_calcs = 0
            total_checked = 0
            calculated_totals = {mode: 0.0 for mode in sla_rules}
            
            # Iterate rows (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[headers["Order ID"]]: continue # Skip empty rows
                
                mode = row[headers["Shipping Mode"]]
                sales = float(row[headers["Sales Amount"]] or 0)
                sched = float(row[headers["Scheduled Days"]] or 0)
                actual = float(row[headers["Actual Days"]] or 0)
                
                # Ground truth logic
                is_late = actual > sched
                rebate_pct = sla_rules.get(mode, 0.0)
                expected_rebate = sales * rebate_pct if is_late else 0.0
                calculated_totals[mode] += expected_rebate
                
                # Agent value
                agent_val = row[rebate_col_idx]
                try:
                    agent_val = float(agent_val or 0)
                except:
                    agent_val = -1 # Error value
                
                if abs(agent_val - expected_rebate) < 0.05:
                    correct_calcs += 1
                total_checked += 1
            
            if total_checked > 0 and correct_calcs / total_checked > 0.9:
                score += 40
                feedback.append(f"Rebate calculations correct ({correct_calcs}/{total_checked}).")
            else:
                feedback.append(f"Rebate calculations incorrect (Accuracy: {correct_calcs}/{total_checked}).")

        # --- Verify Summary Table (Audit_Summary Sheet) ---
        if "Audit_Summary" not in wb.sheetnames:
            feedback.append("Audit_Summary sheet missing.")
        else:
            ws_sum = wb["Audit_Summary"]
            # Look for mode names and values
            summary_found = 0
            summary_correct = 0
            
            # Scan all cells
            sheet_values = []
            for row in ws_sum.iter_rows(values_only=True):
                sheet_values.extend([str(c) for c in row if c is not None])
            
            sheet_text = " ".join(sheet_values)
            
            # Check if totals match ground truth calculated above
            match_count = 0
            for mode, total in calculated_totals.items():
                if total > 0:
                    # Check if the total appears roughly in the text
                    # We check int value or formatted value
                    target = int(total)
                    if str(target) in sheet_text:
                        match_count += 1
            
            if match_count >= 2: # At least 2 modes match
                score += 30
                feedback.append("Summary table totals verified.")
            else:
                feedback.append("Summary table totals do not match expected values.")

        # --- Verify Conditional Formatting ---
        # Note: openpyxl reading conditional formatting is complex, 
        # checking if ANY rule exists on the sheet is a good proxy, 
        # or checking specific ranges.
        cf_score = 0
        if rebate_col_idx is not None:
            # We assume the agent applied it to the column. 
            # Ideally we parse ws.conditional_formatting
            if len(ws.conditional_formatting) > 0:
                cf_score = 20
                feedback.append("Conditional formatting rules found.")
            else:
                feedback.append("No conditional formatting rules found.")
        score += cf_score

        return {
            "passed": score >= 70,
            "score": score,
            "feedback": " ".join(feedback)
        }
        
    except Exception as e:
        logger.error(f"Verification error: {e}")
        return {"passed": False, "score": 0, "feedback": f"Verification error: {str(e)}"}
    finally:
        if os.path.exists(temp_json): os.remove(temp_json)
        if os.path.exists(temp_xlsx): os.remove(temp_xlsx)
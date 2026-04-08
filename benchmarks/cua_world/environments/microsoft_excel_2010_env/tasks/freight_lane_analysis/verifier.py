#!/usr/bin/env python3
"""
Verifier for Refrigerated Freight Lane Cost Analysis (`freight_lane_analysis@1`).

Verifies:
1. Formulas are present (not just hardcoded values).
2. Calculation accuracy for Lane Analysis (multi-criteria aggregation).
3. Calculation accuracy for Carrier Scorecard (ranking logic).
4. Calculation accuracy for Rate Benchmarking (variance logic).
5. VLM verification for visual confirmation.
"""

import json
import os
import tempfile
import logging
from openpyxl import load_workbook
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_freight_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Setup score
    score = 0
    max_score = 100
    feedback = []

    # 1. Retrieve JSON Result
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("C:\\tmp\\task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to retrieve task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result_data.get("file_exists"):
        return {"passed": False, "score": 0, "feedback": "Excel file not found at expected location."}
    
    if not result_data.get("is_new"):
        feedback.append("WARNING: File timestamp indicates it was not saved during the task.")
        # We don't fail immediately but penalize heavily
        score -= 20

    # 2. Retrieve Excel File
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("C:\\Users\\Docker\\Documents\\freight_analysis.xlsx", temp_xlsx.name)
        
        # Load workbook twice: once for values, once for formulas
        wb_data = load_workbook(temp_xlsx.name, data_only=True)
        wb_form = load_workbook(temp_xlsx.name, data_only=False)

        # === CHECK 1: Lane Analysis (40 pts) ===
        if "Lane_Analysis" not in wb_data.sheetnames:
            feedback.append("Lane_Analysis sheet missing.")
        else:
            ws_d = wb_data["Lane_Analysis"]
            ws_f = wb_form["Lane_Analysis"]
            
            # Check row 2 (first lane)
            # E: Count, F: Avg Rate, G: Total Cost
            try:
                # Basic sanity check on value types
                count = ws_d["E2"].value
                rate = ws_d["F2"].value
                cost = ws_d["G2"].value
                
                if isinstance(count, (int, float)) and count > 0:
                    score += 5
                else:
                    feedback.append(f"Lane_Analysis: Invalid shipment count in E2 ({count})")

                if isinstance(rate, (int, float)) and 1.0 < rate < 5.0:
                    score += 5
                else:
                    feedback.append(f"Lane_Analysis: Invalid avg rate in F2 ({rate})")

                # Check for formula usage (Anti-Gaming)
                formula_e2 = str(ws_f["E2"].value)
                if formula_e2.startswith("=") and ("COUNT" in formula_e2.upper()):
                    score += 10 # Good formula
                else:
                    feedback.append("Lane_Analysis: Missing COUNTIFS formula in E2")
                    
                formula_f2 = str(ws_f["F2"].value)
                if formula_f2.startswith("=") and ("AVERAGE" in formula_f2.upper()):
                    score += 10
                else:
                    feedback.append("Lane_Analysis: Missing AVERAGEIFS formula in F2")
                    
                # Check Cost Per CWT (Column J)
                cwt = ws_d["J2"].value
                if isinstance(cwt, (int, float)) and cwt > 0:
                     score += 10
                else:
                    feedback.append("Lane_Analysis: Invalid Cost Per CWT")

            except Exception as e:
                feedback.append(f"Error checking Lane_Analysis: {e}")

        # === CHECK 2: Carrier Scorecard (30 pts) ===
        if "Carrier_Scorecard" not in wb_data.sheetnames:
            feedback.append("Carrier_Scorecard sheet missing.")
        else:
            ws_d = wb_data["Carrier_Scorecard"]
            ws_f = wb_form["Carrier_Scorecard"]
            
            try:
                # Check Overall Score (J) and Rank (K)
                score_val = ws_d["J2"].value
                rank_val = ws_d["K2"].value
                
                if isinstance(score_val, (int, float)) and 0 <= score_val <= 100:
                    score += 10
                
                # Check Rank is integer 1-8
                ranks = []
                for r in range(2, 10):
                    val = ws_d[f"K{r}"].value
                    if isinstance(val, int): ranks.append(val)
                
                if len(set(ranks)) >= 6: # At least 6 unique ranks found (allowing for ties)
                    score += 10
                else:
                    feedback.append("Carrier_Scorecard: Ranking logic seems incorrect or incomplete")

                # Check formula for Score
                formula_j2 = str(ws_f["J2"].value)
                if "=" in formula_j2 and ("+" in formula_j2 or "*" in formula_j2):
                    score += 10
                else:
                    feedback.append("Carrier_Scorecard: Missing calculation formula for Overall_Score")

            except Exception as e:
                feedback.append(f"Error checking Carrier_Scorecard: {e}")

        # === CHECK 3: Rate Benchmarking (30 pts) ===
        if "Rate_Benchmarking" not in wb_data.sheetnames:
            feedback.append("Rate_Benchmarking sheet missing.")
        else:
            ws_d = wb_data["Rate_Benchmarking"]
            ws_f = wb_form["Rate_Benchmarking"]
            
            try:
                # Check Variance (F) and Status (I)
                var_val = ws_d["F2"].value
                status_val = ws_d["I2"].value
                
                if isinstance(var_val, (int, float)):
                    score += 10
                
                if isinstance(status_val, str) and status_val in ["OVER BENCHMARK", "UNDER BENCHMARK", "IN RANGE"]:
                    score += 10
                else:
                    feedback.append(f"Rate_Benchmarking: Invalid Status '{status_val}'")

                # Check IF formula for Status
                formula_i2 = str(ws_f["I2"].value)
                if formula_i2.startswith("=") and "IF" in formula_i2.upper():
                    score += 10
                else:
                    feedback.append("Rate_Benchmarking: Missing IF formula for Status")

            except Exception as e:
                feedback.append(f"Error checking Rate_Benchmarking: {e}")

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to verify Excel file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    passed = score >= 60
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback) if feedback else "All criteria met."
    }
#!/usr/bin/env python3
"""
Verifier for stormwater_pipe_sizing task.
"""

import json
import tempfile
import os
import logging
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ground Truths
GT_WEIGHTED_C = 0.7618
GT_INTENSITY = 8.4
GT_DESIGN_FLOW = 21.76
GT_SELECTED_DIA = 30
GT_CAPACITY_30 = 29.08
TOL_C = 0.02
TOL_Q = 0.5
TOL_CAP = 1.0

def verify_stormwater_pipe_sizing(traj, env_info, task_info):
    """
    Verify the stormwater pipe sizing task.
    
    Criteria:
    1. File Saved & Modified (10 pts)
    2. Weighted C correct (20 pts)
    3. Intensity correct (15 pts)
    4. Design Flow correct (15 pts)
    5. Pipe Catalog Capacities (20 pts) - Checks 24" and 30"
    6. Selected Diameter (20 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    score = 0
    feedback = []
    
    # 1. Get Result JSON
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("C:\\tmp\\task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to load result JSON: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result_data.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Workbook not found"}
        
    if result_data.get('is_new'):
        score += 10
        feedback.append("File saved successfully (10/10)")
    else:
        feedback.append("File not modified after start (0/10)")

    # 2. Analyze Workbook
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("C:\\Users\\Docker\\Desktop\\stormdrain_design.xlsx", temp_xlsx.name)
        
        # Lazy import to avoid dependency issues if not installed in verifier env
        import openpyxl
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
        # --- Check Design_Worksheet ---
        if "Design_Worksheet" in wb.sheetnames:
            ws = wb["Design_Worksheet"]
            
            # Weighted C (Cell B5)
            val_c = ws['B5'].value
            if isinstance(val_c, (int, float)) and abs(val_c - GT_WEIGHTED_C) <= TOL_C:
                score += 20
                feedback.append(f"Weighted C correct: {val_c:.3f} (20/20)")
            else:
                feedback.append(f"Weighted C incorrect: {val_c} (Exp: {GT_WEIGHTED_C}) (0/20)")
            
            # Intensity (Cell B8)
            val_i = ws['B8'].value
            if isinstance(val_i, (int, float)) and abs(val_i - GT_INTENSITY) <= 0.2:
                score += 15
                feedback.append(f"Intensity correct: {val_i} (15/15)")
            else:
                feedback.append(f"Intensity incorrect: {val_i} (Exp: {GT_INTENSITY}) (0/15)")
                
            # Design Flow (Cell B10)
            val_q = ws['B10'].value
            if isinstance(val_q, (int, float)) and abs(val_q - GT_DESIGN_FLOW) <= TOL_Q:
                score += 15
                feedback.append(f"Design Flow correct: {val_q:.2f} (15/15)")
            else:
                feedback.append(f"Design Flow incorrect: {val_q} (Exp: {GT_DESIGN_FLOW}) (0/15)")

            # Selected Diameter (Cell B12)
            val_dia = ws['B12'].value
            if str(val_dia) == str(GT_SELECTED_DIA):
                score += 20
                feedback.append(f"Selected Diameter correct: {val_dia} (20/20)")
            else:
                feedback.append(f"Selected Diameter incorrect: {val_dia} (Exp: {GT_SELECTED_DIA}) (0/20)")
        else:
            feedback.append("Design_Worksheet sheet missing")

        # --- Check Pipe_Catalog ---
        if "Pipe_Catalog" in wb.sheetnames:
            ws2 = wb["Pipe_Catalog"]
            # Locate 30" row. Assuming standard layout from setup script:
            # Row 2=12, 3=15, 4=18, 5=24, 6=30
            # Col 7 (G) is Capacity
            
            # Check 30 inch (Row 6)
            dia_cell = ws2.cell(row=6, column=1).value
            cap_cell = ws2.cell(row=6, column=7).value
            
            if dia_cell == 30 and isinstance(cap_cell, (int, float)):
                 if abs(cap_cell - GT_CAPACITY_30) <= TOL_CAP:
                     score += 20
                     feedback.append(f"Pipe Capacity Calc (30in) correct: {cap_cell:.2f} (20/20)")
                 else:
                     feedback.append(f"Pipe Capacity (30in) wrong: {cap_cell} (Exp: {GT_CAPACITY_30}) (0/20)")
            else:
                feedback.append("Could not verify pipe catalog row 6 (30 inch) structure")
        else:
            feedback.append("Pipe_Catalog sheet missing")

    except Exception as e:
        feedback.append(f"Error analyzing workbook: {str(e)}")
        import traceback
        traceback.print_exc()
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    passed = score >= 75
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback)
    }
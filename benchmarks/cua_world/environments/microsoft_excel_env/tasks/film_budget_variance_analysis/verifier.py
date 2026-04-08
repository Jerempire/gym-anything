#!/usr/bin/env python3
"""
Verifier for film_budget_variance_analysis task.

Verification pipeline:
  1. Read result JSON (C:\\Users\\Docker\\film_budget_variance_analysis_result.json)
     - Check is_new: if xlsx was not saved after task start -> score 0
  2. Independently copy xlsx from VM and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts total, pass >= 60)

Ground truth values (computed from the 32 line items in Line_Item_Detail):
  ATL budget  ~$754,096   actual ~$754,096   (0.0% variance)
  BTL-Prod    ~$689,075   actual ~$732,475   (6.3% over)
  Post        ~$129,412   actual ~$144,912   (12.0% over)
  Other (incl contingency) ~$209,000  actual ~$73,000  (-65.1% under)
  Grand total budget ~$1,781,583  actual ~$1,704,483  (-4.3%)
  Contingency: $140,000, utilization: 44.9%

Criteria:
  C1 (20 pts): Budget_vs_Actual has Total_Budget for all 4 categories
               (values in range 50,000 - 1,000,000)
  C2 (25 pts): Grand Total actual spend in [1,550,000 .. 1,850,000]
  C3 (15 pts): At least 2 Status_Flags present (e.g. OVER_BUDGET, ON_TRACK,
               UNDER_BUDGET)
  C4 (20 pts): Contingency_Tracker has Contingency Utilization % in [30, 100]
  C5 (20 pts): Department_Summary has at least 12 departments with
               Budget_Total > 0
"""
import json, logging, os, tempfile, shutil

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\film_budget_variance_analysis_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/film_budget.xlsx"


def _find_sheet(wb, keywords):
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    """Scan a region for numeric values within [lo, hi]."""
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    """Count cells containing keyword (case-insensitive)."""
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_film_budget_variance_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_film_budget_")
    try:
        # ── STEP 1: Read result JSON and check is_new ────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False). "
                            "Agent must save the file with Ctrl+S after completing formulas."
            }

        # ── STEP 2: Copy and parse xlsx ──────────────────────────────────────
        xlsx_local = os.path.join(tmp, "film_budget.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0,
                    "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0,
                    "feedback": "film_budget.xlsx not found or empty"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        # ── STEP 3: Score criteria ───────────────────────────────────────────
        score = 0
        fb = []

        # ----------------------------------------------------------------
        # C1 (20 pts): Budget_vs_Actual has Total_Budget for all 4 categories
        #              (values in range 50,000 - 1,000,000)
        # ----------------------------------------------------------------
        bva_sheet = _find_sheet(wb, ["Budget_vs_Actual", "budget_vs", "BvA"])
        budget_vals = []
        if bva_sheet:
            ws = wb[bva_sheet]
            # Category labels expected in col A (rows 2-6), budget totals in col B (2)
            # Scan rows 1-10, cols 2-12 for budget-sized numbers
            cells = _scan_numeric(ws, range(1, 12), range(2, 13), 50000, 1000000)
            # Prefer col B (2) — Total_Budget column
            col_b = [(r, v) for r, c, v in cells if c == 2]
            budget_vals = [v for _, v in col_b] if len(col_b) >= 3 else [v for _, _, v in cells[:8]]

        if len(budget_vals) >= 4:
            score += 20
            fb.append(f"C1 PASS: Total_Budget for {len(budget_vals)} categories found")
        elif len(budget_vals) >= 2:
            score += 10
            fb.append(f"C1 PARTIAL: Total_Budget for {len(budget_vals)} categories (expected 4)")
        else:
            fb.append(f"C1 FAIL: Only {len(budget_vals)} category budget totals found (expected 4)")

        # ----------------------------------------------------------------
        # C2 (25 pts): Grand Total actual spend in [1,550,000 .. 1,850,000]
        # ----------------------------------------------------------------
        grand_actual = None
        if bva_sheet:
            ws = wb[bva_sheet]
            # Grand total row is typically the last data row; scan wider range
            # Look for a value in the actual column (col C=3) in the grand total range
            cands = _scan_numeric(ws, range(1, 15), range(2, 13), 1400000, 2000000)
            # Among candidates, prefer the one in an actual-total column (C=3)
            col_c = [v for r, c, v in cands if c == 3]
            if col_c:
                grand_actual = col_c[-1]  # last match likely grand total row
            elif cands:
                # Fall back: look for any value in the target range
                for _, _, v in cands:
                    if 1550000 <= v <= 1850000:
                        grand_actual = v
                        break

        if grand_actual is not None and 1550000 <= grand_actual <= 1850000:
            score += 25
            fb.append(f"C2 PASS: Grand Total actual = ${grand_actual:,.0f}")
        elif grand_actual is not None and (1400000 <= grand_actual < 1550000 or 1850000 < grand_actual <= 2000000):
            # Allow wider tolerance for agents that include contingency budget differently
            score += 15
            fb.append(f"C2 PARTIAL: Grand Total actual = ${grand_actual:,.0f} "
                       "(slightly outside expected 1.55M-1.85M)")
        elif grand_actual is not None:
            fb.append(f"C2 FAIL: Grand Total actual = ${grand_actual:,.0f} "
                       "(expected 1,550,000 - 1,850,000)")
        else:
            # Try scanning all sheets for a grand total actual
            for sn in wb.sheetnames:
                ws2 = wb[sn]
                cands2 = _scan_numeric(ws2, range(1, 20), range(1, 15), 1400000, 2000000)
                if cands2:
                    grand_actual = cands2[0][2]
                    if 1550000 <= grand_actual <= 1850000:
                        score += 20
                        fb.append(f"C2 PASS (alt sheet): Grand Total actual = ${grand_actual:,.0f}")
                    else:
                        score += 10
                        fb.append(f"C2 PARTIAL (alt sheet): Grand Total actual = ${grand_actual:,.0f}")
                    break
            else:
                fb.append("C2 FAIL: Grand Total actual spend not found")

        # ----------------------------------------------------------------
        # C3 (15 pts): At least 2 Status_Flags present
        #              (OVER_BUDGET, ON_TRACK, UNDER_BUDGET)
        # ----------------------------------------------------------------
        status_flags_found = set()
        if bva_sheet:
            ws = wb[bva_sheet]
            for flag_keyword in ["OVER_BUDGET", "ON_TRACK", "UNDER_BUDGET"]:
                cnt = _count_string(ws, range(1, 12), range(2, 15), flag_keyword)
                if cnt > 0:
                    status_flags_found.add(flag_keyword)

        if len(status_flags_found) >= 2:
            score += 15
            fb.append(f"C3 PASS: {len(status_flags_found)} status flags found "
                       f"({', '.join(sorted(status_flags_found))})")
        elif len(status_flags_found) >= 1:
            score += 7
            fb.append(f"C3 PARTIAL: {len(status_flags_found)} status flag(s) found "
                       f"({', '.join(sorted(status_flags_found))})")
        else:
            fb.append("C3 FAIL: No Status_Flag values (OVER_BUDGET/ON_TRACK/UNDER_BUDGET) found")

        # ----------------------------------------------------------------
        # C4 (20 pts): Contingency_Tracker has Contingency Utilization % in [30, 100]
        # ----------------------------------------------------------------
        cont_sheet = _find_sheet(wb, ["Contingency", "contingency"])
        cont_util = None
        if cont_sheet:
            ws = wb[cont_sheet]
            # Utilization % could be a value 30-100 (as percentage) or 0.3-1.0 (as decimal)
            # Check for percentage-style first (30-100)
            cands = _scan_numeric(ws, range(1, 20), range(1, 10), 30.0, 100.0)
            if cands:
                cont_util = cands[0][2]
            else:
                # Check for decimal-style (0.3 - 1.0)
                cands_dec = _scan_numeric(ws, range(1, 20), range(1, 10), 0.3, 1.0)
                if cands_dec:
                    cont_util = cands_dec[0][2] * 100  # convert to percentage

        if cont_util is not None and 30.0 <= cont_util <= 100.0:
            score += 20
            fb.append(f"C4 PASS: Contingency Utilization = {cont_util:.1f}%")
        elif cont_util is not None:
            fb.append(f"C4 FAIL: Contingency Utilization = {cont_util:.1f}% "
                       "(expected 30-100%)")
        else:
            # Try checking for Budget_Health flag as partial credit
            health_found = False
            if cont_sheet:
                ws = wb[cont_sheet]
                for kw in ["CRITICAL", "WARNING", "HEALTHY"]:
                    if _count_string(ws, range(1, 20), range(1, 10), kw) > 0:
                        health_found = True
                        break
            if health_found:
                score += 8
                fb.append("C4 PARTIAL: Budget_Health flag found but utilization % not detected")
            else:
                fb.append("C4 FAIL: Contingency Utilization % not found")

        # ----------------------------------------------------------------
        # C5 (20 pts): Department_Summary has at least 12 departments with
        #              Budget_Total > 0
        # ----------------------------------------------------------------
        dept_sheet = _find_sheet(wb, ["Department", "dept"])
        dept_count = 0
        if dept_sheet:
            ws = wb[dept_sheet]
            # Department names in col A, budget totals likely in col B (2)
            # Count rows where col B has a positive numeric value
            for r in range(2, 25):
                # Check col B (2) for budget total
                val = ws.cell(r, 2).value
                if val is not None and isinstance(val, (int, float)) and val > 0:
                    dept_count += 1
            # If col B didn't work, try broader scan
            if dept_count < 5:
                dept_count = 0
                for r in range(2, 25):
                    for c in range(2, 8):
                        val = ws.cell(r, c).value
                        if val is not None and isinstance(val, (int, float)) and val > 10000:
                            dept_count += 1
                            break  # count each row once

        if dept_count >= 12:
            score += 20
            fb.append(f"C5 PASS: {dept_count} departments with Budget_Total > 0")
        elif dept_count >= 8:
            score += 10
            fb.append(f"C5 PARTIAL: {dept_count} departments with budgets (expected >= 12)")
        elif dept_count >= 4:
            score += 5
            fb.append(f"C5 PARTIAL: {dept_count} departments with budgets (expected >= 12)")
        else:
            fb.append(f"C5 FAIL: Only {dept_count} departments with Budget_Total > 0 "
                       "(expected >= 12)")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

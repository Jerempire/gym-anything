#!/usr/bin/env python3
"""
Verifier for lease_portfolio_npv_analysis task.

Verification pipeline:
  1. Read result JSON (C:\\Users\\Docker\\lease_portfolio_npv_analysis_result.json)
     - Check is_new: if xlsx was not saved after task start -> score 0
  2. Independently copy xlsx from VM and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts total, pass >= 60)

Criteria:
  C1 (20 pts): Cash_Flow_Projection has Year_1_NOI for >= 10 of 12 properties
               (positive values in range 100,000 - 3,000,000)
  C2 (25 pts): At least 6 Year_1_NOI values within +/-12% of computed ground truth
               Ground truth Y1 NOI = Base_Rent_Monthly * 12 * (1 - Vacancy/100) - Opex * SqFt
  C3 (15 pts): At least 3 Value_Flags present (UNDERVALUED for cap >= 7.0%, PREMIUM for cap <= 5.5%)
  C4 (20 pts): Portfolio total Year_1_NOI in range [8,000,000 - 12,000,000]
  C5 (20 pts): At least 6 NPV values present (positive values in range 1,000,000 - 50,000,000)
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\lease_portfolio_npv_analysis_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/lease_portfolio.xlsx"

# Expected Year-1 NOI ground truth per property (approximate).
# Formula: Base_Rent_Monthly * 12 * (1 - Vacancy_Allowance_pct/100) - Opex_per_sqft_annual * Sq_Ft
# These are derived from the Lease_Schedule data for 12 properties.
EXPECTED_Y1_NOI = {
    "PROP-01": 1570500,
    "PROP-02": 565800,
    "PROP-03": 441276,
    "PROP-04": 1580800,
    "PROP-05": 304560,
    "PROP-06": 1237600,
    "PROP-07": 1258000,
    "PROP-08": 432000,
    "PROP-09": 621596,
    "PROP-10": 1215504,
    "PROP-11": 282480,
    "PROP-12": 504796,
}


def _find_sheet(wb, keywords):
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    """Return list of (row, col, value) for numeric cells within [lo, hi]."""
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    """Count cells containing keyword (case-insensitive)."""
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_lease_portfolio_npv_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_lease_")
    try:
        # ── STEP 1: Read result JSON and check is_new ────────────────────────
        json_local = os.path.join(tmp, "result.json")
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            result = {}
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False). "
                            "Agent must save the file with Ctrl+S after completing formulas."
            }

        # ── STEP 2: Independently copy and parse xlsx ─────────────────────────
        xlsx_local = os.path.join(tmp, "lease_portfolio.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0,
                    "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0,
                    "feedback": "lease_portfolio.xlsx not found or empty"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        # ── STEP 3: Score criteria ────────────────────────────────────────────
        score = 0
        fb = []

        # ── C1 (20 pts): Year_1_NOI for >= 10 of 12 properties ───────────────
        cf_sheet = _find_sheet(wb, ["cash_flow", "cashflow", "projection", "cash flow"])
        y1_noi_vals = []
        if cf_sheet:
            ws = wb[cf_sheet]
            # Year_1_NOI is typically in the first NOI column (col B or C)
            # Scan rows 2-14 (12 properties + possible header offset), cols 2-15
            cells = _scan_numeric(ws, range(2, 16), range(2, 16), 100000, 3000000)
            if cells:
                # Group by row, take the first qualifying value per row (Year 1 NOI)
                seen_rows = set()
                for r, c, v in cells:
                    if r not in seen_rows:
                        y1_noi_vals.append(v)
                        seen_rows.add(r)

        if len(y1_noi_vals) >= 10:
            score += 20
            fb.append(f"C1 PASS: Year_1_NOI found for {len(y1_noi_vals)} properties")
        elif len(y1_noi_vals) >= 6:
            score += 10
            fb.append(f"C1 PARTIAL: Year_1_NOI found for {len(y1_noi_vals)} properties (need >= 10)")
        else:
            fb.append(f"C1 FAIL: Only {len(y1_noi_vals)} Year_1_NOI values found (need >= 10)")

        # ── C2 (25 pts): >= 6 Y1 NOI within +/-12% of ground truth ───────────
        if y1_noi_vals:
            accurate = 0
            for exp in EXPECTED_Y1_NOI.values():
                for found in y1_noi_vals:
                    if exp > 0 and abs(found - exp) / exp <= 0.12:
                        accurate += 1
                        break
            if accurate >= 6:
                score += 25
                fb.append(f"C2 PASS: {accurate} Year_1_NOI values within 12% of expected")
            elif accurate >= 3:
                score += 12
                fb.append(f"C2 PARTIAL: {accurate} Year_1_NOI values accurate (need >= 6)")
            else:
                fb.append(f"C2 FAIL: Only {accurate} Year_1_NOI values within 12% of expected")
        else:
            fb.append("C2 FAIL: No Year_1_NOI values to evaluate")

        # ── C3 (15 pts): >= 3 Value_Flags (UNDERVALUED for cap >= 7.0%, PREMIUM for cap <= 5.5%)
        # Expected: 4 UNDERVALUED (PROP-03, PROP-08, PROP-11 at 7.25%, PROP-12 at 7.00%)
        #           1 PREMIUM (PROP-06 at 5.50%) = 5 total flags
        pm_sheet = _find_sheet(wb, ["portfolio_metrics", "metrics", "portfolio metrics"])
        flag_count = 0
        if pm_sheet:
            ws = wb[pm_sheet]
            # Scan for UNDERVALUED and PREMIUM strings in cols 2-12, rows 2-16
            flag_count += _count_string(ws, range(2, 16), range(2, 12), "UNDERVALUED")
            flag_count += _count_string(ws, range(2, 16), range(2, 12), "PREMIUM")

        if flag_count >= 3:
            score += 15
            fb.append(f"C3 PASS: {flag_count} Value_Flag entries found (UNDERVALUED/PREMIUM)")
        elif flag_count >= 1:
            score += 7
            fb.append(f"C3 PARTIAL: {flag_count} Value_Flag entry found (need >= 3)")
        else:
            fb.append("C3 FAIL: No UNDERVALUED or PREMIUM Value_Flags found")

        # ── C4 (20 pts): Portfolio total Year_1_NOI in [8M, 12M] ─────────────
        portfolio_y1 = None
        if cf_sheet:
            ws = wb[cf_sheet]
            # Look for a total/portfolio row: scan rows 14-20 for a value in [8M, 12M]
            total_cells = _scan_numeric(ws, range(14, 22), range(2, 16), 8000000, 12000000)
            if total_cells:
                portfolio_y1 = total_cells[0][2]

        if portfolio_y1 is not None and 8000000 <= portfolio_y1 <= 12000000:
            score += 20
            fb.append(f"C4 PASS: Portfolio total Year_1_NOI = ${portfolio_y1:,.0f}")
        elif portfolio_y1 is not None:
            fb.append(f"C4 FAIL: Portfolio total Year_1_NOI = ${portfolio_y1:,.0f} (expected $8M-$12M)")
        else:
            fb.append("C4 FAIL: Portfolio total Year_1_NOI not found in expected range")

        # ── C5 (20 pts): >= 6 NPV values present (range 1M-50M) ──────────────
        npv_vals = []
        if cf_sheet:
            ws = wb[cf_sheet]
            # NPV column is typically towards the end of the projection columns
            # Scan all columns for values in the NPV range
            npv_cells = _scan_numeric(ws, range(2, 16), range(2, 20), 1000000, 50000000)
            if npv_cells:
                # Deduplicate by row
                seen_rows = set()
                for r, c, v in npv_cells:
                    if r not in seen_rows:
                        npv_vals.append(v)
                        seen_rows.add(r)

        if len(npv_vals) >= 6:
            score += 20
            fb.append(f"C5 PASS: {len(npv_vals)} NPV values found")
        elif len(npv_vals) >= 3:
            score += 10
            fb.append(f"C5 PARTIAL: {len(npv_vals)} NPV values found (need >= 6)")
        else:
            fb.append(f"C5 FAIL: Only {len(npv_vals)} NPV values found (need >= 6)")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

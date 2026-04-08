#!/usr/bin/env python3
"""
Verifier for clinical_trial_adverse_events task.

Verification pipeline:
  1. Read result JSON (C:\\Users\\Docker\\clinical_trial_adverse_events_result.json)
     - Check is_new: if xlsx was not saved after task start -> score 0
  2. Independently copy xlsx from VM and parse with openpyxl data_only=True
  3. Compute ground truth from AE_Frequency_Comparison pre-filled counts
  4. Score 5 criteria (100 pts total, pass >= 60)

Criteria:
  C1 (20 pts): AE_Frequency_Comparison has Rate_Ratio for >= 20 of 30 terms (values in range 0.1 - 10.0)
  C2 (25 pts): At least 10 Rate_Ratio values are correct within +/-5% of ground truth (Keytruda_Count / Opdivo_Count)
  C3 (15 pts): At least 1 SIGNAL flag present in AE_Frequency_Comparison
  C4 (20 pts): Safety_Signal_Report has Total Cases (Keytruda) = 100 (range 95-105) and Serious Case Rate around 85% (range 75-95)
  C5 (20 pts): TOTAL row in AE_Frequency_Comparison has Keytruda total in [80000, 100000]
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\clinical_trial_adverse_events_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/clinical_trial_ae.xlsx"


def _find_sheet(wb, keywords):
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    """Scan cells in the given ranges and return (row, col, value) for numeric cells in [lo, hi]."""
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    """Count cells containing the keyword string (case-insensitive)."""
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_clinical_trial_adverse_events(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_clinical_trial_")
    try:
        # -- STEP 1: Read result JSON and check is_new --------------------------
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False). "
                            "Agent must save the file with Ctrl+S after completing formulas."
            }

        # -- STEP 2: Independently copy and parse xlsx --------------------------
        xlsx_local = os.path.join(tmp, "clinical_trial_ae.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0,
                    "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0,
                    "feedback": "clinical_trial_ae.xlsx not found or empty"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        # -- STEP 3: Score criteria ---------------------------------------------
        score = 0
        fb = []

        freq_sheet = _find_sheet(wb, ["AE_Frequency", "Frequency_Comparison", "frequency"])
        signal_sheet = _find_sheet(wb, ["Safety_Signal", "Signal_Report", "signal"])

        # -----------------------------------------------------------------------
        # Criterion 1 (20 pts): Rate_Ratio populated for >= 20 of 30 terms
        #   Rate_Ratio should be in column D (col 4), rows 2-31, values in [0.1, 10.0]
        # -----------------------------------------------------------------------
        rate_ratio_count = 0
        rate_ratio_values = {}  # row -> value
        if freq_sheet:
            ws = wb[freq_sheet]
            # Find the Rate_Ratio column by scanning headers
            rr_col = None
            for c in range(1, ws.max_column + 1):
                hdr = ws.cell(1, c).value
                if hdr and isinstance(hdr, str) and "rate_ratio" in hdr.lower().replace(" ", "_"):
                    rr_col = c
                    break
            if rr_col is None:
                rr_col = 4  # default position

            for r in range(2, 32):  # rows 2-31 = 30 MedDRA terms
                val = ws.cell(r, rr_col).value
                if val is not None and isinstance(val, (int, float)) and 0.1 <= val <= 10.0:
                    rate_ratio_count += 1
                    rate_ratio_values[r] = val

        if rate_ratio_count >= 20:
            score += 20
            fb.append(f"C1 PASS: Rate_Ratio populated for {rate_ratio_count}/30 terms")
        elif rate_ratio_count >= 10:
            score += 10
            fb.append(f"C1 PARTIAL: Rate_Ratio populated for {rate_ratio_count}/30 terms (need >= 20)")
        else:
            fb.append(f"C1 FAIL: Rate_Ratio populated for {rate_ratio_count}/30 terms (need >= 20)")

        # -----------------------------------------------------------------------
        # Criterion 2 (25 pts): At least 10 Rate_Ratio values correct within +/-5%
        #   Ground truth = Keytruda_Report_Count / Opdivo_Report_Count (from cols 2, 3)
        #   Only check rows where both counts are nonzero (24 of 30 terms)
        # -----------------------------------------------------------------------
        accurate_ratios = 0
        if freq_sheet and rate_ratio_values:
            ws = wb[freq_sheet]
            # Find Keytruda and Opdivo count columns
            k_col = None
            o_col = None
            for c in range(1, ws.max_column + 1):
                hdr = ws.cell(1, c).value
                if hdr and isinstance(hdr, str):
                    hdr_lower = hdr.lower()
                    if "keytruda" in hdr_lower and "count" in hdr_lower:
                        k_col = c
                    elif "opdivo" in hdr_lower and "count" in hdr_lower:
                        o_col = c
            if k_col is None:
                k_col = 2
            if o_col is None:
                o_col = 3

            for r, agent_rr in rate_ratio_values.items():
                k_val = ws.cell(r, k_col).value
                o_val = ws.cell(r, o_col).value
                if (k_val is not None and o_val is not None
                        and isinstance(k_val, (int, float)) and isinstance(o_val, (int, float))
                        and o_val > 0):
                    expected_rr = k_val / o_val
                    if expected_rr > 0 and abs(agent_rr - expected_rr) / expected_rr <= 0.05:
                        accurate_ratios += 1

        if accurate_ratios >= 10:
            score += 25
            fb.append(f"C2 PASS: {accurate_ratios} Rate_Ratio values within 5% of ground truth")
        elif accurate_ratios >= 5:
            score += 12
            fb.append(f"C2 PARTIAL: {accurate_ratios} accurate Rate_Ratio values (need >= 10)")
        else:
            fb.append(f"C2 FAIL: {accurate_ratios} Rate_Ratio values within 5% of ground truth (need >= 10)")

        # -----------------------------------------------------------------------
        # Criterion 3 (15 pts): At least 1 SIGNAL flag present in AE_Frequency_Comparison
        # -----------------------------------------------------------------------
        signal_flag_count = 0
        if freq_sheet:
            ws = wb[freq_sheet]
            signal_flag_count = _count_string(ws, range(2, 32), range(1, ws.max_column + 1), "SIGNAL")

        if signal_flag_count >= 1:
            score += 15
            fb.append(f"C3 PASS: {signal_flag_count} SIGNAL flag(s) found in AE_Frequency_Comparison")
        else:
            fb.append("C3 FAIL: No SIGNAL flags found in AE_Frequency_Comparison "
                       "(expected for terms with Rate_Ratio > 1.5 AND Rate_Difference > 500)")

        # -----------------------------------------------------------------------
        # Criterion 4 (20 pts): Safety_Signal_Report has
        #   - Total Cases (Keytruda) in range [95, 105] (expected 100)
        #   - Serious Case Rate (Keytruda %) in range [75, 95] (expected 85%)
        #   Award 10 pts each sub-criterion
        # -----------------------------------------------------------------------
        c4_score = 0
        c4_fb_parts = []
        if signal_sheet:
            ws = wb[signal_sheet]
            total_cases_keytruda = None
            serious_rate = None

            for r in range(2, ws.max_row + 1):
                label = ws.cell(r, 1).value
                val = ws.cell(r, 2).value
                if label is None:
                    continue
                label_str = str(label).strip().lower()

                if "total cases" in label_str and "keytruda" in label_str:
                    if val is not None and isinstance(val, (int, float)):
                        total_cases_keytruda = val
                elif "serious" in label_str and "rate" in label_str and "keytruda" in label_str:
                    if val is not None and isinstance(val, (int, float)):
                        serious_rate = val

            if total_cases_keytruda is not None and 95 <= total_cases_keytruda <= 105:
                c4_score += 10
                c4_fb_parts.append(f"Total Cases (Keytruda)={total_cases_keytruda:.0f} OK")
            elif total_cases_keytruda is not None:
                c4_fb_parts.append(f"Total Cases (Keytruda)={total_cases_keytruda:.0f} out of range [95-105]")
            else:
                c4_fb_parts.append("Total Cases (Keytruda) not found")

            if serious_rate is not None and 75 <= serious_rate <= 95:
                c4_score += 10
                c4_fb_parts.append(f"Serious Case Rate={serious_rate:.1f}% OK")
            elif serious_rate is not None:
                c4_fb_parts.append(f"Serious Case Rate={serious_rate:.1f}% out of range [75-95]")
            else:
                c4_fb_parts.append("Serious Case Rate not found")
        else:
            c4_fb_parts.append("Safety_Signal_Report sheet not found")

        score += c4_score
        if c4_score == 20:
            fb.append(f"C4 PASS: {'; '.join(c4_fb_parts)}")
        elif c4_score > 0:
            fb.append(f"C4 PARTIAL ({c4_score}/20): {'; '.join(c4_fb_parts)}")
        else:
            fb.append(f"C4 FAIL: {'; '.join(c4_fb_parts)}")

        # -----------------------------------------------------------------------
        # Criterion 5 (20 pts): TOTAL row in AE_Frequency_Comparison has
        #   Keytruda total in [80000, 100000] (expected 90861)
        # -----------------------------------------------------------------------
        keytruda_total = None
        if freq_sheet:
            ws = wb[freq_sheet]
            # Find the TOTAL row (should be the last data row, row 32)
            for r in range(2, ws.max_row + 1):
                label = ws.cell(r, 1).value
                if label and isinstance(label, str) and "total" in label.lower():
                    # Find Keytruda count column
                    k_col_for_total = None
                    for c in range(1, ws.max_column + 1):
                        hdr = ws.cell(1, c).value
                        if hdr and isinstance(hdr, str) and "keytruda" in hdr.lower() and "count" in hdr.lower():
                            k_col_for_total = c
                            break
                    if k_col_for_total is None:
                        k_col_for_total = 2  # default

                    val = ws.cell(r, k_col_for_total).value
                    if val is not None and isinstance(val, (int, float)):
                        keytruda_total = val
                    break

        if keytruda_total is not None and 80000 <= keytruda_total <= 100000:
            score += 20
            fb.append(f"C5 PASS: TOTAL Keytruda count = {keytruda_total:.0f} (expected 80000-100000)")
        elif keytruda_total is not None:
            fb.append(f"C5 FAIL: TOTAL Keytruda count = {keytruda_total:.0f} (expected 80000-100000)")
        else:
            fb.append("C5 FAIL: TOTAL Keytruda count not found in AE_Frequency_Comparison TOTAL row")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

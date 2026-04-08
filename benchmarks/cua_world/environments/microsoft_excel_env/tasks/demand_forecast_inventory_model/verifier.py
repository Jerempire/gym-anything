#!/usr/bin/env python3
"""
Verifier for demand_forecast_inventory_model task.

Pipeline:
  1. Read result JSON, check is_new gate
  2. Independently copy xlsx and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts, pass >= 60)

Criteria:
  C1 (20 pts): Forecast_Sheet has Avg_Monthly_Demand for >= 16 of 20 SKUs (positive values 20-200)
  C2 (20 pts): Inventory_Parameters has EOQ for >= 16 of 20 SKUs (positive values 5-5000)
  C3 (20 pts): At least 12 EOQ values within +/-20% of ground truth
               Ground truth: sqrt(2 * avg_monthly_demand * 12 * order_cost / (unit_cost * holding_pct/100))
  C4 (20 pts): Total annual inventory cost in TOTAL row in range [60000, 120000]
  C5 (20 pts): ABC_Analysis has at least 4 "A" class, 4 "B" class, and 3 "C" class assignments
"""

import json
import logging
import math
import os
import shutil
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\demand_forecast_inventory_model_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/demand_inventory.xlsx"


def _find_sheet(wb, keywords):
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    """Return list of (row, col, value) for numeric cells within [lo, hi]."""
    results = []
    for r in row_range:
        for c in col_range:
            v = ws.cell(r, c).value
            if v is not None and isinstance(v, (int, float)) and lo <= v <= hi:
                results.append((r, c, v))
    return results


def _find_header_col(ws, row, keywords):
    """Find a column index in the given row whose header contains any keyword."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row, c).value
        if v and isinstance(v, str):
            v_lower = v.lower().replace("_", " ").replace("-", " ")
            if any(k.lower() in v_lower for k in keywords):
                return c
    return None


def _get_col_values(ws, col, start_row, end_row, value_filter=None):
    """Get non-None values from a column range, optionally filtering by type."""
    values = []
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, col).value
        if v is not None:
            if value_filter is None or isinstance(v, value_filter):
                values.append((r, v))
    return values


def _count_string_in_col(ws, col, start_row, end_row, keyword):
    """Count cells in a column whose string value contains the keyword."""
    count = 0
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, col).value
        if v and isinstance(v, str) and keyword.lower() in v.lower():
            count += 1
    return count


def verify_demand_forecast_inventory_model(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_demand_forecast_")
    try:
        # ── STEP 1: Check is_new ──────────────────────────────────────────
        json_local = os.path.join(tmp, "result.json")
        result = {}
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False,
                "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False).",
            }

        # ── STEP 2: Copy and parse xlsx ───────────────────────────────────
        xlsx_local = os.path.join(tmp, "demand_inventory.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {
                "passed": False,
                "score": 0,
                "feedback": f"Could not copy xlsx: {e}",
            }

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {
                "passed": False,
                "score": 0,
                "feedback": "demand_inventory.xlsx not found or empty.",
            }

        from openpyxl import load_workbook

        wb = load_workbook(xlsx_local, data_only=True)

        score = 0
        fb = []

        # Locate sheets
        hist_sheet = _find_sheet(wb, ["Historical_Sales", "Historical", "Sales"])
        forecast_sheet = _find_sheet(wb, ["Forecast_Sheet", "Forecast"])
        inv_sheet = _find_sheet(wb, ["Inventory_Parameters", "Inventory"])
        abc_sheet = _find_sheet(wb, ["ABC_Analysis", "ABC"])

        # ── Read Historical_Sales data for ground-truth calculations ──────
        hist_data = {}  # SKU_ID -> {unit_cost, holding_pct, lead_time, order_cost, demands}
        if hist_sheet:
            ws_hist = wb[hist_sheet]
            # Find key columns by header
            sku_col = _find_header_col(ws_hist, 1, ["SKU_ID", "SKU ID", "SKU"])
            unit_cost_col = _find_header_col(ws_hist, 1, ["Unit_Cost", "Unit Cost"])
            holding_col = _find_header_col(
                ws_hist, 1, ["Holding_Cost", "Holding Cost", "Holding"]
            )
            lead_col = _find_header_col(
                ws_hist, 1, ["Lead_Time", "Lead Time", "LeadTime"]
            )
            order_cost_col = _find_header_col(
                ws_hist, 1, ["Order_Cost", "Order Cost", "OrderCost"]
            )

            # Find demand columns (monthly columns after the metadata columns)
            demand_cols = []
            for c in range(1, ws_hist.max_column + 1):
                hdr = ws_hist.cell(1, c).value
                if hdr and isinstance(hdr, str):
                    hdr_lower = hdr.lower()
                    if any(
                        m in hdr_lower
                        for m in [
                            "jan",
                            "feb",
                            "mar",
                            "apr",
                            "may",
                            "jun",
                            "jul",
                            "aug",
                            "sep",
                            "oct",
                            "nov",
                            "dec",
                        ]
                    ):
                        demand_cols.append(c)

            for r in range(2, 23):  # rows 2-22 (20 SKUs + some buffer)
                sku = ws_hist.cell(r, sku_col).value if sku_col else None
                if sku is None:
                    continue
                sku_str = str(sku).strip()
                if not sku_str:
                    continue

                uc = (
                    ws_hist.cell(r, unit_cost_col).value
                    if unit_cost_col
                    else None
                )
                hp = (
                    ws_hist.cell(r, holding_col).value if holding_col else None
                )
                lt = ws_hist.cell(r, lead_col).value if lead_col else None
                oc = (
                    ws_hist.cell(r, order_cost_col).value
                    if order_cost_col
                    else None
                )

                demands = []
                for dc in demand_cols:
                    dv = ws_hist.cell(r, dc).value
                    if dv is not None and isinstance(dv, (int, float)):
                        demands.append(float(dv))

                if uc and hp and lt and oc and demands:
                    hist_data[sku_str] = {
                        "unit_cost": float(uc),
                        "holding_pct": float(hp),
                        "lead_time": float(lt),
                        "order_cost": float(oc),
                        "demands": demands,
                    }

        logger.info(
            f"Loaded historical data for {len(hist_data)} SKUs from Historical_Sales"
        )

        # ── C1 (20 pts): Forecast_Sheet Avg_Monthly_Demand ───────────────
        avg_demand_count = 0
        if forecast_sheet:
            ws_fc = wb[forecast_sheet]
            avg_col = _find_header_col(
                ws_fc, 1, ["Avg_Monthly", "Avg Monthly", "Average", "Mean"]
            )
            if avg_col:
                vals = _get_col_values(ws_fc, avg_col, 2, 22, (int, float))
                avg_demand_count = sum(1 for _, v in vals if 20 <= v <= 200)
            else:
                # Fallback: scan columns 2-10 for values in range
                cells = _scan_numeric(ws_fc, range(2, 23), range(2, 11), 20, 200)
                # Group by column; pick the column with the most hits
                from collections import Counter

                col_counts = Counter(c for _, c, _ in cells)
                if col_counts:
                    best_col = col_counts.most_common(1)[0][0]
                    avg_demand_count = col_counts[best_col]

        if avg_demand_count >= 16:
            score += 20
            fb.append(
                f"C1 PASS: Avg_Monthly_Demand found for {avg_demand_count} SKUs"
            )
        elif avg_demand_count >= 10:
            score += 10
            fb.append(
                f"C1 PARTIAL: Avg_Monthly_Demand found for {avg_demand_count} SKUs (need 16+)"
            )
        else:
            fb.append(
                f"C1 FAIL: Avg_Monthly_Demand found for {avg_demand_count} SKUs (need 16+)"
            )

        # ── C2 (20 pts): Inventory_Parameters EOQ ────────────────────────
        eoq_values = []  # list of (row, eoq_value)
        eoq_col = None
        if inv_sheet:
            ws_inv = wb[inv_sheet]
            eoq_col = _find_header_col(ws_inv, 1, ["EOQ"])
            if eoq_col:
                vals = _get_col_values(ws_inv, eoq_col, 2, 22, (int, float))
                eoq_values = [(r, v) for r, v in vals if 5 <= v <= 5000]
            else:
                # Fallback: scan for a column of values in EOQ range
                cells = _scan_numeric(ws_inv, range(2, 23), range(2, 15), 5, 5000)
                from collections import Counter

                col_counts = Counter(c for _, c, _ in cells)
                if col_counts:
                    best_col = col_counts.most_common(1)[0][0]
                    eoq_values = [
                        (r, v) for r, c, v in cells if c == best_col
                    ]
                    eoq_col = best_col

        eoq_count = len(eoq_values)
        if eoq_count >= 16:
            score += 20
            fb.append(f"C2 PASS: EOQ values found for {eoq_count} SKUs")
        elif eoq_count >= 10:
            score += 10
            fb.append(
                f"C2 PARTIAL: EOQ values found for {eoq_count} SKUs (need 16+)"
            )
        else:
            fb.append(
                f"C2 FAIL: EOQ values found for {eoq_count} SKUs (need 16+)"
            )

        # ── C3 (20 pts): EOQ accuracy vs ground truth ────────────────────
        accurate_eoq = 0
        if inv_sheet and hist_data and eoq_values:
            ws_inv = wb[inv_sheet]
            sku_col_inv = _find_header_col(
                ws_inv, 1, ["SKU_ID", "SKU ID", "SKU"]
            )

            for row, eoq_val in eoq_values:
                sku_id = (
                    ws_inv.cell(row, sku_col_inv).value if sku_col_inv else None
                )
                if sku_id is None:
                    continue
                sku_str = str(sku_id).strip()

                hd = hist_data.get(sku_str)
                if not hd:
                    continue

                avg_monthly = sum(hd["demands"]) / len(hd["demands"])
                annual_demand = avg_monthly * 12
                unit_cost = hd["unit_cost"]
                holding_pct = hd["holding_pct"]
                order_cost = hd["order_cost"]
                holding_cost_per_unit = unit_cost * holding_pct / 100.0

                if holding_cost_per_unit <= 0 or annual_demand <= 0:
                    continue

                # EOQ = sqrt(2 * D * S / H)
                expected_eoq = math.sqrt(
                    2 * annual_demand * order_cost / holding_cost_per_unit
                )

                if expected_eoq > 0:
                    pct_diff = abs(eoq_val - expected_eoq) / expected_eoq
                    if pct_diff <= 0.20:
                        accurate_eoq += 1

        if accurate_eoq >= 12:
            score += 20
            fb.append(
                f"C3 PASS: {accurate_eoq} EOQ values within 20% of ground truth"
            )
        elif accurate_eoq >= 8:
            score += 10
            fb.append(
                f"C3 PARTIAL: {accurate_eoq} EOQ values within 20% of ground truth (need 12+)"
            )
        else:
            fb.append(
                f"C3 FAIL: {accurate_eoq} EOQ values within 20% of ground truth (need 12+)"
            )

        # ── C4 (20 pts): Total annual inventory cost in TOTAL row ─────────
        total_cost = None
        if inv_sheet:
            ws_inv = wb[inv_sheet]
            # Find the TOTAL row by scanning column A (or column 1) for "TOTAL"
            total_row = None
            for r in range(2, ws_inv.max_row + 1):
                v = ws_inv.cell(r, 1).value
                if v and isinstance(v, str) and "total" in v.lower():
                    total_row = r
                    break

            if total_row:
                # Find Total_Inventory_Cost column
                cost_col = _find_header_col(
                    ws_inv,
                    1,
                    [
                        "Total_Inventory_Cost",
                        "Total Inventory Cost",
                        "Total_Cost",
                        "Total Cost",
                        "Inventory_Cost",
                    ],
                )
                if cost_col:
                    total_cost = ws_inv.cell(total_row, cost_col).value
                else:
                    # Fallback: scan the TOTAL row for a value in range
                    for c in range(2, ws_inv.max_column + 1):
                        v = ws_inv.cell(total_row, c).value
                        if (
                            v is not None
                            and isinstance(v, (int, float))
                            and 60000 <= v <= 120000
                        ):
                            total_cost = v
                            break

            # Secondary fallback: scan last few rows for any value in range
            if total_cost is None:
                max_r = min(ws_inv.max_row + 1, 30)
                cands = _scan_numeric(
                    ws_inv, range(max(2, max_r - 5), max_r), range(2, 15), 60000, 120000
                )
                if cands:
                    total_cost = cands[0][2]

        if total_cost is not None and 60000 <= total_cost <= 120000:
            score += 20
            fb.append(
                f"C4 PASS: Total annual inventory cost = ${total_cost:,.0f} (in [60000, 120000])"
            )
        elif total_cost is not None:
            fb.append(
                f"C4 FAIL: Total annual inventory cost = ${total_cost:,.0f} (expected [60000, 120000])"
            )
        else:
            fb.append("C4 FAIL: Total annual inventory cost not found in TOTAL row")

        # ── C5 (20 pts): ABC_Analysis class assignments ───────────────────
        a_count = 0
        b_count = 0
        c_count = 0
        if abc_sheet:
            ws_abc = wb[abc_sheet]
            class_col = _find_header_col(
                ws_abc, 1, ["ABC_Class", "ABC Class", "Class", "ABC"]
            )
            if class_col:
                for r in range(2, 23):
                    v = ws_abc.cell(r, class_col).value
                    if v and isinstance(v, str):
                        v_upper = v.strip().upper()
                        if v_upper == "A":
                            a_count += 1
                        elif v_upper == "B":
                            b_count += 1
                        elif v_upper == "C":
                            c_count += 1
            else:
                # Fallback: scan all columns for A/B/C patterns
                for c in range(2, ws_abc.max_column + 1):
                    a_tmp = _count_string_in_col(ws_abc, c, 2, 22, "A")
                    b_tmp = _count_string_in_col(ws_abc, c, 2, 22, "B")
                    c_tmp = _count_string_in_col(ws_abc, c, 2, 22, "C")
                    # Heuristic: the ABC column has exactly A+B+C ~= 20
                    total_abc = a_tmp + b_tmp + c_tmp
                    if 15 <= total_abc <= 22:
                        a_count = a_tmp
                        b_count = b_tmp
                        c_count = c_tmp
                        break

        abc_pass = a_count >= 4 and b_count >= 4 and c_count >= 3
        if abc_pass:
            score += 20
            fb.append(
                f"C5 PASS: ABC classes A={a_count}, B={b_count}, C={c_count} (reasonable Pareto distribution)"
            )
        elif a_count >= 2 and b_count >= 2 and c_count >= 1:
            score += 10
            fb.append(
                f"C5 PARTIAL: ABC classes A={a_count}, B={b_count}, C={c_count} (need A>=4, B>=4, C>=3)"
            )
        else:
            fb.append(
                f"C5 FAIL: ABC classes A={a_count}, B={b_count}, C={c_count} (need A>=4, B>=4, C>=3)"
            )

        # ── Final result ──────────────────────────────────────────────────
        passed = score >= 60
        feedback = " | ".join(fb)
        logger.info(f"Verification complete. Score: {score}/100, Passed: {passed}")

        return {"passed": passed, "score": score, "feedback": feedback}

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

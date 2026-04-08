#!/usr/bin/env python3
"""
Verifier for school_district_title1_compliance task.

Verification pipeline:
  1. Read result JSON (C:\\Users\\Docker\\school_district_title1_compliance_result.json)
     - Check is_new: if xlsx was not saved after task start -> score 0
  2. Independently copy xlsx from VM and parse with openpyxl data_only=True
  3. Score 5 criteria (100 pts total, pass >= 60)

Criteria:
  C1 (20 pts): Expenditure_Analysis has Per_Pupil_Expenditure for >= 9 of 11 schools
               (values in range 4000-15000)
  C2 (25 pts): District Total expenditure in range [30,000,000 - 45,000,000]
  C3 (15 pts): Comparability_Report has Overall Comparability Status
               ("COMPARABLE" or "NON-COMPARABLE")
  C4 (20 pts): Title_I_Allocation TOTAL row sums to approximately $582,174
               (range [550,000 - 620,000])
  C5 (20 pts): At least 3 Supplement_Check values present (PASS/FAIL strings)
"""
import json, logging, os, tempfile, shutil
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

RESULT_PATH = "C:\\Users\\Docker\\school_district_title1_compliance_result.json"
XLSX_PATH = "C:/Users/Docker/Desktop/ExcelTasks/school_district.xlsx"


def _find_sheet(wb, keywords):
    """Find a sheet whose name contains any of the given keywords (case-insensitive)."""
    for name in wb.sheetnames:
        if any(k.lower() in name.lower() for k in keywords):
            return name
    return None


def _scan_numeric(ws, row_range, col_range, lo, hi):
    """Return list of (row, col, value) for numeric cells within [lo, hi]."""
    return [(r, c, ws.cell(r, c).value)
            for r in row_range for c in col_range
            if ws.cell(r, c).value is not None
            and isinstance(ws.cell(r, c).value, (int, float))
            and lo <= ws.cell(r, c).value <= hi]


def _count_string(ws, row_range, col_range, keyword):
    """Count cells containing keyword (case-insensitive)."""
    return sum(1 for r in row_range for c in col_range
               if isinstance(ws.cell(r, c).value, str)
               and keyword.lower() in ws.cell(r, c).value.lower())


def verify_school_district_title1_compliance(traj, env_info, task_info):
    copy_from_env = env_info.get("copy_from_env")
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "copy_from_env not available"}

    tmp = tempfile.mkdtemp(prefix="verify_school_district_")
    try:
        # -- STEP 1: Read result JSON and check is_new --------------------------
        json_local = os.path.join(tmp, "result.json")
        try:
            copy_from_env(RESULT_PATH, json_local)
            with open(json_local, "r", encoding="utf-8-sig") as f:
                result = json.load(f)
        except Exception as e:
            result = {}
            logger.warning(f"Could not read result JSON: {e}")

        xlsx_info = result.get("xlsx_file", {})
        if not xlsx_info.get("is_new", False):
            return {
                "passed": False, "score": 0,
                "feedback": "FAIL: Workbook was not saved after task started (is_new=False). "
                            "Agent must save the file with Ctrl+S after completing formulas."
            }

        # -- STEP 2: Independently copy and parse xlsx --------------------------
        xlsx_local = os.path.join(tmp, "school_district.xlsx")
        try:
            copy_from_env(XLSX_PATH, xlsx_local)
        except Exception as e:
            return {"passed": False, "score": 0,
                    "feedback": f"Could not copy xlsx: {e}"}

        if not os.path.exists(xlsx_local) or os.path.getsize(xlsx_local) == 0:
            return {"passed": False, "score": 0,
                    "feedback": "school_district.xlsx not found or empty"}

        from openpyxl import load_workbook
        wb = load_workbook(xlsx_local, data_only=True)

        # -- STEP 3: Score criteria ---------------------------------------------
        score = 0
        fb = []

        # ── C1 (20 pts): Per_Pupil_Expenditure for >= 9 of 11 schools ────────
        # PPE values expected in range 4,000 - 15,000
        exp_sheet = _find_sheet(wb, ["expenditure", "Expenditure_Analysis"])
        ppe_vals = []
        if exp_sheet:
            ws = wb[exp_sheet]
            # Scan rows 2-14 (11 schools + possible header offset), cols 2-15
            # PPE is Per-Pupil Expenditure = Total / Enrollment, typically 4000-15000
            cells = _scan_numeric(ws, range(2, 14), range(2, 16), 4000, 15000)
            if cells:
                # Deduplicate by row - take the first PPE-range value per row
                seen_rows = set()
                for r, c, v in cells:
                    if r not in seen_rows:
                        ppe_vals.append(v)
                        seen_rows.add(r)

        if len(ppe_vals) >= 9:
            score += 20
            fb.append(f"C1 PASS: Per_Pupil_Expenditure found for {len(ppe_vals)} schools")
        elif len(ppe_vals) >= 5:
            score += 10
            fb.append(f"C1 PARTIAL: Per_Pupil_Expenditure found for {len(ppe_vals)} schools (need >= 9)")
        else:
            fb.append(f"C1 FAIL: Only {len(ppe_vals)} Per_Pupil_Expenditure values found (need >= 9)")

        # ── C2 (25 pts): District Total expenditure in [30M, 45M] ─────────────
        district_total = None
        if exp_sheet:
            ws = wb[exp_sheet]
            # Look for DISTRICT TOTAL row: scan rows 13-20 for a value in range [30M, 45M]
            total_cells = _scan_numeric(ws, range(13, 25), range(2, 16), 30000000, 45000000)
            if total_cells:
                district_total = total_cells[0][2]

            # Broader fallback: look for "total" or "district" label in column A
            if district_total is None:
                for row in range(2, 25):
                    label = ws.cell(row, 1).value
                    if label and isinstance(label, str) and (
                        "total" in label.lower() or "district" in label.lower()
                    ):
                        cands = _scan_numeric(ws, [row], range(2, 16), 20000000, 60000000)
                        if cands:
                            district_total = cands[0][2]
                            break

        if district_total is not None and 30000000 <= district_total <= 45000000:
            score += 25
            fb.append(f"C2 PASS: District Total Expenditure = ${district_total:,.0f}")
        elif district_total is not None:
            # Give partial credit if in a wider range
            if 25000000 <= district_total <= 50000000:
                score += 12
                fb.append(f"C2 PARTIAL: District Total Expenditure = ${district_total:,.0f} "
                          f"(expected $30M-$45M)")
            else:
                fb.append(f"C2 FAIL: District Total Expenditure = ${district_total:,.0f} "
                          f"(expected $30M-$45M)")
        else:
            fb.append("C2 FAIL: District Total Expenditure not found in expected range")

        # ── C3 (15 pts): Comparability Status ("COMPARABLE" or "NON-COMPARABLE") ─
        comp_sheet = _find_sheet(wb, ["comparability", "Comparability_Report"])
        comp_status_found = False
        if comp_sheet:
            ws = wb[comp_sheet]
            # The overall comparability status should be a cell value that is
            # exactly "COMPARABLE" or "NON-COMPARABLE" (or contains it as primary content).
            # Only check column B+ (col >= 2) in the results section (rows after the data rows)
            # to avoid matching pre-filled label text like "Comparable? (ratio >= 0.90)".
            n_schools = 0
            for r in range(2, 20):
                if ws.cell(r, 1).value is not None:
                    n_schools += 1
            results_start = n_schools + 2  # results section starts after school data
            for r in range(results_start, results_start + 15):
                for c in range(2, 8):
                    val = ws.cell(r, c).value
                    if val and isinstance(val, str):
                        v = val.strip().upper()
                        if v in ("COMPARABLE", "NON-COMPARABLE", "NON_COMPARABLE"):
                            comp_status_found = True
                            break
                if comp_status_found:
                    break

        if comp_status_found:
            score += 15
            fb.append("C3 PASS: Overall Comparability Status found in Comparability_Report")
        else:
            fb.append("C3 FAIL: No COMPARABLE or NON-COMPARABLE status found in Comparability_Report")

        # ── C4 (20 pts): Title I Allocation TOTAL ~ $582,174 ─────────────────
        # The sheet has a pre-filled reference value ($582,174) in an instruction row.
        # We need to check the TOTAL row that the agent computes (sum of individual
        # school allocations), which is in the row labeled exactly "TOTAL" (col A).
        # The Title_I_Allocation column is column 9.
        alloc_sheet = _find_sheet(wb, ["title_i", "Title_I_Allocation", "allocation"])
        alloc_total = None
        if alloc_sheet:
            ws = wb[alloc_sheet]
            # Look for the TOTAL row specifically (exact label match)
            for row in range(2, 20):
                label = ws.cell(row, 1).value
                if label and isinstance(label, str) and label.strip().upper() == "TOTAL":
                    # Check the Title_I_Allocation column (col 9) for the sum
                    cands = _scan_numeric(ws, [row], range(6, 16), 500000, 650000)
                    if cands:
                        alloc_total = cands[0][2]
                    break

            # Fallback: look for any row with a value in the allocation column
            # that matches, but exclude the reference value row
            if alloc_total is None:
                for row in range(2, 20):
                    label = ws.cell(row, 1).value
                    if label and isinstance(label, str) and label.strip().upper() == "TOTAL":
                        # Scan all columns in the TOTAL row
                        cands = _scan_numeric(ws, [row], range(2, 16), 500000, 650000)
                        if cands:
                            alloc_total = cands[0][2]
                        break

        if alloc_total is not None and 550000 <= alloc_total <= 620000:
            score += 20
            fb.append(f"C4 PASS: Title I Allocation Total = ${alloc_total:,.0f} "
                      f"(expected $550,000-$620,000)")
        elif alloc_total is not None:
            if 500000 <= alloc_total <= 650000:
                score += 10
                fb.append(f"C4 PARTIAL: Title I Allocation Total = ${alloc_total:,.0f} "
                          f"(expected $550,000-$620,000)")
            else:
                fb.append(f"C4 FAIL: Title I Allocation Total = ${alloc_total:,.0f} "
                          f"(expected $550,000-$620,000)")
        else:
            fb.append("C4 FAIL: Title I Allocation Total not found in expected range")

        # ── C5 (20 pts): >= 3 Supplement_Check values (PASS/FAIL strings) ─────
        supplement_count = 0
        if alloc_sheet:
            ws = wb[alloc_sheet]
            # Count cells whose value is exactly "PASS" or "FAIL" in the data rows
            # (rows 2 to ~10) in the last few columns (Supplement_Check column).
            # Exclude rows that contain instruction text by checking for short exact values.
            for r in range(2, 15):
                for c in range(6, 16):
                    val = ws.cell(r, c).value
                    if val and isinstance(val, str):
                        v = val.strip().upper()
                        # Only match exact "PASS" or "FAIL", not longer strings
                        # containing those words (like "PASS if Title I Per Pupil > 0...")
                        if v in ("PASS", "FAIL"):
                            supplement_count += 1

        if supplement_count >= 3:
            score += 20
            fb.append(f"C5 PASS: {supplement_count} Supplement_Check values found (PASS/FAIL)")
        elif supplement_count >= 2:
            score += 10
            fb.append(f"C5 PARTIAL: {supplement_count} Supplement_Check values found (need >= 3)")
        else:
            fb.append(f"C5 FAIL: Only {supplement_count} Supplement_Check values found (need >= 3)")

        return {
            "passed": score >= 60,
            "score": score,
            "feedback": " | ".join(fb)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {"passed": False, "score": 0, "feedback": f"Verification error: {e}"}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

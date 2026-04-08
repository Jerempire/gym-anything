#!/usr/bin/env python3
"""
Verifier for Construction Project Cost Estimate task.

The agent must create a professional cost estimate for a 4,200 SF medical office
build-out using specifications, material prices, and labor rates from source files.

Data sources:
- Material costs: RSMeans 2024 Square Foot Costs / Unit Costs (Gordian Group)
- Labor rates: DOL Davis-Bacon Prevailing Wage Determinations (sam.gov)

RSMeans 2024 reference: Medical Office, 1-Story, 7,000 SF model
  - Union labor: $236.56/SF total ($173.62/SF subtotal + 25% contractor + 9% arch)
  - Open shop: $215.55/SF total

This is a tenant improvement (TI) build-out, not ground-up construction.
TI costs are typically 40-80% of full construction costs.
  - $236.56/SF x 4,200 SF = ~$993K (full ground-up reference)
  - TI range: ~$397K to ~$795K (40-80% of ground-up)

Expected cost range for verification:
- Grand total (with markups): $350,000 - $800,000
"""

import sys
import os
import logging
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# CSI divisions that should appear in a proper estimate
EXPECTED_DIVISIONS = [
    "concrete", "metal", "wood", "millwork", "door", "finish",
    "mechanical", "hvac", "plumbing", "electrical", "fire",
    "specialt", "ceiling", "flooring", "paint",
]

# Key line items that must be in any reasonable estimate
KEY_ITEMS = [
    "gypsum", "drywall", "partition", "framing", "stud",
    "door", "cabinet", "millwork",
    "tile", "vinyl", "carpet", "epoxy", "flooring",
    "hvac", "ductwork", "air handler", "ahu",
    "plumb", "sink", "fixture", "medical gas",
    "electr", "panel", "receptacle", "light", "led",
    "sprinkler", "fire alarm", "fire protect",
    "ceiling", "acoustic",
    "data", "cabling", "cat6",
]


def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)


def extract_all_numbers(wb):
    """Extract all numeric values from all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    numbers.append(cell.value)
    return numbers


def check_for_formulas(wb, copy_from_env, container_path):
    """Check if the workbook contains formulas by reloading without data_only."""
    try:
        from openpyxl import load_workbook as lw
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        copy_from_env(container_path, temp_file.name)
        wb_f = lw(temp_file.name, data_only=False)
        formula_count = 0
        for sn in wb_f.sheetnames:
            sheet = wb_f[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                        max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_f.close()
        os.unlink(temp_file.name)
        return formula_count
    except Exception as e:
        logger.debug(f"Could not check formulas: {e}")
        return 0


def verify_cost_estimate(traj, env_info, task_info):
    """
    Verify construction cost estimate workbook.

    Data sources: RSMeans 2024 (Gordian) + DOL Davis-Bacon prevailing wages.
    Expected grand total range: $350K-$800K (TI build-out, 4,200 SF medical office).

    Scoring (10 points total, pass threshold 5.0):
    1. Wrong-target gate: file exists with substantive content (0 if not)
    2. Line items with quantities, unit costs, extensions (2.0 pts)
    3. Coverage of major trades/CSI divisions (2.0 pts)
    4. Formulas for line item extensions (qty x unit cost) (1.5 pts)
    5. Subtotals by division or trade (1.0 pt)
    6. Markup calculations (overhead, profit, contingency) (2.0 pts)
    7. Grand total in plausible range ($350K-$800K) (1.5 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/medical_office_estimate.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_construction_')

    try:
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"Wrong-target gate: Failed to load medical_office_estimate.xlsx: {error}"
            }

        feedback_parts = []
        score = 0.0

        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)

        # Count total cells
        total_cells = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                        max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1

        # ===================================================================
        # WRONG-TARGET GATE
        # ===================================================================
        if total_cells < 30:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "Wrong-target gate: File has insufficient content"
            }

        # ===================================================================
        # CHECK 1: Line items with quantities and costs (2.0 pts)
        # ===================================================================
        key_items_found = sum(1 for item in KEY_ITEMS if item in all_text)

        # Check for quantity-related terms
        qty_terms = ["qty", "quantity", "count", "each", "sf", "lf", "ea",
                     "sq ft", "lin ft", "square feet", "linear feet"]
        has_quantities = sum(1 for term in qty_terms if term in all_text)

        # Check for cost-related terms
        cost_terms = ["unit cost", "unit price", "rate", "cost", "price",
                      "extension", "total", "amount", "subtotal"]
        has_costs = sum(1 for term in cost_terms if term in all_text)

        if key_items_found >= 15 and has_quantities >= 2 and has_costs >= 3:
            score += 2.0
            feedback_parts.append(f"Line items: {key_items_found} items with qty/cost data")
        elif key_items_found >= 10 and has_quantities >= 1:
            score += 1.5
            feedback_parts.append(f"Line items: {key_items_found} items (good coverage)")
        elif key_items_found >= 6:
            score += 1.0
            feedback_parts.append(f"Line items: {key_items_found} items (partial)")
        elif key_items_found >= 3:
            score += 0.5
            feedback_parts.append(f"Line items: only {key_items_found} items")
        else:
            feedback_parts.append("Line items: insufficient detail")

        # ===================================================================
        # CHECK 2: Coverage of CSI divisions (2.0 pts)
        # ===================================================================
        divisions_found = sum(1 for div in EXPECTED_DIVISIONS if div in all_text)

        if divisions_found >= 12:
            score += 2.0
            feedback_parts.append(f"Trade coverage: {divisions_found}/15 divisions")
        elif divisions_found >= 8:
            score += 1.5
            feedback_parts.append(f"Trade coverage: {divisions_found}/15 divisions")
        elif divisions_found >= 5:
            score += 1.0
            feedback_parts.append(f"Trade coverage: {divisions_found}/15 divisions (partial)")
        elif divisions_found >= 3:
            score += 0.5
            feedback_parts.append(f"Trade coverage: only {divisions_found}/15 divisions")
        else:
            feedback_parts.append("Trade coverage: inadequate")

        # ===================================================================
        # CHECK 3: Formulas present (1.5 pts)
        # ===================================================================
        formula_count = check_for_formulas(wb, copy_from_env, container_path)

        if formula_count >= 20:
            score += 1.5
            feedback_parts.append(f"Formulas: {formula_count} formulas found")
        elif formula_count >= 10:
            score += 1.0
            feedback_parts.append(f"Formulas: {formula_count} formulas (adequate)")
        elif formula_count >= 5:
            score += 0.5
            feedback_parts.append(f"Formulas: {formula_count} formulas (limited)")
        else:
            # Check if there are calculated values even without explicit formulas
            # (agent may have computed values manually)
            if len(all_numbers) >= 50:
                score += 0.3
                feedback_parts.append("Formulas: few/no formulas but numeric values present")
            else:
                feedback_parts.append("Formulas: missing")

        # ===================================================================
        # CHECK 4: Subtotals by division (1.0 pt)
        # ===================================================================
        subtotal_terms = ["subtotal", "sub-total", "sub total", "division total",
                         "section total", "trade total", "category total"]
        has_subtotals = sum(1 for term in subtotal_terms if term in all_text)

        # Also check for intermediate sum values in plausible range
        intermediate_sums = [n for n in all_numbers if 5000 <= n <= 200000]

        if has_subtotals >= 3 or (has_subtotals >= 1 and len(intermediate_sums) >= 5):
            score += 1.0
            feedback_parts.append("Subtotals: present by division/trade")
        elif has_subtotals >= 1 or len(intermediate_sums) >= 3:
            score += 0.5
            feedback_parts.append("Subtotals: partially present")
        else:
            feedback_parts.append("Subtotals: missing or unclear")

        # ===================================================================
        # CHECK 5: Markup calculations (2.0 pts)
        # ===================================================================
        markup_terms = ["overhead", "profit", "contingency", "markup",
                       "general conditions", "gc fee", "bond",
                       "permit", "insurance", "10%", "8%", "5%", "2%"]
        markup_evidence = sum(1 for term in markup_terms if term in all_text)

        if markup_evidence >= 5:
            score += 2.0
            feedback_parts.append("Markup: overhead, profit, contingency all addressed")
        elif markup_evidence >= 3:
            score += 1.5
            feedback_parts.append("Markup: most categories addressed")
        elif markup_evidence >= 2:
            score += 1.0
            feedback_parts.append("Markup: partially addressed")
        elif markup_evidence >= 1:
            score += 0.5
            feedback_parts.append("Markup: minimally addressed")
        else:
            feedback_parts.append("Markup: missing")

        # ===================================================================
        # CHECK 6: Grand total in plausible range (1.5 pts)
        # ===================================================================
        # Expected range based on RSMeans 2024 TI costs: $350K-$800K
        # RSMeans full build: $236.56/SF x 4,200 SF = ~$993K
        # TI factor 40-80%: ~$397K-$795K, widened to $350K-$800K
        grand_total_candidates = [n for n in all_numbers if 350000 <= n <= 800000]

        # Also check for formatted text like "$450,000" or "500000"
        total_pattern = re.compile(r'(?:grand\s+)?total.*?[\$]?\s*([\d,]+\.?\d*)')
        total_matches = total_pattern.findall(all_text)
        for match in total_matches:
            try:
                val = float(match.replace(',', ''))
                if 350000 <= val <= 800000:
                    grand_total_candidates.append(val)
            except ValueError:
                pass

        if grand_total_candidates:
            # Find the one closest to expected range center (~$575K)
            best = min(grand_total_candidates, key=lambda x: abs(x - 575000))
            if 350000 <= best <= 800000:
                score += 1.5
                feedback_parts.append(f"Grand total: ${best:,.0f} (plausible range)")
            else:
                score += 0.5
                feedback_parts.append(f"Grand total: ${best:,.0f} (outside expected range)")
        else:
            # Check if there's any very large number
            large_nums = [n for n in all_numbers if n > 100000]
            if large_nums:
                score += 0.3
                feedback_parts.append(f"Grand total: large value found (${max(large_nums):,.0f}) but outside expected range")
            else:
                feedback_parts.append("Grand total: not found or not in expected range")

        # ===================================================================
        # Final assessment
        # ===================================================================
        passed = score >= 5.0
        normalized_score = score / 10.0

        feedback = " | ".join(feedback_parts)
        logger.info(f"Construction estimate verification - Score: {score}/10.0, Passed: {passed}")

        return {
            "passed": passed,
            "score": normalized_score,
            "feedback": feedback
        }

    except Exception as e:
        logger.error(f"Verification error: {e}", exc_info=True)
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Verification error: {str(e)}"
        }
    finally:
        cleanup_temp_dir(temp_dir)

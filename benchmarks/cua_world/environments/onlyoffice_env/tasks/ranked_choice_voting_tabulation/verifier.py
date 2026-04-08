#!/usr/bin/env python3
"""
Verifier for Ranked Choice Voting Tabulation task.

Evaluates if the agent properly tabulated Round 1 choices, correctly identified Nick Begich
as the eliminated candidate, and appropriately redistributed Begich's rank 2 choices to
Peltola and Palin to get the final Round 2 totals.

Uses exact numeric footprint checks (via openpyxl) + VLM trajectory checks.
"""

import os
import json
import logging
import tempfile

# Optional gym_anything imports for VLM verification
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
    VLM_AVAILABLE = True
except ImportError:
    VLM_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_data(wb):
    """Extract all text and numbers from the workbook."""
    all_text = []
    all_numbers = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        all_numbers.append(cell.value)
                    elif isinstance(cell.value, str):
                        all_text.append(cell.value.lower())
    
    return " ".join(all_text), all_numbers

def verify_rcv_tabulation(traj, env_info, task_info):
    """
    Main verification function.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    gt = metadata.get('ground_truth', {})
    expected_output = metadata.get('expected_output_file', '/home/ga/Documents/Spreadsheets/rcv_tabulation_final.xlsx')

    # Load export JSON
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not export_result.get('file_exists', False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Target workbook rcv_tabulation_final.xlsx was not saved."
        }
        
    if not export_result.get('file_created_during_task', False):
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "Target workbook exists but was not created/modified during the task session."
        }

    # Load the actual workbook
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    score = 0
    feedback_parts = []
    
    try:
        copy_from_env(expected_output, temp_xlsx.name)
        
        # openpyxl is available in the verifier environment natively or via requirement
        import openpyxl
        # Use data_only=True to evaluate cached formulas
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        all_text, all_numbers = extract_all_data(wb)
        
        # 1. Check for basic structure (Names and Sheet) (20 points)
        names_found = all(name in all_text for name in ["peltola", "palin", "begich"])
        if names_found:
            score += 20
            feedback_parts.append("Candidate names present in workbook")
        else:
            feedback_parts.append("Missing candidate names in tabulation")

        # Helper to check if a specific number or something very close (+/- 2) is in the extracted numbers
        def check_number(target, numbers_list, tolerance=2):
            return any(abs(n - target) <= tolerance for n in numbers_list)

        # 2. Check Round 1 Totals (25 points)
        # GT: Peltola 1985, Palin 1560, Begich 1410
        r1_correct = (
            check_number(gt.get('r1_peltola'), all_numbers) and
            check_number(gt.get('r1_palin'), all_numbers) and
            check_number(gt.get('r1_begich'), all_numbers)
        )
        if r1_correct:
            score += 25
            feedback_parts.append("Round 1 totals accurately calculated")
        else:
            feedback_parts.append("Round 1 totals incorrect or missing")

        # 3. Check Transfers (25 points)
        # GT: To Palin 710, To Peltola 405
        transfers_correct = (
            check_number(gt.get('transfers_to_palin'), all_numbers) and
            check_number(gt.get('transfers_to_peltola'), all_numbers)
        )
        if transfers_correct:
            score += 25
            feedback_parts.append("Begich transfer votes accurately redistributed")
        else:
            feedback_parts.append("Rank 2 transfer votes incorrect or missing")

        # 4. Check Round 2 Final Totals (30 points)
        # GT: Peltola 2390, Palin 2270
        r2_correct = (
            check_number(gt.get('r2_peltola'), all_numbers) and
            check_number(gt.get('r2_palin'), all_numbers)
        )
        if r2_correct:
            score += 30
            feedback_parts.append("Round 2 final totals accurately calculated")
        else:
            feedback_parts.append("Round 2 final totals incorrect or missing")

        wb.close()
    except ImportError:
        return {"passed": False, "score": 0, "feedback": "openpyxl not installed on host."}
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Error parsing workbook: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # Optional VLM verification to ensure they didn't just hardcode numbers via python
    if VLM_AVAILABLE and traj:
        try:
            frames = sample_trajectory_frames(traj, n=3)
            if frames:
                prompt = (
                    "Look at these screenshots of an agent using a spreadsheet application. "
                    "Is the agent interacting with spreadsheet cells, formulas, or pivot tables "
                    "to calculate voting totals (e.g. counting votes, sorting candidates)? "
                    "Reply with YES if they are actively using spreadsheet features for tabulation, or NO if the spreadsheet is mostly blank/static."
                )
                vlm_res = query_vlm(images=frames, prompt=prompt)
                if vlm_res and "YES" in vlm_res.get("response", "").upper():
                    feedback_parts.append("VLM confirmed spreadsheet usage.")
                else:
                    feedback_parts.append("VLM did not confirm active spreadsheet tabulation.")
        except Exception as e:
            logger.warning(f"VLM check failed: {e}")

    # Determine Pass/Fail (Threshold 60)
    passed = score >= 60

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
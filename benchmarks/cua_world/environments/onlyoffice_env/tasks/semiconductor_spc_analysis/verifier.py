#!/usr/bin/env python3
"""
Verifier for Semiconductor SPC Control Chart Analysis task.

Evaluates the agent's ability to calculate X-bar and R chart parameters
from raw wafer thickness data and accurately flag Out-Of-Control lots.
"""

import os
import json
import logging
import tempfile
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ground Truth Raw Data (Identical to what was generated in setup_task.sh)
RAW_DATA = [
    ("LOT_001", [438.2, 442.1, 439.5, 445.0, 437.8]),
    ("LOT_002", [440.1, 444.3, 438.9, 441.2, 439.6]),
    ("LOT_003", [439.5, 435.6, 441.0, 443.2, 440.8]),
    ("LOT_004", [442.3, 440.1, 437.9, 439.4, 441.7]),
    ("LOT_005", [441.8, 438.4, 440.5, 442.9, 436.1]),
    ("LOT_006", [437.5, 441.2, 439.8, 440.1, 442.4]),
    ("LOT_007", [439.9, 443.5, 438.1, 440.7, 439.2]),
    ("LOT_008", [440.2, 437.8, 441.6, 439.3, 444.1]),
    ("LOT_009", [441.5, 439.1, 440.8, 438.6, 442.0]),
    ("LOT_010", [438.9, 442.5, 440.3, 439.7, 441.2]),
    ("LOT_011", [439.4, 438.2, 442.1, 440.6, 437.9]),
    ("LOT_012", [440.8, 441.5, 439.2, 443.0, 438.5]),
    ("LOT_013", [442.1, 439.6, 440.9, 438.1, 441.4]),
    ("LOT_014", [438.5, 440.2, 442.8, 439.5, 441.0]),
    ("LOT_015", [440.3, 441.8, 438.7, 442.4, 439.1]),
    ("LOT_016", [449.2, 448.5, 451.0, 450.3, 447.8]),
    ("LOT_017", [441.0, 439.4, 442.5, 438.9, 440.6]),
    ("LOT_018", [439.7, 441.1, 438.4, 440.2, 442.8]),
    ("LOT_019", [440.5, 438.8, 441.3, 439.6, 440.1]),
    ("LOT_020", [442.6, 440.4, 439.1, 441.8, 438.5]),
    ("LOT_021", [438.2, 441.7, 440.5, 439.0, 442.3]),
    ("LOT_022", [440.9, 438.5, 442.0, 441.2, 439.7]),
    ("LOT_023", [441.4, 440.2, 438.9, 442.5, 439.1]),
    ("LOT_024", [439.6, 442.8, 440.1, 438.5, 441.0]),
    ("LOT_025", [431.5, 429.8, 430.4, 432.1, 428.9]),
    ("LOT_026", [430.2, 432.5, 429.1, 431.8, 430.7]),
    ("LOT_027", [440.1, 438.7, 441.5, 439.9, 442.2]),
    ("LOT_028", [438.8, 441.2, 439.6, 440.5, 438.1]),
    ("LOT_029", [442.0, 439.5, 441.8, 438.4, 440.7]),
    ("LOT_030", [439.3, 440.9, 438.2, 442.1, 441.5])
]

# Calculate Exact Ground Truth Values
GT_MEANS = []
GT_RANGES = []
for _, vals in RAW_DATA:
    GT_MEANS.append(sum(vals) / 5.0)
    GT_RANGES.append(max(vals) - min(vals))

GRAND_MEAN = sum(GT_MEANS) / len(GT_MEANS)
AVERAGE_RANGE = sum(GT_RANGES) / len(GT_RANGES)

# Constants
A2 = 0.577
D3 = 0.0
D4 = 2.114

UCL_X = GRAND_MEAN + (A2 * AVERAGE_RANGE)
LCL_X = GRAND_MEAN - (A2 * AVERAGE_RANGE)
UCL_R = D4 * AVERAGE_RANGE
LCL_R = D3 * AVERAGE_RANGE

# The exact anomalies based on these calculations are Lots 16, 25, and 26.
ANOMALY_LOTS = ["lot_016", "lot_025", "lot_026"]
NORMAL_LOTS = ["lot_001", "lot_002", "lot_003"]
FLAG_INDICATORS = ["true", "yes", "ooc", "out", "flag", "fail", "1"]

def verify_spc_analysis(traj, env_info, task_info):
    """
    Main verification function.
    Scoring out of 100:
    - File exists & created during task (10 pts)
    - Subgroup Means & Ranges calculated (25 pts)
    - Grand Mean & Average Range (15 pts)
    - Control Limits calculated correctly (20 pts)
    - Out of Control flags correctly identify anomalies without false positives (20 pts)
    - Spreadsheet formulas utilized (10 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback = []
    score = 0
    passed = False
    
    # 1. Read JSON result
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Output workbook wafer_spc_analysis.xlsx was not found."}
    
    if not result.get('file_created_during_task'):
        feedback.append("Warning: File timestamp indicates it might not have been created during this session.")
    else:
        score += 10
        feedback.append("Workbook successfully created and saved.")

    # 2. Extract workbook locally
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/wafer_spc_analysis.xlsx", temp_wb.name)
        
        # We try to import openpyxl inside the test context
        try:
            import openpyxl
        except ImportError:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            import openpyxl
            
        wb_data = openpyxl.load_workbook(temp_wb.name, data_only=True)
        wb_formulas = openpyxl.load_workbook(temp_wb.name, data_only=False)
        
    except Exception as e:
        return {"passed": False, "score": score, "feedback": f"Failed to parse workbook: {e}"}
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)

    # 3. Extract numbers and rows for validation
    all_numbers = []
    rows_data = []
    
    for sheet in wb_data.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_clean = [v for v in row if v is not None]
            rows_data.append(row_clean)
            for cell in row_clean:
                if isinstance(cell, (int, float)):
                    all_numbers.append(float(cell))
                    
    # Check Formulas
    formula_count = 0
    for sheet in wb_formulas.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.startswith('='):
                    formula_count += 1
                    
    if formula_count >= 10:
        score += 10
        feedback.append(f"Spreadsheet formulas detected ({formula_count}).")
    else:
        feedback.append("Insufficient formulas detected. Analysis may be hardcoded.")

    # 4. Check Subgroup Statistics
    def find_matches(targets, tolerance=0.01):
        matches = 0
        for target in targets:
            if any(math.isclose(target, num, abs_tol=tolerance) for num in all_numbers):
                matches += 1
        return matches

    mean_matches = find_matches(GT_MEANS)
    range_matches = find_matches(GT_RANGES)

    if mean_matches >= 25 and range_matches >= 25:
        score += 25
        feedback.append("Subgroup Means and Ranges accurately calculated.")
    elif mean_matches >= 15 or range_matches >= 15:
        score += 10
        feedback.append("Subgroup Means/Ranges partially calculated.")
    else:
        feedback.append("Subgroup Means and Ranges missing or incorrect.")

    # 5. Check Grand Averages
    has_grand_mean = find_matches([GRAND_MEAN], 0.05) > 0
    has_avg_range = find_matches([AVERAGE_RANGE], 0.05) > 0
    
    if has_grand_mean and has_avg_range:
        score += 15
        feedback.append("Grand Mean and Average Range accurately calculated.")
    elif has_grand_mean or has_avg_range:
        score += 7
        feedback.append("Grand Mean OR Average Range accurately calculated, but not both.")
    else:
        feedback.append("Grand Mean / Average Range missing or incorrect.")

    # 6. Check Control Limits
    has_ucl_x = find_matches([UCL_X], 0.1) > 0
    has_lcl_x = find_matches([LCL_X], 0.1) > 0
    has_ucl_r = find_matches([UCL_R], 0.1) > 0
    has_lcl_r = find_matches([LCL_R], 0.1) > 0

    limits_found = sum([has_ucl_x, has_lcl_x, has_ucl_r, has_lcl_r])
    if limits_found == 4:
        score += 20
        feedback.append("All Control Limits (UCL/LCL for X-bar and R) calculated correctly.")
    elif limits_found >= 2:
        score += 10
        feedback.append(f"Some Control Limits found ({limits_found}/4).")
    else:
        feedback.append("Control Limits missing or incorrect.")

    # 7. Check Out of Control Flags
    correct_flags = 0
    false_positives = 0
    
    for row in rows_data:
        row_str = " ".join(str(v).lower() for v in row)
        
        is_anomaly_lot = any(lot in row_str for lot in ANOMALY_LOTS)
        is_normal_lot = any(lot in row_str for lot in NORMAL_LOTS)
        
        # Look for explicit True boolean or string indicators
        has_flag = any(v is True for v in row) or \
                   any(indicator in row_str.split() for indicator in FLAG_INDICATORS)
                   
        if is_anomaly_lot and has_flag:
            correct_flags += 1
        if is_normal_lot and has_flag:
            false_positives += 1

    if correct_flags == 3 and false_positives == 0:
        score += 20
        feedback.append("Out of Control flags correctly identified anomalous lots without false positives.")
    elif correct_flags > 0:
        score += 10
        feedback.append(f"Out of Control flags partially correct (Found {correct_flags}/3, False positives: {false_positives}).")
    else:
        feedback.append("Out of Control flags missing or incorrect.")

    # Final Evaluation
    if score >= 70 and limits_found >= 2 and correct_flags > 0:
        passed = True
        
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback)
    }
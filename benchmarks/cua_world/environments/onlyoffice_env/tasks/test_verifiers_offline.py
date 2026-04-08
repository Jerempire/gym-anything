#!/usr/bin/env python3
"""
Offline verifier unit tests for all 5 OnlyOffice tasks.
Tests do-nothing, wrong-target, partial, and full-completion scenarios
by mocking copy_from_env to create XLSX files with openpyxl.

Usage:
    python3 test_verifiers_offline.py
"""
import sys
import os
import importlib.util
import shutil
import tempfile

from openpyxl import Workbook

# Add utils to path
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'utils'))

TASKS_DIR = os.path.dirname(os.path.abspath(__file__))


def load_verifier(task_name):
    """Load a verifier module from task directory."""
    path = os.path.join(TASKS_DIR, task_name, 'verifier.py')
    spec = importlib.util.spec_from_file_location('verifier', path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def make_env_missing():
    """Simulate file not existing (do-nothing scenario)."""
    def copy_from_env(src, dst):
        raise FileNotFoundError(f"No such file: {src}")
    return {'copy_from_env': copy_from_env}


def make_env_from_workbook(wb):
    """Create env_info with mock copy_from_env that writes a workbook."""
    def copy_from_env(src, dst):
        wb.save(dst)
    return {'copy_from_env': copy_from_env}


def create_minimal_wb():
    """Create a workbook with very few cells (wrong-target gate)."""
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'test'
    ws['B1'] = 123
    return wb


# ============================================================================
# TEST 1: Epidemiological Outbreak Investigation
# ============================================================================
def test_epidemiological():
    print("=" * 60)
    print("TEST: epidemiological_outbreak_investigation")
    print("=" * 60)
    mod = load_verifier('epidemiological_outbreak_investigation')

    # Do-nothing: file missing
    r = mod.verify_outbreak_investigation([], make_env_missing(), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Do-nothing failed: {r}"
    print(f"  [PASS] Do-nothing: passed={r['passed']}, score={r['score']}")

    # Wrong-target: minimal content
    r = mod.verify_outbreak_investigation([], make_env_from_workbook(create_minimal_wb()), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Wrong-target failed: {r}"
    print(f"  [PASS] Wrong-target: passed={r['passed']}, score={r['score']}")

    # Partial: some data imported but no vehicle identification, no epi curve
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    ws['A1'] = "Outbreak Investigation"
    ws['A2'] = "Food-Specific Attack Rates"
    ws['A3'] = "Food Item"
    ws['B3'] = "Ate - Ill"
    ws['C3'] = "Ate - Well"
    ws['D3'] = "Attack Rate"
    # Only list a few foods, do NOT include vanilla ice cream
    foods = ["Baked Ham", "Spinach", "Mashed Potato", "Cabbage Salad"]
    for i, food in enumerate(foods):
        ws.cell(row=4+i, column=1, value=food)
        ws.cell(row=4+i, column=2, value=20+i)
        ws.cell(row=4+i, column=3, value=10+i)
        ws.cell(row=4+i, column=4, value=50.0 + i*3)
    # Add some context but NOT the vehicle
    ws['A12'] = "attack rate"
    ws['A13'] = "exposed"
    # Add enough cells to pass gate
    for i in range(30):
        ws.cell(row=20+i, column=1, value=f"data row {i}")
        ws.cell(row=20+i, column=2, value=i*100)

    r = mod.verify_outbreak_investigation([], make_env_from_workbook(wb), {})
    assert r['passed'] is False, f"Partial should not pass: {r}"
    assert r['score'] > 0.0, f"Partial should have some score: {r}"
    print(f"  [PASS] Partial: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")

    # Full completion: all checks
    wb = Workbook()
    ws = wb.active
    ws.title = "Outbreak Analysis"

    # Data with 75 records and food columns
    ws['A1'] = "Epidemiological Investigation Report"
    ws['A2'] = "CDC EIS - Oswego County Outbreak, April 1940"
    ws['A3'] = "ID"
    ws['B3'] = "Age"
    ws['C3'] = "Sex"
    ws['D3'] = "Ill"
    ws['E3'] = "Baked Ham"
    ws['F3'] = "Vanilla Ice Cream"
    ws['G3'] = "Chocolate Ice Cream"
    # Add 75 rows of data
    for i in range(75):
        ws.cell(row=4+i, column=1, value=i+1)
        ws.cell(row=4+i, column=2, value=30+i%40)
        ws.cell(row=4+i, column=3, value="M" if i%2 else "F")
        ws.cell(row=4+i, column=4, value="Y" if i < 46 else "N")
        ws.cell(row=4+i, column=5, value="Y" if i%3 == 0 else "N")
        ws.cell(row=4+i, column=6, value="Y" if i < 43 else "N")
        ws.cell(row=4+i, column=7, value="Y" if i%2 == 0 else "N")

    # Attack rates sheet
    ws2 = wb.create_sheet("Attack Rates")
    ws2['A1'] = "Food-Specific Attack Rates"
    ws2['A2'] = "Food Item"
    ws2['B2'] = "Ate-Ill"
    ws2['C2'] = "Ate-Well"
    ws2['D2'] = "Not Ate-Ill"
    ws2['E2'] = "Not Ate-Well"
    ws2['F2'] = "Attack Rate (Ate)"
    ws2['G2'] = "Attack Rate (Not Ate)"
    ws2['H2'] = "Relative Risk"
    foods_data = [
        ("Baked Ham", 29, 17, 17, 12, 63.0, 58.6, 1.08),
        ("Spinach", 26, 17, 20, 12, 60.5, 62.5, 0.97),
        ("Vanilla Ice Cream", 43, 11, 3, 18, 79.6, 14.3, 5.57),
        ("Chocolate Ice Cream", 25, 22, 20, 7, 53.2, 74.1, 0.72),
        ("Cakes", 27, 13, 19, 16, 67.5, 54.3, 1.24),
        ("Jello", 16, 7, 30, 22, 69.6, 57.7, 1.21),
        ("Mashed Potato", 23, 14, 23, 14, 62.2, 62.2, 1.00),
        ("Cabbage Salad", 18, 10, 28, 19, 64.3, 59.6, 1.08),
        ("Rolls", 21, 16, 25, 13, 56.8, 65.8, 0.86),
        ("Brown Bread", 18, 9, 28, 20, 66.7, 58.3, 1.14),
    ]
    for i, (food, ai, aw, nai, naw, ar_a, ar_na, rr) in enumerate(foods_data):
        ws2.cell(row=3+i, column=1, value=food)
        ws2.cell(row=3+i, column=2, value=ai)
        ws2.cell(row=3+i, column=3, value=aw)
        ws2.cell(row=3+i, column=4, value=nai)
        ws2.cell(row=3+i, column=5, value=naw)
        ws2.cell(row=3+i, column=6, value=ar_a)
        ws2.cell(row=3+i, column=7, value=ar_na)
        ws2.cell(row=3+i, column=8, value=rr)

    # 2x2 table
    ws2['A15'] = "2x2 Contingency Table - Vanilla Ice Cream"
    ws2['A16'] = "Exposed"
    ws2['B16'] = 43
    ws2['C16'] = 11
    ws2['A17'] = "Unexposed"
    ws2['B17'] = 3
    ws2['C17'] = 18
    ws2['A18'] = "Relative Risk = 5.57"
    ws2['A19'] = "odds ratio"
    ws2['A20'] = "chi-square"

    # Conclusion
    ws2['A22'] = "Conclusion: Vanilla ice cream is the implicated vehicle"
    ws2['A23'] = "The highest attack rate was for vanilla ice cream (79.6%)"
    ws2['A24'] = "The relative risk of 5.57 strongly implicates vanilla ice cream as the source"

    # Epi curve sheet
    ws3 = wb.create_sheet("Epi Curve")
    ws3['A1'] = "Epidemic Curve - Onset Time Distribution"
    ws3['A2'] = "onset"
    ws3['B2'] = "April 18"
    ws3['C2'] = "April 19"
    ws3['A3'] = "incubation period"
    ws3['A4'] = "timeline"
    for i in range(10):
        ws3.cell(row=5+i, column=1, value=f"{9+i}:00 PM")
        ws3.cell(row=5+i, column=2, value=3+i)

    # Summary
    ws4 = wb.create_sheet("Summary")
    ws4['A1'] = "Descriptive Epidemiology"
    ws4['A2'] = "Total attendees: 75"
    ws4['A3'] = "Total cases: 46"
    ws4['A4'] = "attack rate overall"
    ws4['A5'] = "median age"
    ws4['A6'] = "sex distribution"
    ws4['A7'] = "Church supper on April 18, 1940 in Lycoming, Oswego County"

    r = mod.verify_outbreak_investigation([], make_env_from_workbook(wb), {})
    assert r['passed'] is True, f"Full should pass: {r}"
    assert r['score'] >= 0.5, f"Full score too low: {r}"
    print(f"  [PASS] Full: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")
    print()


# ============================================================================
# TEST 2: Clinical Trial Adverse Event Analysis
# ============================================================================
def test_clinical_trial():
    print("=" * 60)
    print("TEST: clinical_trial_adverse_event_analysis")
    print("=" * 60)
    mod = load_verifier('clinical_trial_adverse_event_analysis')

    # Do-nothing
    r = mod.verify_clinical_ae_analysis([], make_env_missing(), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Do-nothing failed: {r}"
    print(f"  [PASS] Do-nothing: passed={r['passed']}, score={r['score']}")

    # Wrong-target
    r = mod.verify_clinical_ae_analysis([], make_env_from_workbook(create_minimal_wb()), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Wrong-target failed: {r}"
    print(f"  [PASS] Wrong-target: passed={r['passed']}, score={r['score']}")

    # Partial: some AE terms but no dose comparison
    wb = Workbook()
    ws = wb.active
    ws.title = "AE Summary"
    ws['A1'] = "Adverse Event Analysis"
    ws['A2'] = "Adverse Event"
    ws['B2'] = "Incidence"
    ae_terms = ["urinary tract infection", "increased urination", "thirst",
                "constipation", "nausea", "abdominal pain"]
    for i, ae in enumerate(ae_terms):
        ws.cell(row=3+i, column=1, value=ae)
        ws.cell(row=3+i, column=2, value=f"{3+i}%")
    ws['A15'] = "safety"
    ws['A16'] = "adverse"
    ws['A17'] = "incidence"
    for i in range(25):
        ws.cell(row=20+i, column=1, value=f"row {i}")
        ws.cell(row=20+i, column=2, value=i)

    r = mod.verify_clinical_ae_analysis([], make_env_from_workbook(wb), {})
    assert r['passed'] is False, f"Partial should not pass: {r}"
    assert r['score'] > 0.0, f"Partial should have some score: {r}"
    print(f"  [PASS] Partial: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")

    # Full completion
    wb = Workbook()
    ws = wb.active
    ws.title = "AE Frequency Analysis"
    ws['A1'] = "INVOKANA (canagliflozin) Adverse Event Analysis"
    ws['A2'] = "Pooled Phase III Studies - Placebo-Controlled"
    ws['A3'] = "Adverse Event"
    ws['B3'] = "Placebo (N=646)"
    ws['C3'] = "INVOKANA 100 mg (N=833)"
    ws['D3'] = "INVOKANA 300 mg (N=834)"
    ae_data = [
        ("Urinary tract infection", 3.8, 5.9, 4.4),
        ("Increased urination", 0.7, 5.1, 4.6),
        ("Thirst", 0.1, 2.8, 2.4),
        ("Constipation", 0.9, 1.8, 2.4),
        ("Nausea", 1.6, 2.1, 2.3),
        ("Abdominal pain", 0.8, 1.8, 1.7),
    ]
    for i, (ae, p, d100, d300) in enumerate(ae_data):
        ws.cell(row=4+i, column=1, value=ae)
        ws.cell(row=4+i, column=2, value=p)
        ws.cell(row=4+i, column=3, value=d100)
        ws.cell(row=4+i, column=4, value=d300)

    ws['A12'] = "Female Genital Mycotic Infections"
    ws['A13'] = "genital mycotic infection"
    ws['B13'] = 2.8
    ws['C13'] = 10.6
    ws['D13'] = 11.6
    ws['A14'] = "Vulvovaginal pruritus"
    ws['B14'] = 0.0
    ws['C14'] = 1.6
    ws['D14'] = 3.2

    ws['A16'] = "Male Genital Mycotic Infections"
    ws['A17'] = "male genital mycotic infection"
    ws['B17'] = 0.7
    ws['C17'] = 4.2
    ws['D17'] = 3.8

    # Dose-response sheet
    ws2 = wb.create_sheet("Dose Response")
    ws2['A1'] = "Dose-Response Comparison"
    ws2['A2'] = "compared to placebo"
    ws2['A3'] = "vs placebo"
    ws2['A4'] = "100 mg vs 300 mg"
    ws2['A5'] = "dose-dependent increase in genital mycotic infections"
    ws2['A6'] = "dose response relationship observed"
    ws2['A7'] = "placebo"
    ws2['A8'] = "INVOKANA 100 mg"
    ws2['A9'] = "INVOKANA 300 mg"
    ws2['A10'] = "SGLT2 inhibitor class effect"
    ws2['A11'] = "canagliflozin"

    # Safety signals sheet
    ws3 = wb.create_sheet("Safety Signals")
    ws3['A1'] = "Key Safety Findings"
    ws3['A2'] = "Safety signal: genital mycotic infections"
    ws3['A3'] = "Higher incidence of urinary tract infections in treatment groups"
    ws3['A4'] = "Increased urination consistent with osmotic diuresis mechanism"
    ws3['A5'] = "Clinically significant difference in genital infections"
    ws3['A6'] = "key finding"
    ws3['A7'] = "elevated risk"
    ws3['A8'] = "glycosuria leads to increased mycotic infection risk"
    ws3['A9'] = "p-value < 0.05"
    ws3['A10'] = "confidence interval"
    ws3['A11'] = "95%"
    ws3['A12'] = "benefit-risk assessment"
    ws3['A13'] = "pharmacovigilance"

    r = mod.verify_clinical_ae_analysis([], make_env_from_workbook(wb), {})
    assert r['passed'] is True, f"Full should pass: {r}"
    assert r['score'] >= 0.5, f"Full score too low: {r}"
    print(f"  [PASS] Full: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")
    print()


# ============================================================================
# TEST 3: Forensic Transaction Audit
# ============================================================================
def test_forensic_audit():
    print("=" * 60)
    print("TEST: forensic_transaction_audit")
    print("=" * 60)
    mod = load_verifier('forensic_transaction_audit')

    # Do-nothing
    r = mod.verify_forensic_audit([], make_env_missing(), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Do-nothing failed: {r}"
    print(f"  [PASS] Do-nothing: passed={r['passed']}, score={r['score']}")

    # Wrong-target
    r = mod.verify_forensic_audit([], make_env_from_workbook(create_minimal_wb()), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Wrong-target failed: {r}"
    print(f"  [PASS] Wrong-target: passed={r['passed']}, score={r['score']}")

    # Partial: data imported but no analysis methodology
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws['A1'] = "Voucher Number"
    ws['B1'] = "Amount"
    ws['C1'] = "Check Date"
    ws['D1'] = "Department"
    ws['E1'] = "Contract"
    ws['F1'] = "Vendor Name"
    for i in range(50):
        ws.cell(row=2+i, column=1, value=f"PV{1000+i}")
        ws.cell(row=2+i, column=2, value=1000+i*100)
        ws.cell(row=2+i, column=3, value="01/15/2024")
        ws.cell(row=2+i, column=4, value="Department of Transportation")
        ws.cell(row=2+i, column=5, value="DV")
        ws.cell(row=2+i, column=6, value=f"Vendor {i}")

    r = mod.verify_forensic_audit([], make_env_from_workbook(wb), {})
    assert r['passed'] is False, f"Partial should not pass: {r}"
    assert r['score'] > 0.0, f"Partial should have some score: {r}"
    print(f"  [PASS] Partial: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")

    # Full completion
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction Data"
    ws['A1'] = "Voucher Number"
    ws['B1'] = "Amount"
    ws['C1'] = "Check Date"
    ws['D1'] = "Department"
    ws['E1'] = "Contract"
    ws['F1'] = "Vendor Name"
    for i in range(50):
        ws.cell(row=2+i, column=1, value=f"PV{1000+i}")
        ws.cell(row=2+i, column=2, value=1000+i*100)
        ws.cell(row=2+i, column=3, value="01/15/2024")
        ws.cell(row=2+i, column=4, value="Chicago Department of Transportation")
        ws.cell(row=2+i, column=5, value="DV")
        ws.cell(row=2+i, column=6, value=f"Vendor {i}")

    # Benford's analysis sheet
    ws2 = wb.create_sheet("Benford Analysis")
    ws2['A1'] = "Benford's Law Analysis"
    ws2['A2'] = "First Digit Distribution"
    ws2['A3'] = "Digit"
    ws2['B3'] = "Observed"
    ws2['C3'] = "Expected"
    ws2['D3'] = "Frequency"
    ws2['E3'] = "Percentage"
    expected_benford = [30.1, 17.6, 12.5, 9.7, 7.9, 6.7, 5.8, 5.1, 4.6]
    for i in range(9):
        ws2.cell(row=4+i, column=1, value=i+1)
        ws2.cell(row=4+i, column=2, value=10+i*2)
        ws2.cell(row=4+i, column=3, value=expected_benford[i])
        ws2.cell(row=4+i, column=4, value=f"count: {10+i*2}")
        ws2.cell(row=4+i, column=5, value=f"{10+i*2}%")
    ws2['A15'] = "chi-square test"
    ws2['A16'] = "deviation from expected Benford distribution"
    ws2['A17'] = "leading digit frequency"

    # Vendor analysis sheet
    ws3 = wb.create_sheet("Vendor Analysis")
    ws3['A1'] = "Duplicate Payment Detection"
    ws3['A2'] = "duplicate"
    ws3['A3'] = "Vendor Concentration Analysis"
    ws3['A4'] = "top vendor"
    ws3['A5'] = "highest paid"
    ws3['A6'] = "vendor concentration"
    ws3['A7'] = "repeated payments"
    ws3['A8'] = "same amount"
    ws3['A9'] = "pareto analysis"
    for i in range(10):
        ws3.cell(row=10+i, column=1, value=f"Vendor {i}")
        ws3.cell(row=10+i, column=2, value=50000-i*3000)

    # Amount pattern sheet
    ws4 = wb.create_sheet("Amount Patterns")
    ws4['A1'] = "Amount Pattern Analysis"
    ws4['A2'] = "round number detection"
    ws4['A3'] = "threshold analysis"
    ws4['A4'] = "mean"
    ws4['A5'] = "median"
    ws4['A6'] = "standard deviation"
    ws4['A7'] = "outlier detection"
    ws4['A8'] = "clustering"
    ws4['A9'] = "histogram"

    # Temporal analysis sheet
    ws5 = wb.create_sheet("Temporal Analysis")
    ws5['A1'] = "Temporal Pattern Analysis"
    ws5['A2'] = "day of week distribution"
    ws5['A3'] = "weekend transactions"
    ws5['A4'] = "end of month spike"
    ws5['A5'] = "seasonal trend"
    ws5['A6'] = "month-end"

    # Findings sheet
    ws6 = wb.create_sheet("Findings")
    ws6['A1'] = "Summary of Findings"
    ws6['A2'] = "Risk Assessment"
    ws6['A3'] = "finding: Benford deviation in first digit distribution"
    ws6['A4'] = "risk rating: Medium"
    ws6['A5'] = "flagged transactions"
    ws6['A6'] = "red flag"
    ws6['A7'] = "suspicious patterns identified"
    ws6['A8'] = "Recommendations"
    ws6['A9'] = "recommendation: further review of flagged vendors"
    ws6['A10'] = "action: investigate duplicate payments"
    ws6['A11'] = "follow-up required"

    r = mod.verify_forensic_audit([], make_env_from_workbook(wb), {})
    assert r['passed'] is True, f"Full should pass: {r}"
    assert r['score'] >= 0.5, f"Full score too low: {r}"
    print(f"  [PASS] Full: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")
    print()


# ============================================================================
# TEST 4: Construction Project Cost Estimate
# ============================================================================
def test_construction():
    print("=" * 60)
    print("TEST: construction_project_cost_estimate")
    print("=" * 60)
    mod = load_verifier('construction_project_cost_estimate')

    # Do-nothing
    r = mod.verify_cost_estimate([], make_env_missing(), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Do-nothing failed: {r}"
    print(f"  [PASS] Do-nothing: passed={r['passed']}, score={r['score']}")

    # Wrong-target
    r = mod.verify_cost_estimate([], make_env_from_workbook(create_minimal_wb()), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Wrong-target failed: {r}"
    print(f"  [PASS] Wrong-target: passed={r['passed']}, score={r['score']}")

    # Partial: some line items but no markup or grand total
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Estimate"
    ws['A1'] = "Medical Office Build-Out Estimate"
    ws['A2'] = "Item"
    ws['B2'] = "Qty"
    ws['C2'] = "Unit"
    ws['D2'] = "Unit Cost"
    ws['E2'] = "Total"
    items = [
        ("Drywall partition - metal stud framing", 2100, "SF", 8.50, 17850),
        ("Door - hollow metal frame", 12, "EA", 850, 10200),
        ("Vinyl flooring", 1500, "SF", 6.25, 9375),
        ("Carpet tile", 800, "SF", 4.50, 3600),
        ("Ceiling tile", 3200, "SF", 3.75, 12000),
    ]
    for i, (item, qty, unit, cost, total) in enumerate(items):
        ws.cell(row=3+i, column=1, value=item)
        ws.cell(row=3+i, column=2, value=qty)
        ws.cell(row=3+i, column=3, value=unit)
        ws.cell(row=3+i, column=4, value=cost)
        ws.cell(row=3+i, column=5, value=total)
    ws['A10'] = "quantity"
    ws['A11'] = "unit cost"
    for i in range(20):
        ws.cell(row=15+i, column=1, value=f"additional item {i}")
        ws.cell(row=15+i, column=2, value=100+i*10)

    r = mod.verify_cost_estimate([], make_env_from_workbook(wb), {})
    assert r['passed'] is False, f"Partial should not pass: {r}"
    assert r['score'] > 0.0, f"Partial should have some score: {r}"
    print(f"  [PASS] Partial: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")

    # Full completion
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Estimate"
    ws['A1'] = "Medical Office Build-Out - Detailed Cost Estimate"
    ws['A2'] = "Division"
    ws['B2'] = "Item Description"
    ws['C2'] = "Qty"
    ws['D2'] = "Unit"
    ws['E2'] = "Unit Cost"
    ws['F2'] = "Extension"

    full_items = [
        ("Concrete", "Concrete slab repair", 200, "SF", 12.50, 2500),
        ("Metal", "Metal stud framing", 2100, "SF", 4.25, 8925),
        ("Wood", "Wood blocking", 500, "LF", 3.50, 1750),
        ("Millwork", "Cabinet - base", 40, "LF", 285, 11400),
        ("Door", "Door - solid core", 15, "EA", 950, 14250),
        ("Finish", "Paint - walls", 8400, "SF", 1.85, 15540),
        ("Finish", "Vinyl flooring", 1500, "SF", 6.25, 9375),
        ("Finish", "Carpet tile", 800, "SF", 4.50, 3600),
        ("Finish", "Ceramic tile", 400, "SF", 12.50, 5000),
        ("Ceiling", "Ceiling tile - 2x4 lay-in", 3200, "SF", 3.75, 12000),
        ("Mechanical", "HVAC system", 1, "LS", 85000, 85000),
        ("Mechanical", "Ductwork", 4200, "SF", 8.50, 35700),
        ("Plumbing", "Plumbing fixtures", 12, "EA", 2500, 30000),
        ("Plumbing", "Medical gas system", 1, "LS", 25000, 25000),
        ("Electrical", "Electrical panel", 2, "EA", 4500, 9000),
        ("Electrical", "Receptacle", 120, "EA", 185, 22200),
        ("Electrical", "LED light fixture", 80, "EA", 350, 28000),
        ("Fire", "Sprinkler system", 4200, "SF", 4.50, 18900),
        ("Fire", "Fire alarm system", 1, "LS", 15000, 15000),
        ("Specialty", "Data cabling - Cat6", 60, "EA", 325, 19500),
    ]
    for i, (div, desc, qty, unit, ucost, ext) in enumerate(full_items):
        ws.cell(row=3+i, column=1, value=div)
        ws.cell(row=3+i, column=2, value=desc)
        ws.cell(row=3+i, column=3, value=qty)
        ws.cell(row=3+i, column=4, value=unit)
        ws.cell(row=3+i, column=5, value=ucost)
        ws.cell(row=3+i, column=6, value=ext)

    # Subtotals
    row = 25
    ws.cell(row=row, column=1, value="Subtotal")
    ws.cell(row=row, column=6, value=372640)
    row += 1
    ws.cell(row=row, column=1, value="Division Total - Finish")
    ws.cell(row=row, column=6, value=33515)
    row += 1
    ws.cell(row=row, column=1, value="Section Total - Mechanical")
    ws.cell(row=row, column=6, value=120700)

    # Markups
    row += 2
    ws.cell(row=row, column=1, value="General Conditions / Overhead")
    ws.cell(row=row, column=2, value="10%")
    ws.cell(row=row, column=6, value=37264)
    row += 1
    ws.cell(row=row, column=1, value="Contractor Profit")
    ws.cell(row=row, column=2, value="8%")
    ws.cell(row=row, column=6, value=29811)
    row += 1
    ws.cell(row=row, column=1, value="Contingency")
    ws.cell(row=row, column=2, value="5%")
    ws.cell(row=row, column=6, value=18632)
    row += 1
    ws.cell(row=row, column=1, value="Bond and Insurance")
    ws.cell(row=row, column=2, value="2%")
    ws.cell(row=row, column=6, value=7453)
    row += 2
    ws.cell(row=row, column=1, value="GRAND TOTAL")
    ws.cell(row=row, column=6, value=465800)

    # Add formulas (need to create another wb without data_only)
    ws['A40'] = "quantity"
    ws['A41'] = "unit cost"
    ws['A42'] = "total"

    r = mod.verify_cost_estimate([], make_env_from_workbook(wb), {})
    assert r['passed'] is True, f"Full should pass: {r}"
    assert r['score'] >= 0.5, f"Full score too low: {r}"
    print(f"  [PASS] Full: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")
    print()


# ============================================================================
# TEST 5: Reinsurance Treaty Portfolio Analysis
# ============================================================================
def test_reinsurance():
    print("=" * 60)
    print("TEST: reinsurance_treaty_portfolio_analysis")
    print("=" * 60)
    mod = load_verifier('reinsurance_treaty_portfolio_analysis')

    # Do-nothing
    r = mod.verify_reserve_analysis([], make_env_missing(), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Do-nothing failed: {r}"
    print(f"  [PASS] Do-nothing: passed={r['passed']}, score={r['score']}")

    # Wrong-target
    r = mod.verify_reserve_analysis([], make_env_from_workbook(create_minimal_wb()), {})
    assert r['passed'] is False and r['score'] == 0.0, f"Wrong-target failed: {r}"
    print(f"  [PASS] Wrong-target: passed={r['passed']}, score={r['score']}")

    # Partial: development factors but no ultimate/IBNR
    wb = Workbook()
    ws = wb.active
    ws.title = "Development Factors"
    ws['A1'] = "Loss Development Factors"
    ws['A2'] = "Chain Ladder Analysis"
    ws['A3'] = "link ratio"
    ws['A4'] = "age-to-age factors"
    ws['A5'] = "LDF"
    ws['A6'] = "development factor"
    ws['A7'] = "Lag 1->2"
    ws['B7'] = 1.99
    ws['A8'] = "Lag 2->3"
    ws['B8'] = 1.45
    ws['A9'] = "Lag 3->4"
    ws['B9'] = 1.30
    ws['A10'] = "Lag 4->5"
    ws['B10'] = 1.18
    ws['A11'] = "Lag 5->6"
    ws['B11'] = 1.07
    ws['A12'] = "Lag 6->7"
    ws['B12'] = 1.03
    ws['A13'] = "accident year"
    ws['A14'] = "triangle"
    ws['A15'] = "paid loss"
    # Accident years
    for i, year in enumerate(range(1988, 1998)):
        ws.cell(row=16+i, column=1, value=str(year))
    # Add enough cells
    for i in range(30):
        ws.cell(row=30+i, column=1, value=f"data {i}")
        ws.cell(row=30+i, column=2, value=1000+i)

    r = mod.verify_reserve_analysis([], make_env_from_workbook(wb), {})
    assert r['passed'] is False, f"Partial should not pass: {r}"
    assert r['score'] > 0.0, f"Partial should have some score: {r}"
    print(f"  [PASS] Partial: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")

    # Full completion
    wb = Workbook()
    ws = wb.active
    ws.title = "Development Factors"
    ws['A1'] = "Chain-Ladder Reserve Analysis"
    ws['A2'] = "Loss Development Factors (Age-to-Age)"
    ws['A3'] = "development factor"
    ws['A4'] = "link ratio"
    ws['A5'] = "LDF"
    ws['A6'] = "loss development factor"
    # LDFs
    ldfs = [1.99, 1.45, 1.30, 1.18, 1.07, 1.035, 1.015, 1.010, 1.005]
    for i, ldf in enumerate(ldfs):
        ws.cell(row=7+i, column=1, value=f"Lag {i+1}->{i+2}")
        ws.cell(row=7+i, column=2, value=ldf)

    # Cumulative factors sheet
    ws2 = wb.create_sheet("Cumulative Factors")
    ws2['A1'] = "Cumulative Development Factors"
    ws2['A2'] = "CDF"
    ws2['A3'] = "cumulative"
    ws2['A4'] = "to ultimate"
    ws2['A5'] = "tail factor"
    cdfs = [5.19, 2.61, 1.80, 1.38, 1.17, 1.10, 1.06, 1.045, 1.035, 1.00]
    for i, cdf in enumerate(cdfs):
        ws2.cell(row=6+i, column=1, value=f"Lag {i+1}")
        ws2.cell(row=6+i, column=2, value=cdf)

    # Ultimate projections
    ws3 = wb.create_sheet("Ultimate Losses")
    ws3['A1'] = "Ultimate Loss Projections by Accident Year"
    ws3['A2'] = "ultimate"
    ws3['A3'] = "projected ultimate"
    ws3['A4'] = "estimated"
    ay_data = [
        (1988, 81094), (1989, 92303), (1990, 101527), (1991, 114831),
        (1992, 128664), (1993, 149284), (1994, 155191), (1995, 179647),
        (1996, 215678), (1997, 241697),
    ]
    for i, (year, ult) in enumerate(ay_data):
        ws3.cell(row=5+i, column=1, value=year)
        ws3.cell(row=5+i, column=2, value=ult)

    # IBNR sheet
    ws4 = wb.create_sheet("IBNR Reserves")
    ws4['A1'] = "IBNR Reserves by Accident Year"
    ws4['A2'] = "ibnr"
    ws4['A3'] = "incurred but not reported"
    ws4['A4'] = "reserve"
    ws4['A5'] = "outstanding"
    ibnr_data = [
        (1988, 0), (1989, 146), (1990, 103), (1991, 469),
        (1992, 575), (1993, 660), (1994, 495), (1995, 419),
        (1996, 1262), (1997, 195098),
    ]
    for i, (year, ibnr) in enumerate(ibnr_data):
        ws4.cell(row=6+i, column=1, value=year)
        ws4.cell(row=6+i, column=2, value=ibnr)
        ws4.cell(row=6+i, column=3, value="reserve")
    ws4['A18'] = "Total IBNR"
    ws4['B18'] = 199227

    # Professional structure
    ws5 = wb.create_sheet("Summary")
    ws5['A1'] = "Actuarial Reserve Summary"
    ws5['A2'] = "chain ladder method"
    ws5['A3'] = "accident year analysis"
    ws5['A4'] = "Schedule P"
    ws5['A5'] = "CAS"
    ws5['A6'] = "New Jersey"
    ws5['A7'] = "development"
    ws5['A8'] = "triangle"
    ws5['A9'] = "selection"
    ws5['A10'] = "paid loss"
    ws5['A11'] = "earned premium"

    r = mod.verify_reserve_analysis([], make_env_from_workbook(wb), {})
    assert r['passed'] is True, f"Full should pass: {r}"
    assert r['score'] >= 0.5, f"Full score too low: {r}"
    print(f"  [PASS] Full: passed={r['passed']}, score={r['score']:.2f}, fb={r['feedback'][:80]}...")
    print()


# ============================================================================
# Main
# ============================================================================
if __name__ == '__main__':
    passed = 0
    failed = 0
    errors = []

    tests = [
        ("epidemiological_outbreak_investigation", test_epidemiological),
        ("clinical_trial_adverse_event_analysis", test_clinical_trial),
        ("forensic_transaction_audit", test_forensic_audit),
        ("construction_project_cost_estimate", test_construction),
        ("reinsurance_treaty_portfolio_analysis", test_reinsurance),
    ]

    for name, test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            failed += 1
            errors.append((name, str(e)))
            print(f"  [FAIL] {name}: {e}")
            import traceback
            traceback.print_exc()
            print()

    print("=" * 60)
    print(f"RESULTS: {passed}/{len(tests)} tasks passed, {failed} failed")
    if errors:
        print("FAILURES:")
        for name, err in errors:
            print(f"  - {name}: {err}")
    print("=" * 60)

    sys.exit(0 if failed == 0 else 1)

#!/usr/bin/env python3
"""
Verifier for Airline OTP Analysis task.
Programmatically parses the expected XLSX report to verify analytical methodology.
"""

import sys
import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
    return numbers

def verify_otp_analysis(traj, env_info, task_info):
    """
    Verify the Airline OTP Analysis workbook.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # Read result.json
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    result = {}
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        logger.error(f"Failed to read result: {e}")
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)
            
    if not result.get("output_file_exists", False):
        return {
            "passed": False,
            "score": 0.0,
            "feedback": "Output file otp_analysis_report.xlsx was not found."
        }

    # Now load the actual XLSX using openpyxl
    container_path = "/home/ga/Documents/Spreadsheets/otp_analysis_report.xlsx"
    temp_wb_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    score = 0.0
    feedback_parts = []
    
    try:
        copy_from_env(container_path, temp_wb_file.name)
        
        import openpyxl
        wb = openpyxl.load_workbook(temp_wb_file.name, data_only=True)
        
        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        num_sheets = len(wb.sheetnames)
        
        # Structure check
        if num_sheets >= 2:
            score += 1.5
            feedback_parts.append(f"Professional structure: {num_sheets} sheets found (1.5/1.5)")
        else:
            feedback_parts.append("Single sheet found, expected multi-sheet (0.0/1.5)")
            
        # 1. Route-level OTP metrics
        airport_codes = ["sea", "pdx", "sfo", "boi", "geg", "mfr"]
        found_airports = sum(1 for ac in airport_codes if ac in all_text)
        if found_airports >= 4:
            score += 2.0
            feedback_parts.append(f"Route-level metrics: {found_airports} airports found (2.0/2.0)")
        elif found_airports >= 2:
            score += 1.0
            feedback_parts.append(f"Route-level metrics: Partial airports found (1.0/2.0)")
        else:
            feedback_parts.append("Route-level metrics missing (0.0/2.0)")
            
        # 2. Delay cause breakdown
        delay_cats = ["carrier", "weather", "nas", "security", "late"]
        found_delays = sum(1 for dc in delay_cats if dc in all_text)
        if found_delays >= 4:
            score += 1.5
            feedback_parts.append(f"Delay cause breakdown: {found_delays} categories found (1.5/1.5)")
        elif found_delays >= 2:
            score += 0.75
            feedback_parts.append(f"Delay cause breakdown: Partial categories found (0.75/1.5)")
        else:
            feedback_parts.append("Delay cause breakdown missing (0.0/1.5)")
            
        # 3. Worst route identified (SEA-BOI)
        if "boi" in all_text and ("worst" in all_text or "lowest" in all_text or "bottom" in all_text or "poor" in all_text):
            score += 2.0
            feedback_parts.append("Worst route (SEA-BOI) identified (2.0/2.0)")
        elif "boi" in all_text:
            score += 1.0
            feedback_parts.append("BOI route present but worst label not clearly identified (1.0/2.0)")
        else:
            feedback_parts.append("Worst route (SEA-BOI) not identified (0.0/2.0)")
            
        # 4. Monthly trend analysis
        months = ["july", "august", "september"]
        found_months = sum(1 for m in months if m in all_text)
        if found_months == 3:
            score += 1.5
            feedback_parts.append("Monthly trend analysis present (1.5/1.5)")
        elif found_months > 0:
            score += 0.75
            feedback_parts.append(f"Partial monthly trend analysis ({found_months} months) (0.75/1.5)")
        else:
            feedback_parts.append("Monthly trend analysis missing (0.0/1.5)")
            
        # 5. Fleet-wide KPIs (OTP ~ 78%, Cancel ~ 3.2%)
        found_otp = False
        found_cancel = False
        
        for n in all_numbers:
            # Check for OTP %
            if (70 <= n <= 85) or (0.70 <= n <= 0.85):
                found_otp = True
            # Check for Cancellation %
            if (1.5 <= n <= 5.0) or (0.015 <= n <= 0.05):
                found_cancel = True
                
        kpi_score = 0.0
        if found_otp: kpi_score += 0.75
        if found_cancel: kpi_score += 0.75
        
        score += kpi_score
        feedback_parts.append(f"Fleet-wide KPIs found: OTP={found_otp}, Cancel={found_cancel} ({kpi_score}/1.5)")
        
    except ImportError:
        logger.error("openpyxl is not installed or available.")
        return {"passed": False, "score": 0.0, "feedback": "Verifier missing openpyxl dependency"}
    except Exception as e:
        logger.error(f"Verification error: {e}")
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Error parsing workbook: {e}"
        }
    finally:
        if os.path.exists(temp_wb_file.name):
            os.unlink(temp_wb_file.name)

    passed = score >= 5.0
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
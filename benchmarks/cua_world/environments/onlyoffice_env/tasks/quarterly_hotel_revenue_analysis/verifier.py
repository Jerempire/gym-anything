#!/usr/bin/env python3
"""
Verifier for the Quarterly Hotel Revenue Performance Analysis task.

Evaluates the completed Excel workbook for hospitality KPI calculations,
data aggregation patterns, structural complexity, and workflow trajectory.
"""

import os
import json
import logging
import tempfile

# Safely import VLM components
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
    VLM_AVAILABLE = True
except ImportError:
    VLM_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_text_and_formulas(wb):
    """Extracts lowercase text and counts formulas."""
    all_text = []
    formula_count = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 300), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value)
                    all_text.append(val_str.lower())
                    if val_str.startswith('='):
                        formula_count += 1
                        
    return " ".join(all_text), formula_count

def extract_numeric_values(wb):
    """Extracts all raw numeric values for range checks."""
    numbers = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 300), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(float(cell.value))
    return numbers

def verify_hotel_revenue_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "System error: copy_from_env not available."}

    # ====================================================================
    # 1. Read task execution metadata
    # ====================================================================
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read execution metadata: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    score = 0
    feedback_parts = []
    
    # Anti-gaming: Ensure file was created during the task
    if not result.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "Output file hotel_revenue_analysis.xlsx not found."}
        
    if not result.get("file_created_during_task", False):
        feedback_parts.append("Warning: File timestamp indicates it may have existed prior to task start.")
    else:
        score += 15
        feedback_parts.append("File created/modified successfully (15/15).")

    # ====================================================================
    # 2. Parse the Workbook
    # ====================================================================
    container_xlsx_path = "/home/ga/Documents/Spreadsheets/hotel_revenue_analysis.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(container_xlsx_path, temp_xlsx.name)
        import openpyxl
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        wb_data = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
    except Exception as e:
        return {"passed": False, "score": score, "feedback": f"Failed to parse workbook: {e}"}
    
    all_text, formula_count = extract_all_text_and_formulas(wb)
    all_numbers = extract_numeric_values(wb_data)
    
    # Check Formulas
    if formula_count >= 10:
        score += 15
        feedback_parts.append(f"Formulas utilized effectively ({formula_count} found) (15/15).")
    elif formula_count > 0:
        score += 7
        feedback_parts.append(f"Some formulas used ({formula_count} found) (7/15).")
    else:
        feedback_parts.append("No Excel formulas found. Data was hardcoded (0/15).")

    # ====================================================================
    # 3. Assess Content & Calculations
    # ====================================================================
    # Check A: Monthly KPIs (20 pts)
    months_found = sum(1 for m in ["july", "august", "september", "jul", "aug", "sep"] if m in all_text)
    kpis_found = sum(1 for k in ["occupancy", "adr", "revpar", "occ"] if k in all_text)
    
    if months_found >= 3 and kpis_found >= 3:
        score += 20
        feedback_parts.append("Monthly KPIs (Occ, ADR, RevPAR) identified (20/20).")
    elif months_found >= 1 and kpis_found >= 1:
        score += 10
        feedback_parts.append("Partial Monthly KPIs identified (10/20).")
        
    # Check B: Analytical Dimensions (20 pts)
    # Channels
    channels_found = sum(1 for c in ["bar", "corporate", "ota", "group", "package"] if c in all_text)
    # Day of week
    dow_found = sum(1 for d in ["weekend", "weekday", "monday", "friday"] if d in all_text)
    # Room types
    rooms_found = sum(1 for r in ["king", "double", "deluxe", "suite"] if r in all_text)
    
    dimension_score = 0
    if channels_found >= 3: dimension_score += 8
    if dow_found >= 2: dimension_score += 6
    if rooms_found >= 2: dimension_score += 6
    
    score += dimension_score
    feedback_parts.append(f"Analytical dimensions assessed: {dimension_score}/20 points.")

    # Check C: Structure & Executive Summary (10 pts)
    if len(wb.sheetnames) >= 3:
        score += 10
        feedback_parts.append(f"Excellent multi-sheet structure ({len(wb.sheetnames)} sheets) (10/10).")
    elif len(wb.sheetnames) == 2 or "summary" in all_text:
        score += 5
        feedback_parts.append("Basic sheet structure/summary present (5/10).")

    # Cleanup openpyxl
    wb.close()
    wb_data.close()
    if os.path.exists(temp_xlsx.name):
        os.unlink(temp_xlsx.name)

    # ====================================================================
    # 4. Trajectory VLM Verification (20 pts) - Anti-Spoofing
    # ====================================================================
    vlm_score = 0
    if VLM_AVAILABLE:
        try:
            frames = sample_trajectory_frames(traj, n=4)
            final_img = get_final_screenshot(traj)
            
            prompt = """
            You are auditing a workflow. The agent was tasked with analyzing hotel revenue data in ONLYOFFICE Spreadsheet Editor.
            Review these trajectory frames and determine:
            1. Did the agent actively navigate and manipulate spreadsheet data?
            2. Are there charts, pivot tables, or aggregated tables visible indicating analysis (not just raw CSV data)?
            
            Respond in JSON:
            {"active_workflow_observed": true/false, "analysis_created": true/false}
            """
            
            vlm_res = query_vlm(images=frames + [final_img], prompt=prompt)
            if vlm_res.get("success"):
                parsed = vlm_res.get("parsed", {})
                if parsed.get("active_workflow_observed"): vlm_score += 10
                if parsed.get("analysis_created"): vlm_score += 10
                feedback_parts.append(f"VLM trajectory verification: {vlm_score}/20 points.")
            else:
                feedback_parts.append("VLM verification failed. Granting partial credit (10/20).")
                vlm_score += 10
        except Exception as e:
            logger.error(f"VLM check failed: {e}")
            vlm_score += 10 # Grace score on framework error
    else:
        logger.info("VLM not available, granting pass-through credit for trajectory.")
        vlm_score += 20
        
    score += vlm_score

    # ====================================================================
    # 5. Final Evaluation
    # ====================================================================
    key_criteria_met = (result.get("file_created_during_task", False) and formula_count > 0)
    passed = score >= 60 and key_criteria_met

    return {
        "passed": bool(passed),
        "score": int(score),
        "feedback": " | ".join(feedback_parts)
    }
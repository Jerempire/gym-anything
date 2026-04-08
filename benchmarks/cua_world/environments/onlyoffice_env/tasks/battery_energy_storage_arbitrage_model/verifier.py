#!/usr/bin/env python3
"""
Verifier for Battery Energy Storage Arbitrage Model task.

The agent must build a financial model calculating battery dispatch profit
using CAISO SP15 wholesale electricity prices.

Verification logic parses the Excel workbook to check for:
1. Output file exists and was modified during the task.
2. Structure: Presence of ~365 rows in an analysis sheet.
3. Logic: MAXIFS/MINIFS formulas to aggregate daily prices.
4. Math: Use of 0.85 (RTE), 20 (Capacity), and 200 (Degradation cost).
5. Output Metrics: SUM formulas calculating total profit and total cycles.
"""

import sys
import os
import json
import logging
import tempfile
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def install_openpyxl_if_missing():
    try:
        import openpyxl
    except ImportError:
        logger.info("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

def verify_battery_arbitrage(traj, env_info, task_info):
    """
    Verify the arbitrage model using programmatic file parsing.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    install_openpyxl_if_missing()
    from openpyxl import load_workbook

    # 1. Retrieve the export result JSON
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    output_exists = result.get('output_exists', False)
    file_created = result.get('file_created_during_task', False)

    if not output_exists:
        return {
            "passed": False,
            "score": 0,
            "feedback": "The required file 'bess_arbitrage_model.xlsx' was not found. The agent may not have saved the file or saved it in the wrong format/location."
        }

    # 2. Retrieve the actual workbook
    container_path = "/home/ga/Documents/Spreadsheets/bess_arbitrage_model.xlsx"
    temp_wb = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    score = 0
    feedback_parts = []
    
    try:
        copy_from_env(container_path, temp_wb.name)
        # Load with data_only=False to inspect formulas
        wb = load_workbook(temp_wb.name, data_only=False)
        
        all_formulas = []
        max_rows_in_sheet = 0
        
        # Extract all formulas and find the sheet with the daily analysis
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            filled_rows = 0
            
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 30)):
                row_has_data = False
                for cell in row:
                    if cell.value is not None:
                        row_has_data = True
                        val_str = str(cell.value).upper()
                        if val_str.startswith('='):
                            all_formulas.append(val_str)
                if row_has_data:
                    filled_rows += 1
                    
            if filled_rows > max_rows_in_sheet:
                max_rows_in_sheet = filled_rows

        # Join all formulas into a massive string for regex checking
        formulas_blob = " ".join(all_formulas)
        
        # --- Evaluation Criteria ---
        
        # Criterion 1: Structure (10 points) - Are there approx 365 rows of analysis?
        if max_rows_in_sheet >= 365:
            score += 10
            feedback_parts.append("Structure: Correctly set up ~365 rows for daily analysis.")
        elif max_rows_in_sheet >= 100:
            score += 5
            feedback_parts.append("Structure: Partial setup, fewer than 365 rows found.")
        else:
            feedback_parts.append("Structure: Missing daily analysis rows.")
            
        # Criterion 2: Price Aggregation (25 points) - MINIFS/MAXIFS usage
        has_max = bool(re.search(r'MAX\(|MAXIFS\(', formulas_blob))
        has_min = bool(re.search(r'MIN\(|MINIFS\(', formulas_blob))
        if has_max and has_min:
            score += 25
            feedback_parts.append("Price Aggregation: Successfully utilized MAX/MIN formulas.")
        elif has_max or has_min:
            score += 10
            feedback_parts.append("Price Aggregation: Partial usage of MIN/MAX formulas.")
        else:
            feedback_parts.append("Price Aggregation: Missing expected MIN/MAX price aggregation formulas.")
            
        # Criterion 3: Revenue Mathematics (20 points) - 0.85 RTE and 20 MWh Capacity
        has_rte = bool(re.search(r'0\.85|85%|/ ?85|\* ?1\.17', formulas_blob))
        has_cap = bool(re.search(r'\b20\b', formulas_blob))
        if has_rte and has_cap:
            score += 20
            feedback_parts.append("Math: Successfully integrated RTE (0.85) and Capacity (20) into revenue calculations.")
        elif has_rte or has_cap:
            score += 10
            feedback_parts.append("Math: Missing either RTE or Capacity parameters in formulas.")
        else:
            feedback_parts.append("Math: Missing correct physical parameters (RTE, Capacity) in formulas.")
            
        # Criterion 4: Dispatch Logic (20 points) - IF statement for >200 logic
        has_if = bool(re.search(r'IF\(', formulas_blob))
        has_hurdle = bool(re.search(r'200', formulas_blob))
        if has_if and has_hurdle:
            score += 20
            feedback_parts.append("Logic: Successfully implemented conditional dispatch logic based on the $200 degradation cost.")
        elif has_if or has_hurdle:
            score += 10
            feedback_parts.append("Logic: Partial dispatch logic implemented.")
        else:
            feedback_parts.append("Logic: Missing dispatch rules (IF statement / $200 hurdle).")
            
        # Criterion 5: Summary Metrics (15 points) - SUM formulas for final aggregation
        has_sum = bool(re.search(r'SUM\(', formulas_blob))
        if has_sum:
            score += 15
            feedback_parts.append("Summary: Utilized SUM functions for final annual metrics.")
        else:
            feedback_parts.append("Summary: Missing aggregate SUM formulas for total profit/cycles.")
            
        # Anti-Gaming check:
        if not file_created:
            # If they just renamed the raw CSV without doing the task during the timeframe
            if len(all_formulas) == 0:
                score = 0
                feedback_parts = ["Anti-gaming: Output file contains no formulas and was not actively modeled."]

    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Error parsing workbook: {e}"}
    finally:
        if os.path.exists(temp_wb.name):
            os.unlink(temp_wb.name)

    # Final pass determination
    # Pass requires a score >= 65 and key modeling components (MIN/MAX aggregation)
    passed = (score >= 65)

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
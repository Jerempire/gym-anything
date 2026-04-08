#!/usr/bin/env python3
"""
Verifier for Earthquake Seismicity Analysis task.

Evaluates the agent's ability to:
1. Import a CSV dataset of earthquake records
2. Perform a Gutenberg-Richter Magnitude-Frequency analysis (b-value)
3. Group data by depth (Depth Distribution)
4. Group data by time (Temporal clustering)
5. Create summary statistics
"""

import sys
import os
import json
import logging
import tempfile
import re

# Ensure we can import the VLM utilities from the framework
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
    VLM_AVAILABLE = True
except ImportError:
    VLM_AVAILABLE = False
    print("Warning: VLM utilities not found, visual verification will be skipped.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)


def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    numbers.append(cell.value)
    return numbers


def extract_formulas(wb, copy_from_env, container_path):
    """Check if formulas are present by reading with data_only=False."""
    formula_count = 0
    try:
        from openpyxl import load_workbook
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        copy_from_env(container_path, temp_file.name)
        wb_f = load_workbook(temp_file.name, data_only=False)
        for sn in wb_f.sheetnames:
            sheet = wb_f[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000),
                                        max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
        wb_f.close()
        os.unlink(temp_file.name)
    except Exception as e:
        logger.warning(f"Could not check formulas: {e}")
    return formula_count


def verify_seismicity_analysis(traj, env_info, task_info):
    """
    Verify the seismicity analysis workbook.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # Load export JSON
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/earthquake_seismicity_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            export_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to load export data: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not export_data.get("output_file_exists", False):
        return {"passed": False, "score": 0.0, "feedback": "Target file seismicity_analysis.xlsx was not saved."}
        
    # Check Anti-gaming (File must be created/modified after task started)
    start_ts = export_data.get("task_start_ts", 0)
    mtime = export_data.get("output_file_mtime", 0)
    if mtime > 0 and start_ts > 0 and mtime < start_ts:
        return {"passed": False, "score": 0.0, "feedback": "Anti-gaming check failed: File was modified before the task started."}

    container_path = "/home/ga/Documents/Spreadsheets/seismicity_analysis.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(container_path, temp_xlsx.name)
        
        # Load workbook via openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(temp_xlsx.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0.0, "feedback": f"Failed to parse XLSX file: {e}"}

        feedback_parts = []
        score = 0.0

        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        num_sheets = len(wb.sheetnames)

        # Count total cells and rows of data
        total_cells = 0
        data_rows = set()
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row_idx, row in enumerate(sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 30))):
                row_has_data = False
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1
                        row_has_data = True
                if row_has_data:
                    data_rows.add(f"{sn}_{row_idx}")

        # Gate check: Needs substantial content
        if total_cells < 20 or len(data_rows) < 10:
            return {"passed": False, "score": 0.0, "feedback": "File has insufficient content (<20 cells or <10 rows)."}

        # Check 1: Data import completeness (1.5 pts)
        has_time = "time" in all_text or "date" in all_text
        has_mag = "mag" in all_text or "magnitude" in all_text
        has_depth = "depth" in all_text
        
        if has_mag and has_depth and len(data_rows) >= 15:
            score += 1.5
            feedback_parts.append("Data import successful")
        else:
            feedback_parts.append("Missing primary data columns or insufficient rows")

        # Check 2: Magnitude-frequency / Gutenberg-Richter (2.0 pts)
        gr_terms = ["cumulative", "frequency", "log", "log10", "gutenberg", "richter"]
        has_gr_concept = sum(1 for term in gr_terms if term in all_text) >= 2
        if has_gr_concept:
            score += 2.0
            feedback_parts.append("Magnitude-frequency analysis present")
        else:
            feedback_parts.append("Missing magnitude-frequency analysis (cumulative/log)")

        # Check 3: b-value estimation (2.0 pts)
        b_terms = ["b-value", "b value", "b =", "b = "]
        has_b_term = any(term in all_text for term in b_terms)
        
        # Look for a number in the plausible tectonic range (0.5 to 1.5)
        has_plausible_b = any(0.5 <= n <= 1.5 for n in all_numbers)
        
        if has_b_term and has_plausible_b:
            score += 2.0
            feedback_parts.append("b-value calculated in plausible range")
        elif has_b_term:
            score += 1.0
            feedback_parts.append("b-value mentioned but numeric result missing/out-of-bounds")
        else:
            feedback_parts.append("Missing b-value estimation")

        # Check 4: Depth distribution analysis (1.5 pts)
        depth_bins = ["0-5", "5-10", "10-15", ">10", ">20", "shallow", "deep"]
        has_depth_bins = any(bin_str in all_text for bin_str in depth_bins)
        has_mean_median = "mean" in all_text or "median" in all_text or "average" in all_text
        
        if has_depth_bins or (has_depth and has_mean_median):
            score += 1.5
            feedback_parts.append("Depth analysis present")
        else:
            feedback_parts.append("Missing depth distribution analysis")

        # Check 5: Temporal analysis (1.5 pts)
        temporal_terms = ["jan", "feb", "mar", "apr", "may", "jun", "month", "week", "january", "february"]
        has_temporal = any(term in all_text for term in temporal_terms)
        if has_temporal:
            score += 1.5
            feedback_parts.append("Temporal analysis present")
        else:
            feedback_parts.append("Missing temporal analysis")

        # Check 6: Summary statistics & Structure (1.5 pts)
        summary_terms = ["total", "max", "largest", "summary", "count"]
        has_summary = sum(1 for term in summary_terms if term in all_text) >= 2
        if has_summary and num_sheets >= 2:
            score += 1.5
            feedback_parts.append("Multi-sheet summary present")
        elif has_summary:
            score += 1.0
            feedback_parts.append("Summary present (single sheet)")
        else:
            feedback_parts.append("Missing overall summary")

        # VLM Trajectory Verification (Optional bonus/confirmation)
        if VLM_AVAILABLE:
            try:
                frames = sample_trajectory_frames(traj, n=3)
                final = get_final_screenshot(traj)
                if frames and final:
                    prompt = (
                        "Look at these screenshots of a spreadsheet application. "
                        "Did the user perform an analysis on earthquake data? "
                        "Look for charts, tables with magnitude bins, depth calculations, or 'b-value'. "
                        "Return JSON with {'performed_analysis': true/false, 'confidence': 'high/medium/low'}."
                    )
                    vlm_result = query_vlm(images=frames + [final], prompt=prompt)
                    if vlm_result.get("success") and vlm_result.get("parsed", {}).get("performed_analysis"):
                        logger.info("VLM confirmed earthquake data analysis workflow.")
                        # Could use this to validate borderline scores, but programmatic rules primarily apply
            except Exception as e:
                logger.warning(f"VLM verification skipped/failed: {e}")

        # Check if formulas were used
        formula_count = extract_formulas(wb, copy_from_env, container_path)
        if formula_count == 0:
            feedback_parts.append("Warning: No formulas detected, only pasted values")
            score = min(score, 7.5) # Cap score if no formulas are used

        passed = score >= 5.0

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)
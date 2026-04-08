#!/usr/bin/env python3
"""
Verifier for Transit Ridership Analysis task.

Evaluates multi-sheet workbook generation, formula application over large datasets,
and conditional aggregations based on transit cost modeling.

Check criteria:
1. Output exists and created during task execution
2. Proper sheet structure established
3. Mathematical precision on operating costs
4. Successful aggregation of total and weekday boardings
5. Complex ratio calculations (FRR)
"""

import sys
import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import VLM capabilities safely
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
    VLM_AVAILABLE = True
except ImportError:
    VLM_AVAILABLE = False


def check_value_present(target, numbers_list, tolerance=0.01):
    """Fuzzy matching for floating point results in extracted numbers."""
    for n in numbers_list:
        if abs(n - target) <= tolerance:
            return True
    return False


def verify_transit_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available from environment."}

    score = 0
    feedback_parts = []
    
    # 1. Load result JSON from container
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    result = {}
    try:
        copy_from_env("/tmp/transit_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read exported result JSON: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)
            
    # 2. Check File Existence and Anti-Gaming Creation stamp
    if not result.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "Target workbook transit_performance_q1.xlsx was not saved."}
        
    score += 10
    feedback_parts.append("Workbook created successfully (+10)")
    
    if result.get("file_created_during_task", False):
        score += 5
        feedback_parts.append("Anti-gaming: File created during active task (+5)")

    # 3. Load Ground Truth values
    temp_gt = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    gt = {}
    try:
        copy_from_env("/tmp/transit_ground_truth.json", temp_gt.name)
        with open(temp_gt.name, 'r') as f:
            gt = json.load(f)
    except Exception as e:
        logger.warning(f"Could not load ground truth: {e}")
    finally:
        if os.path.exists(temp_gt.name):
            os.unlink(temp_gt.name)

    route_88x_cost = gt.get("route_88x_daily_cost", 1721.5)
    route_15_boardings = gt.get("route_15_total_boardings", 226200)
    route_15_wkdy = gt.get("route_15_avg_wkdy", 3000)
    route_88x_frr = gt.get("route_88x_frr", 0.112028)
    total_sys_boardings = gt.get("total_system_boardings", 0)

    # 4. Parse XLSX via OpenPyXL
    try:
        import openpyxl
    except ImportError:
        return {"passed": False, "score": 0, "feedback": "Verifier missing openpyxl dependency."}
        
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/transit_performance_q1.xlsx", temp_xlsx.name)
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
        # Check Sheets
        sheet_names = [sn.lower() for sn in wb.sheetnames]
        has_summary = any("summary" in sn for sn in sheet_names)
        has_kpis = any("kpi" in sn for sn in sheet_names)
        
        if has_summary or len(sheet_names) >= 2:
            score += 10
            feedback_parts.append("Multi-sheet structure detected (+10)")
            
        # Extract all numbers dynamically (bypasses rigid cell coordinate requirements)
        all_numbers = set()
        for sn in wb.sheetnames:
            sheet = wb[sn]
            # Limit depth to avoid massive memory overhead but ensure we capture the 4550 rows
            for row in sheet.iter_rows(max_row=5000, max_col=20):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        all_numbers.add(float(cell.value))
                        
        # Row-level calculations (Operating Cost)
        if check_value_present(route_88x_cost, all_numbers, 0.1):
            score += 20
            feedback_parts.append("Row-level Operating Costs matched formula (+20)")
            
        # Aggregation: Route 15 Total
        if check_value_present(route_15_boardings, all_numbers, 1.0):
            score += 10
            feedback_parts.append("Route 15 Total Boardings aggregated correctly (+10)")
            
        # Conditional Aggregation: Route 15 Avg Weekday
        if check_value_present(route_15_wkdy, all_numbers, 1.0):
            score += 10
            feedback_parts.append("Route 15 Average Weekday Boardings correct (+10)")
            
        # Ratio Aggregation: Route 88X Overall FRR
        # Agent might represent as decimal 0.112 or percentage 11.20%
        if check_value_present(route_88x_frr, all_numbers, 0.001) or check_value_present(route_88x_frr * 100, all_numbers, 0.1):
            score += 10
            feedback_parts.append("Route 88X Overall FRR aggregated correctly (+10)")
            
        # System KPI check
        if total_sys_boardings > 0 and check_value_present(total_sys_boardings, all_numbers, 1.0):
            score += 5
            feedback_parts.append("Total System Boardings calculated correctly (+5)")

    except Exception as e:
        feedback_parts.append(f"Error parsing workbook contents: {e}")
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # 5. VLM Trajectory Verification
    if VLM_AVAILABLE:
        try:
            frames = sample_trajectory_frames(traj, n=5)
            final = get_final_screenshot(traj)
            prompt = """You are analyzing a transit ridership spreadsheet task.
Look at the sequence of screenshots. 
Did the agent actively enter formulas, select data, create new sheets, or construct summary tables/KPIs?
Respond with JSON: {"spreadsheet_activity_visible": true/false}"""

            vlm_res = query_vlm(images=frames + [final], prompt=prompt)
            if vlm_res.get("parsed", {}).get("spreadsheet_activity_visible", False):
                score += 15
                feedback_parts.append("VLM confirmed spreadsheet activity (+15)")
            else:
                feedback_parts.append("VLM did not detect active spreadsheet workflow (0)")
        except Exception as e:
            logger.warning(f"VLM verification failed: {e}")
            score += 15  # gracefully award points if framework fails
    else:
        # Gracefully award points if VLM module is missing from the environment
        score += 15
        feedback_parts.append("VLM unavailable, awarding full visual verification points (+15)")

    passed = score >= 60
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
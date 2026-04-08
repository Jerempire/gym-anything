#!/usr/bin/env python3
"""
Verifier for E-Commerce RFM Customer Segmentation task.

Performs robust programmatic verification against the ground truth dataset:
1. Re-computes exactly what the RFM metrics should be based on the source CSV.
2. Extracts evaluated values from the agent's XLSX output using openpyxl (data_only=True).
3. Evaluates structural components (sheets, LineTotal calculations).
4. Cross-references 20 customers' metrics (R, F, M, Scores, Segment) directly.
5. Verifies aggregate summary counts.
"""

import os
import json
import tempfile
import csv
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def compute_ground_truth(csv_path):
    """Parses raw CSV and calculates the absolute ground truth for RFM rules."""
    gt = {}
    ref_date = datetime(2011, 12, 1)
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            cid = row['CustomerID'].strip()
            if not cid: continue
            
            qty = int(row['Quantity'])
            price = float(row['UnitPrice'])
            date_str = row['InvoiceDate']
            inv_date = datetime.strptime(date_str, '%Y-%m-%d')
            line_total = qty * price
            
            if cid not in gt:
                gt[cid] = {'latest_date': inv_date, 'F': 0, 'M': 0.0}
            
            if inv_date > gt[cid]['latest_date']:
                gt[cid]['latest_date'] = inv_date
                
            gt[cid]['F'] += 1
            gt[cid]['M'] += line_total
            
    summary_counts = {"champions": 0, "loyal customers": 0, "at risk": 0, "lost": 0}
            
    for cid, data in gt.items():
        r_days = (ref_date - data['latest_date']).days
        r_days = max(0, r_days)
        data['R'] = r_days
        
        # Scoring Logic from prompt
        if r_days <= 15: r_score = 4
        elif r_days <= 30: r_score = 3
        elif r_days <= 60: r_score = 2
        else: r_score = 1
        
        f = data['F']
        if f >= 15: f_score = 4
        elif f >= 6: f_score = 3
        elif f >= 2: f_score = 2
        else: f_score = 1
        
        m = data['M']
        if m >= 500: m_score = 4
        elif m >= 150: m_score = 3
        elif m >= 50: m_score = 2
        else: m_score = 1
        
        data['R_Score'] = r_score
        data['F_Score'] = f_score
        data['M_Score'] = m_score
        
        rfm_sum = r_score + f_score + m_score
        data['RFM_Sum'] = rfm_sum
        
        if rfm_sum >= 11: seg = "champions"
        elif rfm_sum >= 8: seg = "loyal customers"
        elif rfm_sum >= 5: seg = "at risk"
        else: seg = "lost"
        
        data['Segment'] = seg
        summary_counts[seg] += 1
        
    return gt, summary_counts

def normalize_hdr(s):
    if not s: return ""
    return str(s).lower().strip().replace('_', '').replace(' ', '')

def find_sheet_by_name(wb, name_hints):
    for sn in wb.sheetnames:
        if any(h.lower() in sn.lower() for h in name_hints):
            return wb[sn]
    return None

def verify_rfm_segmentation(traj, env_info, task_info):
    """
    Evaluates the agent's resulting XLSX against mathematical ground truth.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    temp_dir = tempfile.mkdtemp()
    result_json_path = os.path.join(temp_dir, 'task_result.json')
    xlsx_path = os.path.join(temp_dir, 'rfm_segmentation.xlsx')
    csv_path = os.path.join(temp_dir, 'online_retail_q4.csv')
    
    try:
        copy_from_env("/tmp/task_result.json", result_json_path)
        with open(result_json_path, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}

    # Anti-gaming check
    if not result.get('file_modified_during_task', False):
        return {"passed": False, "score": 0, "feedback": "Anti-gaming failure: target file was not created or modified during the task execution."}
        
    output_exists = result.get('output_exists', False)
    if not output_exists:
        return {"passed": False, "score": 0, "feedback": "rfm_segmentation.xlsx was not saved or could not be found."}
        
    score = 10
    feedback_parts = ["Workbook saved"]
    
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/rfm_segmentation.xlsx", xlsx_path)
        copy_from_env("/home/ga/Documents/Spreadsheets/online_retail_q4.csv", csv_path)
    except Exception as e:
        return {"passed": False, "score": score, "feedback": f"Error copying files from environment: {e}"}

    # Generate Ground Truth
    gt, summary_counts = compute_ground_truth(csv_path)

    try:
        import openpyxl
        # Use data_only=True so formula outputs (cached by ONLYOFFICE) are readable
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {"passed": False, "score": score, "feedback": f"Failed to parse Excel file: {e}"}
        
    # Check LineTotal in Raw Data (Usually Sheet 1)
    raw_sheet = wb.worksheets[0]
    headers_raw = {}
    for row in raw_sheet.iter_rows(min_row=1, max_row=5):
        h = {normalize_hdr(c.value): idx for idx, c in enumerate(row) if c.value}
        if 'quantity' in h and 'unitprice' in h:
            headers_raw = h
            break

    if 'linetotal' in headers_raw or 'total' in headers_raw:
        col_idx = headers_raw.get('linetotal', headers_raw.get('total'))
        qty_idx = headers_raw.get('quantity')
        price_idx = headers_raw.get('unitprice')
        
        valid_linetotals = 0
        for row in raw_sheet.iter_rows(min_row=1, max_row=20):
            if str(row[qty_idx].value).lower() == 'quantity': continue
            try:
                q = float(row[qty_idx].value)
                p = float(row[price_idx].value)
                lt = float(row[col_idx].value)
                if abs(q*p - lt) < 0.01:
                    valid_linetotals += 1
            except:
                pass
                
        if valid_linetotals >= 5:
            score += 10
            feedback_parts.append("LineTotal calculated correctly")
        else:
            feedback_parts.append("LineTotal present but calculations are missing or incorrect")
    else:
        feedback_parts.append("LineTotal column not found in raw data")

    # Check RFM_Analysis sheet
    rfm_sheet = find_sheet_by_name(wb, ['rfm', 'analysis', 'segmentation'])
    if not rfm_sheet:
        feedback_parts.append("RFM_Analysis sheet not found")
        return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}
        
    headers_rfm = {}
    for row in rfm_sheet.iter_rows(min_row=1, max_row=5):
        h = {normalize_hdr(c.value): idx for idx, c in enumerate(row) if c.value}
        if 'customerid' in h or 'id' in h or 'customer' in h:
            headers_rfm = h
            break
    
    cid_keys = ['customerid', 'customer', 'id']
    cid_idx = next((headers_rfm[k] for k in cid_keys if k in headers_rfm), None)
    
    if cid_idx is not None:
        agent_customers = 0
        correct_base_metrics = 0
        correct_scores = 0
        correct_segments = 0
        eval_count = 0
        
        for row in rfm_sheet.iter_rows(min_row=1, max_row=500):
            if not row[cid_idx].value: continue
            cid_val = str(row[cid_idx].value).strip()
            
            if cid_val.lower() in ['customerid', 'customer', 'id']:
                continue
                
            if cid_val.endswith('.0'): cid_val = cid_val[:-2]
            
            agent_customers += 1
            
            if cid_val in gt and eval_count < 20:
                eval_count += 1
                gt_data = gt[cid_val]
                
                # Check Base Metrics (R, F, M)
                try:
                    r_val = float(row[headers_rfm.get('recency', headers_rfm.get('r'))].value)
                    f_val = float(row[headers_rfm.get('frequency', headers_rfm.get('f'))].value)
                    m_val = float(row[headers_rfm.get('monetary', headers_rfm.get('m'))].value)
                    # Use generous tolerance for potential timezone/date math nuances
                    if abs(r_val - gt_data['R']) <= 2 and abs(f_val - gt_data['F']) <= 1 and abs(m_val - gt_data['M']) <= 5.0:
                        correct_base_metrics += 1
                except:
                    pass
                    
                # Check Scores
                try:
                    rs_val = float(row[headers_rfm.get('rscore')].value)
                    fs_val = float(row[headers_rfm.get('fscore')].value)
                    ms_val = float(row[headers_rfm.get('mscore')].value)
                    if rs_val == gt_data['R_Score'] and fs_val == gt_data['F_Score'] and ms_val == gt_data['M_Score']:
                        correct_scores += 1
                except:
                    pass
                    
                # Check Segments
                try:
                    seg_val = str(row[headers_rfm.get('segment')].value).lower().strip()
                    if seg_val == gt_data['Segment']:
                        correct_segments += 1
                except:
                    pass
                    
        # Verify Unique Customers count (Expect around 380)
        if abs(agent_customers - len(gt)) <= 15:
            score += 10
            feedback_parts.append("Correct unique customer count")
        else:
            feedback_parts.append(f"Customer count mismatch: found {agent_customers}, expected {len(gt)}")
            
        # Verify correctness percentages
        if eval_count > 0:
            bm_ratio = correct_base_metrics / eval_count
            score += int(25 * bm_ratio)
            if bm_ratio > 0.8: feedback_parts.append("RFM base metrics mostly correct")
            else: feedback_parts.append(f"RFM base metrics incorrect or missing (ratio: {bm_ratio:.2f})")
            
            score_ratio = correct_scores / eval_count
            score += int(20 * score_ratio)
            if score_ratio > 0.8: feedback_parts.append("RFM scores mostly correct")
            
            seg_ratio = correct_segments / eval_count
            score += int(15 * seg_ratio)
            if seg_ratio > 0.8: feedback_parts.append("Customer segments mostly correct")
            
    else:
        feedback_parts.append("CustomerID column not found in RFM sheet")
        
    # Check Summary Sheet
    summary_sheet = find_sheet_by_name(wb, ['summary', 'sum'])
    if summary_sheet:
        agent_summary = {}
        for row in summary_sheet.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.lower().strip()
                    if val in summary_counts:
                        try:
                            # Usually count is right next to the segment string
                            count_val = summary_sheet.cell(row=cell.row, column=cell.column+1).value
                            if count_val is not None:
                                agent_summary[val] = float(count_val)
                        except:
                            pass
                            
        summary_matches = 0
        for seg, count in summary_counts.items():
            if seg in agent_summary and abs(agent_summary[seg] - count) <= 5:
                summary_matches += 1
                
        if summary_matches == 4:
            score += 10
            feedback_parts.append("Summary sheet counts strongly match ground truth")
        elif summary_matches > 0:
            score += 5
            feedback_parts.append(f"Summary sheet partial match ({summary_matches}/4)")
        else:
            feedback_parts.append("Summary sheet counts incorrect or could not be parsed")
    else:
        feedback_parts.append("Summary sheet not found")
        
    passed = score >= 65
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
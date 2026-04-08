#!/usr/bin/env python3
"""
Verifier for Space Debris Conjunction Assessment Task.
Scores the agent's ability to compute advanced formulas, logically categorize
risk thresholds using nested conditions, and separate high-risk data.
"""

import json
import os
import logging
import tempfile
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_conjunction_assessment(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_path = metadata.get('expected_output_path', '/home/ga/Documents/Spreadsheets/conjunction_assessment.xlsx')

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Fetch task result JSON to verify anti-gaming constraints
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export JSON: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    output_exists = export_result.get('output_exists', False)
    if not output_exists:
        return {
            "passed": False, 
            "score": 0, 
            "feedback": "conjunction_assessment.xlsx was not saved or found."
        }

    # 2. Fetch the actual XLSX workbook
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env(expected_path, temp_xlsx.name)
        import openpyxl
        # Load twice: once to read values, once to read formulas (if needed)
        wb_vals = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        wb_forms = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse XLSX file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    sheet_names = [sn.strip().lower() for sn in wb_vals.sheetnames]
    
    # CRITERION 1: File & Sheets (15 points)
    has_data = "data" in sheet_names
    has_briefing = "maneuver briefing" in sheet_names
    has_summary = "summary" in sheet_names

    if has_data and has_briefing and has_summary:
        score += 15
        feedback_parts.append("All 3 sheets created correctly")
    elif len(sheet_names) >= 3:
        score += 8
        feedback_parts.append("Multiple sheets created but names do not exactly match instructions")
    else:
        feedback_parts.append("Missing required sheets")

    # Locate Data Sheet
    data_sheet = None
    for sn in wb_vals.sheetnames:
        if sn.strip().lower() == "data":
            data_sheet = wb_vals[sn]
            break
    if not data_sheet:
        data_sheet = wb_vals.active  # Fallback

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in data_sheet[1]]
    
    # Helper to find column index (0-based)
    def find_col(possible_names):
        for idx, h in enumerate(headers):
            if any(p in h for p in possible_names):
                return idx
        return -1

    idx_range_km = find_col(['min_range_km'])
    idx_vel = find_col(['rel_velocity_km_s'])
    idx_prob = find_col(['max_prob'])
    
    idx_range_m = find_col(['min_range_m'])
    idx_energy = find_col(['collision_energy', 'energy'])
    idx_risk = find_col(['risk'])

    # CRITERION 2 & 3 & 4: Range Conv, Kinetic Energy, Risk Category
    correct_range_m = 0
    correct_energy = 0
    correct_risk = 0
    checked_rows = 0

    if idx_range_km >= 0 and idx_vel >= 0 and idx_prob >= 0:
        for row in data_sheet.iter_rows(min_row=2, max_row=min(data_sheet.max_row, 50)):
            if not row[idx_range_km].value:
                continue
            
            try:
                km_val = float(row[idx_range_km].value)
                vel_val = float(row[idx_vel].value)
                prob_val = float(row[idx_prob].value)
                checked_rows += 1

                # Range (m) = km * 1000  (10 points)
                if idx_range_m >= 0 and row[idx_range_m].value is not None:
                    try:
                        m_val = float(row[idx_range_m].value)
                        if math.isclose(m_val, km_val * 1000, rel_tol=0.01):
                            correct_range_m += 1
                    except ValueError:
                        pass

                # Energy = 0.5 * 500 * (V*1000)^2 / 1000000000 = 0.25 * V^2  (20 points)
                expected_energy = 0.25 * (vel_val ** 2)
                if idx_energy >= 0 and row[idx_energy].value is not None:
                    try:
                        e_val = float(row[idx_energy].value)
                        if math.isclose(e_val, expected_energy, rel_tol=0.05):
                            correct_energy += 1
                    except ValueError:
                        pass

                # Risk Logic (20 points)
                expected_risk = "LOW"
                if prob_val >= 0.0001 and (km_val * 1000) <= 1000:
                    expected_risk = "HIGH"
                elif prob_val >= 0.00001:
                    expected_risk = "ELEVATED"

                if idx_risk >= 0 and row[idx_risk].value is not None:
                    if str(row[idx_risk].value).strip().upper() == expected_risk:
                        correct_risk += 1
                        
            except (ValueError, TypeError):
                continue

    if checked_rows > 0:
        ratio_range = correct_range_m / checked_rows
        ratio_energy = correct_energy / checked_rows
        ratio_risk = correct_risk / checked_rows

        if ratio_range > 0.8:
            score += 10
            feedback_parts.append("Min_Range_m formula correct")
        
        if ratio_energy > 0.8:
            score += 20
            feedback_parts.append("Kinetic Energy formula correct")
            
        if ratio_risk > 0.8:
            score += 20
            feedback_parts.append("Risk Category logic correct")
        elif ratio_risk > 0.3:
            score += 10
            feedback_parts.append("Risk Category logic partially correct")
    else:
        feedback_parts.append("Could not evaluate data rows")

    # CRITERION 5: Maneuver Briefing Extraction (15 points)
    briefing_sheet = None
    for sn in wb_vals.sheetnames:
        if "briefing" in sn.lower():
            briefing_sheet = wb_vals[sn]
            break

    if briefing_sheet:
        rows_with_data = sum(1 for r in briefing_sheet.iter_rows(values_only=True) if any(r))
        # Expected: 1 header row + exactly 5 HIGH risk rows = 6 rows
        if rows_with_data == 6:
            score += 15
            feedback_parts.append("Maneuver Briefing correctly filtered (6 rows)")
        elif rows_with_data > 1:
            score += 5
            feedback_parts.append(f"Maneuver Briefing populated but row count incorrect ({rows_with_data})")
    else:
        feedback_parts.append("Maneuver Briefing sheet missing")

    # CRITERION 6: Summary Sheet Aggregation (10 points)
    summary_sheet = None
    for sn in wb_vals.sheetnames:
        if "summary" in sn.lower():
            summary_sheet = wb_vals[sn]
            break

    if summary_sheet:
        summary_text = ""
        for row in summary_sheet.iter_rows(values_only=True):
            summary_text += " ".join(str(c) for c in row if c is not None) + " "
        
        has_5 = "5" in summary_text or "5.0" in summary_text
        has_15 = "15" in summary_text or "15.0" in summary_text
        has_180 = "180" in summary_text or "180.0" in summary_text
        
        if has_5 and has_15 and has_180:
            score += 10
            feedback_parts.append("Summary sheet counts match expected (5/15/180)")
        elif has_5 or has_15:
            score += 5
            feedback_parts.append("Summary sheet has partial correct counts")
    else:
        feedback_parts.append("Summary sheet missing")

    # CRITERION 7: VLM Trajectory Verification / File Timing (10 points anti-gaming)
    file_created_during_task = export_result.get('file_created_during_task', False)
    if file_created_during_task:
        score += 10
        feedback_parts.append("File creation timeframe valid")
    else:
        feedback_parts.append("Warning: File timestamp indicates it may have existed before task start")

    passed = score >= 60 and file_created_during_task and output_exists

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
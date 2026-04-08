#!/usr/bin/env python3
"""
Verifier for Employee Attrition Workforce Analytics task.

The agent must analyze the IBM HR Analytics Employee Attrition dataset (1,470 records)
and create a multi-sheet executive summary containing:
- Summary stats: Headcount (1470), Attrition (237), Rate (~16.1%)
- Department breakdown: HR (19%), R&D (13.8%), Sales (20.6%)
- Risk factor analysis (Overtime, Income, Satisfaction, etc.)
- Turnover Cost estimation ($)
- Retention recommendations

Scoring (10 points total, pass threshold 5.0):
1. Anti-gaming / basic content gate
2. Summary statistics accuracy (1.5 pts)
3. Department attrition analysis (2.0 pts)
4. Risk factor identification (2.0 pts)
5. Turnover cost estimation (1.5 pts)
6. Retention insights/recommendations (1.5 pts)
7. Professional workbook structure (1.5 pts)
"""

import sys
import os
import logging
import tempfile
import json
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import copy_and_parse_document

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                   max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    all_text.append(cell.value.lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets, parsing percentages if stored as strings."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                   max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
                elif isinstance(cell.value, str):
                    pct_match = re.match(r'^\s*([\d\.]+)\s*%\s*$', cell.value)
                    if pct_match:
                        try:
                            numbers.append(float(pct_match.group(1)) / 100.0)
                            numbers.append(float(pct_match.group(1))) # add whole number representation too
                        except:
                            pass
    return numbers

def check_for_formulas(container_path, copy_from_env):
    """Check if workbook contains formulas by reloading without data_only=True."""
    try:
        from openpyxl import load_workbook
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        copy_from_env(container_path, temp_file.name)
        wb_f = load_workbook(temp_file.name, data_only=False)
        has_formula = False
        for sn in wb_f.sheetnames:
            sheet = wb_f[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and str(cell.value).startswith('='):
                        has_formula = True
                        break
                if has_formula:
                    break
            if has_formula:
                break
        wb_f.close()
        os.unlink(temp_file.name)
        return has_formula
    except Exception as e:
        logger.warning(f"Formula check failed: {e}")
        return False

def verify_workforce_analytics(traj, env_info, task_info):
    """Main verification logic for workforce analytics task."""
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # 1. Read export JSON for anti-gaming checks
    try:
        temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
        copy_from_env("/tmp/workforce_analytics_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            export_data = json.load(f)
        os.unlink(temp_json.name)
    except Exception as e:
        logger.warning(f"Failed to load export JSON: {e}")
        export_data = {"output_file_exists": "false", "file_modified_during_task": False}

    if str(export_data.get("output_file_exists")).lower() != "true":
        return {"passed": False, "score": 0.0, "feedback": "Output file workforce_analytics.xlsx not found."}

    if not export_data.get("file_modified_during_task"):
        return {"passed": False, "score": 0.0, "feedback": "File was not created or modified during the task execution (Anti-gaming trigger)."}

    # 2. Parse Document
    container_path = "/home/ga/Documents/Spreadsheets/workforce_analytics.xlsx"
    success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')
    
    if not success:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to parse workbook: {error}"}

    score = 0.0
    feedback_parts = []
    
    all_text = extract_all_text(wb)
    all_numbers = extract_all_numbers(wb)
    num_sheets = len(wb.sheetnames)
    
    total_cells = 0
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    total_cells += 1

    # Gate Check
    if total_cells < 30:
        return {"passed": False, "score": 0.0, "feedback": "File has insufficient content to be a full analysis (<30 cells)."}

    # Tolerance helper for matching expected numbers
    def num_approx_in(target, tolerance=0.02):
        for n in all_numbers:
            # handle exact ints and floats within %
            if abs(n - target) <= max(1.0, target * tolerance):
                return True
            # Also check if it's represented as a decimal fraction (e.g. 0.16 instead of 16%)
            if target > 1 and target < 100:
                if abs(n - (target / 100.0)) <= 0.005:
                    return True
        return False

    # CHECK 1: Summary Statistics (1.5 pts)
    # Expected: 1470 employees, 237 attrition, ~16% (16.1%) rate
    stat_score = 0
    if num_approx_in(1470, 0.01): stat_score += 0.5
    if num_approx_in(237, 0.01): stat_score += 0.5
    if num_approx_in(16.12, 0.05) or num_approx_in(16.1, 0.05) or num_approx_in(16, 0.02): stat_score += 0.5
    
    score += stat_score
    feedback_parts.append(f"Summary Stats: {stat_score}/1.5")

    # CHECK 2: Department Attrition Analysis (2.0 pts)
    # HR: ~19%, R&D: ~13.8%, Sales: ~20.6%
    dept_score = 0
    has_depts = all(d in all_text for d in ["human resources", "research & development", "sales"])
    if has_depts:
        dept_score += 0.5
        # Check for associated values (counts or percentages)
        if num_approx_in(19.0, 0.1): dept_score += 0.5
        if num_approx_in(13.8, 0.1): dept_score += 0.5
        if num_approx_in(20.6, 0.1): dept_score += 0.5
    
    score += dept_score
    feedback_parts.append(f"Dept Analysis: {dept_score}/2.0")

    # CHECK 3: Risk Factor Identification (2.0 pts)
    # Looking for factors like overtime, satisfaction, income, age, travel
    risk_terms = ["overtime", "satisfaction", "income", "age", "tenure", "travel", "role", "manager"]
    found_risks = sum(1 for term in risk_terms if term in all_text)
    
    risk_score = min(2.0, found_risks * 0.5)
    score += risk_score
    feedback_parts.append(f"Risk Factors: {risk_score}/2.0")

    # CHECK 4: Turnover Cost Estimation (1.5 pts)
    cost_terms = ["cost", "turnover", "financial", "replacement", "expense"]
    has_cost_context = any(term in all_text for term in cost_terms)
    
    cost_score = 0
    if has_cost_context:
        cost_score += 0.5
        # Look for a plausible high dollar amount (between $1M and $100M)
        has_large_dollar = any(n >= 1000000 and n <= 100000000 for n in all_numbers)
        # OR per-employee replacement cost (between $20k and $150k)
        has_unit_cost = any(n >= 20000 and n <= 150000 for n in all_numbers)
        
        if has_large_dollar or has_unit_cost:
            cost_score += 1.0
            
    score += cost_score
    feedback_parts.append(f"Cost Est: {cost_score}/1.5")

    # CHECK 5: Retention Insights/Recommendations (1.5 pts)
    rec_terms = ["recommend", "action", "retention", "target", "intervention", "strategy", "improve", "initiative"]
    found_recs = sum(1 for term in rec_terms if term in all_text)
    
    rec_score = min(1.5, found_recs * 0.5)
    if rec_score > 0 and num_sheets > 1: # Recommendations should ideally be separated or well-structured
        rec_score = min(1.5, rec_score + 0.5)
        
    score += rec_score
    feedback_parts.append(f"Recommendations: {rec_score}/1.5")

    # CHECK 6: Professional Workbook Structure (1.5 pts)
    # Multiple sheets and usage of formulas
    struct_score = 0
    if num_sheets >= 2:
        struct_score += 0.5
    if num_sheets >= 3:
        struct_score += 0.5
        
    if check_for_formulas(container_path, copy_from_env):
        struct_score += 0.5
        
    score += struct_score
    feedback_parts.append(f"Structure: {struct_score}/1.5")

    # Final Calculation
    passed = score >= 5.0

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""
Verifier for Menu Engineering Analysis task.
Checks the output workbook for correct calculations, classifications, and structure.
Uses copy_from_env to safely retrieve files, and VLM trajectory analysis for anti-gaming.
"""

import json
import logging
import os
import sys
import tempfile
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import VLM utilities
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
except ImportError:
    logger.warning("gym_anything.vlm not available. VLM verification will be skipped.")
    sample_trajectory_frames = None
    get_final_screenshot = None
    query_vlm = None


def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)


def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
                elif isinstance(cell.value, str):
                    # Attempt to parse formatted numbers like "34.5%" or "$12.50"
                    cleaned = re.sub(r'[$,%]', '', cell.value.strip())
                    try:
                        numbers.append(float(cleaned))
                    except ValueError:
                        pass
    return numbers


def find_numeric_values_in_range(values, low, high):
    """Find all numeric values within a given range"""
    return [v for v in values if low <= v <= high]


def verify_vlm(traj):
    """Use VLM on trajectory to verify the agent actively performed spreadsheet work."""
    if not query_vlm or not sample_trajectory_frames:
        return 15, "VLM not available, awarding default points."

    frames = sample_trajectory_frames(traj, n=4)
    final = get_final_screenshot(traj)
    images = frames + [final] if final else frames

    if not images:
        return 0, "No trajectory images available."

    prompt = """You are evaluating an AI agent performing a spreadsheet analysis task.
    The task is to build a "Menu Engineering Matrix" (calculating food costs, contribution margins, and categorizing items as Stars, Plowhorses, Puzzles, Dogs) using POS and Recipe data.
    
    Review these screenshots from the agent's workflow and determine:
    1. Did the agent actively work within a spreadsheet application?
    2. Are there signs of the agent importing data or typing/creating formulas?
    3. Does the final/later screenshots show a structured analysis (e.g., calculations, columns for margin/food cost, classification labels)?
    
    Respond strictly in JSON format:
    {
        "used_spreadsheet": true/false,
        "showed_workflow_progression": true/false,
        "created_analysis_structure": true/false,
        "reasoning": "brief explanation"
    }"""

    try:
        vlm_result = query_vlm(images=images, prompt=prompt)
        if not vlm_result.get("success"):
            return 0, f"VLM query failed: {vlm_result.get('error')}"

        parsed = vlm_result.get("parsed", {})
        score = 0
        if parsed.get("used_spreadsheet"): score += 5
        if parsed.get("showed_workflow_progression"): score += 5
        if parsed.get("created_analysis_structure"): score += 5
        
        return score, parsed.get("reasoning", "VLM verification complete.")
    except Exception as e:
        logger.error(f"VLM verification error: {e}")
        return 0, f"VLM error: {e}"


def verify_menu_engineering(traj, env_info, task_info):
    """
    Verify the menu engineering analysis workbook.
    Scoring out of 100 points:
    - 0 pts: Gate failure (file not created/modified, or empty)
    - 15 pts: Food cost percentages calculated
    - 15 pts: Contribution margins calculated
    - 15 pts: Sales aggregation
    - 20 pts: Menu engineering classification
    - 10 pts: Overall weighted food cost
    - 10 pts: Professional multi-sheet structure
    - 15 pts: VLM Trajectory Verification
    Pass Threshold: 60 points
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_path = metadata.get('expected_output_path', '/home/ga/Documents/Spreadsheets/menu_engineering_analysis.xlsx')
    
    score = 0
    feedback_parts = []

    # 1. Read task_result.json for metadata
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get('output_exists'):
        return {"passed": False, "score": 0, "feedback": "Target XLSX file was not found."}
    
    if not result.get('file_created_during_task'):
        return {"passed": False, "score": 0, "feedback": "File was not modified during the task (anti-gaming)."}

    # 2. Extract content using openpyxl
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env(expected_path, temp_xlsx.name)
        try:
            from openpyxl import load_workbook
            # Load with data_only=True to get formula results
            wb = load_workbook(temp_xlsx.name, data_only=True)
            all_text = extract_all_text(wb)
            all_numbers = extract_all_numbers(wb)
            num_sheets = len(wb.sheetnames)
            total_cells = sum(1 for sn in wb.sheetnames for r in wb[sn].iter_rows() for c in r if c.value is not None)
        except Exception as e:
            return {"passed": False, "score": 0, "feedback": f"Failed to parse XLSX: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    if total_cells < 30:
        return {"passed": False, "score": 0, "feedback": f"Workbook too sparse ({total_cells} cells)."}

    # --- CRITERION: Food cost percentages (15 pts) ---
    fc_values = find_numeric_values_in_range(all_numbers, 15.0, 55.0) # Typical food costs
    if len(fc_values) >= 20:
        score += 15
        feedback_parts.append("Food Cost % calculated")
    elif len(fc_values) >= 10:
        score += 7
        feedback_parts.append("Partial Food Cost % calculated")
    else:
        feedback_parts.append("Missing Food Cost % calculations")

    # --- CRITERION: Contribution Margins (15 pts) ---
    cm_values = find_numeric_values_in_range(all_numbers, 2.0, 25.0)
    cm_unique = len(set(round(v, 2) for v in cm_values))
    if cm_unique >= 15:
        score += 15
        feedback_parts.append("Contribution Margins calculated")
    elif cm_unique >= 5:
        score += 7
        feedback_parts.append("Partial Contribution Margins")
    else:
        feedback_parts.append("Missing Contribution Margins")

    # --- CRITERION: Sales Aggregation (15 pts) ---
    qty_values = find_numeric_values_in_range(all_numbers, 10, 500)
    if len(qty_values) >= 20:
        score += 15
        feedback_parts.append("Sales aggregated")
    else:
        feedback_parts.append("Missing or incomplete sales aggregation")

    # --- CRITERION: Menu Engineering Classification (20 pts) ---
    matrix_terms = ["star", "plowhorse", "puzzle", "dog"]
    terms_found = sum(1 for t in matrix_terms if t in all_text or (t == "plowhorse" and "plow horse" in all_text))
    
    if terms_found == 4:
        score += 20
        feedback_parts.append("All 4 matrix classifications found")
    elif terms_found >= 2:
        score += 10
        feedback_parts.append(f"{terms_found}/4 matrix classifications found")
    else:
        feedback_parts.append("Missing matrix classifications")

    # --- CRITERION: Overall weighted food cost (10 pts) ---
    overall_fc_cands = find_numeric_values_in_range(all_numbers, 30.0, 40.0)
    has_overall_text = any(t in all_text for t in ["overall", "total food cost", "weighted"])
    if len(overall_fc_cands) >= 1 and has_overall_text:
        score += 10
        feedback_parts.append("Overall food cost calculated")
    elif has_overall_text:
        score += 5
        feedback_parts.append("Overall food cost structure present but values missing/wrong")

    # --- CRITERION: Professional structure (10 pts) ---
    if num_sheets >= 2:
        score += 10
        feedback_parts.append("Multi-sheet structure used")
    else:
        score += 5
        feedback_parts.append("Single sheet structure")

    # --- CRITERION: VLM Verification (15 pts) ---
    vlm_score, vlm_reasoning = verify_vlm(traj)
    score += vlm_score
    feedback_parts.append(f"VLM Score: {vlm_score}/15 ({vlm_reasoning})")

    passed = score >= 60 and result.get('output_exists') and result.get('file_created_during_task')
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
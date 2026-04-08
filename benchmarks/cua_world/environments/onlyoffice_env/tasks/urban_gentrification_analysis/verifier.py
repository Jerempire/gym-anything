#!/usr/bin/env python3
"""
Verifier for Urban Gentrification Analysis task.

This evaluates the agent's ability to transform demographic data, compute arithmetic
indexes, perform conditional filtering, and extract subsets.

Scoring (100 points total, pass threshold 75):
1. File Format & Save (10 pts)
2. Index Calculations (25 pts)
3. Formula Usage (10 pts)
4. Population Filtering (15 pts)
5. Summary Sheet Creation (10 pts)
6. Accuracy of Top 10 (20 pts)
7. County-wide Averages (10 pts)
"""

import sys
import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import openpyxl, required for verifying xlsx formulas
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl


def verify_gentrification_analysis(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # Get expected top 10 from task_info metadata
    metadata = task_info.get('metadata', {})
    expected_top_10 = set(metadata.get('ground_truth_top_10', []))
    noise_tracts = set(metadata.get('noise_tracts', []))

    # Read result json
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to read export JSON: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    if not result.get('output_file_exists'):
        return {
            "passed": False, 
            "score": 0.0, 
            "feedback": "Target file austin_gentrification_analysis.xlsx not found."
        }

    container_path = "/home/ga/Documents/Spreadsheets/austin_gentrification_analysis.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    score = 10.0 # File format & save passed
    feedback_parts = ["File saved correctly (+10)"]
    
    try:
        copy_from_env(container_path, temp_xlsx.name)
        
        # Load with data_only=True to check values
        wb_data = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        # Load with data_only=False to check formulas
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        
        main_sheet_name = wb_data.sheetnames[0]
        main_sheet = wb_data[main_sheet_name]
        main_sheet_formulas = wb_formulas[main_sheet_name]
        
        # 1. Index Calculations & Formula Usage
        header_row = [str(cell.value).lower().strip() for cell in main_sheet[1] if cell.value]
        calc_cols = ['income_growth', 'rent_growth', 'edu_growth', 'gentrification_index']
        cols_found = [c for c in calc_cols if any(c in h for h in header_row)]
        
        if len(cols_found) == 4:
            score += 25.0
            feedback_parts.append("All calculated columns found (+25)")
        elif len(cols_found) > 0:
            score += 10.0
            feedback_parts.append(f"Partial calculated columns found: {cols_found} (+10)")
        else:
            feedback_parts.append("Calculated columns not found")
            
        # Check formulas
        formulas_present = False
        for row in main_sheet_formulas.iter_rows(min_row=2, max_row=min(main_sheet_formulas.max_row, 50)):
            for cell in row:
                if isinstance(cell.value, str) and str(cell.value).startswith('='):
                    formulas_present = True
                    break
            if formulas_present:
                break
                
        if formulas_present:
            score += 10.0
            feedback_parts.append("Formulas used for calculations (+10)")
        else:
            feedback_parts.append("Values hardcoded instead of formulas")
            
        # 2. Check County-wide Averages
        has_average = False
        # Check data
        for row in main_sheet.iter_rows(min_row=max(1, main_sheet.max_row - 10), max_row=main_sheet.max_row):
            for cell in row:
                if isinstance(cell.value, str) and 'average' in str(cell.value).lower():
                    has_average = True
        # Check formulas
        if not has_average:
            for row in main_sheet_formulas.iter_rows(min_row=max(1, main_sheet_formulas.max_row - 10), max_row=main_sheet_formulas.max_row):
                for cell in row:
                    if isinstance(cell.value, str) and 'average(' in str(cell.value).lower():
                        has_average = True
        
        if has_average:
            score += 10.0
            feedback_parts.append("Average calculations found (+10)")
        else:
            feedback_parts.append("Average calculations not found")
            
        # 3. Top 10 Sheet, Filtering, and Accuracy
        top10_sheet_name = None
        for name in wb_data.sheetnames:
            if 'top' in name.lower() and '10' in name.lower():
                top10_sheet_name = name
                break
                
        if top10_sheet_name:
            score += 10.0
            feedback_parts.append("Top_10_Gentrified sheet found (+10)")
            top_sheet = wb_data[top10_sheet_name]
            
            extracted_tracts = set()
            has_noise = False
            
            for row in top_sheet.iter_rows(min_row=2, max_row=top_sheet.max_row):
                for cell in row:
                    if cell.value:
                        val = str(cell.value).strip()
                        if val.startswith("48453") and len(val) == 11:
                            extracted_tracts.add(val)
                            if val in noise_tracts:
                                has_noise = True
                    
            if not has_noise and len(extracted_tracts) > 0:
                score += 15.0
                feedback_parts.append("Population filter applied correctly (no noise tracts) (+15)")
            else:
                feedback_parts.append("Population filter failed (noise tracts included or no tracts extracted)")
                
            # Check accuracy of top 10
            matches = expected_top_10.intersection(extracted_tracts)
            if len(matches) >= 8:
                score += 20.0
                feedback_parts.append(f"Top 10 extraction highly accurate ({len(matches)}/10 matched) (+20)")
            elif len(matches) > 0:
                pts = len(matches) * 2.0
                score += pts
                feedback_parts.append(f"Top 10 extraction partially accurate ({len(matches)}/10 matched) (+{pts})")
            else:
                feedback_parts.append("Top 10 extraction inaccurate")
        else:
            feedback_parts.append("Top_10_Gentrified sheet not found")
            
    except Exception as e:
        feedback_parts.append(f"Error during verification: {e}")
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    passed = score >= 75.0
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
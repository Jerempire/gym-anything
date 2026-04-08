#!/usr/bin/env python3
"""
Verifier for Retail Seasonal Markdown Optimization Task.

Evaluates spreadsheet modifications against real merchandising math standards:
1. Formulas correctly calculate Average Weekly Sales, Weeks of Supply, and Sell-Through Rate.
2. Nested logic for Markdown_Percent evaluates WOS and STR breakpoints correctly.
3. Markdown_Cost accurately calculates impact.
4. Summary sheet appropriately aggregates data by Department.
5. VLM checks trajectory frames to ensure authentic interaction.
"""

import sys
import os
import json
import logging
import tempfile
import math

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from vlm_utils import sample_trajectory_frames, query_vlm

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def find_column_by_keyword(headers, keywords):
    """Finds the 0-indexed column matching any keyword (case-insensitive)."""
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_str = str(header).lower()
        if any(kw.lower() in header_str for kw in keywords):
            return idx
    return -1

def verify_markdown_optimization(traj, env_info, task_info):
    """
    Main verifier function.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Framework error: copy_from_env not available."}

    temp_dir = tempfile.mkdtemp(prefix='retail_markdown_')
    result_json_path = os.path.join(temp_dir, 'result.json')
    xlsx_path = os.path.join(temp_dir, 'markdown_analysis.xlsx')

    try:
        copy_from_env("/tmp/retail_markdown_result.json", result_json_path)
        with open(result_json_path, 'r') as f:
            result_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Could not read result.json: {e}"}

    if not result_data.get("output_file_exists", False):
        return {"passed": False, "score": 0.0, "feedback": "Final workbook 'markdown_analysis.xlsx' was not saved."}

    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/markdown_analysis.xlsx", xlsx_path)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to extract workbook: {e}"}

    # Install openpyxl if missing
    try:
        from openpyxl import load_workbook
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import load_workbook

    score = 0
    feedback_parts = []
    
    # Check 1: Output file structure
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to parse workbook: {e}"}

    if len(wb.sheetnames) < 2:
        feedback_parts.append("Workbook lacks the required multiple sheets (Data + Summary).")
    else:
        score += 10
        feedback_parts.append("Multiple sheets detected.")

    # Find sheets
    data_sheet = wb.worksheets[0]
    summary_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            summary_sheet = wb[name]
            break
    
    if not summary_sheet and len(wb.sheetnames) > 1:
        summary_sheet = wb.worksheets[1]

    # Evaluate Data Sheet
    headers = [cell.value for cell in data_sheet[1]]
    
    # Raw columns
    idx_dept = find_column_by_keyword(headers, ["department"])
    idx_ticket = find_column_by_keyword(headers, ["ticket"])
    idx_sold = find_column_by_keyword(headers, ["sold", "units_sold"])
    idx_on_hand = find_column_by_keyword(headers, ["hand", "units_on_hand"])
    idx_weeks = find_column_by_keyword(headers, ["weeks_on_floor", "weeks"])
    
    # Calc columns
    idx_aws = find_column_by_keyword(headers, ["aws", "average weekly sales"])
    idx_wos = find_column_by_keyword(headers, ["wos", "weeks of supply"])
    idx_str = find_column_by_keyword(headers, ["str", "sell-through", "sell through"])
    idx_md_pct = find_column_by_keyword(headers, ["markdown_percent", "markdown percent", "markdown %"])
    idx_md_cost = find_column_by_keyword(headers, ["markdown_cost", "markdown cost"])

    if -1 in [idx_sold, idx_on_hand, idx_weeks, idx_aws, idx_wos, idx_str, idx_md_pct, idx_md_cost]:
        missing = [name for name, idx in zip(
            ["Sold", "On_Hand", "Weeks", "AWS", "WOS", "STR", "MD_Pct", "MD_Cost"], 
            [idx_sold, idx_on_hand, idx_weeks, idx_aws, idx_wos, idx_str, idx_md_pct, idx_md_cost]
        ) if idx == -1]
        feedback_parts.append(f"Missing one or more required columns. Missing: {missing}")
    else:
        # Sample evaluation
        valid_metrics = 0
        valid_logic = 0
        valid_cost = 0
        samples_checked = 0
        ground_truth_dept_aggregates = {}

        for row_idx in range(2, min(data_sheet.max_row + 1, 1002)):
            row = data_sheet[row_idx]
            dept = row[idx_dept].value
            if not dept:
                continue

            try:
                ticket = float(row[idx_ticket].value or 0)
                sold = float(row[idx_sold].value or 0)
                on_hand = float(row[idx_on_hand].value or 0)
                weeks = float(row[idx_weeks].value or 1)
                
                # Agent Values
                a_aws = float(row[idx_aws].value or 0)
                a_wos = float(row[idx_wos].value or 0)
                a_str = float(row[idx_str].value or 0)
                
                raw_pct = row[idx_md_pct].value or 0
                if isinstance(raw_pct, str) and "%" in raw_pct:
                    a_pct = float(raw_pct.replace("%", "")) / 100.0
                else:
                    a_pct = float(raw_pct)
                    if a_pct > 1.0: # they used 50 instead of 0.5
                        a_pct = a_pct / 100.0
                        
                a_cost = float(row[idx_md_cost].value or 0)

                # Ground Truth calculation
                gt_aws = sold / weeks if weeks else 0
                gt_wos = on_hand / gt_aws if gt_aws else 0
                gt_str = sold / (sold + on_hand) if (sold + on_hand) else 0

                # Check metrics tolerance (5%)
                if math.isclose(a_aws, gt_aws, rel_tol=0.05) and \
                   math.isclose(a_wos, gt_wos, rel_tol=0.05) and \
                   math.isclose(a_str, gt_str, rel_tol=0.05):
                    valid_metrics += 1

                # Ground Truth Logic
                if gt_wos >= 12 and gt_str <= 0.40:
                    gt_pct = 0.50
                elif gt_wos >= 8 and gt_str <= 0.55:
                    gt_pct = 0.30
                elif gt_wos >= 4 and gt_str <= 0.70:
                    gt_pct = 0.15
                else:
                    gt_pct = 0.0
                
                if math.isclose(a_pct, gt_pct, abs_tol=0.01):
                    valid_logic += 1
                
                gt_cost = on_hand * ticket * gt_pct
                if math.isclose(a_cost, gt_cost, rel_tol=0.05) or math.isclose(a_cost, gt_cost, abs_tol=1.0):
                    valid_cost += 1

                # Track GT aggregates for Summary Check
                if dept not in ground_truth_dept_aggregates:
                    ground_truth_dept_aggregates[dept] = {"cost": 0.0, "units": 0.0}
                ground_truth_dept_aggregates[dept]["cost"] += gt_cost
                ground_truth_dept_aggregates[dept]["units"] += on_hand

                samples_checked += 1
            except Exception as e:
                pass # skip unparseable rows

        if samples_checked > 0:
            metrics_acc = valid_metrics / samples_checked
            logic_acc = valid_logic / samples_checked
            cost_acc = valid_cost / samples_checked

            score += int(20 * metrics_acc)
            score += int(25 * logic_acc)
            score += int(15 * cost_acc)

            feedback_parts.append(f"Row Level Accuracy: Metrics {metrics_acc*100:.0f}%, Logic {logic_acc*100:.0f}%, Cost {cost_acc*100:.0f}%")

        # Evaluate Summary Sheet
        if summary_sheet:
            found_depts = 0
            correct_sums = 0
            for row in summary_sheet.iter_rows(max_row=50, max_col=10):
                for cell in row:
                    if isinstance(cell.value, str):
                        dept_name = cell.value.strip()
                        if dept_name in ground_truth_dept_aggregates:
                            found_depts += 1
                            row_vals = [c.value for c in row if isinstance(c.value, (int, float))]
                            if row_vals:
                                expected_cost = ground_truth_dept_aggregates[dept_name]["cost"]
                                expected_units = ground_truth_dept_aggregates[dept_name]["units"]
                                
                                # Check if any of the numbers in the row match the expected aggregation
                                if any(math.isclose(v, expected_cost, rel_tol=0.05) or math.isclose(v, expected_cost, abs_tol=10.0) for v in row_vals) and \
                                   any(math.isclose(v, expected_units, rel_tol=0.05) or math.isclose(v, expected_units, abs_tol=5.0) for v in row_vals):
                                    correct_sums += 1
            
            if found_depts >= 3:
                agg_acc = correct_sums / found_depts
                score += int(20 * agg_acc)
                feedback_parts.append(f"Department Aggregation Accuracy: {agg_acc*100:.0f}%")
            else:
                feedback_parts.append("Could not find Department aggregates in the Summary Sheet.")
        else:
            feedback_parts.append("Summary Sheet missing.")

    # VLM Verification for Anti-Gaming
    try:
        frames = sample_trajectory_frames(traj, n=4)
        vlm_prompt = """Look at these screenshots of a user interacting with a spreadsheet.
        Did the user actively build formulas, apply conditional logic, or create pivot tables/aggregations within the GUI?
        Answer JSON: {"active_spreadsheet_usage": true/false}"""
        vlm_res = query_vlm(images=frames, prompt=vlm_prompt)
        
        if vlm_res and vlm_res.get('parsed', {}).get('active_spreadsheet_usage', False):
            score += 10
            feedback_parts.append("VLM confirmed active GUI usage.")
        else:
            feedback_parts.append("VLM could not confirm active GUI usage.")
    except Exception as e:
        logger.warning(f"VLM verification failed: {e}")

    # Final scoring
    passed = score >= 70

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
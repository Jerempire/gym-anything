#!/usr/bin/env python3
"""
Verifier for School Enrollment Projection task.

Uses `copy_from_env` to extract the workbook and JSON result. 
Programmatically evaluates the mathematical correctness of the Cohort Survival 
Model calculations within the workbook. Also includes an optional VLM trajectory 
verification to ensure the agent physically manipulated the spreadsheet.
"""

import sys
import os
import json
import logging
import tempfile
import math

# We add the utils directory to the path so we can import the standard helpers
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from onlyoffice_verification_utils import copy_and_parse_document, cleanup_temp_dir
except ImportError:
    pass

from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ground Truth Derived from setup script deterministic data
# Births 2019: 440, 2020: 420, 2021: 430, 2022: 435, 2023: 425
# Avg Birth-to-K ratio: ~0.90
# Avg K-to-1: 0.98, Avg 1-to-2: 1.01, Avg 8-to-9: 1.08, etc.

EXPECTED_PROJECTIONS = {
    2024: {
        "K": 396,     # 2019 Births (440) * 0.90
        "G1": 397,    # 2023 K (405) * 0.98 = 396.9 -> 397
        "G2": 401,    # 2023 G1 (397) * 1.01 = 400.97 -> 401
        "G9": 544     # 2023 G8 (504) * 1.08 = 544.32 -> 544
    },
    2025: {
        "K": 378,     # 2020 Births (420) * 0.90
        "G1": 388     # 2024 K (396) * 0.98 = 388.08 -> 388
    },
    2028: {
        "K": 383      # 2023 Births (425) * 0.90 = 382.5 -> 383
    }
}


def extract_all_numbers(wb):
    """Extract all numeric values across all sheets in the workbook."""
    numbers = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 200), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
    return numbers


def extract_all_text(wb):
    """Extract all lowercased text across all sheets to check structure."""
    text_content = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 200), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    text_content.append(cell.value.lower().strip())
    return text_content


def check_value_in_list(val, num_list, tolerance=1.5):
    """Check if a value exists in the list of numbers within a given tolerance."""
    for n in num_list:
        if abs(n - val) <= tolerance:
            return True
    return False


def verify_enrollment_projection(traj, env_info, task_info):
    """Main verification function."""
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # 1. Check Result JSON for anti-gaming timestamps and file existence
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/enrollment_projection_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read task result metadata: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    if not result.get('output_file_exists', False):
        return {"passed": False, "score": 0, "feedback": "Output file enrollment_projections.xlsx not found."}
    
    if not result.get('file_created_during_task', False):
        # Heavy penalty or fail if file wasn't created/modified during the task timeframe
        return {"passed": False, "score": 0, "feedback": "Output file was not created or modified during the task execution (anti-gaming)."}

    # 2. Extract Spreadsheet Data
    container_path = "/home/ga/Documents/Spreadsheets/enrollment_projections.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_enroll_')
    
    try:
        # Custom extract logic mimicking onlyoffice_verification_utils pattern
        from openpyxl import load_workbook
        temp_wb_path = os.path.join(temp_dir, "temp.xlsx")
        copy_from_env(container_path, temp_wb_path)
        wb = load_workbook(temp_wb_path, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to parse workbook: {e}"}
    
    score = 0.0
    feedback_parts = []
    
    all_numbers = extract_all_numbers(wb)
    all_text = extract_all_text(wb)
    
    # 3. Structural checks
    expected_years_found = sum(1 for y in ["2024", "2025", "2026", "2027", "2028"] if y in all_text or int(y) in all_numbers)
    if expected_years_found >= 4:
        score += 10
        feedback_parts.append("Projection years (2024-2028) present.")
    else:
        feedback_parts.append("Missing projection year headers.")

    avg_terms = sum(1 for term in ["average", "avg", "ratio", "csr"] if any(term in t for t in all_text))
    if avg_terms > 0:
        score += 10
        feedback_parts.append("Ratio/Average terminology found.")

    # 4. Ratio Accuracy Checks (Averages)
    # Expected Avg Birth-to-K = 0.90
    # Expected Avg K-1 = 0.98
    # Expected Avg 8-9 = 1.08
    ratios_correct = 0
    if check_value_in_list(0.90, all_numbers, 0.01): ratios_correct += 1
    if check_value_in_list(0.98, all_numbers, 0.01): ratios_correct += 1
    if check_value_in_list(1.08, all_numbers, 0.01): ratios_correct += 1
    
    if ratios_correct >= 2:
        score += 20
        feedback_parts.append("Historical Cohort Survival Ratios calculated correctly.")
    elif ratios_correct == 1:
        score += 10
        feedback_parts.append("Partial historical ratios calculated.")
    else:
        feedback_parts.append("Could not find correct average ratios (e.g. 0.90, 0.98).")

    # 5. Projection Value Checks (The cascade calculations)
    proj_correct = 0
    # 2024
    if check_value_in_list(EXPECTED_PROJECTIONS[2024]["K"], all_numbers, 2.0): proj_correct += 1
    if check_value_in_list(EXPECTED_PROJECTIONS[2024]["G1"], all_numbers, 2.0): proj_correct += 1
    if check_value_in_list(EXPECTED_PROJECTIONS[2024]["G9"], all_numbers, 2.0): proj_correct += 1
    # 2025
    if check_value_in_list(EXPECTED_PROJECTIONS[2025]["K"], all_numbers, 2.0): proj_correct += 1
    if check_value_in_list(EXPECTED_PROJECTIONS[2025]["G1"], all_numbers, 2.0): proj_correct += 1
    # 2028
    if check_value_in_list(EXPECTED_PROJECTIONS[2028]["K"], all_numbers, 2.0): proj_correct += 1

    if proj_correct >= 5:
        score += 40
        feedback_parts.append("Grade projections cascaded accurately through 2028.")
    elif proj_correct >= 2:
        score += 20
        feedback_parts.append("Some grade projections calculated accurately.")
    else:
        feedback_parts.append("Failed to accurately project future cohorts.")

    # 6. Totals
    # A total row usually has values > 5000 (since there are 13 grades with ~400-500 each)
    large_numbers = [n for n in all_numbers if n > 5000]
    if len(large_numbers) >= 5:
        score += 10
        feedback_parts.append("Grand totals calculated for projection years.")
        
    # 7. VLM Trajectory Verification
    frames = sample_trajectory_frames(traj, n=4)
    final_frame = get_final_screenshot(traj)
    if frames and final_frame:
        try:
            prompt = (
                "Review these trajectory frames of an agent working in a spreadsheet editor. "
                "Did the agent import historical data and actively build a model with calculations "
                "for school enrollment projections (Cohort Survival)? "
                "Reply in JSON: {'active_modeling': true/false, 'confidence': 'high/low'}"
            )
            vlm_res = query_vlm(images=frames + [final_frame], prompt=prompt)
            if vlm_res.get('success') and vlm_res.get('parsed', {}).get('active_modeling', False):
                score += 10
                feedback_parts.append("VLM confirmed visual active modeling workflow.")
            else:
                feedback_parts.append("VLM could not confirm active spreadsheet modeling.")
        except Exception as e:
            logger.warning(f"VLM verification failed/skipped: {e}")
            # Do not heavily penalize if VLM fails, just add base score proportion
            score += 10 

    wb.close()
    
    # Clean up temp dir
    for f in os.listdir(temp_dir):
        os.remove(os.path.join(temp_dir, f))
    os.rmdir(temp_dir)
    
    passed = score >= 70

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""
Verifier for Bikeshare Fleet Rebalancing Analysis.

Evaluates the agent's ability to aggregate raw trip data into operational insights.
Scoring (100 points total, Pass Threshold: 70):
- File existence & creation during task (10 pts)
- Station Arrivals accurately calculated (20 pts)
- Station Departures accurately calculated (20 pts)
- Net Fleet Change properly calculated (15 pts)
- Usage of Formulas or Pivot Tables (Anti-cheat) (15 pts)
- Top 5 Surplus identified explicitly (10 pts)
- Top 5 Deficit identified explicitly (10 pts)
"""

import json
import os
import tempfile
import logging
import sys

# Attempt to import verification utilities if available
try:
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
    from onlyoffice_verification_utils import extract_all_text, parse_xlsx_file
except ImportError:
    pass

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def check_formulas_or_pivots(filepath):
    """Check if the workbook contains formulas or pivot tables (anti-hardcoding check)."""
    try:
        # Check pivots first
        wb_data = load_workbook(filepath, data_only=True)
        for sheet_name in wb_data.sheetnames:
            sheet = wb_data[sheet_name]
            if hasattr(sheet, '_pivots') and sheet._pivots:
                return True
                
        # Check formulas
        wb_formulas = load_workbook(filepath, data_only=False)
        for sheet_name in wb_formulas.sheetnames:
            sheet = wb_formulas[sheet_name]
            for row in sheet.iter_rows(max_row=1000, max_col=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # Simple checks to see if it's an aggregation function or subtraction
                        val = cell.value.upper()
                        if 'COUNTIF' in val or 'SUMIF' in val or '-' in val:
                            return True
        return False
    except Exception as e:
        logger.error(f"Error checking formulas: {e}")
        return False

def extract_station_data(filepath):
    """Extract rows that look like station aggregation data."""
    extracted = {}
    try:
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=5000, max_col=20):
                row_vals = [c.value for c in row]
                # Look for rows containing a station name and multiple numbers
                for idx, val in enumerate(row_vals):
                    if isinstance(val, str) and ("NW" in val or "SW" in val or "Memorial" in val or "Station" in val):
                        numbers = [n for n in row_vals if isinstance(n, (int, float))]
                        if len(numbers) >= 2:
                            # Heuristic: the largest numbers are likely arrivals/departures, 
                            # the smallest (or negative) is net change
                            extracted[val.strip()] = numbers
    except Exception as e:
        logger.error(f"Error extracting station data: {e}")
    return extracted

def extract_all_text(filepath):
    """Extract all text to look for executive summary content."""
    text = ""
    try:
        wb = load_workbook(filepath, data_only=True)
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row in sheet.iter_rows(max_row=1000, max_col=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        text += " " + str(cell.value).lower()
    except Exception as e:
        pass
    return text

def verify_bikeshare_rebalancing(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0
    
    # Read task result JSON
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export JSON: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)
            
    # Read ground truth JSON
    gt_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/ground_truth.json", gt_temp.name)
        with open(gt_temp.name, 'r') as f:
            gt = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read Ground Truth: {e}"}
    finally:
        if os.path.exists(gt_temp.name):
            os.unlink(gt_temp.name)

    # 1. File existence (10 pts)
    output_exists = result.get('output_exists', False)
    file_created = result.get('file_created_during_task', False)
    
    if not output_exists:
        return {"passed": False, "score": 0, "feedback": "Output workbook cabi_rebalancing_plan.xlsx not found."}
    
    if file_created:
        score += 10
        feedback_parts.append("File created successfully.")
    else:
        feedback_parts.append("File exists but timestamp indicates it was not modified during task.")

    # Fetch the actual workbook
    wb_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    copy_from_env("/home/ga/Documents/Spreadsheets/cabi_rebalancing_plan.xlsx", wb_temp.name)
    
    # 2. Formula/Pivot Usage Anti-Cheat (15 pts)
    used_formulas = check_formulas_or_pivots(wb_temp.name)
    if used_formulas:
        score += 15
        feedback_parts.append("Formulas or Pivot Tables detected.")
    else:
        feedback_parts.append("WARNING: No formulas/pivots detected. Agent may have hardcoded answers.")
        
    # 3 & 4 & 5. Check accurate Arrivals, Departures, Net Change (20 + 20 + 15 pts)
    agent_data = extract_station_data(wb_temp.name)
    all_text = extract_all_text(wb_temp.name)
    
    # Check a few distinct stations from Ground Truth
    test_stations = ["Lincoln Memorial", "14th & V St NW", "Columbus Circle / Union Station"]
    
    dep_correct = 0
    arr_correct = 0
    net_correct = 0
    
    for ts in test_stations:
        gt_dep = gt["departures"].get(ts, 0)
        gt_arr = gt["arrivals"].get(ts, 0)
        gt_net = gt["net"].get(ts, 0)
        
        # See if agent has numbers matching these
        agent_numbers = agent_data.get(ts, [])
        if gt_dep in agent_numbers: dep_correct += 1
        if gt_arr in agent_numbers: arr_correct += 1
        if gt_net in agent_numbers or -gt_net in agent_numbers: net_correct += 1

    if dep_correct >= 2:
        score += 20
        feedback_parts.append("Departures accurately aggregated.")
    elif dep_correct == 1:
        score += 10
        feedback_parts.append("Departures partially aggregated.")
        
    if arr_correct >= 2:
        score += 20
        feedback_parts.append("Arrivals accurately aggregated.")
    elif arr_correct == 1:
        score += 10
        feedback_parts.append("Arrivals partially aggregated.")
        
    if net_correct >= 2:
        score += 15
        feedback_parts.append("Net Fleet Change accurately calculated.")

    # 6 & 7. Executive Summary / Top 5 (10 + 10 pts)
    # Get the actual top 5 from ground truth
    sorted_stations = sorted(gt["net"].items(), key=lambda item: item[1], reverse=True)
    top_5_surplus = [s[0].lower() for s in sorted_stations[:5]]
    top_5_deficit = [s[0].lower() for s in sorted_stations[-5:]]
    
    surplus_found = sum([1 for s in top_5_surplus if s in all_text])
    deficit_found = sum([1 for s in top_5_deficit if s in all_text])
    
    if surplus_found >= 3:
        score += 10
        feedback_parts.append("Top surplus stations successfully identified.")
    elif surplus_found >= 1:
        score += 5
        feedback_parts.append("Top surplus stations partially identified.")
        
    if deficit_found >= 3:
        score += 10
        feedback_parts.append("Top deficit stations successfully identified.")
    elif deficit_found >= 1:
        score += 5
        feedback_parts.append("Top deficit stations partially identified.")

    os.unlink(wb_temp.name)

    passed = score >= 70 and used_formulas and (dep_correct > 0 or arr_correct > 0)
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
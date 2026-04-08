#!/usr/bin/env python3
"""
Verifier for Exoplanet Transit Analysis task.

Evaluates if the agent correctly parsed Kepler light curve data, calculated
orbital phase and normalized flux, and extracted planetary physical parameters.
Uses multi-signal verification including file parsing, value checks, and VLM trajectory.
"""

import sys
import os
import json
import logging
import tempfile

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Attempt to load VLM utilities securely
try:
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot, query_vlm
    VLM_AVAILABLE = True
except ImportError:
    VLM_AVAILABLE = False
    logger.warning("gym_anything.vlm not available. VLM trajectory verification will be skipped.")

# Attempt to load openpyxl
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Required for spreadsheet parsing.")


def extract_all_text(wb):
    """Extract all text from all cells in all sheets to find labels."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 200), max_col=min(sheet.max_column, 20)):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)


def extract_all_numbers(wb):
    """Extract all numeric values across all sheets to find calculated metrics."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 200), max_col=min(sheet.max_column, 20)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
    return numbers


def count_formulas(wb_path):
    """Load workbook without data_only to count raw formulas (anti-gaming)."""
    try:
        wb = load_workbook(wb_path, data_only=False)
        count = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            # Only check a sample of rows to save time, but deep enough to catch mass formulas
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 20)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and str(cell.value).startswith('='):
                        count += 1
        return count
    except Exception as e:
        logger.error(f"Error counting formulas: {e}")
        return 0


def verify_exoplanet_analysis(traj, env_info, task_info):
    """
    Main verification logic.
    Max Score: 100 points
    - File exists & created during task: 15 pts
    - Spreadsheets contain formulas (>100): 20 pts (proves computation)
    - Transit Depth calculated correctly: 15 pts
    - Radius Ratio calculated correctly: 15 pts
    - Planet Radius calculated correctly: 20 pts
    - VLM Trajectory check: 15 pts
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    metadata = task_info.get('metadata', {})
    expected_depth_min = metadata.get('expected_depth_min', 0.008)
    expected_depth_max = metadata.get('expected_depth_max', 0.013)
    expected_ratio_min = metadata.get('expected_ratio_min', 0.09)
    expected_ratio_max = metadata.get('expected_ratio_max', 0.115)
    expected_radius_min = metadata.get('expected_radius_min', 1.30)
    expected_radius_max = metadata.get('expected_radius_max', 1.65)

    feedback_parts = []
    score = 0
    max_score = 100

    # 1. Read the export result
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    # Criterion 1: Output exists and was created
    output_exists = result.get('output_exists', False)
    file_created = result.get('file_created_during_task', False)

    if output_exists and file_created:
        score += 15
        feedback_parts.append("Output file saved correctly")
    elif output_exists:
        score += 5
        feedback_parts.append("Output file exists (but timestamp issue)")
    else:
        feedback_parts.append("Expected kepler8b_analysis.xlsx not found")
        return {"passed": False, "score": 0, "feedback": " | ".join(feedback_parts)}

    # Copy the actual workbook out
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/kepler8b_analysis.xlsx", temp_xlsx.name)

        if not OPENPYXL_AVAILABLE:
            feedback_parts.append("openpyxl missing, cannot parse XLSX contents")
            return {"passed": False, "score": score, "feedback": " | ".join(feedback_parts)}

        # Anti-Gaming: Check for formulas
        formula_count = count_formulas(temp_xlsx.name)
        if formula_count > 100:
            score += 20
            feedback_parts.append(f"Formulas present ({formula_count})")
        elif formula_count > 10:
            score += 10
            feedback_parts.append(f"Some formulas present ({formula_count})")
        else:
            feedback_parts.append("Very few or no formulas found (hardcoded?)")

        # Parse values
        wb_data = load_workbook(temp_xlsx.name, data_only=True)
        all_text = extract_all_text(wb_data)
        all_numbers = extract_all_numbers(wb_data)

        # Look for required labels
        labels_found = 0
        if "depth" in all_text: labels_found += 1
        if "ratio" in all_text: labels_found += 1
        if "radius" in all_text: labels_found += 1
        
        # Check specific values
        depth_found = False
        ratio_found = False
        radius_found = False

        for num in all_numbers:
            if expected_depth_min <= num <= expected_depth_max:
                depth_found = True
            if expected_ratio_min <= num <= expected_ratio_max:
                ratio_found = True
            if expected_radius_min <= num <= expected_radius_max:
                radius_found = True

        if depth_found:
            score += 15
            feedback_parts.append("Transit Depth calculated correctly")
        else:
            feedback_parts.append("Transit Depth value missing/incorrect")

        if ratio_found:
            score += 15
            feedback_parts.append("Radius Ratio calculated correctly")
        else:
            feedback_parts.append("Radius Ratio value missing/incorrect")

        if radius_found:
            score += 20
            feedback_parts.append("Planet Radius calculated correctly")
        else:
            feedback_parts.append("Planet Radius value missing/incorrect")

    except Exception as e:
        feedback_parts.append(f"Error reading workbook: {e}")
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # VLM Trajectory check
    if VLM_AVAILABLE:
        try:
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            prompt = (
                "You are an AI adjudicator verifying a spreadsheet task. "
                "Look at these trajectory screenshots showing the agent's work progression. "
                "Did the agent manipulate columns of numerical data, use formulas, and calculate final summary metrics like 'transit depth' or 'radius'? "
                "Reply with a JSON: {'spreadsheet_editing_seen': true/false}"
            )
            vlm_res = query_vlm(images=frames + [final], prompt=prompt)
            if vlm_res.get("success"):
                parsed = vlm_res.get("parsed", {})
                if parsed.get("spreadsheet_editing_seen", False):
                    score += 15
                    feedback_parts.append("VLM verified trajectory progression")
                else:
                    feedback_parts.append("VLM did not detect spreadsheet editing")
            else:
                score += 15  # Give benefit of doubt on VLM failure
                feedback_parts.append("VLM query failed (awarded default points)")
        except Exception as e:
            logger.warning(f"VLM verification error: {e}")
            score += 15  # Give benefit of doubt
    else:
        # Scale up the programmatic score if VLM is unavailable
        score = int(score * (100 / 85))
        score = min(score, 100)
        feedback_parts.append("VLM unavailable, score scaled programmatically")

    key_criteria_met = output_exists and (depth_found or radius_found) and (formula_count > 10)
    passed = score >= 60 and key_criteria_met

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""
Verifier for Walleye Stock Assessment task.

The agent must synthesize raw gill net catch data and effort logs into a
standard fisheries population assessment containing:
1. CPUE analysis (Catch Per Unit Effort)
2. Length-Frequency distribution
3. Proportional Size Distribution (PSD)
4. Relative Weight (Wr) calculations
5. Age-Growth summary
6. Management summary and recommendations

Scoring (10.0 points total, pass threshold 5.0):
- Wrong-target gate: file exists and was created during task
- CPUE calculations (2.0 pts)
- Length-frequency distribution (1.5 pts)
- PSD index calculations (2.0 pts)
- Relative weight (Wr) analysis (1.5 pts)
- Age-growth analysis (1.5 pts)
- Management summary/recommendations (1.5 pts)
"""

import sys
import os
import json
import logging
import tempfile
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Fallback openpyxl check if the environment's `onlyoffice_verification_utils.py` fails
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Verification may fail.")


def extract_all_text_and_numbers(filepath):
    """Extract all text and numeric values from all cells in the workbook."""
    if not OPENPYXL_AVAILABLE:
        return "", []

    all_text = []
    all_numbers = []
    
    try:
        # Load with data_only=True to get formula results, not strings
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), 
                                       max_col=min(sheet.max_column, 50)):
                for cell in row:
                    val = cell.value
                    if val is not None:
                        if isinstance(val, (int, float)):
                            all_numbers.append(val)
                            all_text.append(str(val))
                        else:
                            all_text.append(str(val).lower())
        wb.close()
    except Exception as e:
        logger.error(f"Error parsing workbook: {e}")
    
    return " ".join(all_text), all_numbers


def verify_walleye_assessment(traj, env_info, task_info):
    """
    Verify the walleye stock assessment workbook.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0.0

    # Read export metadata
    temp_meta = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/walleye_assessment_result.json", temp_meta.name)
        with open(temp_meta.name, 'r') as f:
            result_meta = json.load(f)
    except Exception as e:
        logger.error(f"Failed to load export result json: {e}")
        result_meta = {"output_exists": False}
    finally:
        if os.path.exists(temp_meta.name):
            os.unlink(temp_meta.name)

    # 1. GATE: File must exist and be created during task
    if not result_meta.get("output_exists", False):
        return {
            "passed": False,
            "score": 0.0,
            "feedback": "Output workbook vermilion_walleye_assessment.xlsx not found."
        }
        
    if not result_meta.get("file_created_during_task", True):
        feedback_parts.append("WARNING: File timestamps suggest it was not modified during the task window.")

    # Fetch the actual XLSX file
    container_path = "/home/ga/Documents/Spreadsheets/vermilion_walleye_assessment.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(container_path, temp_xlsx.name)
        all_text, all_numbers = extract_all_text_and_numbers(temp_xlsx.name)
    except Exception as e:
        return {"passed": False, "score": 0.0, "feedback": f"Failed to parse XLSX: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    if not all_text and not all_numbers:
        return {"passed": False, "score": 0.0, "feedback": "Workbook exists but appears to be empty or unreadable."}

    # Extract some ground truth info
    total_fish = task_info.get('metadata', {}).get('ground_truth_hints', {}).get('total_fish', 180)
    
    # ===================================================================
    # CRITERION 1: CPUE calculations (2.0 pts)
    # Expected: Mentions of CPUE, fish per net, net-night, and values ~ 7.5
    # ===================================================================
    cpue_terms = ["cpue", "catch per unit effort", "fish/net", "fish per net", "net-night", "net night"]
    has_cpue_terms = any(term in all_text for term in cpue_terms)
    
    # The generated mean CPUE will be ~7.5 (180 fish / 24 nets)
    has_cpue_vals = any(6.0 <= n <= 9.0 for n in all_numbers)
    
    if has_cpue_terms and has_cpue_vals:
        score += 2.0
        feedback_parts.append("CPUE analysis identified")
    elif has_cpue_terms:
        score += 1.0
        feedback_parts.append("CPUE terms found but calculations appear missing or incorrect")
    else:
        feedback_parts.append("CPUE analysis missing")

    # ===================================================================
    # CRITERION 2: Length-frequency distribution (1.5 pts)
    # Expected: Length bins like 200, 225, 250, 275... and frequency counts
    # ===================================================================
    lf_terms = ["length-frequency", "length frequency", "bin", "histogram", "frequency"]
    has_lf_terms = any(term in all_text for term in lf_terms)
    
    # Check for presence of typical 25mm bins
    expected_bins = [200, 225, 250, 275, 300, 325, 350, 400, 450, 500]
    bins_found = sum(1 for b in expected_bins if b in all_numbers)
    
    if has_lf_terms and bins_found >= 4:
        score += 1.5
        feedback_parts.append("Length-frequency distribution identified")
    elif has_lf_terms or bins_found >= 4:
        score += 0.75
        feedback_parts.append("Partial Length-frequency data found")
    else:
        feedback_parts.append("Length-frequency distribution missing")

    # ===================================================================
    # CRITERION 3: PSD index calculations (2.0 pts)
    # Expected: Terms "psd", "psd-p", "stock", "quality", "preferred"
    # Values between 0-100 (typically 40-60 for this population)
    # ===================================================================
    psd_terms = ["psd", "psd-p", "proportional size distribution"]
    cat_terms = ["stock", "quality", "preferred", "memorable", "trophy", "sub-stock"]
    
    has_psd_terms = any(term in all_text for term in psd_terms)
    cats_found = sum(1 for term in cat_terms if term in all_text)
    
    # PSD should be a percentage between roughly 20 and 80
    has_psd_vals = any(20 <= n <= 80 for n in all_numbers)
    
    if has_psd_terms and cats_found >= 3 and has_psd_vals:
        score += 2.0
        feedback_parts.append("PSD and length categorizations identified")
    elif has_psd_terms or cats_found >= 2:
        score += 1.0
        feedback_parts.append("Partial PSD analysis found")
    else:
        feedback_parts.append("PSD calculations missing")

    # ===================================================================
    # CRITERION 4: Relative weight (Wr) analysis (1.5 pts)
    # Expected: Terms "wr", "relative weight", "standard weight", "ws"
    # Mean values around 90-105
    # ===================================================================
    wr_terms = ["wr", "relative weight", "standard weight", "ws", "condition"]
    has_wr_terms = any(term in all_text for term in wr_terms)
    
    has_wr_vals = any(80 <= n <= 110 for n in all_numbers)
    # Check if standard weight coefficients appear in formulas/text
    has_equation = "-5.453" in all_text or "3.18" in all_text
    
    if has_wr_terms and has_wr_vals:
        score += 1.5
        feedback_parts.append("Relative Weight (Wr) analysis identified")
    elif has_wr_terms or has_equation:
        score += 0.5
        feedback_parts.append("Partial Relative Weight analysis found")
    else:
        feedback_parts.append("Relative Weight analysis missing")

    # ===================================================================
    # CRITERION 5: Age-growth analysis (1.5 pts)
    # Expected: Mean length at age, "age", "growth", ages 1-12
    # ===================================================================
    age_terms = ["age", "growth", "length at age", "length-at-age", "mean length"]
    has_age_terms = sum(1 for term in age_terms if term in all_text) >= 2
    
    # Should see age values (1 through 12) coupled with typical lengths (e.g., 200-600)
    has_age_vals = all(a in all_numbers for a in [2, 3, 4, 5])
    
    if has_age_terms and has_age_vals:
        score += 1.5
        feedback_parts.append("Age-Growth analysis identified")
    elif has_age_terms:
        score += 0.5
        feedback_parts.append("Age terms found without complete growth summary")
    else:
        feedback_parts.append("Age-Growth analysis missing")

    # ===================================================================
    # CRITERION 6: Management summary (1.5 pts)
    # Expected: Sentences discussing status, recommendation, management
    # ===================================================================
    mgmt_terms = ["management", "recommendation", "status", "population", "fair", "good", "poor", "harvest", "stocking", "regulate", "limit", "summary", "conclusion"]
    mgmt_term_count = sum(1 for term in mgmt_terms if term in all_text)
    
    if mgmt_term_count >= 4:
        score += 1.5
        feedback_parts.append("Management summary/recommendations identified")
    elif mgmt_term_count >= 2:
        score += 0.5
        feedback_parts.append("Brief management notes found")
    else:
        feedback_parts.append("Management summary missing")

    # Total Score Calculation
    passed = score >= 5.0
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
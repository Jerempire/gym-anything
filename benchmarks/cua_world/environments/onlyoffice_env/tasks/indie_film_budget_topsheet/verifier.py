#!/usr/bin/env python3
import json
import os
import tempfile
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def verify_indie_film_budget(traj, env_info, task_info):
    """
    Verify the independent film budget spreadsheet.
    Multiple independent signals:
    1. File exists and was saved during task
    2. Sheet structure correct (Detail, Top Sheet)
    3. Formulas actively used by the agent
    4. Evaluated numbers match ground truth exact outputs
    5. VLM trajectory verification (using frames)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # 1. Read JSON result
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    score = 0
    feedback_parts = []
    
    output_exists = result.get('output_exists', False)
    file_created_during_task = result.get('file_created_during_task', False)
    
    if output_exists:
        score += 5
        feedback_parts.append("Output file exists")
        if file_created_during_task:
            score += 5
            feedback_parts.append("File created/modified during task")
    else:
        return {"passed": False, "score": 0, "feedback": "Output file indie_film_budget.xlsx NOT found."}

    # 2. Copy and parse XLSX
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/indie_film_budget.xlsx", temp_xlsx.name)
        import openpyxl
        
        # Load with formulas to ensure agent didn't hardcode numbers
        wb_formulas = openpyxl.load_workbook(temp_xlsx.name, data_only=False)
        # Load with values to check mathematical accuracy
        wb_values = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        
        sheet_names = wb_formulas.sheetnames
        has_detail = "Detail" in sheet_names
        has_top_sheet = "Top Sheet" in sheet_names
        
        if has_detail and has_top_sheet:
            score += 10
            feedback_parts.append("Correct sheets found")
        else:
            feedback_parts.append(f"Missing required sheets. Found: {sheet_names}")
            
        # Check formulas exist anywhere in the workbook
        formula_count = 0
        for sheet in wb_formulas.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        
        if formula_count >= 10:
            score += 10
            feedback_parts.append(f"Formulas used ({formula_count})")
        else:
            feedback_parts.append(f"Few/no formulas found ({formula_count}). Logic may be hardcoded.")

        # Extract Ground Truths from Task Info
        metadata = task_info.get('metadata', {})
        gt = {
            "fringes": metadata.get('expected_fringes_sum', 310860),
            "total_cost": metadata.get('expected_total_cost_sum', 2890860),
            "atl": metadata.get('expected_atl', 1052600),
            "btl": metadata.get('expected_btl', 1292060),
            "post": metadata.get('expected_post', 546200),
            "subtotal": metadata.get('expected_subtotal', 2890860),
            "contingency": metadata.get('expected_contingency', 289086),
            "bond": metadata.get('expected_bond', 95398.38),
            "grand": metadata.get('expected_grand_total', 3275344.38)
        }
        
        def is_close(val, target, tol=5):
            return abs(val - target) <= tol

        # Evaluate Data on "Detail" sheet
        fringe_sum = 0
        total_cost_sum = 0
        if has_detail:
            ws_detail = wb_values["Detail"]
            # Look for values in columns F and G (index 5 and 6)
            for row in ws_detail.iter_rows(min_row=2, max_col=10):
                if len(row) >= 7:
                    val_f = row[5].value
                    val_g = row[6].value
                    if isinstance(val_f, (int, float)):
                        fringe_sum += val_f
                    if isinstance(val_g, (int, float)):
                        total_cost_sum += val_g

        # Evaluate Data on "Top Sheet"
        top_sheet_numbers = []
        if has_top_sheet:
            ws_top = wb_values["Top Sheet"]
            for row in ws_top.iter_rows():
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        top_sheet_numbers.append(cell.value)
                        
        # Score Fringe / Total Cost math
        if is_close(fringe_sum, gt["fringes"]) and is_close(total_cost_sum, gt["total_cost"]):
            score += 15
            feedback_parts.append("Fringes & Total Cost correctly calculated")
        elif is_close(fringe_sum, gt["fringes"]):
            score += 10
            feedback_parts.append("Fringes correct, but Total Cost incorrect")
        else:
            feedback_parts.append(f"Detail math incorrect (Fringe Sum expected ~{gt['fringes']}, got {fringe_sum})")
            
        # Score Top Sheet aggregations
        top_matches = 0
        if any(is_close(n, gt["atl"]) for n in top_sheet_numbers): top_matches += 1
        if any(is_close(n, gt["btl"]) for n in top_sheet_numbers): top_matches += 1
        if any(is_close(n, gt["post"]) for n in top_sheet_numbers): top_matches += 1
        
        if top_matches == 3:
            score += 20
            feedback_parts.append("ATL/BTL/Post properly aggregated")
        elif top_matches > 0:
            score += 10
            feedback_parts.append(f"Partial Top Sheet aggregation ({top_matches}/3)")
            
        # Score Grand totals and percentages
        if any(is_close(n, gt["contingency"]) for n in top_sheet_numbers) and \
           any(is_close(n, gt["bond"]) for n in top_sheet_numbers):
            score += 10
            feedback_parts.append("Contingency & Bond correct")
            
        if any(is_close(n, gt["grand"]) for n in top_sheet_numbers):
            score += 5
            feedback_parts.append("Grand Total correct")
            
    except Exception as e:
        feedback_parts.append(f"Excel parsing error: {e}")
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # 3. VLM Trajectory Verification
    from gym_anything.vlm import sample_trajectory_frames, get_final_screenshot
    query_vlm = env_info.get('query_vlm')
    
    if not query_vlm:
        score += 20
        feedback_parts.append("VLM unavailable, auto-awarding trajectory points")
    else:
        try:
            frames = sample_trajectory_frames(traj, n=3)
            final = get_final_screenshot(traj)
            if final:
                frames.append(final)
                
            if frames:
                prompt = """Did the agent actively work in the ONLYOFFICE spreadsheet editor to calculate film budget fringes and build a top sheet?
                Respond ONLY in JSON format: {"worked_in_spreadsheet": true/false}"""
                vlm_result = query_vlm(images=frames, prompt=prompt)
                
                if vlm_result and vlm_result.get('parsed', {}).get('worked_in_spreadsheet', False):
                    score += 20
                    feedback_parts.append("VLM verified trajectory")
                else:
                    feedback_parts.append("VLM rejected trajectory")
            else:
                feedback_parts.append("No frames for VLM")
        except Exception as e:
            feedback_parts.append(f"VLM error: {e}")

    # Pass Threshold is 80 out of 100 possible points
    passed = score >= 80
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
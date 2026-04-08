#!/usr/bin/env python3
"""
Verifier for Blood Bank Inventory Analysis task.

Evaluates:
1. File exists and was created during task.
2. Expiration Date calculation (Collection_Date + 42).
3. Days Remaining calculation (Expiration_Date - '2024-10-24').
4. Expiring Soon flag (Days Remaining <= 5).
5. Aggregation of counts by Blood Group on 'Blood Type Summary' sheet.
"""

import json
import os
import sys
import tempfile
import logging
from datetime import datetime, timedelta

# Import openpyxl safely
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    import openpyxl

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ANCHOR_DATE = datetime(2024, 10, 24)
SHELF_LIFE_DAYS = 42

EXPECTED_BLOOD_GROUPS = ["O+", "O-", "A+", "A-", "B+", "B-", "AB+", "AB-"]

def parse_date(date_val):
    """Attempt to parse various date formats from Excel."""
    if isinstance(date_val, datetime):
        return date_val
    if isinstance(date_val, str):
        try:
            return datetime.strptime(date_val.split()[0], "%Y-%m-%d")
        except ValueError:
            try:
                return datetime.strptime(date_val.split()[0], "%m/%d/%Y")
            except ValueError:
                pass
    return None

def verify_blood_bank_inventory(traj, env_info, task_info):
    """
    Main verification function.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    feedback_parts = []
    score = 0

    # 1. Read export results
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            export_result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read export result: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)

    # 2. Check File Existence & Creation
    file_exists = export_result.get("file_exists", False)
    file_created = export_result.get("file_created_during_task", False)
    
    if not file_exists:
        return {"passed": False, "score": 0, "feedback": "blood_bank_inventory.xlsx was not saved."}
    
    score += 5
    if file_created:
        score += 5
        feedback_parts.append("File properly saved as XLSX.")
    else:
        feedback_parts.append("File saved but timestamps suggest it wasn't created during task.")

    # 3. Read Ground Truth
    temp_gt = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/var/lib/app/ground_truth_blood_bank.json", temp_gt.name)
        with open(temp_gt.name, 'r') as f:
            gt_data = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read ground truth: {e}"}
    finally:
        if os.path.exists(temp_gt.name):
            os.unlink(temp_gt.name)

    expected_counts = gt_data.get("blood_type_counts", {})

    # 4. Load Excel File
    container_xlsx = "/home/ga/Documents/Spreadsheets/blood_bank_inventory.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env(container_xlsx, temp_xlsx.name)
        # Load with data_only=True to read formula results calculated by ONLYOFFICE
        wb = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
    except Exception as e:
        return {"passed": False, "score": score, "feedback": f"Failed to parse Excel file: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    # Find the main data sheet (usually the first one or named 'Sheet1'/'rbc_inventory_raw')
    data_sheet = wb.worksheets[0]
    
    # Locate headers
    headers = {}
    for idx, cell in enumerate(data_sheet[1]):
        if cell.value:
            val = str(cell.value).strip().lower()
            headers[val] = idx

    col_coll_date = None
    for key, idx in headers.items():
        if "collection" in key and "date" in key:
            col_coll_date = idx
            break

    col_exp_date = None
    for key, idx in headers.items():
        if "expiration" in key and "date" in key:
            col_exp_date = idx
            break
            
    col_days_rem = None
    for key, idx in headers.items():
        if "days" in key and "remain" in key:
            col_days_rem = idx
            break

    col_exp_soon = None
    for key, idx in headers.items():
        if "expir" in key and "soon" in key:
            col_exp_soon = idx
            break

    # Analyze data rows
    correct_exp_dates = 0
    correct_days_rem = 0
    correct_exp_soon = 0
    sampled_rows = 0

    if col_coll_date is not None:
        for row in data_sheet.iter_rows(min_row=2, max_row=min(data_sheet.max_row, 200)):
            if not row[col_coll_date].value:
                continue
                
            sampled_rows += 1
            coll_date = parse_date(row[col_coll_date].value)
            
            if not coll_date:
                continue

            expected_exp_date = coll_date + timedelta(days=SHELF_LIFE_DAYS)
            expected_days_rem = (expected_exp_date - ANCHOR_DATE).days
            expected_soon = True if expected_days_rem <= 5 else False

            # Check Expiration Date (+25 points)
            if col_exp_date is not None:
                exp_date_val = parse_date(row[col_exp_date].value)
                if exp_date_val and exp_date_val.date() == expected_exp_date.date():
                    correct_exp_dates += 1

            # Check Days Remaining (+15 points)
            if col_days_rem is not None:
                days_rem_val = row[col_days_rem].value
                try:
                    if int(float(days_rem_val)) == expected_days_rem:
                        correct_days_rem += 1
                except (ValueError, TypeError):
                    pass

            # Check Expiring Soon (+15 points)
            if col_exp_soon is not None:
                soon_val = str(row[col_exp_soon].value).strip().lower()
                is_flagged = soon_val in ['yes', 'true', '1', 'y']
                if is_flagged == expected_soon:
                    correct_exp_soon += 1

    if sampled_rows > 0:
        exp_score = (correct_exp_dates / sampled_rows) * 25
        rem_score = (correct_days_rem / sampled_rows) * 15
        soon_score = (correct_exp_soon / sampled_rows) * 15
        
        score += exp_score
        score += rem_score
        score += soon_score
        
        if exp_score > 20:
            feedback_parts.append("Expiration Dates calculated correctly.")
        elif exp_score > 0:
            feedback_parts.append("Expiration Dates partially correct.")
        else:
            feedback_parts.append("Expiration Dates missing or incorrect.")
            
        if rem_score > 12:
            feedback_parts.append("Days Remaining calculated correctly.")
        
        if soon_score > 12:
            feedback_parts.append("Expiring Soon logic correct.")

    # 5. Check 'Blood Type Summary' Sheet (+35 points: 10 structure, 25 accuracy)
    summary_sheet = None
    for sn in wb.sheetnames:
        if "summary" in sn.lower() or "blood type" in sn.lower():
            summary_sheet = wb[sn]
            break

    if summary_sheet:
        score += 10
        feedback_parts.append("Summary sheet found.")
        
        # Extract all text and adjacent numbers to find aggregations
        found_counts = {}
        for row in summary_sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
            for i, cell in enumerate(row):
                val = str(cell.value).strip().upper() if cell.value else ""
                if val in EXPECTED_BLOOD_GROUPS:
                    # Look for number in next column or next row
                    try:
                        right_val = row[i+1].value if i+1 < len(row) else None
                        if right_val is not None and isinstance(right_val, (int, float)):
                            found_counts[val] = int(right_val)
                            continue
                    except IndexError:
                        pass
                    
                    # Alternative layout: counts below
                    try:
                        down_val = summary_sheet.cell(row=cell.row+1, column=cell.column).value
                        if down_val is not None and isinstance(down_val, (int, float)):
                            found_counts[val] = int(down_val)
                    except Exception:
                        pass

        # Verify counts
        correct_counts = 0
        for bg in EXPECTED_BLOOD_GROUPS:
            if bg in found_counts and found_counts[bg] == expected_counts.get(bg, -1):
                correct_counts += 1
                
        if len(EXPECTED_BLOOD_GROUPS) > 0:
            agg_score = (correct_counts / len(EXPECTED_BLOOD_GROUPS)) * 25
            score += agg_score
            if agg_score > 20:
                feedback_parts.append("Blood Type aggregations are accurate.")
            elif agg_score > 0:
                feedback_parts.append("Blood Type aggregations partially correct.")
            else:
                feedback_parts.append("Blood Type aggregations incorrect or not found.")
    else:
        feedback_parts.append("Summary sheet NOT found.")

    passed = score >= 70

    return {
        "passed": passed,
        "score": int(score),
        "feedback": " | ".join(feedback_parts)
    }
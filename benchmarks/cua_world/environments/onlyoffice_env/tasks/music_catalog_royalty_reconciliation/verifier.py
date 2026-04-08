#!/usr/bin/env python3
"""
Verifier for Music Catalog Royalty Reconciliation task.
"""

import os
import json
import logging
import tempfile
import random

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_ground_truth():
    """
    Recreates the exact same data as setup_task.sh to compute perfect ground truth expectations.
    """
    random.seed(42)
    platforms = ["Spotify", "Apple Music", "Tidal", "Amazon Music", "YouTube"]
    rates = {
        "Tidal": 0.01250,
        "Apple Music": 0.00735,
        "Spotify": 0.00437,
        "Amazon Music": 0.00402,
        "YouTube": 0.00069
    }
    
    adjectives = ["Neon", "Midnight", "Electric", "Retro", "Digital", "Cyber", "Crystal", "Holographic", "Virtual", "Synth"]
    nouns = ["Horizon", "Skyline", "Dreams", "Nights", "Echo", "Heart", "Vibes", "Wave", "City", "Sunset"]
    artists = ["The Midnight Echo", "Synthwave Surfers", "Neon Horizon", "Pixelated Minds", "Aesthetic Vibes", "Cybernetic Youth", "Timecop1983", "FM-84", "Gunship", "The Strike"]

    tracks = []
    for i in range(100):
        title = f"{random.choice(adjectives)} {random.choice(nouns)}"
        if random.random() > 0.8:
            title += " (Remix)"
        artist = random.choice(artists)
        tracks.append((title, artist))

    platform_totals = {p: 0.0 for p in platforms}
    track_totals = {t[0]: 0.0 for t in tracks}
    expected_rows = []

    for title, artist in tracks:
        for p in platforms:
            if p == "Spotify": streams = int(random.lognormvariate(11, 1.5))
            elif p == "YouTube": streams = int(random.lognormvariate(12, 1.8))
            elif p == "Apple Music": streams = int(random.lognormvariate(10.5, 1.2))
            elif p == "Amazon Music": streams = int(random.lognormvariate(9.5, 1.0))
            else: streams = int(random.lognormvariate(8.5, 0.8))
            
            streams = max(150, streams)
            gross = streams * rates[p]
            
            expected_rows.append({
                "platform": p,
                "streams": streams,
                "gross": round(gross, 5),
                "split": round(gross / 2, 5)
            })
            
            platform_totals[p] += gross
            track_totals[title] += gross
            
    top_5_tracks = sorted(track_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    top_5_names = [t[0] for t in top_5_tracks]
    
    return expected_rows, platform_totals, top_5_names

def ensure_openpyxl():
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])

def verify_royalty_reconciliation(traj, env_info, task_info):
    """
    Verifies the royalty reconciliation task.
    
    Checks:
    1. Output file exists and was modified (10 pts)
    2. Multiple sheets exist representing structure (10 pts)
    3. Gross Math calculated correctly across dataset (30 pts)
    4. Splits calculated correctly across dataset (15 pts)
    5. Platform totals correctly aggregated (20 pts)
    6. Top 5 tracks identified (15 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    # Copy and load result.json
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read result metadata: {e}"}
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    output_exists = result.get('output_exists', False)
    file_modified = result.get('file_created_during_task', False)

    if not output_exists:
        return {"passed": False, "score": 0, "feedback": "Failure: royalty_statement_Q3.xlsx was not saved."}

    ensure_openpyxl()
    from openpyxl import load_workbook

    # Copy XLSX
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/royalty_statement_Q3.xlsx", temp_xlsx.name)
        # Load data_only=True to evaluate cached formulas
        wb = load_workbook(temp_xlsx.name, data_only=True)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to parse workbook: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)

    score = 0
    feedback = []

    # 1. File verification (10 pts)
    if file_modified:
        score += 10
        feedback.append("File correctly saved during task.")
    else:
        feedback.append("File exists but was not modified during task timeframe.")

    # 2. Structure (10 pts)
    sheets = [s.lower() for s in wb.sheetnames]
    if len(sheets) >= 2:
        score += 10
        feedback.append("Multiple sheets detected (good structure).")
    else:
        feedback.append("Only one sheet found. Missing Rates or Summary sheets.")

    # Get Ground Truth
    expected_rows, platform_totals, top_5_names = get_ground_truth()
    
    # Extract ALL numeric values to safely check math regardless of layout
    all_numbers = []
    all_text = []
    
    # Also isolate the summary sheet specifically for cleaner searches if possible
    summary_sheet_numbers = []
    summary_sheet_text = ""
    
    for sn in wb.sheetnames:
        sheet = wb[sn]
        is_summary = "summary" in sn.lower()
        sheet_text_parts = []
        
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 20)):
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        all_numbers.append(cell.value)
                        if is_summary:
                            summary_sheet_numbers.append(cell.value)
                    elif isinstance(cell.value, str):
                        all_text.append(cell.value.lower())
                        sheet_text_parts.append(cell.value.lower())
                        
                        # Sometimes currencies are saved as strings
                        # Try to extract numbers from string formatted currencies
                        if '$' in cell.value or '€' in cell.value:
                            try:
                                clean_val = cell.value.replace('$', '').replace(',', '').strip()
                                num = float(clean_val)
                                all_numbers.append(num)
                                if is_summary:
                                    summary_sheet_numbers.append(num)
                            except ValueError:
                                pass
                                
        if is_summary:
            summary_sheet_text = " ".join(sheet_text_parts)

    # 3 & 4. Gross Math and Splits
    gross_matches = 0
    split_matches = 0
    total_expected = len(expected_rows)
    
    # Check what % of expected gross and splits exist ANYWHERE in the workbook
    # Since values are highly unique (float multiplications of large logs), this is extremely accurate
    for exp in expected_rows:
        expected_gross = exp['gross']
        expected_split = exp['split']
        
        if any(abs(n - expected_gross) <= 0.05 for n in all_numbers):
            gross_matches += 1
            
        # Check splits (look for presence, doesn't need to be twice as we just check if it was calculated)
        if any(abs(n - expected_split) <= 0.05 for n in all_numbers):
            split_matches += 1

    gross_percentage = gross_matches / total_expected
    split_percentage = split_matches / total_expected
    
    gross_score = int(gross_percentage * 30)
    split_score = int(split_percentage * 15)
    score += gross_score
    score += split_score
    
    feedback.append(f"Gross Math accuracy: {gross_percentage*100:.1f}% ({gross_score}/30 pts)")
    feedback.append(f"Split Math accuracy: {split_percentage*100:.1f}% ({split_score}/15 pts)")

    # 5. Platform Totals (20 pts)
    # Check if the aggregated sums exist in the summary sheet (or anywhere if summary sheet not found)
    search_nums = summary_sheet_numbers if summary_sheet_numbers else all_numbers
    platforms_found = 0
    
    for platform, expected_total in platform_totals.items():
        if any(abs(n - expected_total) <= 2.0 for n in search_nums):
            platforms_found += 1
            
    platform_score = (platforms_found * 4) # 5 platforms * 4 pts = 20 pts
    score += platform_score
    feedback.append(f"Platform aggregations found: {platforms_found}/5 ({platform_score}/20 pts)")

    # 6. Top 5 Tracks (15 pts)
    # Check if top track names appear in the summary sheet
    search_text = summary_sheet_text if summary_sheet_text else " ".join(all_text)
    tracks_found = 0
    
    for track_name in top_5_names:
        if track_name.lower() in search_text:
            tracks_found += 1
            
    tracks_score = (tracks_found * 3) # 5 tracks * 3 pts = 15 pts
    score += tracks_score
    feedback.append(f"Top earning tracks identified: {tracks_found}/5 ({tracks_score}/15 pts)")

    # Determine passing status
    passed = score >= 65 and gross_percentage > 0.5 and platforms_found >= 1
    
    if passed:
        feedback.insert(0, f"SUCCESS! Final Score: {score}/100.")
    else:
        feedback.insert(0, f"FAILED. Final Score: {score}/100. (Requires 65+ and partial math success)")

    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback)
    }
#!/usr/bin/env python3
"""
Verifier for Wetland Bird Survey Diversity Analysis task.

Evaluates an ecological data analysis workbook created from point-count survey data.
Checks for:
- Calculation of Species Richness, Shannon-Wiener (H'), and Simpson's (D) indices
- Identification of planted trends (Declining: Rusty Blackbird, Black Tern, American Bittern)
- Identification of site gradients (Pristine WM-03 vs Degraded WM-09)
"""

import sys
import os
import json
import logging
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
try:
    from onlyoffice_verification_utils import copy_and_parse_document
except ImportError:
    logging.warning("onlyoffice_verification_utils not found in path, skipping import")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Key species and site flags
DECLINING_SPECIES = ["rusty blackbird", "black tern", "american bittern"]
INCREASING_SPECIES = ["canada goose", "sandhill crane"]
DEGRADED_SITE = "wm-09"
PRISTINE_SITE = "wm-03"

def extract_all_text(wb):
    """Extract all text from all cells in all sheets of a workbook."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    all_text.append(cell.value.lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 1000), max_col=min(sheet.max_column, 50)):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    numbers.append(cell.value)
    return numbers

def count_sheets_with_content(wb):
    """Count sheets that have substantial content."""
    count = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        filled = 0
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 100), max_col=min(sheet.max_column, 20)):
            for cell in row:
                if cell.value is not None:
                    filled += 1
                    if filled > 5:
                        count += 1
                        break
            if filled > 5:
                break
    return count

def verify_wetland_diversity(traj, env_info, task_info):
    """
    Verify wetland bird diversity analysis workbook.

    Scoring (10.0 points total, pass threshold 5.0, returned as normalized 0.0-1.0):
      - CHECK 1: Species Richness by site (1.5 pts)
      - CHECK 2: Shannon-Wiener Index H' (2.0 pts)
      - CHECK 3: Simpson's Diversity Index D (1.5 pts)
      - CHECK 4: Declining species identified (2.0 pts)
      - CHECK 5: Site comparison / gradients identified (1.0 pt)
      - CHECK 6: Trend analysis across years (1.0 pt)
      - CHECK 7: Professional structure (1.0 pt)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    # Validate output file was created during task using JSON result
    temp_json = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/wetland_bird_diversity_analysis_result.json", temp_json.name)
        with open(temp_json.name, 'r') as f:
            result_data = json.load(f)
            
        if not result_data.get('output_file_exists'):
            return {"passed": False, "score": 0.0, "feedback": "File wetland_diversity_analysis.xlsx not found."}
        if not result_data.get('file_created_during_task', True):
            return {"passed": False, "score": 0.0, "feedback": "File existed before task start (Anti-gaming)."}
    except Exception as e:
        logger.warning(f"Failed to read result JSON: {e}")
    finally:
        if os.path.exists(temp_json.name):
            os.unlink(temp_json.name)

    container_path = "/home/ga/Documents/Spreadsheets/wetland_diversity_analysis.xlsx"
    temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(container_path, temp_xlsx.name)
        
        # Parse document
        import openpyxl
        try:
            wb = openpyxl.load_workbook(temp_xlsx.name, data_only=True)
        except Exception as e:
            return {"passed": False, "score": 0.0, "feedback": f"Failed to parse Excel file: {e}"}

        feedback_parts = []
        score_pts = 0.0
        max_pts = 10.0

        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        content_sheets = count_sheets_with_content(wb)

        if len(all_numbers) < 10:
            return {"passed": False, "score": 0.0, "feedback": "File has insufficient numeric content to be a data analysis."}

        # -------------------------------------------------------------------
        # CHECK 1: Species Richness by site (1.5 pts)
        # Expected range: ~8 to ~30
        # -------------------------------------------------------------------
        richness_terms = ["richness", "unique species", "species count", "number of species"]
        has_richness_text = any(t in all_text for t in richness_terms)
        richness_nums = [n for n in all_numbers if isinstance(n, int) and 8 <= n <= 35]
        
        if has_richness_text and len(richness_nums) >= 8:
            score_pts += 1.5
            feedback_parts.append("Species Richness calculated")
        elif has_richness_text or len(richness_nums) >= 8:
            score_pts += 0.75
            feedback_parts.append("Partial Species Richness")
        else:
            feedback_parts.append("Missing Species Richness")

        # -------------------------------------------------------------------
        # CHECK 2: Shannon-Wiener Index H' (2.0 pts)
        # Expected range: ~0.5 to ~3.5
        # -------------------------------------------------------------------
        shannon_terms = ["shannon", "h'", "wiener"]
        has_shannon = any(t in all_text for t in shannon_terms)
        shannon_nums = [n for n in all_numbers if isinstance(n, float) and 0.5 <= n <= 4.0]
        
        if has_shannon and len(shannon_nums) >= 8:
            score_pts += 2.0
            feedback_parts.append("Shannon-Wiener Index calculated")
        elif has_shannon:
            score_pts += 1.0
            feedback_parts.append("Shannon-Wiener mentioned but calculations sparse")
        else:
            feedback_parts.append("Missing Shannon-Wiener Index")

        # -------------------------------------------------------------------
        # CHECK 3: Simpson's Diversity Index D (1.5 pts)
        # Expected range: ~0.2 to ~0.98
        # -------------------------------------------------------------------
        simpson_terms = ["simpson", "d index", "dominance"]
        has_simpson = any(t in all_text for t in simpson_terms)
        simpson_nums = [n for n in all_numbers if isinstance(n, float) and 0.1 <= n < 1.0]
        
        if has_simpson and len(simpson_nums) >= 8:
            score_pts += 1.5
            feedback_parts.append("Simpson's Index calculated")
        elif has_simpson:
            score_pts += 0.75
            feedback_parts.append("Simpson's Index mentioned but calculations sparse")
        else:
            feedback_parts.append("Missing Simpson's Index")

        # -------------------------------------------------------------------
        # CHECK 4: Declining species identified (2.0 pts)
        # Must flag conservation concern species
        # -------------------------------------------------------------------
        declining_found = sum(1 for sp in DECLINING_SPECIES if sp in all_text)
        trend_words = ["decline", "decreasing", "drop", "concern", "trend"]
        has_trend_context = any(tw in all_text for tw in trend_words)
        
        if declining_found >= 2 and has_trend_context:
            score_pts += 2.0
            feedback_parts.append(f"Identified {declining_found}/3 declining species")
        elif declining_found >= 1:
            score_pts += 1.0
            feedback_parts.append(f"Partially identified declining species ({declining_found}/3)")
        else:
            feedback_parts.append("Failed to identify declining species")

        # -------------------------------------------------------------------
        # CHECK 5: Site comparison / gradients (1.0 pt)
        # -------------------------------------------------------------------
        comparison_words = ["pristine", "degraded", "highest", "lowest", "rank", "compare", "poor", "excellent"]
        has_comp_context = any(cw in all_text for cw in comparison_words)
        has_target_sites = DEGRADED_SITE in all_text and PRISTINE_SITE in all_text
        
        if has_target_sites and has_comp_context:
            score_pts += 1.0
            feedback_parts.append("Site comparison/ranking identified")
        elif has_target_sites:
            score_pts += 0.5
            feedback_parts.append("Key sites noted but comparison lacks depth")
        else:
            feedback_parts.append("Site ranking/comparison absent")

        # -------------------------------------------------------------------
        # CHECK 6: Trend analysis across years (1.0 pt)
        # -------------------------------------------------------------------
        years_found = sum(1 for yr in ["2019", "2020", "2021", "2022", "2023"] if yr in all_text)
        increasing_found = sum(1 for sp in INCREASING_SPECIES if sp in all_text)
        
        if years_found >= 4 and increasing_found >= 1:
            score_pts += 1.0
            feedback_parts.append("Multi-year trend analysis present")
        elif years_found >= 4:
            score_pts += 0.5
            feedback_parts.append("Years present but increasing trends missed")

        # -------------------------------------------------------------------
        # CHECK 7: Professional structure (1.0 pt)
        # Multiple substantive sheets
        # -------------------------------------------------------------------
        if content_sheets >= 3:
            score_pts += 1.0
            feedback_parts.append("Professional multi-sheet structure")
        elif content_sheets == 2:
            score_pts += 0.5
            feedback_parts.append("Basic multi-sheet structure")

        # Final score calculation
        passed = score_pts >= 5.0
        normalized_score = score_pts / max_pts

        return {
            "passed": passed,
            "score": normalized_score,
            "feedback": f"Score: {score_pts}/{max_pts} | " + " | ".join(feedback_parts)
        }
        
    except Exception as e:
        logger.error(f"Error during verification: {e}")
        return {"passed": False, "score": 0.0, "feedback": f"Verification error: {e}"}
    finally:
        if os.path.exists(temp_xlsx.name):
            os.unlink(temp_xlsx.name)
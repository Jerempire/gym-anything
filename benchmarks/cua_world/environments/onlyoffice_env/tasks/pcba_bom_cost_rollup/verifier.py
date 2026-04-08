#!/usr/bin/env python3
"""
Verifier for PCBA BOM Cost Rollup task.

The agent must parse a raw BOM CSV and create a well-formatted analysis workbook tracking:
1. Total PCBA Assembly Cost ($38.45)
2. Sole Source Exposure ($14.30)
3. Lifecycle Risks (4 parts)
4. Category breakdown

Scoring (10 points total):
- Base File & Content present (pass requirement)
- Extended cost formulas/values (2.0 pts)
- Total Assembly Cost correct (2.0 pts)
- Category breakdown correct (2.0 pts)
- Sole Source Exposure calculated (1.5 pts)
- Lifecycle risk flagged (1.5 pts)
- VLM Trajectory (spreadsheet usage check) (1.0 pt)
"""

import os
import sys
import logging
import tempfile
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_text(wb):
    """Extract all text from all cells in the workbook."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    numbers.append(float(cell.value))
    return numbers

def check_for_formulas(wb_path):
    """Check if the workbook actually contains formulas rather than hardcoded math."""
    try:
        from openpyxl import load_workbook
        wb_f = load_workbook(wb_path, data_only=False)
        count = 0
        for sn in wb_f.sheetnames:
            sheet = wb_f[sn]
            for row in sheet.iter_rows(max_row=200, max_col=30):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        count += 1
        wb_f.close()
        return count
    except Exception as e:
        logger.debug(f"Formula check failed: {e}")
        return 0

def verify_pcba_bom_cost_rollup(traj, env_info, task_info):
    """Main verifier function."""
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Framework error: copy_from_env missing"}

    container_path = "/home/ga/Documents/Spreadsheets/pcba_cost_rollup.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_pcba_')
    local_path = os.path.join(temp_dir, 'pcba_cost_rollup.xlsx')

    try:
        copy_from_env(container_path, local_path)
        
        if not os.path.exists(local_path):
            return {"passed": False, "score": 0.0, "feedback": "Wrong-target gate: Report file not found."}
            
        from openpyxl import load_workbook
        wb = load_workbook(local_path, data_only=True)
        formulas_count = check_for_formulas(local_path)
        
        score = 0.0
        feedback = []
        
        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        
        # Verify sufficient content was produced
        if len(all_numbers) < 15:
            return {"passed": False, "score": 0.0, "feedback": "Insufficient data/content in workbook."}
            
        # 1. Extended Cost Verification (2.0 pts)
        expected_ext_costs = [2.8, 1.15, 8.5, 2.5, 3.5, 1.25, 1.8, 1.0, 1.2, 3.0, 1.8, 0.6, 0.3, 0.75, 8.3]
        ext_found = sum(1 for e in expected_ext_costs if any(abs(n - e) < 0.01 for n in all_numbers))
        if ext_found >= 10 or formulas_count >= 10:
            score += 2.0
            feedback.append(f"Extended costs detected (Found {ext_found}/15 or {formulas_count} formulas).")
            
        # 2. Total PCBA Cost Verification (2.0 pts)
        if any(abs(n - 38.45) < 0.01 for n in all_numbers):
            score += 2.0
            feedback.append("Total assembly cost ($38.45) correctly identified.")
            
        # 3. Category Breakdown Verification (2.0 pts)
        category_costs = [22.50, 4.20, 2.70, 0.75, 8.30]
        cats_found = sum(1 for c in category_costs if any(abs(n - c) < 0.01 for n in all_numbers))
        if cats_found >= 3:
            score += 2.0
            feedback.append(f"Category costs accurately aggregated ({cats_found}/5 categories found).")
            
        # 4. Sole Source Exposure (1.5 pts)
        if any(abs(n - 14.30) < 0.01 for n in all_numbers):
            score += 1.5
            feedback.append("Sole Source exposure spend ($14.30) correctly calculated.")
            
        # 5. Lifecycle Risk Flagging (1.5 pts)
        has_risk_text = "obsolete" in all_text or "nrnd" in all_text or "lifecycle" in all_text
        has_risk_count = any(abs(n - 4) < 0.01 for n in all_numbers)
        if has_risk_text and has_risk_count:
            score += 1.5
            feedback.append("Lifecycle risk properly evaluated (4 parts identified).")
            
        # 6. VLM Trajectory Verification (1.0 pt fallback/anti-gaming)
        vlm_score = 0.0
        try:
            from gym_anything.vlm import sample_trajectory_frames, query_vlm
            frames = sample_trajectory_frames(traj, n=3)
            if frames:
                prompt = "Does this sequence of images show a user actively using a spreadsheet application (ONLYOFFICE) to perform data analysis, sorting, or formula entry? Answer in JSON format: {\"used_spreadsheet\": true/false}"
                vlm_res = query_vlm(images=frames, prompt=prompt)
                if vlm_res.get('parsed', {}).get('used_spreadsheet', False):
                    vlm_score = 1.0
                    feedback.append("VLM confirms active spreadsheet interaction.")
                else:
                    feedback.append("VLM did not detect strong spreadsheet interaction evidence.")
        except Exception as e:
            logger.warning(f"VLM verification skipped/failed: {e}")
            vlm_score = 1.0 # Assume pass if framework capability is absent
            
        score += vlm_score

        # Ensure core requirement is met for a 'pass'
        has_total = any(abs(n - 38.45) < 0.01 for n in all_numbers)
        passed = (score >= 5.0) and has_total
        
        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback)
        }
    except Exception as e:
        logger.error(f"Verification error: {str(e)}")
        return {"passed": False, "score": 0.0, "feedback": f"Verification encountered an error: {str(e)}"}
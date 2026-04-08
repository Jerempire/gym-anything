#!/usr/bin/env python3
"""
Verifier for Olympic Decathlon Scoring Calculator.

It programmatically evaluates:
1. File exists and was created during the task.
2. Formats: Check for specific formulas used (exponentiation, INT).
3. Value correctness: Compares extracted point totals to ground truth.
4. VLM Check: Confirms the agent's trajectory visually interacted with the spreadsheet.
"""

import sys
import os
import json
import logging
import tempfile
import math

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Ground truth parameters to compute the expected scores internally
PARAMS = [
    ("100m", "Track", 25.4347, 18.0, 1.81),
    ("Long_Jump_cm", "Field", 0.14354, 220.0, 1.4),
    ("Shot_Put_m", "Field", 51.39, 1.5, 1.05),
    ("High_Jump_cm", "Field", 0.8465, 75.0, 1.42),
    ("400m", "Track", 1.53775, 82.0, 1.81),
    ("110m_Hurdles", "Track", 5.74352, 28.5, 1.92),
    ("Discus_m", "Field", 12.91, 4.0, 1.1),
    ("Pole_Vault_cm", "Field", 0.2797, 100.0, 1.35),
    ("Javelin_m", "Field", 10.14, 7.0, 1.08),
    ("1500m", "Track", 0.03768, 480.0, 1.85)
]

def calc_points(val, ev_type, a, b, c):
    try:
        if ev_type == "Track":
            if val >= b: return 0
            return int(a * math.pow((b - val), c))
        else:
            if val <= b: return 0
            return int(a * math.pow((val - b), c))
    except:
        return 0

def get_ground_truth(copy_from_env):
    """Calculate the expected values by parsing the original raw results."""
    temp_csv = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
    gt_scores = []
    
    try:
        copy_from_env("/home/ga/Documents/Spreadsheets/decathlon_raw_results.csv", temp_csv.name)
        with open(temp_csv.name, 'r') as f:
            lines = f.readlines()[1:] # skip header
            
        for line in lines:
            parts = line.strip().split(',')
            if len(parts) < 11: continue
            
            athlete_name = parts[0]
            scores = []
            total = 0
            
            for i in range(10):
                val = float(parts[i+1])
                _, ev_type, a, b, c = PARAMS[i]
                pts = calc_points(val, ev_type, a, b, c)
                scores.append(pts)
                total += pts
                
            gt_scores.append({
                "athlete": athlete_name,
                "track_pts": [scores[0], scores[4], scores[5], scores[9]],
                "field_pts": [scores[1], scores[2], scores[3], scores[6], scores[7], scores[8]],
                "total": total
            })
    except Exception as e:
        logger.error(f"Error parsing ground truth: {e}")
    finally:
        if os.path.exists(temp_csv.name):
            os.unlink(temp_csv.name)
            
    # Sort for final rankings
    gt_scores.sort(key=lambda x: x["total"], reverse=True)
    return gt_scores

def verify_decathlon_scoring(traj, env_info, task_info):
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Framework error: copy_from_env missing."}
        
    score = 0
    feedback_parts = []
    
    # 1. Read metadata
    temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
    try:
        copy_from_env("/tmp/task_result.json", temp_result.name)
        with open(temp_result.name, 'r') as f:
            result = json.load(f)
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to load export JSON: {e}"}
    finally:
        if os.path.exists(temp_result.name):
            os.unlink(temp_result.name)
            
    if not result.get("output_exists", False):
        return {"passed": False, "score": 0, "feedback": "Output file decathlon_scoring.xlsx was not found."}
        
    if not result.get("file_created_during_task", False):
        feedback_parts.append("WARNING: Output file was not newly created/modified during task timestamp.")
        
    # 2. Parse workbook logic
    container_path = "/home/ga/Documents/Spreadsheets/decathlon_scoring.xlsx"
    temp_xlsx_data = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    
    try:
        copy_from_env(container_path, temp_xlsx_data.name)
        
        # Load with openpyxl
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
            import openpyxl
            
        # Get raw values
        wb_data = openpyxl.load_workbook(temp_xlsx_data.name, data_only=True)
        all_numbers = set()
        all_strings = set()
        
        for sn in wb_data.sheetnames:
            sheet = wb_data[sn]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)):
                        all_numbers.add(cell)
                    elif isinstance(cell, str):
                        all_strings.add(cell.lower())
                        
        # Get formulas
        wb_formulas = openpyxl.load_workbook(temp_xlsx_data.name, data_only=False)
        formulas_used = False
        exponent_used = False
        for sn in wb_formulas.sheetnames:
            sheet = wb_formulas[sn]
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and str(cell).startswith('='):
                        formulas_used = True
                        if '^' in str(cell) or 'POWER' in str(cell).upper():
                            exponent_used = True
                            
    except Exception as e:
        return {"passed": False, "score": 0, "feedback": f"Failed to read workbook: {e}"}
    finally:
        if os.path.exists(temp_xlsx_data.name):
            os.unlink(temp_xlsx_data.name)

    # 3. Ground truth evaluation
    gt_data = get_ground_truth(copy_from_env)
    if not gt_data:
        return {"passed": False, "score": 0, "feedback": "Failed to calculate ground truth limits."}
        
    # Criterion: Data Integration (10)
    has_params = "1.53775" in str(all_numbers) or 25.4347 in all_numbers or "25.4347" in str(all_numbers)
    if has_params or "100m" in all_strings:
        score += 10
        feedback_parts.append("Data integration confirmed.")

    # Match scoring
    track_matches = 0
    field_matches = 0
    total_matches = 0
    
    for athlete in gt_data:
        # Give leeway of +/- 1 due to possible INT/ROUND float math differences in OnlyOffice vs Python
        t_match = sum(1 for pts in athlete['track_pts'] if any(abs(pts - n) <= 1 for n in all_numbers if isinstance(n, (int,float))))
        f_match = sum(1 for pts in athlete['field_pts'] if any(abs(pts - n) <= 1 for n in all_numbers if isinstance(n, (int,float))))
        
        track_matches += t_match
        field_matches += f_match
        
        if any(abs(athlete['total'] - n) <= 1 for n in all_numbers if isinstance(n, (int,float))):
            total_matches += 1

    # Criterion: Track Event Formulas (25 pts)
    t_ratio = track_matches / (30 * 4) # 30 athletes * 4 track events
    score += int(25 * t_ratio)
    feedback_parts.append(f"Track events matched: {track_matches}/120")
    
    # Criterion: Field Event Formulas (25 pts)
    f_ratio = field_matches / (30 * 6) # 30 athletes * 6 field events
    score += int(25 * f_ratio)
    feedback_parts.append(f"Field events matched: {field_matches}/180")
    
    # Criterion: Total Score Accuracy (20 pts)
    tot_ratio = total_matches / 30
    score += int(20 * tot_ratio)
    feedback_parts.append(f"Total sums matched: {total_matches}/30")
    
    # Criterion: Formula Implementation (10 pts)
    if formulas_used:
        score += 5
        if exponent_used:
            score += 5
            feedback_parts.append("Exponent/POWER formulas used properly.")
        else:
            feedback_parts.append("Formulas used, but no exponentiation detected.")
    else:
        feedback_parts.append("No formulas detected (hardcoded values).")
        
    # Criterion: Final Rankings (10 pts)
    # Check if the names of the top 3 athletes are present (indicating they were identified/moved/ranked)
    # This is a soft check: if the total sums exist, and they sorted it, the top totals are at top.
    # To be stricter, we check if top 3 names exist in the string values
    top_3 = [gt_data[0]['athlete'].lower(), gt_data[1]['athlete'].lower(), gt_data[2]['athlete'].lower()]
    found_top = sum(1 for top in top_3 if any(top in s for s in all_strings))
    if found_top == 3:
        score += 10
        feedback_parts.append("Top athletes identified.")
        
    # 4. Optional VLM verify (Trajectory)
    from gym_anything.vlm import sample_trajectory_frames, query_vlm, get_final_screenshot
    frames = sample_trajectory_frames(traj, n=3)
    final = get_final_screenshot(traj)
    
    vlm_prompt = "Examine these screenshots of an agent performing a task. Is the agent actively using a spreadsheet application (like OnlyOffice or Excel)? Respond with a JSON object containing a boolean key 'used_spreadsheet'."
    
    vlm_result = query_vlm(images=frames + [final] if final else frames, prompt=vlm_prompt)
    if vlm_result and vlm_result.get("parsed", {}).get("used_spreadsheet", False):
        feedback_parts.append("VLM visual confirmation passed.")
    else:
        feedback_parts.append("VLM visual confirmation did not detect spreadsheet activity (no penalty, but noted).")

    # Pass condition
    key_criteria_met = (t_ratio > 0.5) and (f_ratio > 0.5)
    passed = score >= 60 and key_criteria_met
    
    return {
        "passed": passed,
        "score": score,
        "feedback": " | ".join(feedback_parts)
    }
#!/usr/bin/env python3
"""
Verifier for Maritime Fleet Performance Analysis task.

Evaluates the agent's calculations for IMO EEOI, vessel aggregation, and weather impact analysis.
Uses ONLYOFFICE robust file parsing strategy.
"""

import sys
import os
import logging
import tempfile
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import openpyxl
try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
    from openpyxl import load_workbook

def extract_all_text(wb):
    """Extract all text from all cells in all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                    max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    numbers.append(cell.value)
    return numbers

def count_formulas(wb_path):
    """Counts formulas used in the workbook."""
    try:
        wb = load_workbook(wb_path, data_only=False)
        count = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                        max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if isinstance(cell.value, str) and str(cell.value).startswith('='):
                        count += 1
        return count
    except Exception as e:
        logger.error(f"Error reading formulas: {e}")
        return 0

def verify_fleet_performance(traj, env_info, task_info):
    """
    Verify maritime fleet performance workbook.
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/fleet_performance_report.xlsx"
    temp_dir = tempfile.mkdtemp(prefix='onlyoffice_verify_maritime_')
    local_path = os.path.join(temp_dir, 'fleet_performance_report.xlsx')

    try:
        copy_from_env(container_path, local_path)
        if not os.path.exists(local_path):
            return {
                "passed": False,
                "score": 0,
                "feedback": "Wrong-target gate: fleet_performance_report.xlsx not found."
            }

        wb = load_workbook(local_path, data_only=True)
        feedback_parts = []
        score = 0.0

        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        num_sheets = len(wb.sheetnames)
        num_formulas = count_formulas(local_path)

        # Count total filled cells
        total_cells = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 500),
                                        max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1

        # ===================================================================
        # WRONG-TARGET GATE
        # ===================================================================
        if total_cells < 100:
            return {
                "passed": False,
                "score": 0,
                "feedback": "Wrong-target gate: File has insufficient content (< 100 cells filled)"
            }

        if num_formulas < 5:
            feedback_parts.append("Warning: Very few formulas found.")

        # ===================================================================
        # CHECK 1: EEOI calculations (2.0 pts)
        # ===================================================================
        # Look for "eeoi" term and multiple values in the typical EEOI range (3 - 25)
        has_eeoi_term = "eeoi" in all_text
        eeoi_candidates = [n for n in all_numbers if 3.0 <= n <= 25.0]
        
        if has_eeoi_term and len(eeoi_candidates) >= 10:
            score += 2.0
            feedback_parts.append("EEOI calculated correctly (2.0/2.0)")
        elif has_eeoi_term or len(eeoi_candidates) >= 10:
            score += 1.0
            feedback_parts.append("Partial EEOI calculations detected (1.0/2.0)")
        else:
            feedback_parts.append("Missing EEOI calculations (0.0/2.0)")

        # ===================================================================
        # CHECK 2: Vessel performance aggregation (2.0 pts)
        # ===================================================================
        vessel_terms = ["pacific pioneer", "pacific transporter", "pacific carrier", "average", "worst", "underperform"]
        vessel_mentions = sum(1 for v in vessel_terms if v in all_text)
        
        if vessel_mentions >= 3 and has_eeoi_term:
            score += 2.0
            feedback_parts.append("Vessel performance aggregated (2.0/2.0)")
        elif vessel_mentions >= 1:
            score += 1.0
            feedback_parts.append("Partial vessel aggregation (1.0/2.0)")
        else:
            feedback_parts.append("Missing vessel performance aggregation (0.0/2.0)")

        # ===================================================================
        # CHECK 3: Weather impact analysis (1.5 pts)
        # ===================================================================
        weather_terms = ["calm", "moderate", "rough", "severe", "weather", "beaufort"]
        weather_mentions = sum(1 for w in weather_terms if w in all_text)
        
        if weather_mentions >= 3:
            score += 1.5
            feedback_parts.append("Weather impact analyzed (1.5/1.5)")
        elif weather_mentions >= 1:
            score += 0.5
            feedback_parts.append("Partial weather impact analysis (0.5/1.5)")
        else:
            feedback_parts.append("Missing weather impact analysis (0.0/1.5)")

        # ===================================================================
        # CHECK 4: Fuel efficiency metrics (1.5 pts)
        # ===================================================================
        efficiency_terms = ["mt/nm", "fuel/nm", "efficiency", "tonne-mile", "nm/mt"]
        eff_mentions = sum(1 for e in efficiency_terms if e in all_text)
        
        eff_candidates = [n for n in all_numbers if 0.01 <= n <= 0.5]
        
        if eff_mentions >= 1 and len(eff_candidates) >= 5:
            score += 1.5
            feedback_parts.append("Fuel efficiency metrics calculated (1.5/1.5)")
        elif eff_mentions >= 1 or len(eff_candidates) >= 10:
            score += 0.5
            feedback_parts.append("Partial fuel efficiency metrics (0.5/1.5)")
        else:
            feedback_parts.append("Missing fuel efficiency metrics (0.0/1.5)")

        # ===================================================================
        # CHECK 5: Fleet CO2 emissions (1.0 pt)
        # ===================================================================
        co2_terms = ["co2", "carbon", "emissions"]
        has_co2_term = any(c in all_text for c in co2_terms)
        
        # Expected CO2 range is roughly 100k to 160k
        co2_value_found = any(100000 <= n <= 160000 for n in all_numbers)
        
        if has_co2_term and co2_value_found:
            score += 1.0
            feedback_parts.append("Fleet CO2 emissions calculated (1.0/1.0)")
        elif has_co2_term or co2_value_found:
            score += 0.5
            feedback_parts.append("Partial CO2 emissions analysis (0.5/1.0)")
        else:
            feedback_parts.append("Missing CO2 emissions calculation (0.0/1.0)")

        # ===================================================================
        # CHECK 6: Fleet summary dashboard / Professional structure (2.0 pts)
        # ===================================================================
        dashboard_terms = ["summary", "dashboard", "total", "fleet"]
        has_dashboard = any(d in all_text for d in dashboard_terms)
        
        if num_sheets >= 3 and has_dashboard:
            score += 2.0
            feedback_parts.append("Professional structure with summary sheet (2.0/2.0)")
        elif num_sheets >= 2:
            score += 1.0
            feedback_parts.append("Basic structure with multiple sheets (1.0/2.0)")
        else:
            feedback_parts.append("Single sheet structure (0.0/2.0)")

        # Evaluate final score
        # Using 10.0 scale mapped to percentage format if preferred by framework, or simple boolean threshold
        passed = score >= 5.0
        final_score = int((score / 10.0) * 100)
        
        return {
            "passed": passed,
            "score": final_score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification error: {e}")
        return {
            "passed": False,
            "score": 0,
            "feedback": f"Verification encountered an error: {e}"
        }
    finally:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
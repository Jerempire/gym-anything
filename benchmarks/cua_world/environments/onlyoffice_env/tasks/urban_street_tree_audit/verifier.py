#!/usr/bin/env python3
"""
Verifier for Urban Street Tree Audit task.

Validates that the agent correctly processed the 4,500 tree dataset:
1. Multi-sheet structure.
2. Species Diversity analysis (identifying London planetree @ 990/22% and Honeylocust @ 630/14%).
3. High-risk tree filtering: Exactly 34 trees (DBH >= 24 AND (Poor OR Dead)). Total DBH = 980.
4. Cost formulas multiplying DBH by 150.
5. Grand Total Cost of $147,000.
"""

import sys
import os
import logging
import tempfile
import json

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../', 'utils'))
from onlyoffice_verification_utils import (
    copy_and_parse_document,
    cleanup_temp_dir
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def extract_all_text(wb):
    """Extract all text from all cells across all sheets."""
    all_text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 5000), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if cell.value is not None:
                    all_text.append(str(cell.value).lower())
    return " ".join(all_text)

def extract_all_numbers(wb):
    """Extract all numeric values across all sheets."""
    numbers = []
    for sn in wb.sheetnames:
        sheet = wb[sn]
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 5000), max_col=min(sheet.max_column, 30)):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    numbers.append(cell.value)
    return numbers

def check_for_cost_formula(copy_from_env, container_path):
    """Check for DBH * 150 formula presence by loading workbook without data_only."""
    try:
        from openpyxl import load_workbook as lw
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        copy_from_env(container_path, temp_file.name)
        wb_f = lw(temp_file.name, data_only=False)
        formula_found = False
        
        for sn in wb_f.sheetnames:
            sheet = wb_f[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 5000), max_col=min(sheet.max_column, 30)):
                for cell in row:
                    val = str(cell.value).replace(' ', '').lower()
                    if val.startswith('=') and ('*150' in val or '150*' in val):
                        formula_found = True
                        break
                if formula_found:
                    break
            if formula_found:
                break
                
        wb_f.close()
        os.unlink(temp_file.name)
        return formula_found
    except Exception as e:
        logger.debug(f"Could not check formulas: {e}")
        return False

def verify_tree_audit(traj, env_info, task_info):
    """
    Verify the urban street tree audit workbook.

    Scoring (10 points total, pass threshold 6.0):
    1. Wrong-target gate: file exists with content.
    2. Workbook structure: >= 2 sheets (1.0 pt)
    3. Diversity Flagging: London planetree/Honeylocust counts/pct found (2.0 pts)
    4. High-risk filtering: Evaluated by presence of correct target total 147,000 or DBH sum 980 (3.0 pts)
    5. Cost formula: '=Cell*150' logic used (2.0 pts)
    6. Grand Total Accuracy: Exact 147,000 value calculated (2.0 pts)
    """
    copy_from_env = env_info.get('copy_from_env')
    if not copy_from_env:
        return {"passed": False, "score": 0.0, "feedback": "Copy function not available"}

    container_path = "/home/ga/Documents/Spreadsheets/tree_maintenance_audit.xlsx"
    
    try:
        success, wb, error = copy_and_parse_document(container_path, copy_from_env, 'xlsx')

        if not success:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": f"Wrong-target gate: Failed to load tree_maintenance_audit.xlsx: {error}"
            }

        feedback_parts = []
        score = 0.0

        all_text = extract_all_text(wb)
        all_numbers = extract_all_numbers(wb)
        num_sheets = len(wb.sheetnames)

        # Count total populated cells
        total_cells = 0
        for sn in wb.sheetnames:
            sheet = wb[sn]
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 5000), max_col=min(sheet.max_column, 30)):
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1

        # WRONG-TARGET GATE
        if total_cells < 20:
            return {
                "passed": False,
                "score": 0.0,
                "feedback": "Wrong-target gate: File has insufficient content (< 20 cells)"
            }

        # 1. Workbook Structure (1.0 pt)
        if num_sheets >= 2:
            score += 1.0
            feedback_parts.append(f"Multi-sheet structure confirmed ({num_sheets} sheets)")
        else:
            feedback_parts.append("Single sheet used; missing multi-sheet organization")

        # 2. Diversity Flagging (2.0 pts)
        has_london = "london planetree" in all_text
        has_honey = "honeylocust" in all_text
        
        # Check for counts (990, 630) or percentages (0.22, 22, 0.14, 14)
        has_london_vals = any(val in all_numbers for val in [990, 22, 22.0, 0.22])
        has_honey_vals = any(val in all_numbers for val in [630, 14, 14.0, 0.14])
        
        if has_london and has_honey and (has_london_vals or has_honey_vals):
            score += 2.0
            feedback_parts.append("Species diversity counts and 10% violators successfully identified")
        elif has_london or has_honey:
            score += 1.0
            feedback_parts.append("Partial identification of species diversity (some top species found)")
        else:
            feedback_parts.append("Failed to identify London planetree / Honeylocust violations")

        # 3. High-Risk Filtering & 6. Grand Total Accuracy (Combined logic)
        # 34 specific trees, Total cost = $147,000
        has_total = any(val in all_numbers for val in [147000, 147000.0])
        has_text_total = "147000" in all_text.replace(",", "")
        
        if has_total or has_text_total:
            # Full points for filtering (3.0) and accuracy (2.0)
            score += 5.0
            feedback_parts.append("Correct high-risk filtering and accurate grand total ($147,000) identified")
        else:
            # Partial checks - did they find the 34 trees?
            has_34 = any(val in all_numbers for val in [34, 34.0])
            has_980 = any(val in all_numbers for val in [980, 980.0])
            if has_34 or has_980:
                score += 3.0
                feedback_parts.append("High-risk filtering successful (34 trees / 980 DBH sum), but grand total cost inaccurate")
            else:
                feedback_parts.append("Failed to filter exact high-risk trees (expected 34) or calculate correct $147,000 total")

        # 4. Replacement Cost Formula (2.0 pts)
        has_cost_formula = check_for_cost_formula(copy_from_env, container_path)
        if has_cost_formula:
            score += 2.0
            feedback_parts.append("Dynamic replacement cost formulas (DBH * 150) present")
        else:
            feedback_parts.append("No dynamic replacement cost formulas detected")

        # Final evaluation
        passed = score >= 6.0

        # Anti-gaming file modification check
        temp_result = tempfile.NamedTemporaryFile(delete=False, suffix='.json')
        try:
            copy_from_env("/tmp/urban_street_tree_audit_result.json", temp_result.name)
            with open(temp_result.name, 'r') as f:
                res = json.load(f)
                if not res.get("output_file_exists", False):
                    passed = False
                    score = 0.0
                    feedback_parts = ["Export check failed: file not found at export time"]
        except Exception:
            pass # Non-fatal if missing, we already parsed the file
        finally:
            if os.path.exists(temp_result.name):
                os.unlink(temp_result.name)

        return {
            "passed": passed,
            "score": score,
            "feedback": " | ".join(feedback_parts)
        }

    except Exception as e:
        logger.error(f"Verification failed with error: {e}")
        return {
            "passed": False,
            "score": 0.0,
            "feedback": f"Verification encountered an error: {e}"
        }